# coding=utf-8

import re
import os
import sys
import time
import json
import xlrd
import math
import copy
import chardet
import random
import urllib
import pymysql
import logging
import calendar
import datetime
import requests
import paramiko
import operator
import threading
import traceback
import func_timeout
from jira import JIRA
from func_timeout import func_set_timeout
from dateutil import parser
from openpyxl import load_workbook
reload(sys)
sys.setdefaultencoding("utf8")
from xLibrary.chunyu.persent import query_persent





"""API0000.py文件中全部类:
       API0000_diy()   非系统自定义接口类
       API0000_nb()  新契约接口类
       API0000_pa()  保全接口类
       API0000_payment()   收付费接口类
       API0000_uw()   核保接口类
       API0000_queryPolicy()   保单综合查询接口类
       API0000_clm()   理赔接口类
       API0000_expert_sys()   checkdata-专家系统相关类
       API0000_sys_otherapi()   系统其它接口类 
       API0000_sys_notherapi()   非系统其它接口类 
       API0000_other()   其它类 
"""


"""非系统自定义接口类"""
class API0000_diy():
    # 获取当期服务器时间 并转换为正常格式:年-月-日
    def get_date(self, env_name):
        try:
            #print '****get_date:begin*********************'
            service_time = ""
            env_name = env_name.encode('utf-8')
            #获取excel数据
            if 'win' in sys.platform:
                filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
            else:
                filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx'
            wb = load_workbook(filename)
            sh = wb['ssh']
            # step1:定义服务器ip、服务器用户名、服务器密码
            if env_name == 'uat4':
                rows = '3'
            elif env_name == 'uat6':
                rows = '4'
            elif env_name == 'uat7':
                rows = '5'
            elif env_name == 'uat8':
                rows = '6'
            elif env_name == '预生产':
                rows = '7'
            service_username = sh['B' + rows].value   # 定义service_username服务器用户名
            service_ip = sh['C' + rows].value         # 定义service_ip服务器ip
            service_password = sh['D' + rows].value   # 定义service_password服务器密码
            #print service_ip
                # step2获取当前服务器时间
                # 创建ssh对象
            # 连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=service_ip, port=22, username=service_username, password=service_password, timeout=300,allow_agent=False,look_for_keys=False)
            # 获取服务器时间
            stdin, stdout, stderr = ssh.exec_command('date')   #执行ssh命令
            service_time = stdout.read() or stderr.read()   #读取系统时间
            service_time = service_time.strip()   #去除空白行
            ssh.close()   #关闭服务器
            #print service_time
            time_struct = time.strptime(service_time,'%a %b %d %H:%M:%S CST %Y')  #CST中国标准时间格式转换
            str_time = time.strftime('%Y-%m-%d %H:%M:%S', time_struct)
            #print str_time
        except Exception, e:
            print traceback.format_exc()
        finally:
            pass
            #print '****get_date:end*********************'
        return str_time



    def set_time(self, ip, user_name, password, to_modify_date):
        try:
            #print '****set_date_' + ip + ':begin****'
            #初始化
            service_time_new = ""
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=ip, port=22, username=user_name, password=password, timeout=300,allow_agent=False,look_for_keys=False)
            #修改服务器时间并获取服务器时间
            cmd = 'date -s "%s"' % to_modify_date   #ssh修复服务器时间命令
            stdin, stdout, stderr = ssh.exec_command(cmd)   #执行ssh命令
            service_time = stdout.read() or stderr.read()   #获取修复服务器时间后的系统时间
            service_time = service_time.strip()    #去重空白行
            service_time = time.strptime(service_time,'%a %b %d %H:%M:%S CST %Y')   #CST中国标准时间格式转换
            service_time = time.strftime('%Y-%m-%d', service_time)   #去除空行
            """step2:得到base和bcp全部进程的列表"""
            # 查询base的进程,获取每一行的进程id,存入列表中
        
            stdin, stdout, stderr = ssh.exec_command("docker ps|grep base-batch|awk '{print $1}'")
            result1 = stdout.read()     #获取进程
            list1 = result1.split()   #拆分进程存入列表中
            #print 'list1:'
            #print list1
            #查询bcp的进程,获取每一行的进程id,存入列表中
            stdin, stdout, stderr = ssh.exec_command("docker ps|grep bcp|awk '{print $1}'")
            result2 = stdout.read()     #获取进程
            list2 = result2.split()   #拆分进程存入列表中
            #print 'list2:'
            #print list2
            list1.extend(list2)   #base和bcp进程列表合并
            #print 'list3:'
            #print list1
            """step3:遍历base和bcp的每个进程，重启"""
            for a in list1:
                cmd = 'docker restart ' + a   #ssh命令
                stdin, stdout, stderr = ssh.exec_command(cmd)   #执行命令
                info_cmd = stdout.read() or stderr.read()   #得到执行命令后的结果
                info_cmd.strip()   #去除空行
                #print info_cmd
            """step4:重新获取服务器时间"""
            stdin, stdout, stderr = ssh.exec_command('date')
            service_time_new = stdout.read() or stderr.read()
            service_time_new = service_time_new.strip()
            service_time_new = time.strptime(service_time_new,'%a %b %d %H:%M:%S CST %Y')   #CST中国标准时间格式转换
            service_time_new = time.strftime('%Y-%m-%d %H:%M', service_time_new)   #去除空行
            #print service_time_new
            ssh.close()
        except Exception, e:
            #print '---err----'
            print traceback.format_exc()
        finally:
            pass
            #print '****set_date_' + ip + ':end****'
        return service_time_new


  

    # 修改系统时间-原子级脚本
    def set_time1(self, ip, user_name, password, to_modify_date):
        try:
            #print '****set_date_' + ip + ':begin****'
            #初始化
            service_time_new = ""
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=ip, port=22, username=user_name, password=password, timeout=300,allow_agent=False,look_for_keys=False)
            #修改服务器时间并获取服务器时间
            cmd = 'date -s "%s"' % to_modify_date   #ssh修复服务器时间命令
            stdin, stdout, stderr = ssh.exec_command(cmd)   #执行ssh命令
            service_time = stdout.read() or stderr.read()   #获取修复服务器时间后的系统时间
            service_time = service_time.strip()    #去重空白行
            service_time = time.strptime(service_time,'%a %b %d %H:%M:%S CST %Y')   #CST中国标准时间格式转换
            service_time = time.strftime('%Y-%m-%d', service_time)   #去除空行
            #print '****set_date_' + ip + ':begin****'
            #初始化
            service_time_new = ""
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=ip, port=22, username=user_name, password=password, timeout=300,allow_agent=False,look_for_keys=False)
            #修改服务器时间并获取服务器时间
            cmd = 'date -s "%s"' % to_modify_date   #ssh修复服务器时间命令
            #print cmd
            stdin, stdout, stderr = ssh.exec_command(cmd)   #执行ssh命令
            service_time = stdout.read() or stderr.read()   #获取修复服务器时间后的系统时间
            service_time = service_time.strip()    #去重空白行
            service_time = time.strptime(service_time,'%a %b %d %H:%M:%S CST %Y')   #CST中国标准时间格式转换
            service_time = time.strftime('%Y-%m-%d', service_time)   #去除空行
            """step2:重启服务器"""
            cmd = 'docker ps | egrep "policy|base-batch" | awk "{print $1}" | while read line;do docker restart $line ;done'
            #print cmd
            stdin, stdout, stderr = ssh.exec_command(cmd)
            result = stdout.read()     #获取进程
            #print result
            cmd = 'docker ps | egrep "bcp" | awk "{print $1}" | while read line;do docker restart $line ;done'
            #print cmd
            stdin, stdout, stderr = ssh.exec_command(cmd)
            result = stdout.read()     #获取进程
            #print result
            """step3:重新获取服务器时间"""
            stdin, stdout, stderr = ssh.exec_command('date')
            service_time_new = stdout.read() or stderr.read()
            service_time_new = service_time_new.strip()
            service_time_new = time.strptime(service_time_new,'%a %b %d %H:%M:%S CST %Y')   #CST中国标准时间格式转换
            service_time_new = time.strftime('%Y-%m-%d %H:%M', service_time_new)   #去除空行
            #print service_time_new
            ssh.close()
        except Exception, e:
            #print '---err----'
            print traceback.format_exc()
        finally:
            pass
            #print '****set_date_' + ip + ':end****'
        return service_time_new


    # 修复服务器时间(单线程)
    def modify_date(self, to_modify_date, env_name):  # to_modify_date想要修改的日期，env_name环境名称
        #print '****modify_date:begin*********************'
        env_name = env_name.encode('utf-8')
        #获取excel数据
        if 'win' in sys.platform:
            filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
        else:
            filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx'
        wb = load_workbook(filename)
        sh = wb['ssh']
        # step1:定义服务器ip、服务器用户名、服务器密码
        if env_name == 'uat4':
            rows = '3'
        elif env_name == 'uat6':
            rows = '4'
        elif env_name == 'uat7':
            rows = '5'
        elif env_name == 'uat8':
            rows = '6'
        elif env_name == '预生产':
            rows = '7'
        service_username = sh['B' + rows].value   # 定义service_username服务器用户名
        service_ip1 = sh['C' + rows].value         # 定义service_i1p服务器ip
        service_password1 = sh['D' + rows].value   # 定义service_password1服务器密码

        service_ip2 = sh['E' + rows].value         # 定义service_ip2服务器ip
        service_password2 = sh['F' + rows].value   # 定义service_password2服务器密码

        # setp2修改服务器时间
        service1_date = API0000_diy().set_time(service_ip1, service_username, service_password1, to_modify_date)
        service2_date = API0000_diy().set_time(service_ip2, service_username, service_password2, to_modify_date)
        #print service1_date
        #print type(service1_date)
        #print '****modify_date:end*************************'
        return service1_date


    # 修复服务器时间（多线程）
    def modify_date1(self, to_modify_date, env_name):  # to_modify_date想要修改的日期，env_name环境名称
        #print '****modify_date:begin*********************'
        env_name = env_name.encode('utf-8')
        #获取excel数据
        if 'win' in sys.platform:
            filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
        else:
            filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx'
        wb = load_workbook(filename)
        sh = wb['ssh']
        # step1:定义服务器ip、服务器用户名、服务器密码
        if env_name == 'uat4':
            rows = '3'
        elif env_name == 'uat6':
            rows = '4'
        elif env_name == 'uat7':
            rows = '5'
        elif env_name == 'uat8':
            rows = '6'
        elif env_name == '预生产':
            rows = '7'
        service_username = sh['B' + rows].value   # 定义service_username服务器用户名
        service_ip1 = sh['C' + rows].value         # 定义service_i1p服务器ip
        service_password1 = sh['D' + rows].value   # 定义service_password1服务器密码

        service_ip2 = sh['E' + rows].value         # 定义service_ip2服务器ip
        service_password2 = sh['F' + rows].value   # 定义service_password2服务器密码

        service_ip3 = sh['G' + rows].value         # 定义service_ip3服务器ip
        service_password3 = sh['H' + rows].value   # 定义service_password3服务器密码
        # setp2修改服务器时间
        threads = []
        #"""
        thread1 = myThread(env_name, service_ip1, service_username, service_password1, to_modify_date)    # 创建新线程
        threads.append(thread1)
        thread2 = myThread(env_name, service_ip2, service_username, service_password2, to_modify_date)    # 创建新线程
        threads.append(thread2)

        for t in threads:
            t.setDaemon(True)
            t.start()

        #join()方法，用于等待线程终止
        for t in threads:
             t.join()
        #修复服务器时间后查询当前时间
        service1_date = API0000_diy().get_date(env_name)
        #print service1_date
        #print type(service1_date)
        #print '****modify_date:end*************************'
        return service1_date



    def define_pa_name(self, name):
        if name == '客户基本资料变更':
            pa_name = 'BD'    
        elif name == '客户重要资料变更':
            pa_name = 'CD' 
        elif name == '客户联系方式变更':
            pa_name = 'LD' 
        elif name == '职业类别变更':
            pa_name = 'IO' 
        elif name == '投保人变更':
            pa_name = 'AE'
        elif name == '受益人变更':
            pa_name = '' 
        elif name == '增补健康告知':
            pa_name = 'HI' 
        elif name == '保费自垫状态变更':
            pa_name = 'AG' 
        elif name == '红利领取方式变更':
            pa_name = 'HG'
        elif name == '年金/满期金领取方式变更':
            pa_name = '' 
        elif name == '续期缴费方式及账号变更':
            pa_name = 'renew_change' 
        elif name == '年金领取年龄/频率变更':
            pa_name = '' 
        elif name == '保单补发':
            pa_name = 'reissue'
        elif name == '保单挂失、解挂':
            pa_name = 'GS' 
        elif name == '保单迁移':
            pa_name = 'QY' 
        elif name == '保险合同要约确认':
            pa_name = 'contract_determine' 
        elif name == '保费自垫清偿':
            pa_name = ''
        elif name == '增补健康告知':
            pa_name = '' 
        elif name == '减额交清':
            pa_name = '' 
        elif name == '增加险种':
            pa_name = '' 
        elif name == '取消险种':
            pa_name = ''
        elif name == '减保':
            pa_name = 'reduce_premium' 
        elif name == '加保':
            pa_name = 'increase_premium' 
        elif name == '红利领取':
            pa_name = '' 
        elif name == '部分领取':
            pa_name = ''
        elif name == '保单贷款':
            pa_name = 'loan' 
        elif name == '保单还款':
            pa_name = 'loan_repayment' 
        elif name == '险种转换':
            pa_name = '' 
        elif name == '保单质押贷款登记':
            pa_name = ''
        elif name == '解除保单质押贷款登记':
            pa_name = '' 
        elif name == '保单复效':
            pa_name = 'revival' 
        elif name == '退保':
            pa_name = 'tb' 
        elif name == '犹豫期内退保':
            pa_name = 'hesitation_tb'
        elif name == '年金、满期金给付':
            pa_name = '' 
        elif name == '保全重新支付':
            pa_name = 'repeat_pay' 
        elif name == '保全回退':
            pa_name = '' 
        elif name == '协议退保':
            pa_name = 'agreement_tb'
        elif name == '追加保费':
            pa_name = 'additional_premium' 
        elif name == '协议减保':
            pa_name = '' 
        elif name == '公司解约':
            pa_name = '' 
        elif name == '保全通知书寄送标识修改':
            pa_name = 'notice_send_change' 
        elif name == '顺位投保人变更':
            pa_name = 'sequence_AE' 
        return pa_name



    # 数据库配置脚本。env_name环境名称。db_name数据库库名（非预生产环境库名后追加_uat）
    def db_conf(self, env_name, db_name):
        env_name = env_name.encode('utf-8')
        #获取excel数据
        if 'win' in sys.platform:
            filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
        else:
            filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx'
        wb = load_workbook(filename)
        sh = wb['database']
        # step1:定义数据库ip、数据库用户名、数据库密码
        if env_name == 'uat4':
            rows = '3'
        elif env_name == 'uat6':
            rows = '4'
        elif env_name == 'uat7':
            rows = '5'
        elif env_name == 'uat8':
            rows = '6'
        elif env_name == '预生产':
            rows = '7'
        db_ip = sh['B' + rows].value         # 定义数据库ip
        db_user = sh['C' + rows].value       # 定义数据库用户名
        db_password = sh['D' + rows].value   # 定义数据库密码

        # 定义db_pa数据库库名
        if env_name != '预生产':
            db_name = db_name + '_uat'
        elif env_name == '预生产':
            db_name = db_name

        # 连接数据库
        db = API0000_diy().connect_db(db_ip, db_user, db_password, db_name)
        cursor = db.cursor()
        return cursor



    # 数据库配置脚本。env_name环境名称。db_name预生产环境库名（非预生产环境库名后追加_uat）
    def db_conf_update(self, env_name, db_name):
        env_name = env_name.encode('utf-8')
        #获取excel数据
        if 'win' in sys.platform:
            filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
        else:
            filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx' 
        wb = load_workbook(filename)
        sh = wb['database']
        # step1:定义数据库ip、数据库用户名、数据库密码
        if env_name == 'uat4':
            rows = '3'
        elif env_name == 'uat6':
            rows = '4'
        elif env_name == 'uat7':
            rows = '5'
        elif env_name == 'uat8':
            rows = '6'
        elif env_name == '预生产':
            rows = '7'
        db_ip = sh['B' + rows].value         # 定义数据库ip
        db_user = sh['C' + rows].value       # 定义数据库用户名
        db_password = sh['D' + rows].value   # 定义数据库密码

        # 定义db_pa数据库库名
        if env_name != '预生产':
            db_name = db_name + '_uat'
        elif env_name == '预生产':
            db_name = db_name

        # 连接数据库
        db = API0000_diy().connect_db(db_ip, db_user, db_password, db_name)
        cursor = db.cursor()
        return db, cursor



    # 连接数据库脚本
    def connect_db(self, db_ip, db_user, db_password, db_name):
        db = pymysql.connect(host=db_ip, user=db_user, passwd=db_password, db=db_name, charset='utf8')
        return db



    # 接口ip配置脚本。
    def ip_conf(self, env_name):
        env_name = env_name.encode('utf-8')   #env_name环境名称
        #获取excel数据
        if 'win' in sys.platform:
            filename = r'D:\xLibrary\chunyu\doc\password.xlsx'
        else:
            filename = '/data/xServer/xLibrary/chunyu/doc/password.xlsx'

        wb = load_workbook(filename)
        sh = wb['interface_ip']
        # step1:定义数据库ip、数据库用户名、数据库密码
        if env_name == 'uat4':
            rows = '3'
        elif env_name == 'uat6':
            rows = '4'
        elif env_name == 'uat7':
            rows = '5'
        elif env_name == 'uat8':
            rows = '6'
        elif env_name == '预生产':
            rows = '7'
        interface_ip = sh['B' + rows].value         # 定义数据库ip    
        return interface_ip


    # 发送post请求
    def send_post(self, url, headers, data):
        # 定义响应response
        response = {
            "success": False,
            "message": "参数不完整"
        }
        #print url
        #print data
        if (not url) or (not data):
            response["message"] = "参数不完整"
        else:
            response = requests.post(url, data=data, headers=headers)
            #print response.status_code
            if response.status_code == 200:
                response = response.json()
            elif str(response.status_code) in ('500', '501', '502', '503', '504', '505'):   #返回502重新发送post请求
                for i in range(0, 10):
                    response = requests.post(url, data=data, headers=headers)
                    #print response.status_code
                    if response.status_code == 200:
                        response = response.json()
                        break    
                    else:
                        time.sleep(5)
                        #print response.status_code
            else:
                response["message"] = "response error,status_code=" + str(response.status_code)
        #print response
        return response



    # 发送get请求
    def send_get(self, url, data):
        # 定义响应response
        response = {
            "success": False,
            "message": "参数不完整"
        }
        #print url
        #print data
        if (not url) or (not data):
            response["message"] = "参数不完整"
        else:
            response = requests.get(url=url, params=data)
            #print response.status_code

            if response.status_code == 200:
                response = response.json()
            else:
                response["message"] = "response error,status_code=" + str(response.status_code)
        #print response
        return response


    #写入自定义报告
    def report(self, content):
        #print type(content)
        if type(content) == list:
            content = json.dumps(content)
        #文件路径
        file_path = r'C:\Users\9an02\Desktop\report.txt'
        #1.创建.py文件
        file = open(file_path, 'w')
        #2.写入内容
        file.write(content)
        #3.保存并关闭文件
        file.close()
        os.system(file_path)




    # 随机生成邮箱
    def random_email(self):
        __emailtype = ["@qq.com", "@163.com", "@126.com", "@189.com"]
        randomEmail = random.choice(__emailtype)
        rang = random.randint(4, 12)
        Number = "0123456789qbcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPWRSTUVWXYZ"
        randomNumber = "".join(random.choice(Number) for i in range(rang))
        email = randomNumber + randomEmail
        return email


    # 随机生成16位投保单号:110102月日时分秒
    def random_proposalNo(self):
        now_time = datetime.datetime.now()  # 获取当前时间
        mrhms = now_time.strftime('%m%d%H%M%S')  # 获取当前时间的 月日时分 10位
        proposalNo = '110102' + mrhms  # 定义16位投保单号
        return proposalNo





    # 随机生成期间内日期
    def random_date(self, start_date, end_date):
        #重新调试大小日期
        temp = ''
        if start_date > end_date:
            temp = end_date 
            end_date = start_date
            start_date = temp

        start_date_year = start_date[0:4]
        start_date_month = start_date[5:7]
        start_date_day = start_date[8:10]

        end_date_year = end_date[0:4]
        end_date_month = end_date[5:7]
        end_date_day = end_date[8:10]
        #随机年
        random_year = str(random.randint(int(start_date_year), int(end_date_year)))    
        random_month = str(random.randint(1, 12))    
        #随机月
        if len(str(random_month)) == 1:
            random_month = '0' + str(random_month)
        random_day = str(random.randint(0, 31))    
        #随机日
        if len(str(random_day)) == 1:
            random_day = '0' + str(random_day)
        random_date = random_year + '-' + random_month + '-' + random_day
        if (random_date < start_date) or (random_date > end_date) or API0000_diy().isvaild_date(random_date) == False:
            random_date = API0000_diy().random_date(start_date, end_date)
        return random_date




    #定义系统dict
    def define_dict(self):
        """列表格式定义:
        1.被保人列表格式
        "insuRelation": "",   被保人与投保人关系
        "insuName": "",   被保人姓名
        "insuGender": "",   被保人性别
        "insuBirth": "",   被保人出生日期
        "insuHeight": "",   身高/CM
        "insuWeight": "",   体重/KG
        "insuMarriage": "",   婚姻状况
        "insuCertiType": "",   被保人证件类型
        "insuCertiCode": "",   被保人证件号码
        "insuCertiValidEndDate": "",   被保人证件止期
        "insuWorkCode": "",   职业代码/名称
        "insujobClass": "",   职业等级
        "insuEmail": "",   电子邮箱
        "insuWorkType": "",   工作行业
        "insuCompany": "",   工作单位/学校名称
        "insuMobile": "",   被保人手机号
        "insuNationatiy": "",   国籍/地区
        "insunationnality1": "",   民族
        "officeTel": "",   固定电话
        "insumedicalInsType": "",   医保类型
        "driverLicenseType": "",   驾照类型
        "homeTel": "",   家庭电话
        "incomeSource": "",   收入来源
        "incomeSourceNote": "",   收入来源备注
        "annualIncome": "",   年收入/万元
        "insuresidentType": "",   居民类型
        "taxPayerType": "",   税收居民类型
        "insueducationId": "",   学位
        "insuPostalCode": "",   邮编
        "insuProvinceCode": "",   省
        "insuCityCode": "",  市
        "insuDistrictCode": "",   区
        "insuAddrDetail": "",   详细地址
        2.主险信息列表格式:
        "main_risk_num": "",   第x主险
        "productCode": "",   产品代码
        "chargeMode": "",   缴费频率
        "coverPeriodType": "",   保障期间类型
        "coveragePeriod": "",   保障期间
        "chargePeriodType": "",   缴费期间类型
        "chargePeriod": "",   缴费期间
        "sumAssured": "",   保额
        "stdPremBf": "",   保费
        "Units": "",   份数
        "planFreq": "",   年金领取方式 或 年金/生存金领取方式
        "payOption": "",   年金使用方式 
        "bonusYearOrAge": ""   领取年期或年龄
        3.附加险信息列表格式:
        "belong_mainrisk_productId": "",   所属主险的productId
        "productCode": "",   产品代码
        "chargeMode": "",   缴费频率
        "coverPeriodType": "",   保障期间类型
        "coveragePeriod": "",   保障期间
        "chargePeriodType": "",   缴费期间类型
        "chargePeriod": "",   缴费期间
        "sumAssured": "",   保额
        "stdPremBf": "",   保费
        "Units": "",   份数
        "planFreq": "",   年金领取方式 或 年金/生存金领取方式
        "payOption": "",   年金使用方式 
        "bonusYearOrAge": ""   领取年期或年龄
        4.轨迹信息列表格式:
        "trackType": "",   轨迹类型
        "trackTime": "",   轨迹发生时间
        "trackData": {}   S轨迹数据 
        """
        dict = {
            #保单信息
            "policy_info": {
                "proposalNo": "",   #投保单号
                "policyNo": "",   #保单号
                "applicationDate": "",   #投保日期
                "callBackDate": "",   #回执日期
                "policyType": "",   #保单类型
                "effectiveDate": "",   #生效日期
                "organId": "",   #所属机构
                "policyChannel": "",   #公司渠道
                "suspend":"N",   #保单挂失状态。Y-保单挂失；N-保单为解挂
                "suspend_reason":"",   #保单挂失原因
                "policy_status": "",   #保单状态
                "invalid_reason": "",   #失效原因
                "invalid_date": "",   #最近一次失效日期
                "end_reason": "",   #终止原因
                "end_date": "",   #终止日期
                #投保人信息
                "holder_info":{
                    "sameASInsurd": "",   #是否与被保人为同一人
                    "name": "",   #姓名
                    "gender": "",   #性别
                    "pbHoldBirth": "",   #出生日期
                    "certiType": "",   #证件类型
                    "certiCode": "",   #证件号码
                    "pbCertiValidEndDate": "",   #证件止期
                    "isMarried": "",   #婚姻状况
                    "height": "",   #身高/CM
                    "weight": "",   #体重 /KG
                    "mobilePhone": "",   #移动电话
                    "officeTel": "",   #固定电话
                    "homeTel": "",   #家庭电话
                    "Industry": "",   #工作行业
                    "workCompany": "",   #工作单位/学校名称
                    "Email": "",   #电子邮箱
                    "jobCode": "",   #职业代码/名称
                    "jobClass": "",   #职业等级
                    "driverLicenseType": "",   #驾照类型
                    "Nationality": "",   #国籍/地区
                    "nationnality1": "",   #民族
                    "educationId": "",   #学位
                    "medicalInsType": "",   #医保类型
                    "incomeSource": "",   #收入来源
                    "incomeSourceNote": "",   #收入来源备注
                    "annualIncome": "",   #年收入/万元
                    "familyIncome": "",   #家庭年收入/万元
                    "premBudget": "",   #保费预算/元
                    "residentType": "",   #居民类型
                    "taxPayerType": "",   #税收居民类型
                    "postalCode": "",   #邮编
                    "provinceCode": "",   #省份
                    "cityCode": "",   #市区
                    "districtCode": "",   #地区
                    "addrDetail": "",   #详细地址
                },
                # 被保人信息
                "insured_info": [],
                # 新契约收费信息
                "payment_nb":{
                    "payMode": "",   #缴费方式
                    "payment": ""   #收付费金额
                },
                # 受益人信息                
                "benf_info": [],
                #代理人信息
                "agent_info":{
                    "agent_name": "",   #代理人姓名
                    "agent_gender": "",   #代理人性别
                    "agent_birthday": "",   #代理人出生日期
                    "channelId": "",   #销售组织
                    "salesChannelText": "",   #销售渠道
                    "headPartner": "",   #合作机构
                    "partnerId": ""   #合作网点
                },
                #主险信息
                "main_risk_info": [],
                #附加险信息
                "sub_risk_info": [],
                #累计生息账号
                "cbSbAccount":{"cbSbAccount_list":[], "cashAmount":0},
                #万能结息账户
                "investAmount":{"productCode":"", "investAmount_list":[], "interestCapital":0}
            },
           #轨迹信息
           "track_info": [],
           #时间变化轨迹
           "track_change": [],
           #日志信息
           "logInfo":{
                "msg": "",  #接口异常信息
                "err": "",  #日志报错信息
                "code": "0",  #是否报错
                "job":"",
                "result": ""  #输出信息
            },
           #公共信息
            "public": {
                #"autotest_type": "prduct"
                "execNo": "",   #脚本执行编号
                "magicX":"",
                "is_2risk":"",   #是否双主险
                "env_name": "",   #环境名称
                "interface_ip": "",   #环境ip
                "username_entry": "",   #录入用户姓名
                "username_review": "",   #复核用户姓名
                "filename": "",  #excel文件
                "agent": "",  #代理人
                "partner": "",  #合作网点
                "apply_date": "",  #申请日期
                "is_HolderExemption": "",  #是否投保人豁免
                "total_times_renew": 0   #总续期次数（年交）
            },
            #检查点信息
            "check": {
                "calc_nb": {    #新契约计算
                    'product':[],   #产品信息
                    'apply_date':'',   #投保日期
                    'payment':''   #新契约收费
                },   
                "calc_renew": [],    #续期计算              
                "calc_survivalFee": [],    #生存金派发计算  
                "calc_hesitation_tb": {'apply_date':'','payment':''},    #犹豫期计算
                "calc_tb":[],   #退保计算
                "calc_trial_tb":[],   #退保试算计算
                "calc_loan":[],    #贷款计算
                "calc_CD":[],   #客户重要资料变更计算
                "calc_revival":[],   #保单复效计算
                "calc_continuous_bonus": [],   #持续奖金计算
                "calc_additional_premium": [],   #计算追加保费
                "calc_revival": [],   #计算保单复效
                "calc_reduce_amount":[],    #减保
                "calc_suspend":[],   #保单挂失/解挂
                "calc_suspend_loan":[],   #保单质押贷款登记/解除保单质押贷款登记
                "calc_agreement_tb":[],   #协议退保
                "calc_reissue":[],   #保单补发
                "calc_noticeConfig":[],    #保全通知书寄送标识变更
                'calc_csReversal':[]   #保全回退
            }
        }
        return dict


    #dict存入接口报错信息
    def result(self, dict, is_success, message, unit):
        """dict加入程序阻断或错误信息:
           (1)各接口的返回值is_success为True时:dict['logInfo']['result']存入True；dict['logInfo']['msg']存入None
           (2)各接口的返回值is_success为False时:dict['logInfo']['result']存入False；dict['logInfo']['msg']存入message
           (3)dict['logInfo']['err']存入脚本报错日志信息。不在此方法中实现，由各接口的try/catch捕获存入
        """ 
        if message == None:
            message = ''
        if is_success == True:
            message = ''
            dict['logInfo']['msg'] = message
            dict['logInfo']['code'] = '1'              
        else:
            message = unit + ':' + message
            dict['logInfo']['msg'] = message
            dict['logInfo']['code'] = '0'
        return dict


    """excel数据存储到手工计算dict中"""
    def store_nbdata(self, test_type, applicationDate, dict):
        #函数初始化
        is_success = True
        message = ''
        magicX = ''
        unit = ''
        try:
            """step1:获取excel值存入dict中"""
            unit = '节点:excel数据存储到dict中:'  # 节点
            if 'win' in sys.platform:
                filename = r'D:\xLibrary\chunyu\doc\product_life_cycle\product_life_cycle_data.xlsx'
            else:
                filename = dict['public']['filename']
            wb = load_workbook(filename)
            sh = wb['nb_param']
            # 公共信息
            dict['public']['is_HolderExemption'] = sh['B7'].value  # 是否投保人豁免
            dict['public']['is_2risk'] = sh['B25'].value  #是否双主险
            # 保单信息:
            dict['policy_info']['policyChannel'] = sh['B5'].value  # 公司渠道
            dict['policy_info']['applicationDate'] = applicationDate  # 投保日期
            dict['policy_info']['payment_nb']['payMode'] = sh['B24'].value  # 缴费方式
            dict['policy_info']['policy_status'] = '01'  # 保单状态
            dict['policy_info']['effectiveDate'] = (datetime.datetime.strptime(applicationDate, '%Y-%m-%d') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')   #计算生效日期 = 投保日期 + 1天
            #回执日期
            dict['policy_info']['callBackDate'] = applicationDate
            # 保单信息-投保人信息
            dict['policy_info']['holder_info']['sameASInsurd'] = sh['B8'].value  # 投被保人是否同一人
            dict['policy_info']['holder_info']['name'] = sh['B10'].value  # 投保人姓名
            dict['policy_info']['holder_info']['gender'] = sh['B11'].value  # 投保人性别
            dict['policy_info']['holder_info']['pbHoldBirth'] = sh['B12'].value  # 投保人出生日期
            dict['policy_info']['holder_info']['certiType'] = sh['B13'].value  # 投保人证件类型
            dict['policy_info']['holder_info']['certiCode'] = sh['B14'].value  # 投保人证件号码
            dict['policy_info']['holder_info']['pbCertiValidEndDate'] = sh['B15'].value  # 投保人证件止期
            dict['policy_info']['holder_info']['mobilePhone'] = sh['B16'].value  # 投保人手机号
            # 保单信息-被保人信息
                #非投保人豁免，投被保人不为同一人时:
            if (dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0'):
                insured1 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                insured1['insuName'] = sh['B17'].value  # 被保人姓名
                insured1['insuGender'] = sh['B18'].value  # 被保人性别
                insured1['insuBirth'] = sh['B19'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B20'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B21'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B22'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B23'].value  # 被保人手机号
                insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #非投保人豁免，投被保人为同一人时:
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                insured1 = {}
                insured1['insuRelation'] = '00'  # 被保人与投保人关系
                insured1['insuName'] = sh['B10'].value  # 被保人姓名
                insured1['insuGender'] = sh['B11'].value  # 被保人性别
                insured1['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B16'].value  # 被保人手机号
                insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #投保人豁免时:
            elif dict['public']['is_HolderExemption'] == '1':
                insured1 = {}
                insured2 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                insured1['insuName'] = sh['B17'].value  # 被保人姓名
                insured1['insuGender'] = sh['B18'].value  # 被保人性别
                insured1['insuBirth'] = sh['B19'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B20'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B21'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B22'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B23'].value  # 被保人手机号
                insured1['no'] = '1'  # 第1被保人
                insured2['insuRelation'] = '00'  # 被保人与投保人关系
                insured2['insuName'] = sh['B10'].value  # 被保人姓名
                insured2['insuGender'] = sh['B11'].value  # 被保人性别
                insured2['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured2['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured2['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured2['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured2['insuMobile'] = sh['B16'].value  # 被保人手机号
                insured2['no'] = '2'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                dict['policy_info']['insured_info'].append(insured2)
                
            # 保单信息-主险信息
                #判断是否双主险。得到主险个数
            if sh['B25'].value == '0':
                main_risk_num = 1   #主险个数
            elif sh['B25'].value == '1':
                main_risk_num = 2   #主险个数
            
                #根据主险个数，组装主险信息列表。
            for i in range(0, main_risk_num):
                main_risk_info = {}
                if i == 0:
                    main_risk_column = 2   #excel中主险所在的列
                elif i ==1:
                    main_risk_column = 5   #excel中主险所在的列
                main_risk_info['main_risk_num'] = str(i+1) # 第x主险
                main_risk_info['Main_Rider'] = 'M' # 主附险标识
                main_risk_info['productCode'] = str(sh.cell(row=28, column=main_risk_column).value)   #险种代码
                main_risk_info['liability_state'] = '01'   #险种状态
                main_risk_info['chargeMode'] = sh.cell(row=29, column=main_risk_column).value  # 缴费频率
                main_risk_info['coverPeriodType'] = sh.cell(row=30, column=main_risk_column).value  # 保障期间类型
                main_risk_info['coveragePeriod'] = sh.cell(row=31, column=main_risk_column).value  # 保障期间
                main_risk_info['chargePeriodType'] = sh.cell(row=32, column=main_risk_column).value  # 缴费期间类型
                main_risk_info['chargePeriod'] = sh.cell(row=33, column=main_risk_column).value  # 缴费期间
                main_risk_info['sumAssured'] = sh.cell(row=34, column=main_risk_column).value  # 保额
                main_risk_info['stdPremBf'] = sh.cell(row=35, column=main_risk_column).value  # 保费
                main_risk_info['units'] = sh.cell(row=36, column=main_risk_column).value  # 份数
                main_risk_info['planFreq'] = sh.cell(row=37, column=main_risk_column).value  # 年金领取方式 或 年金/生存金领取方式
                main_risk_info['payOption'] = sh.cell(row=38, column=main_risk_column).value  # 年金使用方式
                main_risk_info['bonusYearOrAge'] = sh.cell(row=39, column=main_risk_column).value  # 领取年期或年龄
                    #计算险种下期应缴日+计算险种剩余续期次数
                if main_risk_info['chargeMode'] == 'S':  #趸交
                    main_risk_info['pay_due_date'] = '9999-12-31'
                    main_risk_info['surplus_times_renew'] = 0   #险种剩余续期次数
                elif main_risk_info['chargePeriodType'] in ('A', 'Y') and main_risk_info['chargeMode'] == 'Y':   #缴费期间类型(按年交/交至某年龄)+缴费频率按年交
                    #获取生效日期前一天
                    effectiveDate1 = API0000_diy().date_add_subtract(dict['policy_info']['effectiveDate'], -1)
                    #获取保单生效日期前一日年份
                    effectiveDate1_year = effectiveDate1[0:4]   
                    #初步组装:下期应缴日
                    pay_due_date = str(int(effectiveDate1_year) + 1) + effectiveDate1[4:10]
                    #最终得到:下期应缴日
                    if API0000_diy().isvaild_date(pay_due_date):
                        #日期有效:下期应缴日 = 下期应缴日字符串
                        pay_due_date = pay_due_date
                    else:
                        #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                        pay_due_date = pay_due_date[0:4] + '-02-28'
                    main_risk_info['pay_due_date'] = pay_due_date
                    main_risk_info['surplus_times_renew'] = int(main_risk_info['chargePeriod']) - 1   #险种剩余续期次数
                elif main_risk_info['chargePeriodType'] in ('A', 'M') and main_risk_info['chargeMode'] == 'M':   #缴费期间类型（按月交/交至某年龄)+缴费频率按月交
                    #获取生效日期前一天
                    effectiveDate1 = API0000_diy().date_add_subtract(dict['policy_info']['effectiveDate'], -1)
                    effectiveDate1_year = effectiveDate1[0:4]    #年份
                    effectiveDate1_month = effectiveDate1[5:7]    #月份
                    effectiveDate1_day = effectiveDate1[8:10]    #日期
                    #初步组装:下期应缴日
                    if int(effectiveDate1_month)+1 < 10:
                        pay_due_date = effectiveDate1_year + '-0' + str(int(effectiveDate1_month)+1) + '-' + effectiveDate1_day
                    elif int(effectiveDate1_month)+1 >= 10 and int(effectiveDate1_month)+1 < 13:
                        pay_due_date = effectiveDate1_year + '-' + str(int(effectiveDate1_month)+1) + '-' + effectiveDate1_day
                    elif int(effectiveDate1_month)+1 == 13:
                        pay_due_date = str(int(effectiveDate1_year)+1) + '-01' + '-' + effectiveDate1_day
                    #最终得到:下期应缴日
                    if API0000_diy().isvaild_date(pay_due_date):
                        #日期有效:下期应缴日 = 下期应缴日字符串
                        pay_due_date = pay_due_date
                    else:
                        #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                        pay_due_date = API0000_diy().date_add_subtract(pay_due_date[0:4] + '-03-01', -1)
                    main_risk_info['pay_due_date'] = pay_due_date
                    main_risk_info['surplus_times_renew'] = int(main_risk_info['chargePeriod']) * 12 - 1   #险种剩余续期次数

                # 存入dict
                dict['policy_info']['main_risk_info'].append(main_risk_info)
                
            # 保单信息-附加险信息
            for i in range(3, int(sh['B26'].value)+3):
                sub_risk_info = {} 
                sub_risk_info['belong_mainrisk_productId'] = sh.cell(row=28, column=2).value # 所属主险的productId
                sub_risk_info['Main_Rider'] = 'R' # 主附险标识
                sub_risk_info['productCode'] = str(sh.cell(row=28, column=i).value)   #险种代码
                sub_risk_info['liability_state'] = '01'   #险种状态
                sub_risk_info['chargeMode'] = sh.cell(row=29, column=i).value  # 缴费频率
                sub_risk_info['coverPeriodType'] = sh.cell(row=30, column=i).value  # 保障期间类型
                sub_risk_info['coveragePeriod'] = sh.cell(row=31, column=i).value  # 保障期间
                sub_risk_info['chargePeriodType'] = sh.cell(row=32, column=i).value  # 缴费期间类型
                sub_risk_info['chargePeriod'] = sh.cell(row=33, column=i).value  # 缴费期间
                sub_risk_info['sumAssured'] = sh.cell(row=34, column=i).value  # 保额
                sub_risk_info['stdPremBf'] = sh.cell(row=35, column=i).value  # 保费
                sub_risk_info['units'] = sh.cell(row=36, column=i).value  # 份数
                sub_risk_info['planFreq'] = sh.cell(row=37, column=i).value  # 年金领取方式 或 年金/生存金领取方式
                sub_risk_info['payOption'] = sh.cell(row=38, column=i).value  # 年金使用方式
                sub_risk_info['bonusYearOrAge'] = sh.cell(row=39, column=i).value  # 领取年期或年龄
                    #计算险种下期应缴日+计算险种剩余续期次数
                if sub_risk_info['chargeMode'] == 'S':  #趸交
                    sub_risk_info['pay_due_date'] = '9999-12-31'
                    sub_risk_info['surplus_times_renew'] = 0   #险种剩余续期次数
                elif sub_risk_info['chargePeriodType'] in ('A', 'Y') and sub_risk_info['chargeMode'] == 'Y':   #缴费期间类型(按年交/交至某年龄)+缴费频率按年交
                    #获取生效日期前一天
                    effectiveDate1 = API0000_diy().date_add_subtract(dict['policy_info']['effectiveDate'], -1)
                    #获取保单生效日期前一日年份
                    effectiveDate1_year = effectiveDate1[0:4]   
                    #初步组装:下期应缴日
                    pay_due_date = str(int(effectiveDate1_year) + 1) + effectiveDate1[4:10]
                    #最终得到:下期应缴日
                    if API0000_diy().isvaild_date(pay_due_date):
                        #日期有效:下期应缴日 = 下期应缴日字符串
                        pay_due_date = pay_due_date
                    else:
                        #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                        pay_due_date = pay_due_date[0:4] + '-02-28'
                    sub_risk_info['pay_due_date'] = pay_due_date
                    sub_risk_info['surplus_times_renew'] = int(sub_risk_info['chargePeriod']) - 1   #险种剩余续期次数
                elif sub_risk_info['chargePeriodType'] in ('A', 'M') and sub_risk_info['chargeMode'] == 'M':   #缴费期间类型（按月交/交至某年龄)+缴费频率按月交
                    #获取生效日期前一天
                    effectiveDate1 = API0000_diy().date_add_subtract(dict['policy_info']['effectiveDate'], -1)
                    effectiveDate1_year = effectiveDate1[0:4]    #年份
                    effectiveDate1_month = effectiveDate1[5:7]    #月份
                    effectiveDate1_day = effectiveDate1[8:10]    #日期
                    #初步组装:下期应缴日
                    if int(effectiveDate1_month)+1 < 10:
                        pay_due_date = effectiveDate1_year + '-0' + str(int(effectiveDate1_month)+1) + '-' + effectiveDate1_day
                    elif int(effectiveDate1_month)+1 >= 10 and int(effectiveDate1_month)+1 < 13:
                        pay_due_date = effectiveDate1_year + '-' + str(int(effectiveDate1_month)+1) + '-' + effectiveDate1_day
                    elif int(effectiveDate1_month)+1 == 13:
                        pay_due_date = str(int(effectiveDate1_year)+1) + '-01' + '-' + effectiveDate1_day
                    #最终得到:下期应缴日
                    if API0000_diy().isvaild_date(pay_due_date):
                        #日期有效:下期应缴日 = 下期应缴日字符串
                        pay_due_date = pay_due_date
                    else:
                        #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                        pay_due_date = API0000_diy().date_add_subtract(pay_due_date[0:4] + '-03-01', -1)
                    sub_risk_info['pay_due_date'] = pay_due_date
                    sub_risk_info['surplus_times_renew'] = int(sub_risk_info['chargePeriod']) * 12 - 1   #险种剩余续期次数
                # 存入dict
                dict['policy_info']['sub_risk_info'].append(sub_risk_info) 
                """step2:自定义其他dict内数据"""
            Holder_isMarried = '20'  # 婚姻状况
            Holder_height = '167'  # 身高/CM
            Holder_weight = '67'  # 体重 /KG
            Holder_officeTel = '010-98343543'  # 固定电话
            Holder_homeTel = '010-98343543'  # 家庭电话
            Holder_industry = '03'  # 工作行业
            Holder_workCompany = '东方小清华'  # 工作单位/学校名称
            Holder_email = API0000_diy().random_email()  # 电子邮箱
            Holder_jobCode = '0001001'  # 职业代码/名称
            Holder_jobClass = '1'  # 职业等级
            Holder_driverLicenseType = '06'  # 驾照类型
            Holder_nationality = 'CHN'  # 国籍/地区
            Holder_nationnality1 = '01'  # 民族
            Holder_educationId = '40'  # 学位
            Holder_medicalInsType = '01'  # 医保类型
            Holder_incomeSource = '01'  # 收入来源
            Holder_incomeSourceNote = 'wu'  # 收入来源备注
            Holder_annualIncome = '100000'  # 年收入/万元（系统自动转换为30万元）
            Holder_familyIncome = '100000'  # 家庭年收入/万元（系统自动转换为30万元）
            Holder_premBudget = '30000'  # 保费预算/元
            Holder_residentType = '1'  # 居民类型
            Holder_taxPayerType = '01'  # 税收居民类型
            Holder_postalCode = '121200'  # 邮编
            Holder_provinceCode = '110000'  # 省份
            Holder_cityCode = '110100'  # 市区
            Holder_districtCode = '110101'  # 地区
            Holder_addrDetail = '建国门东大街22号'  # 详细地址

            insuHeight = '178'  # 身高/CM
            insuWeight = '78'  # 体重/KG
            insuMarriage = '20'  # 婚姻状况
            insuWorkCode = '0001001'  # 职业代码/名称
            insujobClass = '1'  # 职业等级
            insuEmail = API0000_diy().random_email()  # 电子邮箱
            insuWorkType = '02'  # 工作行业
            insuCompany = 'apple'  # 工作单位/学校名称
            insuNationatiy = 'CHN'  # 国籍/地区
            insunationnality1 = '01'  # 民族
            insuofficeTel = '010-93456444'  # 固定电话
            insumedicalInsType = '01'  # 医保类型
            insudriverLicenseType = '06'  # 驾照类型
            insuhomeTel = '010-993433344'  # 家庭电话
            insuincomeSource = '01'  # 收入来源
            insuincomeSourceNote = 'wu'  # 收入来源备注
            insuannualIncome = '5000000'  # 年收入/万元
            insuresidentType = '1'  # 居民类型
            insutaxPayerType = '01'  # 税收居民类型
            insueducationId = '40'  # 学位
            insuPostalCode = '120000'  # 邮编
            insuProvinceCode = '110000'  # 省
            insuCityCode = '110100'  # 市
            insuDistrictCode = '110101'  # 区
            insuAddrDetail = '建国门东大街22号'  # 详细地址

                #投保人信息
            dict['policy_info']['holder_info']['isMarried'] = Holder_isMarried  # 婚姻状况
            dict['policy_info']['holder_info']['height'] = Holder_height  # 身高/CM
            dict['policy_info']['holder_info']['weight'] = Holder_weight  # 体重 /KG
            dict['policy_info']['holder_info']['officeTel'] = Holder_officeTel  # 固定电话
            dict['policy_info']['holder_info']['homeTel'] = Holder_homeTel  # 家庭电话
            dict['policy_info']['holder_info']['industry'] = Holder_industry  # 工作行业
            dict['policy_info']['holder_info']['workCompany'] = Holder_workCompany  # 工作单位/学校名称
            dict['policy_info']['holder_info']['email'] = Holder_email  # 电子邮箱
            dict['policy_info']['holder_info']['jobCode'] = Holder_jobCode  # 职业代码/名称
            dict['policy_info']['holder_info']['jobClass'] = Holder_jobClass  # 职业等级
            dict['policy_info']['holder_info']['driverLicenseType'] = Holder_driverLicenseType  # 驾照类型
            dict['policy_info']['holder_info']['nationality'] = Holder_nationality  # 国籍/地区
            dict['policy_info']['holder_info']['nationnality1'] = Holder_nationnality1  # 民族
            dict['policy_info']['holder_info']['educationId'] = Holder_educationId  # 学位
            dict['policy_info']['holder_info']['medicalInsType'] = Holder_medicalInsType  # 医保类型
            dict['policy_info']['holder_info']['incomeSource'] = Holder_incomeSource  # 收入来源
            dict['policy_info']['holder_info']['incomeSourceNote'] = Holder_incomeSourceNote  # 收入来源备注
            dict['policy_info']['holder_info']['annualIncome'] = Holder_annualIncome  # 年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['familyIncome'] = Holder_familyIncome  # 家庭年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['premBudget'] = Holder_premBudget  # 保费预算/元
            dict['policy_info']['holder_info']['residentType'] = Holder_residentType  # 居民类型
            dict['policy_info']['holder_info']['taxPayerType'] = Holder_taxPayerType  # 税收居民类型
            dict['policy_info']['holder_info']['postalCode'] = Holder_postalCode  # 邮编
            dict['policy_info']['holder_info']['provinceCode'] = Holder_provinceCode  # 省份
            dict['policy_info']['holder_info']['cityCode'] = Holder_cityCode  # 市区
            dict['policy_info']['holder_info']['districtCode'] = Holder_districtCode  # 地区
            dict['policy_info']['holder_info']['addrDetail'] = Holder_addrDetail   # 详细地址
                #被保人信息
                #①非投保人豁免，投被保人不为同一人时
            if dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0':
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = insuofficeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                #②非投保人豁免，投被保人为同一人时:
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                dict['policy_info']['insured_info'][0]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = Holder_officeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
                #③非投保人豁免，投被保人为同一人时:
            elif dict['public']['is_HolderExemption'] == '1':
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = insuofficeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                dict['policy_info']['insured_info'][1]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][1]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][1]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][1]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][1]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][1]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][1]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][1]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][1]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][1]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][1]['insuofficeTel'] = Holder_officeTel   # 固定电话
                dict['policy_info']['insured_info'][1]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][1]['insudriverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][1]['insuhomeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][1]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][1]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][1]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][1]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][1]['insutaxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][1]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][1]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][1]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][1]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][1]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][1]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
        except Exception, e:
            dict['logInfo']['err'] = unit + str(e)
            #print '---err------'
            print traceback.format_exc()
            #print dict
            exit()
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict



    #验证用户是否可正常登录、退出系统
    def verify_userlogin(self, env_name, userName, interface_ip):
        #函数初始化
        is_success = True
        message = ''
        unit = '验证用户是否可正常登录、退出系统:'
        magicX = ''
        #登录系统
        rs = API0000_sys_otherapi().login(env_name,userName,interface_ip)   
        is_success = rs[0]
        if is_success == True:
            message = ''
            magicX = rs[2]
            rs = API0000_sys_otherapi().logout(magicX,interface_ip)
            if rs[0] == True:
                message = ''
            else:
                is_success = False
                message = '验证用户是否可成功登录退出系统，登录成功、退出失败'
        else:
            message = rs[1] 
        #print 'message:' + message
        return is_success, message, unit


    #保全通用规则
    def pa_general_rules(self, pa_name, apply_date, dict):
        #初始化
        is_success = True
        message = ''
        #获取数据
        effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
        callBackDate = dict['policy_info']['callBackDate']   #回执日期
        policy_status = dict['policy_info']['policy_status']   #保单状态;01有效；02中止；03终止
        end_reason = dict['policy_info']['end_reason']   #终止原因
        insuBirth = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人生日
        pbCertiValidEndDate = dict['policy_info']['holder_info']['pbCertiValidEndDate']   #投保人证件有效期
        insuCertiValidEndDate_list = []#被保人证件有效期列表
        for i in range(0, len(dict['policy_info']['insured_info'])):
            insuCertiValidEndDate_list.append(dict['policy_info']['insured_info'][i]['insuCertiValidEndDate'])
        #是否欠款标识
        if len(dict['check']['calc_loan']) == 0:
            loan_flag = 'N'
        else:
            if dict['check']['calc_loan'][-1]['loan_type'] == 'loan':
                loan_flag = 'Y'
            else:
                loan_flag = 'N'
        #规则1:当保单处在回执未回销时，不允许操作任何保全项目（申请来源为“内部转办”的，不校验此规则）
        #规则2:当保单续期转账在途时，不允许操作客户重要信息变更、投保人变更、减保、加保、减额交清、犹豫期退保、退保、取消险种、协议退保、公司解约、续期交费方式及账号变更
        #规则3:保单其他保全操作中时，仅可以操作保全重新支付
        #规则4:保单有理赔操作中时，不允许操作任何保全项目
        #规则5:保单处于终止状态时，仅可以操作保全重新支付
        #规则8:保单处于终止状态且终止原因为“保单失效满2年”时，只允许操作退保、协议退保、客户基本信息变更、联系方式变更、保全重新支付
        if policy_status == '03' and end_reason != '11' and pa_name != 'repeat_pay':
            is_success = False
            message = '保单状态非有效'
            return is_success, message
        elif policy_status == '03' and end_reason =='11' and pa_name not in ('tb', 'agreement_tb', 'BD', 'LD', 'repeat_pay'):
            is_success = False
            message = '保单状态非有效'
            return is_success, message

        #规则6: 保单处在挂失状态时，只允许操作挂失解除及保单补发
        #规则7:当保单处在满期终止状态时，只允许操作生存给付柜面领取、红利领取、保全重新支付
        #规则9:当保单处于失效状态时，只允许操作客户基本信息变更、联系方式变更、续期缴费方式及账号变更、保险合同要约确认、退保、减保、协议退保、协议减保、保单复效、保单挂失与解除、保单补发、保全重新支付
        if policy_status == '02' and pa_name not in ('BD', 'LD', 'renew_change', 'contract_determine', 'tb', 'reduce_premium', 'increase_premium', 'agreement_tb', 'revival', 'GS', 'reissue', 'repeat_pay'):
            is_success = False
            message = '保单状态非有效'
            return is_success, message
        #规则10:当保单有欠款未归还时，只允许操作客户基本信息变更、联系方式变更、受益人变更、续期缴费方式及账号变更、保单复效、保险合同要约确认、红利领取方式变更、补充告知、退保、协议退保、协议减保、保单挂失与解除、保单补发、保单贷款还款、保全重新支付、顺位投保人变更、保全通知书寄送标识修改
        if loan_flag == 'Y' and pa_name not in ('BD', 'LD', 'BE', 'renew_change', 'revival', 'contract_determine', 'HG', 'bonus_type_change', 'HI', 'tb', 'trial_tb', 'agreement_tb', 'agreement_rp', 'GS', 'reissue', 'loan', 'loan_repayment', 'repeat_pay', 'sequence_AE', 'notice_send_change'):
            is_success = False
            message = '当保单有欠款未归还时，只允许操作客户基本信息变更、联系方式变更、受益人变更、续期缴费方式及账号变更、保单复效、保险合同要约确认、红利领取方式变更、补充告知、退保、协议退保、协议减保、保单挂失与解除、保单补发、保单贷款还款、保全重新支付、顺位投保人变更、保全通知书寄送标识修改'
            return is_success, message
        #规则11:当保单处于质押贷款登记状态时，只允许客户基本信息变更、联系方式变更、续期缴费方式及账号变更、解除保单质押贷款登记、复效
        #规则12:对于所有证件类型的投保人及被保人，保全（客户基本信息变更、联系方式变更除外）受理时增加证件有效期过期的校验（即系统当前时间>=证件有效期止期）
        if apply_date > pbCertiValidEndDate:
            is_success = False
            message = '客户证件有效期无效'
            return is_success, message
        for i in range(0, len(insuCertiValidEndDate_list)):
            if apply_date > insuCertiValidEndDate_list[i]:
                is_success = False
                message = '客户证件有效期无效'
                return is_success, message     
        #线下保全是提示性校验；
        #规则13:线上保全（银保通、君易保APP、微信）对投保人为阻断性校验，对被保人不校验；
        #规则14:线上保全（官网）不进行以上校验；
        #规则15:证件有效期校验在线上保全仅在申请时进行校验，试算时无需进行校验
        #规则16:保单申请减额交清后，只允许客户基本信息变更、联系方式变更、受益人变更、保险合同要约确认、退保、协议退保、保单挂失与解除、保单补发、保全重新支付、生存给付计划变更、生存给付领取及授权、投保人变更、公司解约、保单贷款、保单贷款还款
        #规则17:保全申请（申请提交）时，如为期交保单且续期保费尚未缴纳，则需判断，提交当时时间距宽限期结束是否≤10个自然日，若是则给予提示
        #规则18:对收付费保全，校验若申请资格人（<生存给付领取及授权>代领为投保人、非代领为被保人，其他保全均为投保人）为“中/较高/最高风险”等级客户，则申请来源须为“内部转办”或申请方式为“客户亲办”，否则给予提示
        return is_success, message



    #检验生存金派发
    #def check_send_survivalFee(self, productCode, apply_date, effectiveDate):




    #根据职业编码判断职业等级
    def get_jobClass(self, jobCode):
        """1.定义职业编码库"""
            #职业等级:1级
        jobCode_1_Class = ['0000001','0000003','0001001','0001003','0001004','0001005','0002001','0002002','0101001','0102001','0201001','0201008','0302019','0303006','0401013','0402001','0402002','0501001','0501003','0501014','0501029','0502001','0502002','0502003','0502004','0502005','0502006','0502015','0502016','0504001','0504002','0504003','0504004','0504005','0504006','0504019','0504020','0504021','0504023','0504030','0504034','0601001','0602001','0602002','0602006','0602007','0603001','0603002','0603003','0603004','0603012','0701001','0701002','0701003','0705001','0706018','0706025','0803046','0803051','0803055','0807034','0810002','0810005','0810006','0810007','0810025','0810045','0813005','0815008','0815042','0818001','0818004','0819001','0901001','0901011','0901012','0901013','0901014','0901015','0901016','0901017','0901018','0901019','0901021','0902001','0902003','1','1001001','1001002','1001015','1002001','1002002','1101001','1101002','1101003','1101007','1101016','1101020','1101021','1101022','1101024','1101026','1101034','1101035','1102005','1103001','1103002','1105001','1105002','1105004','1106001','1106002','1106003','1107001','1107002','1108001','1108002','1108003','1108004','1108009','1108010','1108012','1110002','1201001','1201002','1202001','1202002','1202005','1202006','1202008','1300001','1300002','1300003','1300004','1301001','1301002','1301003','1303003','1401001','1401006','1402001','1405001','1405003','1405004','1501001','1501002','1501003','1501004','1501005','1501006','1501008','1501016','1501017','1501018','1501019','1501022','1501023','1501024','1501025','1501026','1501030','1501031','1501034','1501035','1501036','1501045','1501046','1501047','1501050','1601001','1601007','1601008','1601009','1601011','1602001','1602002','1602003','1602004','1603003','1603004','1603005','1603008','1603012','1603021','1603022','1603025','1701001','1701002','1701003','1801001','1900003','1900006','1901001','1901005','1901006','2','2000002','2000003','2000005','2000006','2001001','2001002','2001003','2101001','2101002','2101003','2147001','2147003','9999999','A0000','B0000']
            #职业等级:2级
        jobCode_2_Class = ['0000004','0000009','0001002','0001006','0101002','0101003','0101004','0101005','0101006','0101008','0101010','0101011','0101012','0101014','0101015','0101016','0101021','0102003','0102004','0102005','0201005','0201012','0302020','0303005','0303007','0401012','0401018','0401019','0401020','0401024','0401025','0401027','0402003','0405001','0501002','0501004','0501008','0501009','0501010','0501015','0501017','0501022','0501027','0501030','0502007','0502008','0502009','0502010','0502017','0502021','0503025','0503028','0503034','0504007','0504008','0504010','0504011','0504015','0504022','0504024','0504025','0601002','0601003','0602003','0602004','0603005','0603007','0603008','0603009','0603011','0701007','0701008','0701010','0701016','0701022','0702003','0704003','0705002','0705005','0706002','0706011','0706019','0706021','0706024','0706026','0706027','0801001','0801002','0802003','0802019','0803001','0803002','0803003','0803004','0803007','0803009','0803018','0803019','0803023','0803029','0803031','0803032','0803042','0803044','0803045','0803047','0803048','0803049','0803050','0803056','0804001','0804002','0804003','0804004','0804025','0805001','0805002','0805003','0805004','0805009','0806001','0806002','0806014','0806021','0807001','0807002','0807006','0807032','0807033','0807069','0807070','0807074','0807082','0809001','0809002','0809007','0810001','0810003','0810004','0810008','0810009','0810010','0810011','0810012','0810013','0810014','0810016','0810017','0810018','0810019','0810020','0810021','0810022','0810023','0810024','0810026','0810027','0810028','0810029','0810030','0810031','0810032','0810049','0810057','0811001','0812002','0812003','0813001','0813007','0814001','0815002','0815007','0815010','0815011','0815016','0815024','0815036','0815037','0815039','0815041','0816001','0816035','0816037','0816038','0816039','0816040','0816042','0816043','0816049','0816050','0816051','0817005','0817006','0817007','0818002','0818003','0818005','0818006','0818007','0818008','0818011','0818013','0819002','0819003','0819004','0819005','0819006','0819008','0819010','0819011','0819014','0819016','0819017','0820001','0820002','0820003','0820004','0820005','0820006','0820008','0820009','0820010','0820011','0820012','0820015','0820017','0820018','0820019','0820020','0820021','0820022','0820023','0820024','0820025','0821001','0821002','0821003','0821004','0821005','0821006','0821007','0821009','0821010','0821011','0822011','0901002','0901003','0901005','0901006','0901007','0901008','0901009','0901020','0901022','0901023','0901024','0902002','0902004','1001004','1001005','1001006','1001007','1001008','1001009','1001011','1001012','1001013','1001014','1001016','1001017','1002005','1002006','1101004','1101008','1101009','1101010','1101011','1101012','1101013','1101017','1101018','1101023','1101027','1101028','1101029','1101030','1101033','1102001','1102002','1102003','1102004','1103004','1104001','1104002','1105003','1107003','1107004','1107009','1108005','1108006','1108007','1108013','1108014','1109001','1109002','1109003','1109004','1110001','1110003','1201003','1201004','1201005','1201008','1202003','1202004','1401002','1401003','1402002','1402010','1402012','1402021','1402054','1403001','1403005','1403012','1403014','1403015','1403016','1404001','1404003','1404004','1404006','1405002','1405005','1405006','1501007','1501009','1501010','1501011','1501012','1501013','1501014','1501015','1501020','1501029','1501033','1501039','1501040','1501048','1501051','1501053','1501054','1501057','1501059','1601002','1601003','1601004','1601005','1601006','1601010','1602005','1603001','1603002','1603006','1603007','1603009','1603011','1603015','1603017','1603018','1603019','1603023','1603026','1700001','1700002','1800001','1800019','1801002','1801003','1801004','1801009','1801010','1801011','1901002','2000001','2000004','2102001','2102002','2103001','2103002','2104001','2104002','2105001','2105002','2106001','2106002','2107001','2107002','2108001','2108002','2108003','2108004','2109001','2109002','2110001','2110002','2111001','2111002','2112001','2113001','2114001','2115001','2116001','2117001','2120001','2122001','2123001','2126001','2127001','2129001','2130001','2147002','2148002','2148012','2148014','2148017','3','4','5','C0000']
            #职业等级:3级
        jobCode_3_Class = ['0000005','0001007','0101007','0101009','0101013','0101017','0102002','0102006','0201002','0201003','0201007','0201010','0201011','0201013','0302002','0302003','0302004','0302005','0302006','0302011','0302015','0302021','0303001','0401011','0401016','0401017','0401026','0401030','0401031','0401032','0401033','0401034','0401035','0401036','0401037','0402007','0404003','0405002','0405011','0405012','0405014','0405015','0405016','0405017','0501005','0501006','0501007','0501011','0501016','0501023','0501024','0501026','0502011','0502012','0502018','0502022','0502024','0502027','0502028','0503029','0503038','0503039','0504009','0504012','0504016','0504026','0504031','0504033','0602005','0603010','0701004','0701005','0701006','0701011','0701023','0701024','0701045','0702001','0702002','0703001','0703009','0704004','0704010','0704011','0704012','0705007','0706020','0706022','0706023','0801003','0801005','0801020','0801048','0802001','0802040','0802041','0803005','0803006','0803008','0803013','0803015','0803016','0803017','0803024','0803025','0803026','0803027','0803028','0803033','0803034','0803052','0803053','0804009','0804019','0804024','0804026','0804027','0804028','0804029','0805005','0805006','0805008','0806003','0806010','0806012','0806013','0806016','0806017','0806022','0806023','0807003','0807005','0807017','0807018','0807019','0807021','0807022','0807023','0807024','0807025','0807026','0807027','0807028','0807029','0807031','0807035','0807036','0807037','0807052','0807064','0807065','0807066','0807067','0807068','0807071','0807072','0807073','0807077','0807078','0807079','0807085','0807086','0807087','0807088','0809004','0809006','0810015','0810033','0810035','0810042','0810044','0810046','0810047','0810048','0810050','0810052','0810054','0810055','0810056','0811002','0811008','0812001','0813002','0813003','0813004','0813006','0814003','0815001','0815003','0815004','0815005','0815006','0815009','0815012','0815013','0815014','0815015','0815017','0815018','0815019','0815020','0815021','0815022','0815023','0815025','0815026','0815027','0815028','0815029','0815030','0815031','0815032','0815033','0815034','0815035','0815038','0815040','0816003','0816004','0816015','0816016','0816023','0816028','0816029','0816030','0816032','0816033','0816044','0816045','0816046','0816047','0816048','0817001','0817011','0817012','0817013','0817014','0817015','0817016','0817017','0817019','0817020','0818009','0818010','0818012','0818014','0818015','0818016','0819007','0819009','0819012','0819013','0819015','0820007','0820014','0820016','0821008','0821012','0821013','0821014','0821015','0821016','0821017','0821018','0821019','0822004','0822005','0822008','0822009','0901010','1001003','1001010','1002003','1002007','1101019','1101025','1101032','1103003','1108008','1108011','1109008','1109009','1109010','1109011','1109012','1109014','1109015','1110004','1201007','1202007','1401004','1401007','1402003','1402009','1402020','1402024','1402025','1402026','1402027','1402028','1402029','1402045','1402052','1402053','1402056','1403002','1403004','1403006','1403011','1501021','1501032','1501037','1501038','1501043','1501044','1501049','1501052','1501055','1601012','1603013','1603024','1800003','1800011','1800012','1800013','1800017','1800018','1800020','1801006','1900001','1900005','1900007','2112002','2113002','2114002','2115002','2116002','2117002','2118001','2118002','2120002','2121001','2121002','2122002','2123002','2128001','2130002','2131001','2132001','2133001','2134001','2135001','2136001','2137001','2148009','2148010','2148016','99','D0000']
            #职业等级:4级
        jobCode_4_Class = ['0000006','0101018','0101019','0101023','0101024','0201004','0201006','0302001','0302008','0302009','0302012','0302013','0302014','0302016','0302017','0302018','0303002','0303003','0401008','0401010','0401021','0401022','0401023','0402004','0402006','0404002','0404004','0501028','0502013','0502014','0502019','0502020','0502025','0502026','0503001','0503002','0503024','0503026','0503027','0503031','0503032','0504013','0504014','0504017','0504018','0504027','0504028','0603006','0701009','0701013','0701014','0701015','0701017','0701027','0701028','0701039','0701040','0701042','0701043','0701046','0701047','0702004','0702005','0702007','0702008','0702010','0702015','0703002','0703006','0703008','0704005','0704006','0704007','0704008','0705003','0705006','0706003','0706005','0706012','0706014','0706016','0801004','0801006','0801007','0801008','0801009','0801022','0801023','0801024','0801025','0801026','0801028','0801029','0801031','0801032','0801033','0801034','0801035','0801036','0801037','0801040','0801047','0801051','0801053','0801054','0801055','0802002','0802005','0802011','0802012','0802015','0802031','0802036','0802037','0802038','0802039','0803010','0803011','0803012','0803014','0803020','0803021','0803022','0803030','0803035','0803036','0803037','0803038','0803039','0803040','0803041','0803043','0803054','0804007','0804010','0804011','0804013','0804014','0804015','0804017','0804018','0804020','0804021','0804022','0804023','0805007','0805010','0805011','0805012','0805013','0806004','0806005','0806018','0806019','0806020','0806024','0806025','0806026','0807011','0807014','0807016','0807020','0807039','0807040','0807044','0807045','0807046','0807047','0807050','0807051','0807053','0807055','0807056','0807059','0807061','0807062','0807063','0807075','0807076','0807080','0807081','0807083','0807084','0807089','0807090','0807091','0807092','0807093','0807094','0807095','0807096','0807097','0807098','0807099','0807100','0807101','0807102','0807103','0807104','0807106','0807108','0807109','0807111','0807113','0809003','0809005','0809008','0810034','0810036','0810037','0810038','0810039','0810040','0810041','0810043','0810051','0810053','0811003','0811005','0811009','0811010','0812004','0812005','0812006','0812007','0812008','0812009','0812010','0814002','0816002','0816012','0816013','0816014','0816021','0816022','0816024','0816025','0816026','0816027','0816031','0816034','0816036','0816052','0816053','0816058','0816059','0816060','0817002','0817003','0817004','0817008','0817009','0817010','0817018','0820013','0822001','0822002','0822003','0822006','1002004','1101014','1101015','1101031','1105005','1107005','1107006','1109005','1109006','1109007','1109013','1109016','1201006','1401005','1401009','1402004','1402011','1402013','1402014','1402015','1402016','1402017','1402018','1402019','1402022','1402023','1402032','1402033','1402034','1402035','1402036','1402041','1402042','1402043','1402044','1402046','1402050','1402051','1402055','1403003','1403008','1403009','1403010','1403013','1404002','1404005','1404007','1501027','1501028','1501041','1501056','1501058','1603010','1603014','1800002','1800009','1800010','1801007','1900004','2119001','2119002','2124001','2124002','2125001','2125002','2126002','2148003','2148004','2148006','2148011','2148013','2148015','2148019','E0000']
            #职业等级:5级
        jobCode_5_Class = ['0101020','0101022','0301001','0301002','0302007','0302010','0401014','0404001','0405003','0405004','0405005','0405006','0405007','0405008','0405009','0405013','0501012','0501013','0501018','0501020','0501025','0502023','0503003','0503004','0503005','0503006','0503007','0503008','0503009','0503010','0503011','0503012','0503013','0503014','0503015','0503016','0503017','0503018','0503019','0503020','0503021','0503022','0503023','0503030','0503041','0602008','0701019','0701020','0701021','0701026','0701037','0701038','0701041','0702006','0702009','0702011','0702012','0702013','0702014','0703003','0703005','0703007','0703010','0704001','0704002','0706004','0706006','0706010','0801010','0801012','0801013','0801014','0801015','0801016','0801017','0801018','0801019','0801021','0801027','0801030','0801038','0801039','0801041','0801043','0801044','0801045','0801046','0801049','0801050','0801052','0802004','0802006','0802007','0802008','0802009','0802010','0802013','0802014','0802016','0802017','0802018','0802020','0802021','0802022','0802023','0802024','0802025','0802026','0802027','0802028','0802029','0802030','0802032','0802033','0802034','0802035','0804008','0804012','0804016','0806015','0807007','0807012','0807013','0807015','0807030','0807043','0807048','0807049','0807057','0807058','0807060','0807105','0807112','0809009','0809010','0809011','0809012','0809013','0811004','0811006','0811007','0816017','0816018','0816019','0816020','0816054','0816056','0816057','1101005','1106004','1402006','1402008','1402030','1402031','1402038','1402039','1402040','1402048','1402049','1403007','1501042','1603016','1800004','1800005','1801008','2127002','2148005','2148007','2148008','2148018','2148020']
            #职业等级:6级
        jobCode_6_Class = ['0201009','0202001','0202002','0301003','0301004','0301005','0301006','0301007','0303004','0401005','0401006','0401007','0401028','0402005','0501019','0501021','0503033','0503036','0503037','0503040','0504029','0504032','0701012','0701025','0701029','0701030','0701034','0703004','0704009','0705004','0706001','0706007','0706013','0706028','0801042','0802042','0804005','0806006','0806007','0806008','0806009','0806011','0807038','0807041','0807042','0807054','0807107','0816055','0822010','0902005','0902006','1107008','1201009','1201010','1401008','1402037','1800006','1800014','1800016','2128002','2129002']
            #职业等级:9级
        jobCode_9_Class = ['0401001','0401002','0401003','0401004','0401009','0401015','0401029','0403001','0403002','0404005','0405010','0503035','0503042','0701018','0701031','0701033','0701035','0701036','0701044','0706008','0706009','0706015','0706017','0801011','0804006','0807004','0807008','0807009','0807010','0807110','0808001','0808002','0816041','0822007','0901004','1101006','1107007','1201011','1402007','1402047','1603020','1800007','1800008','1800015','1801005','1900002','1900008','1900009','1901003','1901004','1901007','2131002','2132002','2133002','2134002','2135002','2136002','2137002','2138001','2138002','2139001','2139002','2140001','2140002','2141001','2141002','2142001','2142002','2143001','2143002','2144001','2144002','2145001','2145002','2146001','2146002','2148001','6','7','F0000','Y0000']
        
        """2.判断职业等级"""
        jobClass = '1'   #默认职业等级为1级
        if jobCode in jobCode_1_Class:
            jobClass = '1'
        elif jobCode in jobCode_2_Class:
            jobClass = '2'
        elif jobCode in jobCode_3_Class:
            jobClass = '3'
        elif jobCode in jobCode_4_Class:
            jobClass = '4'
        elif jobCode in jobCode_5_Class:
            jobClass = '5'
        elif jobCode in jobCode_6_Class:
            jobClass = '6'
        elif jobCode in jobCode_9_Class:
            jobClass = '9'
        return jobClass


    #配置产品现金价值计算利率
    def query_product_presentrate(self, productCode):
        if productCode == '3237':
            rate = 0.04025
        elif productCode == '1217CB':
            rate = 0.055
    
        return rate
        


    #配置产品犹豫期天数
    def query_product_hesitation_days(self, productCode):
        if productCode in ('1217CA','3237','3254','3240','3267CB1','3275','3271','3273','1035','1034','1033','3257CA','3274','3218CA','8234','8235','8236','8237','8240','9201','5211','5202','3201','5201','5207','5206','3245','3241','3239B','3239','3238','3236','3211','3209','3208','3207','3205','3204','1210','1208','3253','3247B','3247','3256','3255','3260','3259','3252','3258','1206','1205','3218','1216CA','1216','1215CB','1214CB','1213CB','1217','3267','3267CO','3254CO','3266','3265','3265CB','3264','8209CB','3257','3261','3268','3269','3270'):
            hesitation_days = 15
        elif productCode in ('1217CB','1211','8232','8228','8231','8230','8233','8224','8225','8229','8214','8227','5214','5213','5241','5242','8208','9201CA','1026','1025','1024','1023','8217','8216','8218','1212CA','1212','1215','1214','1213','8221','8220','1217CA','8223','8222','8206','8206CA','8205','8205CA','8209CA','8209','8210','8210CA','8219','8219CA','8219CA1','8226'):
            hesitation_days = 10
        elif productCode in ('1032','2032','2034D','2034C','2034B','2034','1027','8003'):
            hesitation_days = 0
        return hesitation_days


    #产品贷款利率配置
    def query_product_loan_rate(self, productCode, apply_date):
        if productCode in ('1217CA','3275','8231','8232','8234','8235','8236'):
            loanRate = float(0.05)
        elif productCode in ('3239','3239B'):
            loanRate = float(0.0418)
        elif productCode in ('1217CB','3208','3209','3211','3218CA','3247','3247B','3255','3256','3264','3265','3265CB','3267','3267CB1','3267CO','3270','8237'):
            loanRate = float(0.0535)
        elif productCode in ('8233'):
            loanRate = float(0.06)
        elif productCode in ('5211','8214'):
            loanRate = float(0.0735) 
        elif productCode in ('1211','1212','1212CA','1213','1214','1213CB','1214CB','1215','1215CB','1216','1216CA','3201','3204','3205','3207','3218','3236','3237','3238','3240','3252','3254','3254CO','3259','3261','3266','3268','8205','8205CA','8206','8206CA','8208','8209','8209CA','8209CB','8210','8210CA','8218','8219','8219CA','8219CA1','8220','8221','8222','8223','8224','8226','8227','8228'):
            if apply_date < '2019-12-01':
                loanRate = float(0.0535)
            elif apply_date >= '2019-12-01':
                loanRate = float(0.05)
        elif productCode in ('1217','3241','3253','3260','3269','3271','3273','3274','8225','8229','8230','9201','9201CA'):
            if apply_date < '2020-05-19':
                loanRate = float(0.0535)
            elif apply_date >= '2020-05-19':
                loanRate = float(0.05)  
        elif productCode in ('1205','1206','3257','5202','5206','5207','5214','8216','8217','3257CA'):
            if apply_date < '2019-12-01':
                loanRate = float(0.0735)
            elif apply_date >= '2019-12-01':
                loanRate = float(0.06) 
        return loanRate


    #产品复效利率配置（取产品最近日期的贷款利率）
    def query_product_revival_rate(self, productCode):
        if productCode in ('1217CA','3275','8231','8232','8234','8235','8236'):
            loanRate = float(0.05)
        elif productCode in ('3239','3239B'):
            loanRate = float(0.0418)
        elif productCode in ('1217CB','3208','3209','3211','3218CA','3247','3247B','3255','3256','3264','3265','3265CB','3267','3267CB1','3267CO','3270','8237'):
            loanRate = float(0.0535)
        elif productCode in ('8233'):
            loanRate = float(0.06)
        elif productCode in ('5211','8214'):
            loanRate = float(0.0735) 
        elif productCode in ('1211','1212','1212CA','1213','1214','1213CB','1214CB','1215','1215CB','1216','1216CA','3201','3204','3205','3207','3218','3236','3237','3238','3240','3252','3254','3254CO','3259','3261','3266','3268','8205','8205CA','8206','8206CA','8208','8209','8209CA','8209CB','8210','8210CA','8218','8219','8219CA','8219CA1','8220','8221','8222','8223','8224','8226','8227','8228'):
            loanRate = float(0.05)
        elif productCode in ('1217','3241','3253','3260','3269','3271','3273','3274','8225','8229','8230','9201','9201CA'):
            loanRate = float(0.05)  
        elif productCode in ('1205','1206','3257','5202','5206','5207','5214','8216','8217','3257CA'):
            loanRate = float(0.06) 
        return loanRate



    #根据产品code查询产品id
    def query_productId(self, env_name, productCode):
        productId = ''
        if env_name in('uat4', 'uat6', 'uat7','uat8'):
            sql = "select * from wift_iiws_uat.t_product where PRODUCT_CODE='" + productCode +"'"
        else:
            sql = "select * from wift_iiws.t_product where PRODUCT_CODE='" + productCode +"'"
        #调用原子脚本
        cursor = API0000_diy().db_conf(env_name,'wift_iiws')
        #查询保单的生效日期、下期应缴日、缴费期间
        cursor.execute(sql)
        result = cursor.fetchall()
        if len(result) != 0:
            productId = result[0][0]
        return productId


    #根据产品id查询产品code
    def query_productCode(self, env_name, productId):
        productcode = ''
        if env_name in('uat4', 'uat6', 'uat7','uat8'):
            sql = "select * from wift_iiws_uat.t_product where product_id='" + productId +"'"
        else:
            sql = "select * from wift_iiws.t_product where product_id='" + productId +"'"
        #调用原子脚本
        cursor = API0000_diy().db_conf(env_name,'wift_iiws')
        #查询保单的生效日期、下期应缴日、缴费期间
        cursor.execute(sql)
        result = cursor.fetchall()
        if len(result) != 0:
            productcode = result[0][1]
        return productcode
       
        
    #计算被保人投保年龄
    def calc_policy_insured_age(self, effectiveDate, insuBirth):
        #获取被保人出生日期年份、月日，保单生效日期年份
        insuBirth_year = insuBirth[0:4]   #被保人出生日期年份
        insuBirth_month_day = insuBirth[4:10]   #被保人出生日期月日
        effectiveDate_year = effectiveDate[0:4]   #保单生效日期年份
        #计算被保人生日期年份与保单生效日期年份的差
        year = int(effectiveDate_year) - int(insuBirth_year)
        #拼接生效日期年份、被保人生日_月-日，得到一个新的日期
        date = effectiveDate_year + insuBirth_month_day
        if API0000_diy().isvaild_date(date):
            date = date   #若拼接的日期有效。日期不变
        else:
            date = effectiveDate_year + '-03-01'   #若拼接的日期无效。得到新日期xxxx-03-01（百度获知）。
        #生效日期>=date，投保年龄等于year；否则投保年龄等于year-1
        if datetime.datetime.strptime(effectiveDate, '%Y-%m-%d')>=datetime.datetime.strptime(date, '%Y-%m-%d'):
            policy_insured_age = year
        else:
            policy_insured_age = year-1
        return policy_insured_age
    
    
    #保费计算保额
    def premiumrate_calc_amount(self, FA, BA, premiumrate):
        FA = float(FA)
        BA = float(BA)
        premiumrate = float(premiumrate)
        result = round(float(premiumrate*FA/BA),2)
        return result


    #计算产品保额/保费
    def calc_product_premium(self, productCode, stdPremBf, sumAssured, units, dict):
            #手工计算投保年龄
            effectiveDate = dict['policy_info']['effectiveDate']   #保单生效日期
            insuBirth = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人出生日期
            policy_insured_age = API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth)
            #3237产品:
            if productCode == '3237':
                #获取产品FA、BA
                FA = stdPremBf
                BA = '1000'
                #获取费率premiumrate
                #filename =  sys.path[0] + r'\doc\data.exec.xlsx'  #获取录单excel路径
                if 'win' in sys.platform:
                    filename = r'D:\xLibrary\chunyu\doc\premium_rate\3237_premium_rate.xls'  #获取产品费率表的路径
                else:
                    filename = '/data/xServer/xLibrary/chunyu/doc/premium_rate/3237_premium_rate.xls'
                #连接excel费率文件
                wb = xlrd.open_workbook(filename)    
                sh = wb.sheets()[0]
                #获取费率
                premiumrate1 = sh.cell(5,2).value   #0至50岁;以年交1,000元保险费为计算单位,基本保险金额（元）
                premiumrate2 = sh.cell(6,2).value   #51至60岁;以年交1,000元保险费为计算单位,基本保险金额（元）
                premiumrate3 = sh.cell(7,2).value   #61至70岁;以年交1,000元保险费为计算单位,基本保险金额（元）
                if int(policy_insured_age)<=50:
                    premiumrate = premiumrate1
                elif (int(policy_insured_age)>50 and int(policy_insured_age)<=60):
                    premiumrate = premiumrate2
                elif (int(policy_insured_age)>60 and int(policy_insured_age)<=70):
                    premiumrate = premiumrate3
                else:
                    premiumrate = 0
                """4.保费计算保额"""
                #调用原子脚本:获取公式手工计算结
                sumAssured = API0000_diy().premiumrate_calc_amount(FA, BA, premiumrate)
            elif productCode in ('3257CA','8233','8237'):
                sumAssured = stdPremBf
            elif productCode == '3267CB1':
                stdPremBf = '10'
                sumAssured = '0'
            elif productCode == '1217CB':
                FA = stdPremBf
                BA = '1000'
                bonusYearOrAge = dict['policy_info']['main_risk_info'][0]['bonusYearOrAge']  # 年金领取年龄
                insuGender = dict['policy_info']['insured_info'][0]['insuGender']  # 被保人性别
                chargePeriod = dict['policy_info']['main_risk_info'][0]['chargePeriod']  # 缴费期间
                if 'win' in sys.platform:
                    filename = r'D:\xLibrary\chunyu\doc\premium_rate\1217CB_premium_rate.xls'  #获取产品费率表的路径
                else:
                    filename = '/data/xServer/xLibrary/chunyu/doc/premium_rate/1217CB_premium_rate.xls'
                #连接excel费率文件
                wb = xlrd.open_workbook(filename)    
                sh = wb.sheets()[0]
                #获取费率所在单元格的行数 
                if bonusYearOrAge == '55':
                    row = policy_insured_age + 8                     
                elif bonusYearOrAge == '60':
                    row = policy_insured_age + 67
                elif bonusYearOrAge == '65':
                    row = policy_insured_age + 131   
                #获取费率所在单元格的列数
                if chargePeriod == '1' and insuGender == '0':   #获取单元格列
                    column = 3
                elif chargePeriod == '1' and insuGender == '1':
                    column = 4
                elif chargePeriod == '3' and insuGender == '0':
                    column = 5
                elif chargePeriod == '3' and insuGender == '1':
                    column = 6
                elif chargePeriod == '5' and insuGender == '0':
                    column = 7
                elif chargePeriod == '5' and insuGender == '1':
                    column = 8
                elif chargePeriod == '10' and insuGender == '0':
                    column = 9
                elif chargePeriod == '10' and insuGender == '1':
                    column = 10
                elif chargePeriod == '20' and insuGender == '0':
                    column = 11
                elif chargePeriod == '20' and insuGender == '1':
                    column = 12
                #获取单元格值-费率
                premiumrate = sh.cell(row-1,column-1).value 
                """4.保费计算保额"""
                #调用原子脚本:获取公式手工计算结
                sumAssured = API0000_diy().premiumrate_calc_amount(FA, BA, premiumrate)                
            else:
                stdPremBf = float(stdPremBf)
                sumAssured = float(sumAssured)
            return stdPremBf, sumAssured
                

    #计算产品生存金
    def calc_product_survivalFee(self, sendDate, productCode, stdPremBf, sumAssured, units, dict):
        #3237产品
        if productCode == '3237':
            policy_status == dict['policy_info']['policy_status']   #保单状态
            total_times_renew = dict['public']['total_times_renew']   #续期次数
            if policy_status == '01' and total_times_renew == 2:
                survivalFee = round((round(float(stdPremBf), 2) * 0.1), 2)   #3237产品生存金=保费*10%
        elif productCode == '1217CB':
            a = 1
        else:
            #非3237产品
            survivalFee = 0
        return survivalFee

                
                
    #计算保单下期应缴日
    def calc_pay_due_date(self, total_times_renew, effectiveDate, chargePeriod, chargeMode):
        """1.趸交时:下期应缴日 = '9999-12-31'
           2.年交时:（1）续期次数>=交费期间时:下期应缴日 = '9999-12-31'
                     （2）续期次数<交费期间时:下期应缴日字符串=（生效日期年份 + 续期次数 + 1年） 拼接 （生效日期月-日）
                                               ①若下期应缴日字符串为有效日期:下期应缴日 = 下期应缴日字符串
                                               ②若下期应缴日字符串为无效日期:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
        """
        
        if chargeMode == 'S':   #趸交时
            pay_due_date = '9999-12-31'
        elif chargeMode == 'Y' and ((int(total_times_renew) + 1) >= int(chargePeriod)):   #年交，续期次数>=交费期间时
            pay_due_date = '9999-12-31'
        elif chargeMode == 'Y' and ((int(total_times_renew) + 1) < int(chargePeriod)):   #年交，续期次数<交费期间时
            #获取生效日期前一天
            effectiveDate1 = API0000_diy().date_add_subtract(effectiveDate, -1)
            #获取保单生效日期前一日年份
            effectiveDate1_year = effectiveDate1[0:4]   
            #初步组装:下期应缴日
            pay_due_date = str(int(effectiveDate1_year) + int(total_times_renew) + 1) + effectiveDate1[4:10]
            #最终得到:下期应缴日
            if API0000_diy().isvaild_date(pay_due_date):
                #日期有效:下期应缴日 = 下期应缴日字符串
                pay_due_date = pay_due_date
            else:
                #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                pay_due_date = pay_due_date[0:4] + '-02-28'
        return pay_due_date
            

    #计算申请日期所处的保单年度
    def calc_policy_years(self, effectiveDate , pay_due_date, apply_date):
        """1.续期未缴满时:保单年度 = 续期次数 + 1
           2.续期缴满时:
                （1）获取申请日期所在年份的保单周年日
                （2）判读日期大小:申请日期、申请日期所在年份的保单周年日:
                        ①申请日期 >= 申请日期所在年份的保单周年日:保单年度 = 申请日期年份 - 生效日期年份 + 1
                        ②申请日期 < 申请日期所在年份的保单周年日:保单年度 = 申请日期年份 - 生效日期年份
        """
        if pay_due_date != '9999-12-31':
            policy_years = str((datetime.datetime.strptime(pay_due_date, '%Y-%m-%d') - datetime.datetime.strptime(effectiveDate, '%Y-%m-%d')).days/364)
        else:
            #获取申请日期年份
            apply_date_year = apply_date[0:4]
            #获取保单生效日期年份、生效日期月-日
            effectiveDate_year = effectiveDate[0:4]
            effectiveDate_month_day = effectiveDate[4:10]
            #初步组装:申请日期所在年份的保单周年日
            date = apply_date_year + effectiveDate_month_day   #拼接申请日期年份、生效日期月-日
            #最终得到:申请日期所在年份的保单周年日
            if API0000_diy().isvaild_date(date):
                date = date   #若拼接的日期有效。日期不变
            else:
                date = apply_date_year + '-02-28'   #若拼接的日期无效。得到新日期xxxx-02-28
            #计算申请日期所处的保单年度
            if datetime.datetime.strptime(apply_date, '%Y-%m-%d') >= datetime.datetime.strptime(date, '%Y-%m-%d'):
                #申请日期 >= 申请日期所在年份的保单周年日:保单年度 = 申请日期年份 - 生效日期年份 + 1
                policy_years = str(int(apply_date_year) - int(effectiveDate_year) + 1)
            else:
                #申请日期 < 申请日期所在年份的保单周年日:保单年度 = 申请日期年份 - 生效日期年份
                policy_years = str(int(apply_date_year) - int(effectiveDate_year))
        return policy_years


    #计算申请日时-保单周年日
    def calc_policy_anniversary(self, effectiveDate, policy_years):
        #获取保单生效日期年份、生效日期月-日
        effectiveDate_year = effectiveDate[0:4]
        effectiveDate_month_day = effectiveDate[4:10]
        #计算申请日时保单周年日-所在的年份字符串
        year = str(int(effectiveDate_year) + int(policy_years) - 1)
        #初步组装:申请日时-保单周年日
        policy_anniversary = year + effectiveDate_month_day
        #最终得到:申请日时-保单周年日
        if API0000_diy().isvaild_date(policy_anniversary):
            policy_anniversary = policy_anniversary   #若拼接的日期有效。日期不变
        else:
            policy_anniversary = year + '-02-28'   #若拼接的日期无效。得到新日期xxxx-02-28
        return policy_anniversary


    #获取产品期初、年末现价
    def get_present_rate(self, productCode, policy_insured_age, policy_years, dict):
        #初始化
        list_begin = []   #定义产品期初现价列表
        list_end = []   #定义产品年末现价列表
        if productCode =='3237':
            #获取产品期初现价表的路径
            if 'win' in sys.platform:
                filename_begin = r'D:\xLibrary\chunyu\doc\present_rate\3237_present_begin_rate.xls'  #期初
            else:
                filename_begin = '/data/xServer/xLibrary/chunyu/doc/present_rate/3237_present_begin_rate.xls'
            #连接excel文件
            wb = xlrd.open_workbook(filename_begin)    
            sh = wb.sheets()[0]
            #获取现价
            for i in range(2,5):
                list = []
                for j in range(6,16):
                    list.append(float(sh.cell(j,i).value))
                list_begin.append(list)
            #获取3237产品年末现价表的路径
            if 'win' in sys.platform:
                filename_end =  r'D:\xLibrary\chunyu\doc\present_rate\3237_present_end_rate.xls'  #年末
            else:
                filename_end = '/data/xServer/xLibrary/chunyu/doc/present_rate/3237_present_end_rate.xls'
            #连接excel费率文件
            wb = xlrd.open_workbook(filename_end)    
            sh = wb.sheets()[0]
            #获取费率
            for i in range(2,5):
                list = []
                for j in range(6,16):
                    list.append(float(sh.cell(j,i).value))
                list_end.append(list)
            #获取被保人投保年龄对应的3237产品期初、年末现价    
            if int(policy_insured_age)<=50:
                begin_present_rate = list_begin[0][int(policy_years)-1]   #期初现价
                end_present_rate = list_end[0][int(policy_years)-1]   #年末现价
            elif (int(policy_insured_age)>50 and int(policy_insured_age)<=60):
                begin_present_rate = list_begin[1][int(policy_years)-1]   #期初现价
                end_present_rate = list_end[1][int(policy_years)-1]   #年末现价
            elif (int(policy_insured_age)>60 and int(policy_insured_age)<=70):
                begin_present_rate = list_begin[2][int(policy_years)-1]   #期初现价
                end_present_rate = list_end[2][int(policy_years)-1]   #年末现价
            else:
                begin_present_rate = 0
                end_present_rate = 0
        elif productCode =='1217CB':
            bonusYearOrAge = dict['policy_info']['main_risk_info'][0]['bonusYearOrAge']  # 年金领取年龄
            insuGender = dict['policy_info']['insured_info'][0]['insuGender']  # 被保人性别
            chargePeriod = dict['policy_info']['main_risk_info'][0]['chargePeriod']  # 缴费期间
            #获取产品现价表的路径
            #filename =  r'D:\xLibrary\chunyu\doc\present_rate\1217CB_present_rate.xls'
            #连接excel文件
            #wb = xlrd.open_workbook(filename)    
            #sh = wb.sheets()[0]
            #获取现价
                #获取现价所在单元格的行数 
            insuGender_list = [0,1]  #性别列表
            bonusYearOrAge_list = [55, 60, 65]  # 年金领取年龄列表
            chargePeriod_list = [1,3,5,10,20]  # 缴费期间列表
            sum = 0
            for i in range(0, len(insuGender_list)):  #性别
                for j in range(0, len(bonusYearOrAge_list)):  # 年金领取年龄   
                    for x in range(0, len(chargePeriod_list)):   # 缴费期间
                        if chargePeriod_list[x] in (1,3,5):  
                            age = bonusYearOrAge_list[j] - 5
                        else:
                            age = bonusYearOrAge_list[j] - chargePeriod_list[x]
                        for y in range(0, age+1):   #年龄
                            if insuGender_list[i] == 0 and bonusYearOrAge_list[j] == 65:
                                if (chargePeriod_list[x] == 1 and y == 50) or (chargePeriod_list[x] == 1 and y == 57) or (chargePeriod_list[x] == 3 and y == 31) or (chargePeriod_list[x] == 5 and y == 19)or (chargePeriod_list[x] == 5 and y == 34) or (chargePeriod_list[x] == 5 and y == 43) or (chargePeriod_list[x] == 10 and y == 46):
                                    policy_years1 = 107 - y
                                else:
                                    policy_years1 = 106 - y
                            elif insuGender_list[i] == 1 and bonusYearOrAge_list[j] == 60:
                                if (chargePeriod_list[x] == 3 and y == 2) or (chargePeriod_list[x] == 3 and y == 18) or (chargePeriod_list[x] == 3 and y == 26) or (chargePeriod_list[x] == 3 and y == 44) or (chargePeriod_list[x] == 3 and y == 50) or (chargePeriod_list[x] == 3 and y == 54) or (chargePeriod_list[x] == 5 and y == 31) or (chargePeriod_list[x] == 20 and y == 2) or (chargePeriod_list[x] == 20 and y == 26):
                                    policy_years1 = 107 - y
                                else:
                                    policy_years1 = 106 - y
                            elif insuGender_list[i] == 1 and bonusYearOrAge_list[j] == 65:
                                if (chargePeriod_list[x] == 1 and y == 16) or (chargePeriod_list[x] == 1 and y == 20):
                                    policy_years1 = 107 - y
                                else:
                                    policy_years1 = 106 - y
                            else:
                                policy_years1 = 106 - y   #保单年度
                            for z in range(1, policy_years1+1):
                                sum = sum+1
                                if str(insuGender_list[i]) == insuGender and str(bonusYearOrAge_list[j]) == bonusYearOrAge and str(chargePeriod_list[x]) == chargePeriod and str(y) == str(policy_insured_age) and str(z) == str(policy_years): 
                                    #begin_present_rate = float(sh.cell(sum, 9).value)  #获取单元格值-期初现价
                                    #end_present_rate = float(sh.cell(sum, 8).value)  ##获取单元格值-年末现价 
                                    rs = query_persent().query_1217CB_persent(sum-1)
                                    begin_present_rate = rs[0]  #获取单元格值-期初现价
                                    end_present_rate = rs[1]  ##获取单元格值-年末现价
        return begin_present_rate, end_present_rate
        



    #判断日期是否有效。有效返回True；无效返回False
    def isvaild_date(self, str_time):
        try:
            if parser.parse(str_time):
                isvaild = True
            else:
                isvaild = False
        except:
            isvaild = False
        return isvaild


    #日期加减天数得到日期
    def date_add_subtract(self, date, days):
        date_old = datetime.datetime.strptime(date, '%Y-%m-%d')
        delta = datetime.timedelta(days=days)   
        date_new = date_old + delta   #时间的datetime格式
        date_new = date_new.strftime('%Y-%m-%d')   #时间的str格式
        return date_new



    #计算两个日期字符串相差天数
    def cals_different_days(self, date1, date2):
        begintime=datetime.date(int(date1[0:4]),int(date1[5:7]),int(date1[8:10]))
        endtime=datetime.date(int(date2[0:4]),int(date2[5:7]),int(date2[8:10]))
        days = (begintime-endtime).days
        return days


    #计算现价公式中Days
    def calc_Days(self, apply_date, pay_due_date, policy_anniversary):
        #逻辑:1.缴费未满两种情况①申请日期早于下期应缴日②申请日期晚于下期应缴日2.缴费期满或趸交)
        if pay_due_date !='9999-12-31' and datetime.datetime.strptime(apply_date, '%Y-%m-%d') <= datetime.datetime.strptime(pay_due_date, '%Y-%m-%d'):
            days = (datetime.datetime.strptime(apply_date, '%Y-%m-%d') - datetime.datetime.strptime(policy_anniversary, '%Y-%m-%d')).days 
        elif pay_due_date !='9999-12-31' and datetime.datetime.strptime(apply_date, '%Y-%m-%d') > datetime.datetime.strptime(pay_due_date, '%Y-%m-%d'):
            days = int(365)
        elif pay_due_date == '9999-12-31':
            days = (datetime.datetime.strptime(apply_date, '%Y-%m-%d') - datetime.datetime.strptime(policy_anniversary, '%Y-%m-%d')).days 
        return days 
            
        

    #手工计算产品现价
    def calc_present(self, dict, apply_date):
        #长期险（传统非分红保险和分红保险）
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #主副险列表
        for i in range(0, len(product_list)):
            productCode = product_list[i]['productCode']   #产品代码
            #计算长期险（传统非分红保险和分红保险）现价
            if productCode in ('1217CB', '3237'):
                """1.计算保单下期应缴日。             API0000_diy().calc_pay_due_date(total_times_renew, effectiveDate, chargePeriod, chargeMode)
                   2.计算被保人投保年龄。             API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth)
                   3.计算申请日期所处的保单年度。     API0000_diy().calc_policy_years(effectiveDate , pay_due_date, apply_date)
                   4.计算申请日时的保单周年日。       API0000_diy().calc_policy_anniversary(effectiveDate, policy_years)
                   5.计算Days。                       API0000_diy().calc_Days(apply_date, pay_due_date, policy_anniversary)
                 **6.获取产品期初、年末现价。         API0000_diy().get_present_rate(policy_insured_age, policy_years)
                   7.计算3237产品现价。               
                """
                #初始化
                product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #组装保单产品列表
                effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
                callBackDate = dict['policy_info']['callBackDate']   #回执日期
                policy_status = dict['policy_info']['policy_status']   #保单状态
                end_reason = dict['policy_info']['end_reason']   #终止原因
                insuBirth = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人生日
                total_times_renew = dict['public']['total_times_renew']   #已续期次数
                FA = float(product_list[i]['stdPremBf'])   #FA
                days_hesitation = int(15)   #犹豫期天数
                chargeMode = 'Y'   #缴费方式
                chargePeriod = '3'   #缴费期间
                present_rate = float(API0000_diy().query_product_presentrate(productCode))   #现价计算利率
                BA = float(1000)   #单位保额/保费
                #1.计算保单下期应缴日
                pay_due_date = API0000_diy().calc_pay_due_date(total_times_renew, effectiveDate, chargePeriod, chargeMode)
                #2.计算被保人投保年龄
                policy_insured_age = API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth)
                #3.计算申请日期所处的保单年度
                policy_years = API0000_diy().calc_policy_years(effectiveDate , pay_due_date, apply_date)
                #print policy_years
                #4.计算申请日时的保单周年日
                policy_anniversary = API0000_diy().calc_policy_anniversary(effectiveDate, policy_years)
                #5.计算Days
                days = API0000_diy().calc_Days(apply_date, pay_due_date, policy_anniversary)
                #6.获取产品期初、年末现价
                rs = API0000_diy().get_present_rate(productCode, policy_insured_age, policy_years, dict)
                begin_present_rate = rs[0]   #期初现价
                end_present_rate = rs[1]   #年末现价
                #7.计算3237产品现价
                #情况1:保单非（有效、失效、终止且终止原因为失效两年后终止）时:现价为0
                if (policy_status in('01', '02') or (policy_status == '03' and end_reason == '11')) == False:
                    present = 0
                #情况2:（非情况1）保费申请日期处于犹豫期内是:现价为0
                elif datetime.datetime.strptime(apply_date, '%Y-%m-%d') <= datetime.datetime.strptime(API0000_diy().date_add_subtract(callBackDate, int(days_hesitation)), '%Y-%m-%d'):
                    present = 0
                #情况3:（非情况1、2）缴费未缴满,调用期初现价公式:${年末现价率}/pow((1+${现价计算利率}),(365-${Days})/365)*${FA}/${BA}
                elif pay_due_date != '9999-12-31':
                    present = round(end_present_rate/math.pow(1+present_rate,float(365-days)/365)*float(FA)/BA, 2)
                #情况4:（非情况1、2、3）缴费已经缴满，年末现价公式:(${期初现价}*(365-${Days})+${年末现价率}*${Days})/365*${FA}/${BA}
                elif pay_due_date == '9999-12-31':
                    present = round((begin_present_rate*float(365-days)+end_present_rate*days)/365*float(FA)/BA, 2)
            elif productCode in ('8233','8237'):
                present = 0
            else:
                present = 0
        return present


    #手工计算累计生息利息
    def calc_interest_account(self, dict, sendDate):
        #利率
        rate = float(0.035) 
        total_interest_account = 0   #定义本次累计生息账号利息  
        #如果无累积生息账号为空，累计生息利息返回为0
        if len(dict['policy_info']['cbSbAccount']['cbSbAccount_list']) == 0:
            total_interest_account = 0  
        else:
            #累计生息账号最后一笔交易记录日期
            dealTime = dict['policy_info']['cbSbAccount']['cbSbAccount_list'][-1]['dealTime']
            #累计生息账号余额
            cashAmount = float(dict['policy_info']['cbSbAccount']['cashAmount'])
            #累计生息账号最后一笔交易记录日期的年份、月日
            dealTime_year = dealTime[0:4]   #年份
            dealTime_month_day = dealTime[4:10]   #月日
            #此次计算累计生息利息日期的年份
            sendDate_year = sendDate[0:4]   #年份
            #初步组装:拼接sendDate_year年份、dealTime_month_day月-日
            date = sendDate_year + dealTime_month_day   #拼接sendDate_year年份、dealTime_month_day月-日
            #最终得到:拼接sendDate_year年份、dealTime_month_day月-日
            if API0000_diy().isvaild_date(date):
                date = date   #若拼接的日期有效。日期不变
            else:
                date = sendDate_year + '-03-01'   #若拼接的日期无效。得到新日期xxxx-02-28  
            #计算申请日期所处的保单年度
            if datetime.datetime.strptime(sendDate, '%Y-%m-%d') > datetime.datetime.strptime(date, '%Y-%m-%d'):
                #sendDate >= 拼接sendDate_year年份、dealTime_month_day月-日
                policy_years = int(sendDate_year) - int(dealTime_year) + 1
                #最后不足整年的起期
                last_date = date
            else:
                #sendDate < 拼接sendDate_year年份、dealTime_month_day月-日
                policy_years = int(sendDate_year) - int(dealTime_year)
                if policy_years > 0:
                    if API0000_diy().isvaild_date(str(int(date[0:4])-1) + date[4:10]):
                        last_date = str(int(date[0:4])-1) + date[4:10]   #若拼接的日期有效。日期不变
                    else:
                        last_date = str(int(date[0:4])-1) + '-03-01'   #若拼接的日期无效。得到新日期xxxx-02-28
                elif policy_years == 0:
                    last_date = dealTime        
            #计算累计生息账户利息    
            for i in range(0,(policy_years-1)):
                interest_account = round((cashAmount*math.pow(1+rate,365/float(365)))- cashAmount, 2) 
                cashAmount = cashAmount + interest_account
                total_interest_account = round((total_interest_account + interest_account), 2)
            #获取最后一年累计生息利息
            days1 = (datetime.datetime.strptime(sendDate, '%Y-%m-%d') - datetime.datetime.strptime(last_date, '%Y-%m-%d')).days
            if days1 == 366:
                days1 = 365
            interest_account = round(cashAmount*math.pow(1+rate,days1/float(365))-cashAmount, 2)
            total_interest_account = round((total_interest_account + interest_account), 2)
        return total_interest_account



    #手工退保-欠款余额
    def calc_tb_loanAccountAmount(self, dict, apply_date):
        loanAccountAmount = 0   #初始化贷款欠款余额
        index = -1
        list = []
        if len(dict['check']['calc_loan']) == 0:
            return loanAccountAmount
            #（2）贷款列表长度>0时,获取最后一次还款在贷款列表中的下标（默认为-1）
        elif len(dict['check']['calc_loan']) > 0:
            for i in range(len(dict['check']['calc_loan'])-1, -1, -1):
                if dict['check']['calc_loan'][i]['loan_type'] == 'loan_repayment':
                    index = i   #得到最后一次还款在贷款列表中的下标index
                    break
                #①最后一次还款在贷款列表中的下标=贷款列表长度-1，无贷款待还款记录:还款信息列表为空
            if index == (len(dict['check']['calc_loan'])-1):
                return loanAccountAmount
                #②还款下标=-1，无还款记录。截取待还款的贷款列表
            elif index == -1:
                list = dict['check']['calc_loan']
            elif index != -1:
                #③还款下标不等于-1、贷款列表长度-1时，截取待还款的贷款列表
                list = dict['check']['calc_loan'][int(index+1):int(len(dict['check']['calc_loan'])+1)]
                
        #2.遍历待还款列表，针对每个待还款记录操作如下
            for j in range(0,len(list)):
            #（1）获取贷款起期
                loan_begin_date = list[j]['apply_date']
                #print 'loan_begin_date:' + str(loan_begin_date)
            #（2）计算贷款已过日期
                loan_passed_days = (datetime.datetime.strptime(apply_date, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                #print 'loan_passed_days:' + str(loan_passed_days)
            #（3）计算贷款第一个6个月的贷款止期
                year = loan_begin_date[0:4]   #年
                month = loan_begin_date[5:7]   #月
                day = loan_begin_date[8:10]   #日
                if (int(month) + 6)>12:
                    year = str(int(year) + 1)
                    month = str(int(month) + 6 - 12)
                    if len(month) == 1:
                        month = '0' + month
                    day = day
                else:
                    year = year
                    month = str(int(month) + 6)
                    if len(month) == 1:
                        month = '0' + month
                    day = day
                loan_begin_end1 = year + '-' + month + '-' + day
                if API0000_diy().isvaild_date(loan_begin_end1):
                    #日期有效:下期应缴日 = 下期应缴日字符串
                    loan_begin_end1 = loan_begin_end1
                else:
                    #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                    loan_begin_end1 = loan_begin_end1[0:4] + '-03-01'
                #print 'loan_begin_end1:' + str(loan_begin_end1)
            #（4）计算贷款第一个6个月的天数
                loan_days1 = (datetime.datetime.strptime(loan_begin_end1, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                #print 'loan_days:' + str(loan_days)
            #（5）遍历得到贷款时产品的现价和
                total_presentPrice = 0   #每次贷款各产品的总现价和
                for i in range(0,len(list[j]['product'])):
                    total_presentPrice = total_presentPrice + round(float(list[j]['product'][i]['presentPrice']), 2)
                    #print round(float(list[j]['product'][i]['presentPrice']), 2)
                #print total_presentPrice
                #print "len(list[j]['product']):" + str(len(list[j]['product']))
            #（6）遍历贷款的产品列表:
                for i in range(0,len(list[j]['product'])):
                    productCode = list[j]['product'][i]['productCode']   #产品代码
                    #①获取实际贷款金额、产品贷款时的现价、产品贷款利率
                    payment1 = round(float(list[j]['payment']), 2)   #每次贷款的实际贷款金额 
                    presentPrice_each_loan = round(float(list[j]['product'][i]['presentPrice']), 2)   #每次贷款时产品现价
                    #loanRate = float(list[j]['product'][i]['loanRate'])   #产品贷款利率
                    #②计算产品贷款本金:现价/总现价*实际贷款金额
                    if total_presentPrice == 0:
                        product_loan_prin = 0
                    else:
                        product_loan_prin = round(presentPrice_each_loan/total_presentPrice*payment1, 2)    #计算产品贷款本金
                    #print 'product_loan_prin:' + str(product_loan_prin)
                    #③计算产品还款时现价
                        #3237产品
                    if list[j]['product'][i]['productCode'] == '3237':
                        FA = float(list[j]['product'][i]['stdPremBf'])   #FA
                        #1.计算现金价值
                        presentPrice = float(API0000_diy().calc_present(dict, apply_date))   #计算现价
                    #④计算还款利息:
                        #ⅰ:还款日期<=贷款日期6个月:计算贷款利息、贷款本息和
                    if int(loan_passed_days)<int(loan_days1):
                        #贷款利率
                        loanRate = float(API0000_diy().query_product_loan_rate(list[j]['product'][i]['productCode'], loan_begin_date))
                        if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                            loan_interest = round(product_loan_prin *(math.pow(1+loanRate, float(loan_passed_days)/float(365))-1), 2)
                        else:
                            loan_interest = round((product_loan_prin*loanRate*loan_passed_days/float(365)), 2)#计算贷款利息
                        #print 'loan_interest:' + str(loan_interest)
                        principalInterest = round(product_loan_prin + loan_interest, 2)    #贷款本息和
                        #ⅱ:还款日期>贷款日期6个月:计算资本化利息、未结利息、产品贷款本息和
                    else:
                        #计算资本化利息
                        interestCapital = 0   #初始化资本化利息
                        for k in range(0,1000):
                            #贷款利率
                            loanRate = float(API0000_diy().query_product_loan_rate(list[j]['product'][i]['productCode'], loan_begin_date))
                            #print k
                            #j:计算本次贷款每隔6个月的贷款止期
                            year = loan_begin_date[0:4]   #年
                            month = loan_begin_date[5:7]   #月
                            day = loan_begin_date[8:10]   #日
                            if (int(month) + 6)>12:
                                year = str(int(year) + 1)
                                month = str(int(month) + 6 - 12)
                                if len(month) == 1:
                                    month = '0' + month
                                day = day
                            else:
                                year = year
                                month = str(int(month) + 6)
                                if len(month) == 1:
                                    month = '0' + month
                                day = day
                            loan_begin_end = year + '-' + month + '-' + day
                            if API0000_diy().isvaild_date(loan_begin_end):
                                #日期有效:下期应缴日 = 下期应缴日字符串
                                loan_begin_end = loan_begin_end
                            else:
                                month = str(int(month)+1)
                                day = '01'
                                loan_begin_end = year + '-' + month + '-' + day
                                loan_begin_end = API0000_diy().date_add_subtract(loan_begin_end, -1)
                            #print 'loan_begin_end:' + str(loan_begin_end)
                            
                            #jj:计算本次贷款每6个月的贷款天数
                            loan_days = (datetime.datetime.strptime(loan_begin_end, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                            #print 'loan_days:' + str(loan_days)
                            #iii:计算本次贷款每6个月资本化利息，并合计
                            if int(loan_passed_days)>=int(loan_days):
                                #计算本次贷款每6个月的资本化利息
                                if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                                    interestCapital = round(interestCapital + round((product_loan_prin + interestCapital) *(math.pow(1+loanRate, float(loan_days)/float(365))-1), 2), 2)
                                else:
                                    interestCapital = round((interestCapital + round(((product_loan_prin + interestCapital) * loanRate * loan_days/float(365)), 2)), 2)
                                #重新计算贷款剩余天数
                                loan_passed_days = int(loan_passed_days) - int(loan_days)
                                #重新计算6个月的贷款起期
                                loan_begin_date = loan_begin_end
                            else:
                                break
                            #print interestCapital
                        if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                            loan_interest = round((product_loan_prin+interestCapital) *(math.pow(1+loanRate, float(loan_passed_days)/float(365))-1), 2)
                        else:
                            loan_interest = round(((product_loan_prin+interestCapital)*loanRate*int(loan_passed_days)/float(365)), 2)   #未结利息
                        #print 'loan_interest:' + str(loan_interest)
                        principalInterest = round(product_loan_prin + interestCapital + loan_interest, 2)    #险种贷款合计
            #（7）计算还款总金额:各产品贷款合计之和
                    loanAccountAmount = round((loanAccountAmount + principalInterest), 2)
        return loanAccountAmount


    #手工计算贷款-产品欠款余额列表
    def calc_loanAccountAmount(self, dict, apply_date):
        """计算欠款余额:
             step1:判断贷款列表长度为0时:产品欠款余额列表为空。结束函数
             step2:判断贷款列表长度>0时,获取最后一次还款在贷款列表中的下标，截取待还款的贷款列表:
                      1.还款下标=贷款列表长度-1，无贷款待还款记录:产品欠款余额列表为空。结束函数
                      2.还款下标=-1，无还款记录。截取待还款的贷款列表
                      3.还款下标不等于-1、贷款列表长度-1时，截取待还款的贷款列表
             step3:计算各产品的欠款余额:
                      1.定义贷款险种欠款列表
                      2.遍历未还款的贷款列表
                          2.1获取本次贷款起期
                          2.2计算本次贷款已过日期
                          2.3计算本次贷款6个月的天数
                          2.4遍历得到本次贷款的产品现价和
                          2.5计算产品欠款余额,加入产品欠款余额列表对应的产品欠款余额中
                               (1)计算产品贷款本金
                               (2)判断贷款已过日期小于6个月,计算产品欠款余额
                                    ①计算贷款利息
                                    ②计算产品欠款余额
                                    ③产品欠款余额加入产品欠款余额列表对应的产品欠款余额中
                               (3)判断贷款日期大于6个月,计算产品欠款余额
                                    ①计算资本化利息
                                        i:计算本次贷款每隔6个月的贷款止期
                                        ii:计算本次贷款每6个月的贷款天数
                                        #iii:计算本次贷款每6个月资本化利息，并合计
                                    ②未结利息
                                    ③计算产品欠款余额  
                                    ④产品欠款余额加入产品欠款余额列表对应的产品欠款余额中
             出参:产品欠款余额列表
        """
        #初始化
        product_loanBalance_list = []   #定义贷款险种欠款列表默认为[]
        list = []   #定义未还款的贷款列表默认为[]
        index = -1   #定义最后一次还款在贷款列表中的下标默认为-1
        """step1:判断贷款列表长度为0时:产品欠款余额列表为空。结束函数"""
        if len(dict['check']['calc_loan']) == 0:
            return product_loanBalance_list
            """step2:判断贷款列表长度>0时,获取最后一次还款在贷款列表中的下标，截取待还款的贷款列表:"""
        elif len(dict['check']['calc_loan']) > 0:
            #获取最后一次还款在贷款列表中的下标index
            for i in range(len(dict['check']['calc_loan'])-1, -1, -1):
                if dict['check']['calc_loan'][i]['loan_type'] == 'loan_repayment':
                    index = i   #得到最后一次还款在贷款列表中的下标index
                    break
                #1.还款下标=贷款列表长度-1，无待还款记录:产品欠款余额列表为空。结束函数
            #print 'i:' + str(i)
            #print len(dict['check']['calc_loan'])
            if index == (len(dict['check']['calc_loan'])-1):
                return product_loanBalance_list
                #2.还款下标=-1，无还款记录。截取待还款的贷款列表
            elif index == -1:
                list = dict['check']['calc_loan']   #待还款的贷款列表
            elif index != -1:
                #3.还款下标不等于-1、贷款列表长度-1时，截取待还款的贷款列表
                list = dict['check']['calc_loan'][int(index+1):int(len(dict['check']['calc_loan'])+1)]   #截取待还款的贷款列表
                """step3:计算欠款余额"""
                #1.定义贷款险种欠款列表（列表长度=贷款险种长度）
            for j in range(0,len(list[0]['product'])):
                product_loanBalance_list.append(0)   #定义列表长度=贷款险种长度，列表中所有值为0
                #2.遍历未还款的贷款列表
            for j in range(0,len(list)):
                #2.1获取本次贷款起期
                loan_begin_date = list[j]['apply_date']
                #print 'loan_begin_date:' + str(loan_begin_date)
                #2.2计算本次贷款已过日期
                loan_passed_days = (datetime.datetime.strptime(apply_date, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                #print 'loan_passed_days:' + str(loan_passed_days)
                #2.3计算贷款第一个6个月的贷款止期
                year = loan_begin_date[0:4]   #年
                month = loan_begin_date[5:7]   #月
                day = loan_begin_date[8:10]   #日
                if (int(month) + 6)>12:
                    year = str(int(year) + 1)
                    month = str(int(month) + 6 - 12)
                    if len(month) == 1:
                        month = '0' + month
                    day = day
                else:
                    year = year
                    month = str(int(month) + 6)
                    if len(month) == 1:
                        month = '0' + month
                    day = day
                loan_begin_end1 = year + '-' + month + '-' + day
                if API0000_diy().isvaild_date(loan_begin_end1):
                    #日期有效:下期应缴日 = 下期应缴日字符串
                    loan_begin_end1 = loan_begin_end1
                else:
                    #日期有效:下期应缴日 = （下期应缴日字符串年份） 拼接 （02-28）
                    loan_begin_end1 = loan_begin_end1[0:4] + '-03-01'
                #print 'loan_begin_end1:' + str(loan_begin_end1)
                #2.4计算本次贷款第一个6个月的天数
                loan_days1 = (datetime.datetime.strptime(loan_begin_end1, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                #print 'loan_days:' + str(loan_days)
                #2.4遍历得到本次贷款的产品现价和
                total_presentPrice = 0   #每次贷款各产品的总现价和
                for i in range(0,len(list[j]['product'])):
                    total_presentPrice = total_presentPrice + round(float(list[j]['product'][i]['presentPrice']), 2)
                    #print round(float(list[j]['product'][i]['presentPrice']), 2)
                #print total_presentPrice
                #2.5计算产品欠款余额,加入产品欠款余额列表对应的产品欠款余额中
                payment = round(float(list[j]['payment']), 2)   #贷款本金 
                product_loanBalance = 0   #定义产品欠款余额
                loanBalance = 0   #定义本次贷款的各产品欠款余额之和
                    #计算各产品的欠款余额
                for i in range(0,len(list[j]['product'])):
                    productCode = list[j]['product'][i]['productCode']   #产品代码
                    #(1)计算产品贷款本金
                    presentPrice = round(float(list[j]['product'][i]['presentPrice']), 2)   #产品现价
                    if total_presentPrice == 0:
                        product_loan_prin = 0
                    else:
                        product_loan_prin = presentPrice/total_presentPrice*payment   #计算产品贷款本金
                    #print 'product_loan_prin:' + str(product_loan_prin)
                    #loanRate = float(list[j]['product'][i]['loanRate'])   #产品贷款利率
                    #(2)判断贷款已过日期小于6个月,计算产品欠款余额
                    if int(loan_passed_days)<int(loan_days1):
                        #贷款利率
                        loanRate = float(API0000_diy().query_product_loan_rate(list[j]['product'][i]['productCode'], loan_begin_date))
                        #print loanRate
                        #①计算贷款利息（万能/非万能产品）
                        if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                            loan_interest = round(product_loan_prin *(math.pow(1+loanRate, float(loan_passed_days)/float(365))-1), 2)
                        else:
                            loan_interest = round((product_loan_prin*loanRate*loan_passed_days/float(365)), 2)
                        #print 'loan_interest:' + str(loan_interest)
                        #②计算产品欠款余额
                        product_loanBalance = round((product_loan_prin + loan_interest), 2) 
                        #print 'product_loanBalance' + str(product_loanBalance)
                        #③产品欠款余额加入产品欠款余额列表对应的产品欠款余额中
                        product_loanBalance_list[i] = round(float(product_loanBalance_list[i]), 2) + product_loanBalance
                    #(3)判断贷款日期大于6个月,计算产品欠款余额
                    else:
                        #①计算资本化利息
                        interestCapital = 0   #初始化资本化利息
                        for k in range(0,1000):
                            #贷款利率
                            loanRate = float(API0000_diy().query_product_loan_rate(list[j]['product'][i]['productCode'], loan_begin_date))  
                                #print loanRate
                            #i:计算本次贷款每隔6个月的贷款止期
                            year = loan_begin_date[0:4]   #年
                            month = loan_begin_date[5:7]   #月
                            day = loan_begin_date[8:10]   #日
                            if (int(month) + 6)>12:
                                year = str(int(year) + 1)
                                month = str(int(month) + 6 - 12)
                                if len(month) == 1:
                                    month = '0' + month
                                day = day
                            else:
                                year = year
                                month = str(int(month) + 6)
                                if len(month) == 1:
                                    month = '0' + month
                                day = day
                            loan_begin_end = year + '-' + month + '-' + day
                            if API0000_diy().isvaild_date(loan_begin_end):
                                #日期有效:下期应缴日 = 下期应缴日字符串
                                loan_begin_end = loan_begin_end
                            else:
                                month = str(int(month)+1)
                                day = '01'
                                loan_begin_end = year + '-' + month + '-' + day
                                loan_begin_end = API0000_diy().date_add_subtract(loan_begin_end, -1)
                            #print 'loan_begin_end:' + str(loan_begin_end)
                            
                            #ii:计算本次贷款每6个月的贷款天数
                            loan_days = (datetime.datetime.strptime(loan_begin_end, '%Y-%m-%d') - datetime.datetime.strptime(loan_begin_date, '%Y-%m-%d')).days
                            #iii:计算本次贷款每6个月资本化利息，并合计
                            if int(loan_passed_days)>=int(loan_days):
                                #计算本次贷款每6个月的资本化利息
                                if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                                    interestCapital = round(interestCapital + round((product_loan_prin + interestCapital) *(math.pow(1+loanRate, float(loan_days)/float(365))-1), 2), 2)
                                else:
                                    interestCapital = round((interestCapital + round(((product_loan_prin + interestCapital) * loanRate * loan_days/float(365)), 2)), 2)
                                #重新计算贷款剩余天数
                                loan_passed_days = int(loan_passed_days) - int(loan_days)
                                #重新计算6个月的贷款起期
                                #loan_begin_date = API0000_diy().date_add_subtract(loan_begin_end, 1)
                                loan_begin_date = loan_begin_end
                            else:
                                break
                        #print 'interestCapital:' + str(interestCapital)
                        #②未结利息
                        if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                            loan_interest = round((product_loan_prin+interestCapital) *(math.pow(1+loanRate, float(loan_passed_days)/float(365))-1), 2)
                        else:
                            loan_interest = round(((product_loan_prin+interestCapital)*loanRate*int(loan_passed_days)/float(365)), 2)
                          
                        #print 'loan_interest:' + str(loan_interest)
                        #③计算产品欠款余额
                        product_loanBalance = round((product_loan_prin + interestCapital + loan_interest), 2) 
                        #④产品欠款余额加入产品欠款余额列表对应的产品欠款余额中
                        product_loanBalance_list[i] = round(float(product_loanBalance_list[i]), 2) + product_loanBalance
                #print '----calc_loanBalance_eachloan:' + str(i) + '--------------------'
        #print product_loanBalance_list         
        return product_loanBalance_list



    #计算万能险风险保费-到达年龄
    def calc_arrival_age(self, effectiveDate, birthday, settle_date):
        #到达年龄=投保年龄+保单年度-1
            #计算投保年龄
        age = API0000_diy().calc_policy_insured_age(effectiveDate, birthday) 
            #风险保费计算区间开始日期(区分计算首期风险保费与非首期)
        if effectiveDate[0:7] == settle_date[0:7]:
            settle_date_begin = effectiveDate
        else:
            settle_date_begin = settle_date[0:7] + '-' + '01'
            #计算保单所处年度
        policy_years = API0000_diy().calc_policy_years(effectiveDate , '9999-12-31', settle_date_begin)
        #print policy_years
        #print 'age:' + str(age)
        #print 'policy_years:' + str(policy_years)
        arrival_age = int(age) + int(policy_years) - 1
        return arrival_age
        

    #计算万能险风险保费-实际年龄
    def calc_actual_age(self, birthday, settle_date):
        actual_age = API0000_diy().calc_policy_insured_age(settle_date, birthday)
        return actual_age



    #计算初始扣费
    def calc_initial_deduct(self, productCode, deduct_type, payment):
        if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
            if deduct_type == 'nb_premium':   #初始类型:缴费保险费
                rate = float(0.02)
            elif deduct_type == 'add_premium':   #初始类型:追加保险费
                rate = float(0.01)
        elif productCode == '3267CB1':
            rate = float(0.01)
        else:
            rate = 0
        initial_deduct = round(float(payment) * rate, 2)   #初始扣费金额
        return initial_deduct
        

    #计算有效保险金额
    def calc_valid_amount(self, productCode, value, interestCapital, settle_date, birthday, effectiveDate, dict):
        if productCode in ('8233','8237'):
            continuous_bonus = 0   #初始化持续奖金=0
            age = API0000_diy().calc_arrival_age(effectiveDate, birthday, settle_date)   #到达年龄
            #age = API0000_diy().calc_actual_age(birthday, settle_date)   #实际年龄
            #计算给付系数
            if age <= 17:
                rate = float(1)
            elif age > 17 and age <= 40:
                rate = float(1.6)
            elif age > 40 and age <= 60:
                rate = float(1.4)
            elif age > 60:
                rate = float(1.2)
            #获取万能账户派发的持续奖金
            for i in range(0, len(dict['check']['calc_continuous_bonus'])):
                continuous_bonus = round((continuous_bonus + float(dict['check']['calc_continuous_bonus'][i]['payment'])), 2)
            #计算基本保险金额乘以被保险人到达年龄所对应的给付系数
            #fee = round(float(value + continuous_bonus) * rate, 2)
            fee = round(float(value) * rate, 2)
            #计算有效保险金额计算
            interestCapital = float(interestCapital)
            if interestCapital >= fee:
                valid_amount = interestCapital
            else:
                valid_amount = fee
        else:
            valid_amount = 0
        return valid_amount
   

    #计算风险保险金额
    def calc_riskamount(self, productCode, valid_amount, interestCapital):
        if productCode in ('8233','8237'):      
            riskamount = round(float(valid_amount) - float(interestCapital), 2)
        else:
            riskamount = 0
        if riskamount < 0:
            riskamount = 0
        return riskamount
        


     
    #计算月风险保险费
    def calc_riskpremium(self, productCode, riskamount, birthday, gender, settle_date, effectiveDate):
        if productCode in ('8233','8237'): 
            age = API0000_diy().calc_arrival_age(effectiveDate, birthday, settle_date)   #到达年龄
            #获取风险保险费费率
            if 'win' in sys.platform:
                filename = r'D:\xLibrary\chunyu\doc\riskpremium_rate\8237.xlsx'
            else:
                filename = '/data/xServer/xLibrary/chunyu/doc/riskpremium_rate/8237.xlsx'               
            wb = load_workbook(filename)
            sh = wb['Sheet1']
            if gender == '0':
                riskpremium_rate = float(sh['B' + str(age+2)].value)  # 风险保险费费率
            elif gender == '1':
                riskpremium_rate = float(sh['C' + str(age+2)].value)  # 风险保险费费率
            #print 'age1:' + str(age)
            #print 'settle_date:' + settle_date
            #print 'riskpremium_rate:' + str(riskpremium_rate)
            
            #2.计算当月实际天数
            if settle_date[0:7] == effectiveDate[0:7]:   #结算日与生效日同年月时，当月实际天数为生效日期到下月1号日期的天数
                #组装新日期（下月1号日期）
                effectiveDate_year = effectiveDate[0:4]   #生效日期年份
                effectiveDate_month = effectiveDate[5:7]   #生效日期月份
                
                if int(effectiveDate_month)+1 < 10:
                    year = effectiveDate_year
                    month = '0' + str(int(effectiveDate[5:7])+1)
                elif int(effectiveDate_month)+1 >= 10 and int(effectiveDate_month)+1 <13:
                    year = effectiveDate_year
                    month = str(int(effectiveDate[5:7])+1)
                elif int(effectiveDate_month)+1 == 13:
                    year = str(int(effectiveDate_year) + 1)
                    month = '01'
                date = year + '-' + month + '-01'
                #计算相差天数
                days = API0000_diy().cals_different_days(date, effectiveDate)
            elif settle_date[0:7] != effectiveDate[0:7] and settle_date > effectiveDate:   #结算日与申请日非同年月时
                #计算结算日期当月实际天数 
                #monthRange = calendar.monthrange(int(settle_date_begin[0:4]), int(settle_date_begin[5:7]))
                monthRange = calendar.monthrange(int(settle_date[0:4]), int(settle_date[5:7]))
                days = monthRange[1]
                #if settle_date[0:7] == '2024-04':
                    #print days
            else:
                days = 0
            #3.公式计算月风险保险费
            riskpremium = round(float(riskamount) / float(1000) * riskpremium_rate * float(days) / float(365), 2)
        else:
            riskpremium = 0
        #if settle_date[0:7] == '2027-02':
            #print 'riskamount:' + str(riskamount)
            #print 'riskpremium_rate:' + str(riskpremium_rate)
            #print 'days:' + str(days)
        #print settle_date + ':' + str(riskpremium)
        return riskpremium



    #计算每月万能账户结息
    def calc_investAmount_interest(self, productCode, interestCapital, settle_date, last_settle_date):
        #利率
        if productCode in ('8237'):
            rate_ensure = float(0.03)   #保证结息利率
            rate = float(0.049)   #公布结息利率
            #利率不能低于保证利率
            if rate > rate_ensure:
                rate = rate
            else:
                rate = rate_ensure
        elif productCode in ('3267CB1', '8233'):
            rate = float(0.03)
        elif productCode in ('3257CA'):
            rate_ensure = float(0.03)   #保证结息利率
            if settle_date < '2018-06-01':
                rate = float(0.06)   #公布结息利率
            else:
                rate = float(0.059)   #公布结息利率
            #利率不能低于保证利率
            if rate > rate_ensure:
                rate = rate
            else:
                rate = rate_ensure
        else:
            rate = 0   #公布结息利率  
        #计算本次结算天数
        days = API0000_diy().cals_different_days(settle_date, last_settle_date)
        #if settle_date[0:7] == '2024-04':
            #print days
        #计算公式
        #print 'interestCapital:' + str(interestCapital)
        investAmount_interest = round(float(interestCapital) * float(math.pow(float(1 + rate), float(days)/float(365))-1), 2)
        #if settle_date[0:7] == '2027-04':
            #print 'interestCapital:' + str(interestCapital)
            #print 'days:' + str(days)
            #print 'investAmount_interest:' + str(investAmount_interest)
            #print 'rate:' + str(rate)
        return investAmount_interest


    #首期存入万能账户金额
    def calc_first_investAmount(self, dict, deduct_type):
        riskamount = 0
        birthday = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人出生日期
        gender = dict['policy_info']['insured_info'][0]['insuGender']   #被保人性别
        effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']
        for i in range(0, len(product_list)):
            interestCapital = 0
            productCode = product_list[i]['productCode']   #产品代码
            stdPremBf = float(product_list[i]['stdPremBf'])   #保费
            sumAssured = float(product_list[i]['sumAssured'])   #保额
            if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                #首期万能账户金额 = 保费-初始扣费-风险保费
                #a. #计算初始扣费
                initial_deduct = API0000_diy().calc_initial_deduct(productCode, deduct_type, stdPremBf)  
                #print 'initial_deduct:' + str(initial_deduct)
                #b.计算有效保险金额
                interestCapital = round((float(stdPremBf) - initial_deduct), 2)
                valid_amount = API0000_diy().calc_valid_amount(productCode, sumAssured, interestCapital, effectiveDate, birthday, effectiveDate, dict) 
                #print 'valid_amount:' + str(valid_amount)
                #c.计算风险保险金额
                riskamount = API0000_diy().calc_riskamount(productCode, valid_amount, interestCapital)   
                #print 'riskamount:' + str(riskamount)
                #d.计算月风险保险费
                riskpremium = API0000_diy().calc_riskpremium(productCode, riskamount, birthday, gender, effectiveDate, effectiveDate)
                #print 'riskpremium:' + str(riskpremium)
                #f.计算首期万能账户金额
                interestCapital = round((float(stdPremBf) - initial_deduct - riskpremium), 2)   
                #存入dict中
                #dict['policy_info']['investAmount']['investAmount_list'][-1]['investAmoun_detail_list']['transAmount'] = interestCapital
        return interestCapital





    #计算退保_万能账户价值（万能账户余额)
    def calc_tb_investAmount(self, dict, settle_date):
        riskamount_list = []   #初始化风险保额列表，为了计算退还风险保费
        total_interestCapital = 0
        birthday = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人出生日期
        gender = dict['policy_info']['insured_info'][0]['insuGender']   #被保人性别
        effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']
        for i in range(0, len(product_list)):
            interestCapital = 0   #初始化账户余额
            productCode = product_list[i]['productCode']   #产品代码
            stdPremBf = float(product_list[i]['stdPremBf'])   #保费
            sumAssured = float(product_list[i]['sumAssured'])   #保额
            if productCode in ('3267CB1', '3257CA', '8233', '8237'):
                #遍历全部账户（首期账户、持续奖金账户、追加保费账户）
                for ii in range(0, len(dict['policy_info']['investAmount']['investAmount_list'])):
                    riskamount_list.append([])
                    #账户类型
                    transCode = dict['policy_info']['investAmount']['investAmount_list'][ii]['transCode']
                    #获取上一次结息日期
                    last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime'] 
                    #结算日期与上一次结息日期的年份差
                    years = int(settle_date[0:4]) - int(last_settle_date[0:4])
                    #结算日期与上一次结息日期的月份差（存在负值情况）
                    months = int(settle_date[5:7]) - int(last_settle_date[5:7])  
                    #判断是否存在最后非整月的结息
                    if settle_date[8:10] == '01':  
                        flag_last_month = 'N'
                    else:
                        flag_last_month = 'Y'
                        
                    #(2)计算每月万能账户(不考虑最后的非整月是否存在)
                        #a.计算万能账户结息非最后一个月的总月数
                    settle_months = years*12 + months   #万能账户结息（不考虑最后的非整月是否存在）的月数
                    #print 'months:' + str(months)
                    for j in range(0, settle_months):
                        #print '-------'
                        last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime'] 
                        interestCapital = float(dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'])   #获取每次结息前-万能账户余额
                        #b.获取每月结息的日期
                        if int(last_settle_date[5:7]) <= 8:
                            each_settle_date_year = last_settle_date[0:4]   #每月结息的日期的年份
                            each_settle_date_month = '0' + str(int(last_settle_date[5:7])+1)    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        elif int(last_settle_date[5:7]) > 8 and int(last_settle_date[5:7]) < 12:
                            each_settle_date_year = last_settle_date[0:4]   #每月结息的日期的年份
                            each_settle_date_month = str(int(last_settle_date[5:7])+1)    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        elif int(last_settle_date[5:7]) == 12:
                            each_settle_date_year = str(int(last_settle_date[0:4])+1)   #每月结息的日期的年份
                            each_settle_date_month = '01'    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        each_settle_date = each_settle_date_year + '-' + each_settle_date_month + '-' + each_settle_date_day
                        #c.计算每次结息的万能账户利息
                        investAmount_interest = API0000_diy().calc_investAmount_interest(productCode, interestCapital, each_settle_date, last_settle_date)
                        #print 'investAmount_interest:' + str(investAmount_interest)
                        #d.计算每次结息的月风险保险费 
                            #获取参与计算有效保额的基数。首期账户为基本保额、持续奖金为派发的金额、追加保费为申请追加的金额  
                        if transCode == '00':   #首期账户
                            value = sumAssured
                        elif transCode == '01':   #持续奖金
                            value = round(dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][0]['transAmount'], 2)
                        elif transCode == '02':   #追加保费
                            value = round(dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][0]['transAmount']/0.99, 2)
                        valid_amount = API0000_diy().calc_valid_amount(productCode, value, interestCapital, settle_date, birthday, effectiveDate, dict)   #计算有效保险金额
                        riskamount = API0000_diy().calc_riskamount(productCode, valid_amount, (interestCapital+investAmount_interest))   #计算风险保险金额
                        riskpremium = API0000_diy().calc_riskpremium(productCode, riskamount, birthday, gender, each_settle_date, effectiveDate)   #计算月风险保险费
                        #print 'riskpremium:' + str(riskpremium)
                        #f.计算万能账户余额
                        interestCapital = round((interestCapital + investAmount_interest - riskpremium), 2)   #计算万能账户金额
                        #存入dict-万能账户结息列表中
                        investAmoun_detail_list = {'transAmount':interestCapital, 'dealTime':each_settle_date,'riskpremium':riskpremium,'investAmount_interest':investAmount_interest}
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'].append(investAmoun_detail_list)
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'] = round(interestCapital, 2)   #每个账户的余额
                        #存储各最后一期账户风险保额
                        if len(riskamount_list[-1]) == 0:
                            riskamount_list[-1].append(riskamount)
                        else:
                            riskamount_list[-1][-1] = riskamount
                        #存入riskamount_list风险保额列表
                        #print '-------------'
                    #(3)计算万能账户结息-最后的非整月（最后一月不计算风险保费）
                    if flag_last_month == 'Y':
                        last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime']   #获取上一次结息日期
                        interestCapital = dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital']   #获取每次结息前-万能账户余额
                        #a.计算最后的非整月结息的万能账户利息
                        investAmount_interest = API0000_diy().calc_investAmount_interest(productCode, interestCapital, settle_date, last_settle_date)
                        #c.计算万能账户余额
                        interestCapital = round((interestCapital + investAmount_interest), 2)   #计算最后的非整月的万能账户金额
                        #d.存入dict-万能账户结息列表中
                        investAmoun_detail_list = {'transAmount':interestCapital, 'dealTime':settle_date,'riskpremium':0,'investAmount_interest':investAmount_interest}
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'].append(investAmoun_detail_list)
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'] = round(interestCapital, 2)   #每个账户的余额
                
                #计算总账户余额
                for jj in range(0, len(dict['policy_info']['investAmount']['investAmount_list'])):
                    interestCapital = dict['policy_info']['investAmount']['investAmount_list'][jj]['interestCapital']
                    total_interestCapital = round((total_interestCapital + interestCapital), 2)
                dict['policy_info']['investAmount']['total_interestCapital'] = total_interestCapital
            else:
                interestCapital = 0
                riskamount_list = []
        return total_interestCapital, riskamount_list, dict



    #计算退保试算/贷款_万能账户价值（万能账户余额)
    def calc_trial_tb_investAmount(self, dict, settle_date):
        riskamount_list = []   #初始化风险保额列表，为了计算退还风险保费
        total_interestCapital = 0   #初始化总账户余额（仅计算整月）
        birthday = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人出生日期
        gender = dict['policy_info']['insured_info'][0]['insuGender']   #被保人性别
        effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']
        for i in range(0, len(product_list)):
            interestCapital = 0   #初始化账户余额
            productCode = product_list[i]['productCode']   #产品代码
            stdPremBf = float(product_list[i]['stdPremBf'])   #保费
            sumAssured = float(product_list[i]['sumAssured'])   #保额
            if productCode in ('3267CB1', '3257CA', '8233', '8237'):
                #遍历全部账户（首期账户、持续奖金账户、追加保费账户）
                for ii in range(0, len(dict['policy_info']['investAmount']['investAmount_list'])):
                    riskamount_list.append([])
                    #账户类型
                    transCode = dict['policy_info']['investAmount']['investAmount_list'][ii]['transCode']
                    #获取上一次结息日期
                    last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime'] 
                    #结算日期与上一次结息日期的年份差
                    years = int(settle_date[0:4]) - int(last_settle_date[0:4])
                    #结算日期与上一次结息日期的月份差（存在负值情况）
                    months = int(settle_date[5:7]) - int(last_settle_date[5:7])  
                    #判断是否存在最后非整月的结息
                    if settle_date[8:10] == '01':  
                        flag_last_month = 'N'
                    else:
                        flag_last_month = 'Y'
                        
                    #(2)计算每月万能账户(不考虑最后的非整月是否存在)
                        #a.计算万能账户结息非最后一个月的总月数
                    settle_months = years*12 + months   #万能账户结息（不考虑最后的非整月是否存在）的月数
                    #print 'months:' + str(months)
                    for j in range(0, settle_months):
                        #print '-------'
                        last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime'] 
                        interestCapital = float(dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'])   #获取每次结息前-万能账户余额
                        #b.获取每月结息的日期
                        if int(last_settle_date[5:7]) <= 8:
                            each_settle_date_year = last_settle_date[0:4]   #每月结息的日期的年份
                            each_settle_date_month = '0' + str(int(last_settle_date[5:7])+1)    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        elif int(last_settle_date[5:7]) > 8 and int(last_settle_date[5:7]) < 12:
                            each_settle_date_year = last_settle_date[0:4]   #每月结息的日期的年份
                            each_settle_date_month = str(int(last_settle_date[5:7])+1)    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        elif int(last_settle_date[5:7]) == 12:
                            each_settle_date_year = str(int(last_settle_date[0:4])+1)   #每月结息的日期的年份
                            each_settle_date_month = '01'    #每月结息日期的月份
                            each_settle_date_day = '01' 
                        each_settle_date = each_settle_date_year + '-' + each_settle_date_month + '-' + each_settle_date_day
                        #c.计算每次结息的万能账户利息
                        investAmount_interest = API0000_diy().calc_investAmount_interest(productCode, interestCapital, each_settle_date, last_settle_date)
                        #print 'investAmount_interest:' + str(investAmount_interest)
                        #d.计算每次结息的月风险保险费 
                            #获取参与计算有效保额的基数。首期账户为基本保额、持续奖金为派发的金额、追加保费为申请追加的金额  
                        if transCode == '00':   #首期账户
                            value = sumAssured
                        elif transCode == '01':   #持续奖金
                            value = round(dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][0]['transAmount'], 2)
                        elif transCode == '02':   #追加保费
                            value = round(dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][0]['transAmount']/0.99, 2)
                        valid_amount = API0000_diy().calc_valid_amount(productCode, value, interestCapital, settle_date, birthday, effectiveDate, dict)   #计算有效保险金额
                        riskamount = API0000_diy().calc_riskamount(productCode, valid_amount, (interestCapital+investAmount_interest))   #计算风险保险金额
                        riskpremium = API0000_diy().calc_riskpremium(productCode, riskamount, birthday, gender, each_settle_date, effectiveDate)   #计算月风险保险费
                        #print 'riskpremium:' + str(riskpremium)
                        #f.计算万能账户余额
                        interestCapital = round((interestCapital + investAmount_interest - riskpremium), 2)   #计算万能账户金额
                        #存入dict-万能账户结息列表中
                        investAmoun_detail_list = {'transAmount':interestCapital, 'dealTime':each_settle_date,'riskpremium':riskpremium,'investAmount_interest':investAmount_interest}
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'].append(investAmoun_detail_list)
                        dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'] = round(interestCapital, 2)   #每个账户的余额
                        #存储各最后一期账户风险保额
                        if len(riskamount_list[-1]) == 0:
                            riskamount_list[-1].append(riskamount)
                        else:
                            riskamount_list[-1][-1] = riskamount
                        #total_interestCapital1 = total_interestCapital1 + interestCapital
                        #print '-------------'
                    #(3)计算万能账户结息-最后的非整月（最后一月不计算风险保费）
                    if flag_last_month == 'Y':
                        last_settle_date = dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'][-1]['dealTime']   #获取上一次结息日期
                        interestCapital = dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital']   #获取每次结息前-万能账户余额
                        #a.计算最后的非整月结息的万能账户利息
                        investAmount_interest = API0000_diy().calc_investAmount_interest(productCode, interestCapital, settle_date, last_settle_date)
                        #c.计算万能账户余额
                        interestCapital = round((interestCapital + investAmount_interest), 2)   #计算最后的非整月的万能账户金额
                        #d.存入dict-万能账户结息列表中
                        investAmoun_detail_list = {'transAmount':interestCapital, 'dealTime':settle_date,'riskpremium':0,'investAmount_interest':investAmount_interest}
                        #dict['policy_info']['investAmount']['investAmount_list'][ii]['investAmoun_detail_list'].append(investAmoun_detail_list)
                        #dict['policy_info']['investAmount']['investAmount_list'][ii]['interestCapital'] = round(interestCapital, 2)   #每个账户的余额
                        
                    total_interestCapital = total_interestCapital + interestCapital     

            else:
                total_interestCapital = 0
                riskamount_list = []
        return total_interestCapital, riskamount_list




    #计算退保手续费
    def calc_tb_chargeAmount(self, dict, settle_date):
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #主副险列表
        for i in range(0, len(product_list)):
            productCode = product_list[i]['productCode']   #产品代码
            if productCode in ('3267CB1', '3257CA','8233','8237'):   #退保手续费-退保费用
                total_interestCapital = dict['policy_info']['investAmount']['total_interestCapital']   #获取每次结息前-万能账户余额
                effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
                #1.获取结算日期所处的保单周年日
                pay_due_date = '9999-12-31'   #下期应缴日。8237产品为趸交产品
                policy_years = API0000_diy().calc_policy_years(effectiveDate, pay_due_date, settle_date)  
                #2.获取退保费用率
                if policy_years == '1':
                    rate = 0.05
                elif policy_years == '2':
                    rate = 0.04
                elif policy_years == '3':
                    rate = 0.03
                elif policy_years == '4':
                    rate = 0.02
                elif policy_years == '5':
                    rate = 0.01
                else:
                    rate = 0
                #3.计算退保手续费
                chargeAmount = round(float(total_interestCapital) * float(rate), 2)
            else:
                chargeAmount = 0
        return chargeAmount


    #计算贷款手续费
    def calc_loan_chargeAmount(self, dict, settle_date, interestCapital):
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #主副险列表
        for i in range(0, len(product_list)):
            productCode = product_list[i]['productCode']   #产品代码
            if productCode in ('3267CB1', '3257CA','8233','8237'):   #退保手续费-退保费用
                effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
                #1.获取结算日期所处的保单周年日
                pay_due_date = '9999-12-31'   #下期应缴日。8237产品为趸交产品
                policy_years = API0000_diy().calc_policy_years(effectiveDate, pay_due_date, settle_date)  
                #2.获取退保费用率
                if policy_years == '1':
                    rate = 0.05
                elif policy_years == '2':
                    rate = 0.04
                elif policy_years == '3':
                    rate = 0.03
                elif policy_years == '4':
                    rate = 0.02
                elif policy_years == '5':
                    rate = 0.01
                else:
                    rate = 0
                #3.计算退保手续费
                chargeAmount = round(float(interestCapital) * float(rate), 2)
            else:
                chargeAmount = 0
        return chargeAmount



    #计算退保-退还未到期风险保费
    def calc_tb_riskChargeFee(self, dict, riskamount_list, settle_date):
        total_riskChargeFee = 0   #初始化总退还未到期风险保费
        product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #主副险列表
        for i in range(0, len(product_list)):
            productCode = product_list[i]['productCode']   #产品代码
            if productCode in ('8233','8237'):   #退保手续费-退保费用
                for j in range(0, len(dict['policy_info']['investAmount']['investAmount_list'])):
                    riskamount = float(riskamount_list[j][-1])   #获取对应风险保额
                    #获取dict数据
                    effectiveDate = dict['policy_info']['effectiveDate']   #生效日期
                    birthday = dict['policy_info']['insured_info'][0]['insuBirth']   #被保人出生日期
                    gender = dict['policy_info']['insured_info'][0]['insuGender']   #被保人性别
                    sumAssured = product_list[i]['sumAssured']   #保额

                    #1.计算本月实际结算天数+计算结算日期已过天数
                    if settle_date[0:7] == effectiveDate[0:7]:#如果结算日期与生效日期为同一个月时
                        #计算当月日期天数
                        monthRange = calendar.monthrange(int(settle_date[0:4]), int(settle_date[5:7]))
                        days1 = monthRange[1]
                        #计算生效日期前本月的天数
                        days2 = API0000_diy().cals_different_days(effectiveDate, effectiveDate[0:7] + '-01')
                        #计算本月实际结算天数=生效日期后本月天数
                        days = days1 - days2
                        #计算结算日期已过天数
                        pass_days = API0000_diy().cals_different_days(settle_date, effectiveDate)
                    else:
                        #计算本月实际结算天数=当月日期天数
                        monthRange = calendar.monthrange(int(settle_date[0:4]), int(settle_date[5:7]))
                        days = monthRange[1]
                        #print 'days:' + str(days)
                        #计算结算日期已过天数
                        pass_days = API0000_diy().cals_different_days(settle_date, settle_date[0:7] + '-01')
                    #2.计算退保-退还未到期风险保费:未满期风险保险费
                    age = API0000_diy().calc_arrival_age(effectiveDate, birthday, settle_date)   #到达年龄

                    if 'win' in sys.platform:
                        filename = r'D:\xLibrary\chunyu\doc\riskpremium_rate\8237.xlsx'
                    else:
                        filename = '/data/xServer/xLibrary/chunyu/doc/riskpremium_rate/8237.xlsx'                        
                    wb = load_workbook(filename)
                    sh = wb['Sheet1']
                    if gender == '0':
                        riskpremium_rate = float(sh['B' + str(age+2)].value)  # 风险保险费费率
                    elif gender == '1':
                        riskpremium_rate = float(sh['C' + str(age+2)].value)  # 风险保险费费率
                        #计算退保-退还未到期风险保费
                    riskChargeFee = round((round(float(riskamount) / float(1000) * riskpremium_rate * float(days) / float(365), 2) - round(float(riskamount) / float(1000) * riskpremium_rate * float(pass_days) / float(365), 2)), 2)
                    #print 'riskamount:' + str(riskamount)
                    #print 'riskpremium_rate:' + str(riskpremium_rate)
                    #print 'pass_days:' + str(pass_days)
                    #print 'riskChargeFee:' + str(riskChargeFee)
                    total_riskChargeFee = total_riskChargeFee + riskChargeFee
            else:
                total_riskChargeFee = 0
        return total_riskChargeFee
     

    #判断字符串编码格式。出参为字符串utf8、gbk_gb2312、unicode
    def whichEncode(self, text):
        text0 = text[0]
        try:
            text0.decode('utf8')
        except Exception, e:
            if "unexpected end of data" in str(e):
                return "utf8"
            elif "invalid start byte" in str(e):
                return "gbk_gb2312"
            elif "ascii" in str(e):
                return "unicode"
        return "utf8"


    #解决windows终端字符串乱码问题
    def solve_terminal_garbled(self, text):
        if text != None and text !='':
            if API0000_diy().whichEncode(text) == 'utf8':
                text = text.decode('utf-8')
            if API0000_diy().whichEncode(text) == 'unicode':
                text = text.encode("gbk")
        print text




    #win平台中文转换gbk格式
    def text_conversion(self, text):
        if 'win' in sys.platform:
            text = text.decode('utf-8').encode('gbk')
        return text


    #判断两个列表是否为包含关系
    def get_sub_list_index(self, parent_list, sub_list):   #parent_list大列表   sub_list小列表
        #初始化
        is_success = ''
        message = ''
        #执行
        for i in range(0,len(sub_list)):
            if sub_list[i] in parent_list:
                is_success = True
                message = ''
            else:
                is_success = False
                break
        return is_success


    #两个时间格式为'xxxx年xx月xx日 xx时xx分xx秒'的时间，计算差值后再转化为x时x分x秒格式
    def logger_count_time(self, begin_time, end_time):
        consume_time = ''   #初始化
        #计算相差结果
        time = str(end_time - begin_time)
        #提取时分秒
        hours = int(time.split(':')[0])   #时
        minutes = int(time.split(':')[1])   #分
        seconds = round(float(time.split(':')[2]), 2)   #秒
        #格式：x时xx分xx.xx秒(时分为0时，不展示)
        if hours != 0:
            consume_time = consume_time + str(hours) + '时'
        if  minutes != 0:
            consume_time = consume_time + str(minutes) + '分'
        consume_time = consume_time + str(seconds) + '秒'
        return consume_time


    #两个时间格式为'x时x分x秒'的时间，计算差值后再转化为x时x分x秒格式
    def logger_count_time1(self, time_max, time_min):
        """step1:得到time_max总秒数"""
        #time_max替换
        time_max = time_max.replace('时', '-')
        time_max = time_max.replace('分', '-')
        time_max = time_max.replace('秒', '-')
        list_time_max = time_max.split('-')[::-1] 
        del list_time_max[0]
        if len(list_time_max) == 1:
            list_time_max.append('0')
            list_time_max.append('0')
        elif len(list_time_max) == 2:
            list_time_max.append('0')
        #得到time_max总秒数
        time_max_second = float(list_time_max[0]) + float(list_time_max[1])*60 + float(list_time_max[2])*3600
        
        """step2:得到time_min总秒数"""
        #time_min替换
        time_min = time_min.replace('时', '-')
        time_min = time_min.replace('分', '-')
        time_min = time_min.replace('秒', '-')
        list_time_min = time_min.split('-')[::-1] 
        del list_time_min[0]
        if len(list_time_min) == 1:
            list_time_min.append('0')
            list_time_min.append('0')
        elif len(list_time_min) == 2:
            list_time_min.append('0')
        #得到time_max总秒数
        time_min_second = float(list_time_min[0]) + float(list_time_min[1])*60 + float(list_time_min[2])*3600

        """step3:计算time_max, time_min总秒数差，并转换为时分秒"""
        #时
        hour = (time_max_second-time_min_second)/3600
        if hour < 1 and hour > 0:
            hour = ''
        else:
            hour = str(int(hour)) + '时'
        #分
        minute = round((time_max_second-time_min_second)%3600/60, 2)
        if minute < 1 and minute > 0:
            minute = ''
        else:
            minute = str(int(minute)) + '分'
        #秒
        second = str(round((time_max_second-time_min_second)%3600%60, 2)) + '秒'
        return hour+minute+second


    #YYYY-DD-MM日期字段正则校验
    def regular_date(self, date_name, date):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + date_name +'格式是否符合规范'
        #校验1：YYYY-MM-DD格式
        pattern = r'^\d{4}-\d{1,2}-\d{1,2}$'   #正则表达式:YYYY-MM-DD
        if re.match(pattern, date) == None:
            is_success = False
            message = '格式有误，正确格式为YYYY-MM-DD！'
            return is_success, message, unit

        #日期拆分为年、月、日存入list列表。月日位数不足自动补充
        list = date.split('-')
        year = list[0]   #年
        month = list[1]   #月
        day = list[2]   #日
            #补充月、日
        if len(month) == 1:
            month = '0' + month
        if len(day) == 1:
            day = '0' + day

        #校验2：判断日期-年范围(1990~2099)
        if int(year) < 1990 or int(year) > 2099:
            is_success = False
            message = '录入的年超出范围！'
            return is_success, message, unit

        #校验3：判断日期-月范围(1~12)
        if int(month) < 1 or int(month) > 12:
            is_success = False
            message = '录入的月超出范围！'
            return is_success, message, unit

        #校验4：判断日期-日范围(1~31)
        if int(day) < 1 or int(day) > 31:
            is_success = False
            message = '录入的日超出范围！'
            return is_success, message, unit

        #组装YYYY-MM-DD
        date = year + '-' + month + '-' + day

        #校验5：判断日期是否有效
        if API0000_diy().isvaild_date(date) == False:
            is_success = False
            message = '录入的日期不存在！'
            return is_success, message, unit
        return is_success, message, unit


    #金额字段正则校验
    def regular_amount(self, amount_name, amount):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + amount_name +'格式是否符合规范'
        #校验1：YYYY-MM-DD格式
        pattern = r'^\d+(\.\d+)?$'   #正则表达式:整型、小数
        if re.match(pattern, amount) == None:
            is_success = False
            message = '格式有误，正确格式为整型或小数！'
            return is_success, message, unit
        return is_success, message, unit


    #匹配手机号并返回匹配结果
    def regular_telephone(self, tellphone_name, tellphone):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + tellphone_name +'格式是否符合规范'
        #正则表达式
        pattern = r'^1[35678]\d{9}$'   
        #执行正则匹配得到匹配结果
        if re.match(pattern, tellphone) == None:
            is_success = False
            message = '录入的格式有误！'
            return is_success, message, unit
        return is_success, message, unit


    #匹配邮箱账号并返回匹配结果
    def regular_email(self, email_name, email):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + email_name +'格式是否符合规范'
        #正则表达式
        pattern = r'^(\w)+(.\w+)*@(\w)+((.\w+)+)$'   
        #执行正则匹配得到匹配结果
        if re.match(pattern, email) == None:
            is_success = False
            message = '录入的格式有误！'
            return is_success, message, unit
        return is_success, message, unit


    #匹配登录用户名（字母、数字、下划线）并返回匹配结果
    def regular_sys_name(self, sysname_name, sys_name):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + sys_name +'格式是否符合规范'
        #正则表达式
        pattern = r'^[0-9a-zA-Z_]{1,}$'     #字母、数字、下划线
        #执行正则匹配得到匹配结果
        if re.match(pattern, sys_name) == None:
            is_success = False
            message = '录入的格式有误！'
            return is_success, message, unit
        return is_success, message, unit



    #匹配次数（无范围）并返回匹配结果
    def regular_times(self, times_name, times):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + times_name +'格式是否符合规范'
        #正则表达式
        pattern = r'^\d+$'   
        #执行正则匹配得到匹配结果
        if re.match(pattern, times) == None:
            is_success = False
            message = '录入的格式有误！'
            return is_success, message, unit
        return is_success, message, unit


    #匹配出单次数（小于等于10次）并返回匹配结果
    def regular_nb_times(self, times_name, times):
        #初始化
        is_success = True
        message = ''
        unit = '校验录入的' + times_name +'格式是否符合规范'
        #正则表达式
        pattern = r'^([1-9]|10)$'    
        #执行正则匹配得到匹配结果
        if re.match(pattern, times) == None:
            is_success = False
            print is_success
            message = '录入的内容有误，只支持录入0~10！'
            return is_success, message, unit
        return is_success, message, unit


    #checkdata界面录入字段-后台传值json-脚本，字段是否存在及读取字段值
    def checkdata_json_check(self,input_dict,key_list):
        try:
            #key_list为后台传值json串中字段
            #初始化
            is_run = True
            msg = ''
            list1 = []   #读取json中字段值

            try:
                #校验json串类型是否有误
                input_dict = json.loads(input_dict)   #json转dict格式
            except:
                is_run = False
                msg = msg + '后台传值第二个参数（json串）：json串类型有误！'
            else:
                #遍历后台传值json串中字段
                for i in range(0, len(key_list)):
                    #判断json串是否存在xxx字段
                    if input_dict.has_key(key_list[i]) == False:
                        is_run = False
                        msg = msg + '后台传值第二个参数-json串中不存在' + key_list[i] + '字段！'
                        value = ''
                    else:
                        value = input_dict[key_list[i]]
                    list1.append(value)
        except Exception, e:   #常见异常的捕捉
            print traceback.format_exc()
        return is_run, msg, list1


    #checkdata界面录入字段-后台传值json-脚本，字段是否存在及读取字段值
    def checkdata_json_check1(self,key_dict):
        try:
            #key_list为前端录入字段中待校验列表
            #初始化
            is_run = True
            msg = ''
            #验证环境类型
            if key_dict.has_key('env_name'):
                if key_dict['env_name'] not in ('uat4','uat6','uat7','uat8','预生产'):
                    is_run = False
                    msg = msg + '录入的环境类型有误！'
                    return is_run, msg
            #验证承保原因
            if key_dict.has_key('withdrawalReason'):
                if key_dict['withdrawalReason'] not in ('01','03'):
                    is_run = False
                    msg = msg + '录入的承保原因有误！'
                    return is_run, msg
            #验证解约退费类型
            if key_dict.has_key('cancelContractType'):
                if key_dict['cancelContractType'] not in ('1-解约且退还全部保费','2-解约且退还现金价值','3-解约且不退费'):
                    is_run = False
                    msg = msg + '录入的解约退费类型有误！'
                    return is_run, msg
            #验证解约原因
            if key_dict.has_key('cancelReason'):
                if key_dict['cancelReason'] not in ('1-核保结论解约','2-客户重要信息变更解约','3-职业类别变更解约','4-其他'):
                    is_run = False
                    msg = msg + '录入的解约原因有误！'
                    return is_run, msg
            #续保方式
            if key_dict.has_key('renew_flag'):
                if key_dict['renew_flag'] not in ('0-满期不续保','1-满期续保'):
                    is_run = False
                    msg = msg + '录入的续保方式有误！'
                    return is_run, msg   
            #保费垫交/抵交选项
            if key_dict.has_key('autoPaddingPrem'):
                if key_dict['autoPaddingPrem'] not in ('0-否','1-现金价值自垫','2-万能账户抵交'):
                    is_run = False
                    msg = msg + '录入的保费垫交/抵交选项有误！'
                    return is_run, msg   
            #保全试算类型
            if key_dict.has_key('type_trial'):  
                if key_dict['type_trial'] not in ('退保试算','贷款试算'):
                    is_run = False
                    msg = msg + '录入的保全试算类型有误！'
                    return is_run, msg
            #补发原因
            if key_dict.has_key('reissueReason'):  
                if key_dict['reissueReason'] not in ('1-客户遗失','2-业务员遗失','3-保单污损','4-其他'):
                    is_run = False
                    msg = msg + '录入的补发原因有误！'
                    return is_run, msg
            #补发类型
            if key_dict.has_key('reissueType'):  
                if key_dict['reissueType'] not in ('1-纸质保单','2-电子保单'):
                    is_run = False
                    msg = msg + '录入的补发类型有误！'
                    return is_run, msg
            #红利领取方式
            if key_dict.has_key('bonusPayMode'):  
                if key_dict['bonusPayMode'] not in ('0-现金领取', '2-累计生息', '3-转入万能账户'):
                    is_run = False
                    msg = msg + '录入的红利领取方式有误！'
                    return is_run, msg
            #验证保全申请日期
            if key_dict.has_key('apply_date'):
                rs = API0000_diy().regular_date('保全申请日期', key_dict['apply_date'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证派发日期
            if key_dict.has_key('sendDate'):
                rs = API0000_diy().regular_date('派发日期', key_dict['sendDate'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证满期终止日期
            if key_dict.has_key('expiryDate'):
                rs = API0000_diy().regular_date('满期终止日期', key_dict['expiryDate'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证投保日期
            if key_dict.has_key('applicationDate'):
                rs = API0000_diy().regular_date('投保日期', key_dict['applicationDate'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证保全试算日期
            if key_dict.has_key('trial_day'):
                rs = API0000_diy().regular_date('保全试算日期', key_dict['trial_day'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证贷款金额
            if key_dict.has_key('loanApplyAmount'):
                rs = API0000_diy().regular_amount('贷款申请金额', key_dict['loanApplyAmount'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证贷款还款金额
            if key_dict.has_key('repayment_amount'):
                rs = API0000_diy().regular_amount('贷款还款金额', key_dict['repayment_amount'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证追加保费金额
            if key_dict.has_key('addInvestAmount'):
                rs = API0000_diy().regular_amount('追加保费金额', key_dict['addInvestAmount'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证减保金额
            if key_dict.has_key('adjustAmount'):
                rs = API0000_diy().regular_amount('减保金额', key_dict['adjustAmount'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #验证协议退保金额
            if key_dict.has_key('agreementSurrenderAmount'):
                rs = API0000_diy().regular_amount('协议退保金额', key_dict['agreementSurrenderAmount'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的回执日期格式是否符合规范
            if key_dict.has_key('callBackDate'):
                rs = API0000_diy().regular_date('保单回执日期', key_dict['callBackDate'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的手机号格式是否符合规范
            if key_dict.has_key('telephone'):
                rs = API0000_diy().regular_telephone('新增雇员手机号', key_dict['telephone'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的邮箱格式是否符合规范
            if key_dict.has_key('email'):
                rs = API0000_diy().regular_email('新增雇员邮箱', key_dict['email'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的登录用户名是否符合规范
            if key_dict.has_key('sys_name'):
                rs = API0000_diy().regular_sys_name('登录用户名', key_dict['sys_name'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的续期次数是否符合规范
            if key_dict.has_key('times_renew'):
                rs = API0000_diy().regular_times('续期次数', key_dict['times_renew'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
            #校验录入的出单次数是否符合规范
            if key_dict.has_key('nb_time'):
                rs = API0000_diy().regular_nb_times('出单次数', key_dict['nb_time'])
                if rs[0] == False:
                    is_run = False
                    msg = msg + rs[2] + ':' + rs[1]
                    return is_run, msg
        except Exception, e:   #常见异常的捕捉
            print traceback.format_exc()
        return is_run, msg





"""新契约接口类"""
class API0000_nb():
        # 投保单号提交接口。policyType:1个单2团单；proposalNo投保单号
    def entry_newProposal(self, magicX, interface_ip, policyType, proposalNo):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/propProcess/entry/newProposal.do"
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }

        # 组装请求参数
        data = {
            "policyType": policyType,
            "proposalNo": proposalNo,
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保单基本信息录入接口
    def proposal_updateProposalInfo(self, magicX, interface_ip, proposalNo, policyNo, proposalId, organId, policyChannel, channelId, salesChannelText, salesChannel, headPartner, partnerId, tellerCode, sellerCode, sellerName, agentId, submitChannel, subSalesChannel, applicationDate, effectiveDate, autoPaddingPrem, disputedType, policyDeliveryMode):

        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/proposal/updateProposalInfo.do?magicX=" + magicX
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId,
            "propStatus": '',
            "commentId": '',
            "organId": organId,
            "policyChannel": policyChannel,
            "prepolicyChannel": policyChannel,
            "channelId": channelId,
            "salesChannelText": salesChannelText,
            "salesChannel": salesChannel,
            "headPartner": headPartner,
            "partnerId": partnerId,
            "tellerCode": tellerCode,
            "sellerCode": sellerCode,
            "sellerName": sellerName,
            "agentId": agentId,
            "submitChannel": submitChannel,
            "subSalesChannel": subSalesChannel,
            "applicationDate": applicationDate,
            "preapplicationDate": applicationDate,
            "effectiveDate": effectiveDate,
            "autoPaddingPrem": autoPaddingPrem,
            "disputedType": disputedType,
            "policyDeliveryMode": policyDeliveryMode,
            "proposalNo": proposalNo,
            "policyNo": policyNo
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 投保人信息录入接口
    def savePersonalHolder(self, magicX, interface_ip, proposalNo, policyChannel, sameASInsurd, name, gender, pbHoldBirth, certiType, certiCode, pbCertiValidEndDate, isMarried, height, weight, mobilePhone, officeTel, homeTel, industry, workCompany, email, jobCode, jobClass, driverLicenseType, nationality, nationnality1, educationId, medicalInsType, incomeSource, incomeSourceNote, annualIncome, familyIncome, premBudget, residentType, taxPayerType, postalCode, provinceCode, cityCode, districtCode, addrDetail):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/holder/savePersonalHolder.do?magicX=" + magicX
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "customerId": '',
            "policyChannelinfo": policyChannel,
            "applicationDate": '',
            "subSalesChannel": '',
            "sameASInsurd": sameASInsurd,
            "name": name,
            "prename": '',
            "gender": gender,
            "pregender": '',
            "pbHoldBirth": pbHoldBirth,
            "prepbHoldBirth": '',
            "certiType": certiType,
            "precertiType": '',
            "certiCode": certiCode,
            "precertiCode": '',
            "pbCertiValidEndDate": pbCertiValidEndDate,
            "perpetual": 'on',
            "isMarried": isMarried,
            "height": height,
            "weight": weight,
            "mobilePhone": mobilePhone,
            "officeTel": officeTel,
            "homeTel": homeTel,
            "industry": industry,
            "workCompany": workCompany,
            "email": email,
            "jobCode": jobCode,
            "jobClass": jobClass,
            "driverLicenseType": driverLicenseType,
            "nationality": nationality,
            "nationnality1": nationnality1,
            "educationId": educationId,
            "medicalInsType": medicalInsType,
            "incomeSource": incomeSource,
            "incomeSourceNote": incomeSourceNote,
            "annualIncome": annualIncome,
            "familyIncome": familyIncome,
            "premBudget": premBudget,
            "residentType": residentType,
            "taxPayerType": taxPayerType,
            "postalCode": postalCode,
            "provinceCode": provinceCode,
            "cityCode": cityCode,
            "districtCode": districtCode,
            "addrDetail": addrDetail,
            "firstName": '',
            "lastName": '',
            "birthArea": '',
            "birthCountry": '',
            "inputBirthProvince": '',
            "selectBirthProvince": '',
            "inputBirthCity": '',
            "selectBirthCity": '',
            "birthDetailadd": '',
            "v": [],
            "holderProposalNo": proposalNo
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 投保人告知接口
    def Holder_saveNotifyInfo(self, magicX, interface_ip, proposalId, result, customerId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personinsurance/saveNotifyInfo.do?magicX=" + magicX
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = '{"policyId":"' + proposalId + '","result":' + result + ',"customerid":"' + customerId + '"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response




        # 被保人信息录入接口
    def savePersonalInsured(self, magicX, interface_ip, proposalId, insucustomerId, policyChannel, insuRelation, insuName, insuGender, insuBirth, insuHeight, insuWeight, insuMarriage, insuCertiType, insuCertiCode, insuCertiValidEndDate, insuWorkCode, insujobClass, insuEmail, insuWorkType, insuCompany, insuMobile, insuNationatiy, insunationnality1, officeTel, insumedicalInsType, driverLicenseType, homeTel, incomeSource, incomeSourceNote, annualIncome, insuresidentType, taxPayerType, insueducationId, insuPostalCode, insuProvinceCode, insuCityCode, insuDistrictCode, insuAddrDetail):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/insured/savePersonalInsured.do?magicX=" + magicX
        headers = {
             "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "insuredProposalId": proposalId,
            "insucustomerId": insucustomerId,
            "insucustAddrId": '',
            "insupolicyChannelinfo": policyChannel,
            "subSalesChannel":'',
            "insuId":'',
            "insuRelation": insuRelation,
            "insuName": insuName,
            "preinsuName":'',
            "insuGender": insuGender,
            "preinsuGender": '',
            "insuBirth": insuBirth,
            "preinsuBirth": '',
            "insuHeight": insuHeight,
            "insuWeight": insuWeight,
            "insuMarriage": insuMarriage,
            "insuCertiType": insuCertiType,
            "preinsuCertiType":'',
            "insuCertiCode": insuCertiCode,
            "preinsuCertiCode":'',
            "insuCertiValidEndDate": insuCertiValidEndDate,
            "insurPerpetual": 'on',
            "insuWorkCode": insuWorkCode,
            "insujobClass": insujobClass,
            "insuEmail": insuEmail,
            "insuWorkType": insuWorkType,
            "insuCompany": insuCompany,
            "insuMobile": insuMobile,
            "insuNationatiy": insuNationatiy,
            "insunationnality1": insunationnality1,
            "officeTel": officeTel,
            "insumedicalInsType": insumedicalInsType,
            "driverLicenseType": driverLicenseType,
            "homeTel": homeTel,
            "incomeSource": incomeSource,
            "incomeSourceNote": incomeSourceNote,
            "annualIncome": annualIncome,
            "insuresidentType": insuresidentType,
            "taxPayerType": taxPayerType,
            "insueducationId": insueducationId,
            "insuPostalCode": insuPostalCode,
            "insuProvinceCode": insuProvinceCode,
            "insuCityCode": insuCityCode,
            "insuDistrictCode": insuDistrictCode,
            "insuAddrDetail": insuAddrDetail,
            "firstName":'',
            "lastName":'',
            "birthArea":'',
            "birthCountry":'',
            "inputBirthProvince":'',
            "selectBirthProvince":'',
            "inputBirthCity":'',
            "selectBirthCity":'',
            "birthDetailadd":'',
            "taxInfoList":[]
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 被保人查询接口
    def queryPersonalInsured(self, magicX, interface_ip, proposalNo):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/insured/queryPersonalInsured.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalNo": proposalNo
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 被保人告知接口
    def Insured_saveNotifyInfo(self, magicX, interface_ip, proposalId, result, insucustomerId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personinsurance/saveNotifyInfo.do?magicX=" + magicX
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = '{"policyId":"' + proposalId + '","result":' + result + ',"customerid":"' + insucustomerId + '"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 受益人信息录入接口
    def savePersonalBenf(self, magicX, interface_ip, proposalId, benfcustomerId,holderCustomerId,isLegal,benIsHolder,insured,benfRelation,benfType,benfOrder,benfPercent,benfName,benfGender,benfBirth,benfCertiType,benfCertiCode,benfCertiValidDate,benfMobile,officeTel,homeTel,benfNationatiy,taxPayerType,benfEmail,benfWorkCode,benfPostalCode,benfProvinceCode,benfCityCode,benfDistrictCode,benfAddrDetail):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/beneficiary/savePersonalBenf.do?magicX=" + magicX
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "beneficiaryId": '',
            "benfcustomerId": benfcustomerId,
            "selectedBenf": '',
            "holderCustomerId": holderCustomerId,
            "isLegal": isLegal,
            "benIsHolder": benIsHolder,
            "insured": insured,
            "insuId": '',
            "benfRelation": benfRelation,
            "benfType": benfType,
            "benfOrder": benfOrder,
            "benfPercent": benfPercent,
            "benfName": benfName,
            "benfGender": benfGender,
            "benfBirth": benfBirth,
            "benfCertiType": benfCertiType,
            "benfCertiCode": benfCertiCode,
            "benfCertiValidDate": benfCertiValidDate,
            "benefPerpetual": 'on',
            "benfMobile": benfMobile,
            "officeTel": officeTel,
            "homeTel": homeTel,
            "benfNationatiy": benfNationatiy,
            "taxPayerType": taxPayerType,
            "benfEmail": benfEmail,
            "benfWorkCode": benfWorkCode,
            "benfPostalCode": benfPostalCode,
            "benfProvinceCode": benfProvinceCode,
            "benfCityCode": benfCityCode,
            "benfDistrictCode": benfDistrictCode,
            "benfAddrDetail": benfAddrDetail,
            "benfClienteleName": '',
            "benfPapersType": '',
            "benfPapersNo": '',
            "proposalId": proposalId
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 收费方式信息录入接口
    def payModeSave(self, magicX, interface_ip, proposalId, payMode, name, bankCode, bankAccount, reserveMobile):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personinsurance/payModeSave.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "bankProposalId": proposalId,
            "payMode": payMode,
            "accountName": name,
            "bankCode": bankCode,
            "bankAccount": bankAccount,
            "reserveMobile": reserveMobile,
            "historyData_length": '5',
            "proposalId": proposalId,
            "accountId":''
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        #考虑docker应用重启报错
        if response['success'] == False and response['message'] == '银行账号查询失败，请稍候重试！':
            for i in range(0,30):
                response = API0000_diy().send_post(url, headers, data)
                if response['success'] == False and response['message'] == '银行账号查询失败，请稍候重试！':
                    time.sleep(5)
                else:
                    break
        return response


        # 险种信息录入接口
    def saveProductInfo(self, magicX, interface_ip, proposalId, productId, mainItemId, insuId, chargeMode, coverPeriodType, coveragePeriod, chargePeriodType, chargePeriod, units, sumAssured, stdPremBf, planFreq, payOption, bonusYearOrAge, headPartner, organId, renew):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/product/saveProductInfo.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
            #nb1
        if productId in ('10001','10002','10003','10004','10006','10007','10008','10009','10011','10012','10014','10017','10020','10021','10035','10036','10037','10038','10042','10044','10046','10047','10048','10049','10050','10053','10054','10063','10065','10071','10073','10076','10077','10078','10079','10080','10081','10082','10083','10097','10098','10099','10100','10106','10107','10108','10109','10110','10111','10112','10113','10118','10130','10131','10132','10133','10134','10135','10136','10138','10139','10140','10141','10142','10148','10159','10160','10161','10162','10163','10164','10166','10168','10169','10170','10171','10173','10174','10175','10177','10184','10189','10193','10194'):   #常规
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "units": units,  # 份数
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq": planFreq,  # 年金领取方式 或 年金/生存金领取方式
                "payOption": payOption,  # 年金使用方式
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "headPartner": headPartner,
                "proposalId": proposalId,
                "organId": organId
            }
            ##nb2
        elif productId in ('10200','10204','10207','10211','10214','10223','10224','10226','10129','10231','10240','10241','10245','10250','10251','10252','10253'):   #8230-10200、3267CB-10204、8233-10207、8234-10211、3275-10214、8237-10223、8233CA1-10224、8237CA1-10226、3257CA-10129、8239-10245、8240-10231、8242-10240、8242CA-10241
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb3
        elif productId in ('10143','10144','10203','10205','10206','10232','10246','10248'):   #1217-10143、1217CA-10144、3273-10203、3274CB-10248、8231-10205、8232-10206、1217CB-10232、8241-10246
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq":planFreq,  # 年金领取方式 或 年金/生存金领取方式
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb4
        elif productId in ('10128','xxx'):   #3260-10128
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq": planFreq,  # 年金领取方式 或 年金/生存金领取方式
                "payOption": payOption,  # 年金使用方式
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "_payOption": payOption,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb5
        elif productId in ('10202','xxx'):   #3274-10202
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq": planFreq,  # 年金领取方式 或 年金/生存金领取方式
                "payOption": payOption,  # 年金使用方式
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb6.1
        elif productId in ('10185', '10186', '10187','10188'):   #续保标识产品1
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "renew": renew,   #一年期产品是否续保。0-否；1-是
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb6.2
        elif productId in ('10227','10228'):   #1032-10227、1033-10228
            if productId == '10227':
                mainItemId = ''
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "renew": renew,   #一年期产品是否续保。0-否；1-是
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb6.3
        elif productId in ('10229','10230'):   #1034-10229、1035-10230
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "units": units,  #份数
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_units": units,
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "renew": renew,   #一年期产品是否续保。0-否；1-是
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb6.4
        elif productId in ('10119', 'xxx'):   #1027-10119
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "renew": renew,   #一年期产品是否续保。0-否；1-是
                "units": units,  #份数
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq": planFreq,
                "payOption": payOption,
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "headPartner": headPartner,
                "proposalId": proposalId,
                "organId": organId
            }
            #nb7
        elif productId in ('10209','xxx'):   #4.8003:君康畅行保交通综合意外伤害保险
            data = {
                "id": '',
                "productId": productId,
                "undefined": '1000000',
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "units": '1',  # 份数
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_units": '1',
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId,
                "liabIds": '1020901%2C1020902%2C1020903%2C1020904%2C1020905&amounts=200000%2C200000%2C200000%2C200000%2C1000000'   #保障责任
            }
            #nb8
        elif productId in ('10222', 'xxx'):   #8235:君康臻爱传家终身寿险
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_financeStatus": '823504',   #财务状况
                "headPartner": headPartner,
                "financeStatus": '823504',   #财务状况
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb9.1
        elif productId in ('10180','10181'):   #(1)9201-10180、9201CA-10181红利领取方式
            if productId == '10180':
                dividendChoice = '3'   #红利领取方式。1.现价领取；2-累计生息；3-转入万能账户
            elif productId == '10181':
                dividendChoice = '2'   #红利领取方式。1.现价领取；2-累计生息；3-转入万能账户
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "planFreq": planFreq,  #年金领取方式 或 年金/生存金领取方式
                "payOption": payOption,  # 年金使用方式
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "dividendChoice": dividendChoice,   #红利领取方式
                "proposalId": proposalId,
                "organId": organId
            }
            #nb9.2
        elif productId in ('10127','10172'):   #(2)3259-10127红利领取方式
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": sumAssured,  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_sumAssured": sumAssured,
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "dividendChoice": '2',   #红利领取方式
                "proposalId": proposalId,
                "organId": organId
            }
            #nb10
        elif productId in ('10233','10225'):   #3267CB1-10233
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }
            #nb11
        elif productId in ('10120','10121','10122'):   #2034B-10120、2034C-10121、2034D-10122
            if productId == '10120':
                liabIds = '1012001'   #保障责任
                amounts = '200000'   #保额
            elif productId == '10121':
                liabIds = '101210101,101210102'   #保障责任
                amounts = '200000,200000'   #保额
            elif productId == '10122':
                liabIds = '1012201'   #保障责任
                amounts = '200000'   #保额
            data = {
                "id": '',
                "productId": productId,
                "undefined":'200000',
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "units":'',
                "sumAssured":sumAssured,
                "stdPremBf": stdPremBf,  # 保费
                "planFreq":planFreq,
                "payOption":payOption,
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "headPartner": headPartner,
                "proposalId": proposalId,
                "organId": organId,
                "liabIds":liabIds,
                "amounts":amounts
            }
            ##nb12
        elif productId in ('10216','xxx'):   #8236-10216
            data = {
                "id": '',
                "productId": productId,
                "mainItemId": mainItemId,    # 主险id
                "insurantId": insuId,   # 被保人id
                "chargeMode": chargeMode,   # 缴费频率
                "coverPeriodType": coverPeriodType,   # 保障期间类型
                "coveragePeriod": coveragePeriod,   # 保障期间
                "chargePeriodType": chargePeriodType,  # 缴费期间类型
                "chargePeriod": chargePeriod,  # 缴费期间
                "sumAssured": '3000000',  # 保额
                "stdPremBf": stdPremBf,  # 保费
                "bonusYearOrAge": bonusYearOrAge,  # 领取年期或年龄
                "_stdPremBf": stdPremBf,
                "headPartner": headPartner,
                "giftFlag": '00',
                "proposalId": proposalId,
                "organId": organId
            }

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 新契约录入-特别说明录入接口
    def nb_specialExplain(self, magicX, interface_ip, proposalId, content):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/specialExplain/saveSpecialExplain.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = '{"policyId":"' + proposalId + '","content":"' + 'pass' + '","commentId":"","commentType":"01"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


        # 新契约录入提交前规则校验接口
    def valiDateRule(self, magicX, interface_ip, proposalId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/nBEntryRule/valiDateRule.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 新契约录入提交前规则结果接口
    def valiDateRule(self, magicX, interface_ip, proposalId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/nBEntryRule/valiDateRule.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 新契约录入提交接口
    def reviewProposal(self, magicX, interface_ip, proposalId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/proposal/reviewProposal.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 新契约复核updatePropStatus接口
    def updatePropStatus(self, magicX, interface_ip, proposalId):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/propProcess/updatePropStatus.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId,
            "oldStatus": '21',
            "newStatus": '22',
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 新契约复核-特被说明录入接口
    def fh_saveSpecialExplain(self, magicX, interface_ip, proposalId, content):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/specialExplain/saveSpecialExplain.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = '{"policyId":"' + proposalId + '","content":"' + 'pass' + '","commentId":"","commentType":"01"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


        # 新契约复核提交接口
    def confirmProposal(self, magicX, interface_ip, proposalId, personUW):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/personal/entry/proposal/confirmProposal.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "proposalId": proposalId,
            "personUW": personUW
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保单回执确认查询接口
    def query_policyReceiptConfirm(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/example/policyReceiptConfirm/q/policyReceiptConfirm.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=proposalNo&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=proposalType&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=applicationDate&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=proposalStatus&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=organId&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=policyChannel&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=&sSearch_8=&bRegex_8=false&bSearchable_8=false&bSortable_8=false&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&proposalNo=&policyNo=' + policyNo + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保单回执确认提交接口
    def policyReceiptConfirm(self, magicX, interface_ip, policyNo, proposalNo, callBackDate, callBackWay, appointVisitStartTime, appointVisitEndTime):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/example/policyReceiptConfirm/q/policyReceiptConfirmSaveAction.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'proposalNo=' + proposalNo + '&callBackDate=' + callBackDate + '&callBackWay=' + callBackWay + '&appointVisitStartTime=' + appointVisitStartTime + '&appointVisitEndTime=' + appointVisitEndTime + '&policyNo=' + policyNo

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 承保前撤单查询接口
    def queryPolicyCancle(self, magicX, interface_ip, proposalNo, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/policyCancle/queryPolicyCancle.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=proposalNo&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=policyType&sSearch_3=&bRegex_3=false&bSearchable_3=false&bSortable_3=false&mDataProp_4=applicationDate&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=policyStatus&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=organId&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=policyChannel&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=withDraw&sSearch_8=&bRegex_8=false&bSearchable_8=false&bSortable_8=false&mDataProp_9=&sSearch_9=&bRegex_9=false&bSearchable_9=false&bSortable_9=false&mDataProp_10=&sSearch_10=&bRegex_10=false&bSearchable_10=false&bSortable_10=false&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&proposalNo=' + proposalNo + '&policyNo=' + policyNo + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 承保前撤单提交接口
    def withDrawproposal(self, magicX, interface_ip, proposalNo, policyNo, withDraw, policyType, withdrawalReason):
        # 定义url
        url = "http://" + interface_ip + "/pos-commu/policyCancle/withDrawproposal.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'proposalNo=' + proposalNo + '&policyNo=' + policyNo + '&withDraw=' + withDraw + '&policyType=' + policyType + '&withdrawalReason=' + withdrawalReason + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



"""核保接口类"""
class API0000_uw():
        # 核保共享池-新契约查询接口
    def uwPolicyQuery_nb(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uWShrpool/uwPolicyQuery.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=mainId&sSearch_0=&bRegex_0=false&bSearchable_0=false&bSortable_0=false&mDataProp_1=%23&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=busiNo&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=policyNo&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=proposalNo&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=uwStatus&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=policyType&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=policyChannel&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=subSalesChannel&sSearch_8=&bRegex_8=false&bSearchable_8=false&bSortable_8=false&mDataProp_9=noticeStatus&sSearch_9=&bRegex_9=false&bSearchable_9=false&bSortable_9=false&mDataProp_10=cusName&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=uwLevel&sSearch_11=&bRegex_11=false&bSearchable_11=false&bSortable_11=false&mDataProp_12=organId&sSearch_12=&bRegex_12=false&bSearchable_12=false&bSortable_12=false&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&params={"policyNo":"' + policyNo + '","policyType":"","proposalNo":"","uwStatus":"","salesManCode":"","organId":"","policyChannel":"","salesChannel":"","subSalesChannel":"","startDate":"","endDate":"","uwLevel":"","cusName":"","certiType":"","certiNo":"","uwSource":"01","sourcePage":"shrPool"}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 核保共享池-保全查询接口
    def uwPolicyQuery_pa(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uWShrpool/uwPolicyQuery.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=mainId&sSearch_0=&bRegex_0=false&bSearchable_0=false&bSortable_0=false&mDataProp_1=%23&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=busiNo&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=policyNo&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=serviceId&sSearch_4=&bRegex_4=false&bSearchable_4=false&bSortable_4=false&mDataProp_5=subSalesChannel&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=uwStatus&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=noticeStatus&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=organId&sSearch_8=&bRegex_8=false&bSearchable_8=false&bSortable_8=false&mDataProp_9=busiTime&sSearch_9=&bRegex_9=false&bSearchable_9=false&bSortable_9=false&mDataProp_10=submitDate&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&params={"uwStatus":"","serviceId":"","busiNo":"","subSalesChannel":"","policyNo":"' + policyNo + '","proposalNo":"","organId":"","startDate":"","endDate":"","uwSource":"02","sourcePage":"shrPool"}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 核保-新契约锁定任务接口
    def uwLock_nb(self, magicX, interface_ip, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwProposal/uwLock.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "uwId": mainId,
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



        # 核保-保全任务领取提交接口
    def uwLock_pa(self, magicX, interface_ip, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwLock/processLockUwMain.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "uwMainIds": mainId,
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全-查询产品承保核保决定接口
    def queryLastProductUnderwriting(self, magicX, interface_ip, policyNo, productId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/personalUW/queryLastProductUnderwriting.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'policyNo=' + policyNo + '&productId=' + str(productId) + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 核保获取保单信息接口
    def uw_queryPersonalPolicy(self, magicX, interface_ip, uwId, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/personalUW/queryPersonalPolicy.do"

        # 组装请求参数
        data = {
            "uwMainId": mainId,
            "uwId": uwId,
            "magicX": magicX
        }
        # 调用发送get请求函数
        response = API0000_diy().send_get(url, data)
        return response



    # 核保限额保存接口
    def saveUwProductLimit(self, magicX, interface_ip, uwItemId, limitsumAssured, limitunits):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/premLimit/saveUwProductLimit.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "uwItemId": uwItemId,
            "sumAssured": limitsumAssured,
            "units": limitunits,
            "premCalWay": '2',
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        others = response['others']
        stdPrem = others['stdPrem']
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message, stdPrem





    # 新契约-险种层提交核保决定
    def updateUwDecision_nb(self, magicX, interface_ip, uwItemId, uwId, uwDecision, decisionComment, uwExceptTxt, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwDecision/updateUwDecision.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'params={"uwId":"' + uwId + '","uwItemId":"' + uwItemId + '","uwDecision":"' + uwDecision + '","decisionComment":"pass",' + '"uwSource":"01","benefitLevel":0,"uwExceptTxt":"' + '' + '","policyNo":"' + policyNo + '"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全-险种层提交核保决定
    def updateUwDecision_pa(self, magicX, interface_ip, uwItemId, uwId, uwDecision, decisionComment, uwItemIdProp):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwDecision/updateUwDecision.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'params={"uwId":"' + uwId + '","uwItemId":"' + uwItemId + '","uwDecision":"' + uwDecision + '","decisionComment":"' + decisionComment + '",' + '"benefitLevel":0,"delUwExtraPrems":[],"uwItemIdProp":"' + uwItemIdProp + '","uwSource":"02","uwExceptTxt":""}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 核保特别说明录入接口
    def commitUWComment(self, magicX, interface_ip, uwId, proposalId, content):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uWShrpool/commitUWComment.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'params={"commentId":"","uwId":"' + uwId + '","policyId":"' + proposalId + '","commentType":"03","content":"pass"}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


        # 契约核保结论提交接口
    def commitUwComplete(self, magicX, interface_ip, uwId, uwPolicyDecision):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwComplete/s/commitUwComplete.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = '{"uwId":"' + uwId + '","uwPolicyDecision":"' + uwPolicyDecision + '","magicX":"' + magicX + '"}'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 保单保全结论判定-提交接口
    def savePolicyDecision(self, magicX, interface_ip, mainId, uwId, decisionCode, propUwId, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwDecision/cs/savePolicyDecision.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'uwMainId=' + mainId + '&uwId=' + uwId + '&decisionCode=' + decisionCode + '&propUwId=' + propUwId + '&serviceCode=' + serviceCode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 保单保全结论判定层-保单核保特别说明接口
    def commitUWComment_pa1(self, magicX, interface_ip, mainId, uwId, policyId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uWShrpool/commitUWComment.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'params={"commentId":"","uwMainId":"' + mainId + '","uwId":"' + uwId + '","policyId":"' + policyId + '",' + '"commentType":"03","content":"pass"}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 核保-保全核保结论提交接口
    def saveBusiDecision(self, magicX, interface_ip, mainId, decisionCode, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwDecision/cs/saveBusiDecision.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'uwMainId=' + mainId + '&decisionCode=' + decisionCode + '&serviceCode=' + serviceCode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 保全核保层-保单核保特别说明接口
    def commitUWComment_pa2(self, magicX, interface_ip, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uWShrpool/commitUWComment.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'params={"commentId":"","uwMainId":"' + mainId +  '","commentType":"03","content":"pass"}&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


        # 保全核保解锁接口
    def uwUnLock(self, magicX, interface_ip, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwProposal/uwUnLock.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'uwId=' + mainId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response

        # 保全核保提交接口
    def commitUwComplete_uwComplete(self, magicX, interface_ip, mainId):
        # 定义url
        url = "http://" + interface_ip + "/uw2-commu/uwComplete/endor/commitUwComplete.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "uwId": mainId
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



"""保全保单接口类"""
class API0000_pa():
    # 保全申请接口_保单层 
    def saveCsApplylication(self, magicX, interface_ip, applyTime, applyWay, applyType, serviceCode, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csApply/saveCsApplylication.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'applyTime=' + applyTime + '&applyWay=' + applyWay + '&applyType=' + applyType + '&serviceCode=' + serviceCode + '&applyTime=' + applyTime + '&policyNo=' + policyNo + '&policyCodes=&customerId=&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        """
        #考虑docker应用重启报错
        if response['success'] == False and response['message'] == '未查询到相关保单':
            for i in range(0,30):
                response = API0000_diy().send_post(url, headers, data)
                if response['success'] == False and response['message'] == '未查询到相关保单':
                    time.sleep(5)
                else:
                    break
        """
        #考虑：1.生成保全受理号重复2.提示java.lang.RuntimeException
        Duplicate_msg = '### Error updating database.  Cause: com.mysql.jdbc.exceptions.jdbc4.MySQLIntegrityConstraintViolationException'
        RuntimeException_msg = 'java.lang.RuntimeException'
        if (response['success'] == False and Duplicate_msg in response['message']) or (RuntimeException_msg in response['message']):
            for i in range(0,30):
                response = API0000_diy().send_post(url, headers, data)
                if response['success'] == False and Duplicate_msg in response['message']:
                    time.sleep(3)
                else:
                    break
        return response


    # 保全申请接口_客户层   
    def saveCsApplylication_CD(self, magicX, interface_ip, applyTime, applyWay, applyType, serviceCode, policyNo, policyCodes, customerId):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csApply/saveCsApplylication.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'applyTime=' + applyTime + '&applyWay=' + applyWay + '&applyType=' + applyType + '&serviceCode=' + serviceCode + '&applyTime=' + applyTime + '&policyNo=' + policyNo + '&policyCodes=' + policyCodes + '&customerId=' + customerId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-查询保单客户角色接口(客户层) 
    def queryPolicyRoleCustomer(self, magicX, interface_ip, policyCode, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/customerInfo/queryPolicyRoleCustomer.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=customerId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=name&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=gender&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=birthday&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=certiType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=certiCode&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=roleType&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=jobCateId&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=jobCateId&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=height&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=weight&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=income&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyCode=' + policyCode +  '&serviceCode=' + serviceCode + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-查询被选择改变的客户信息(客户层)
    def queryCustomerChangeSelected(self, magicX, interface_ip, policyNo, partitionIndi, applyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/query/queryCustomerChangeSelected.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=name&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=gender&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=birthday&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=certiType&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=certiCode&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=jobCateId&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=jobCateId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=income&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=incomeSource&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=incomeSourceNote&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=customerId&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&partitionIndi=' + partitionIndi + '&applyNo=' + applyNo + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-客户重要资料修改接口(客户层)
    def updateCustomerImpWithSimilar(self, magicX, interface_ip, name, applyNo, changeId, customerId, gender, birthday, certiType, certiCode, jobCateId, jobClass, income, familyIncome, incomeSource, incomeSourceNote, companyName):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/query/updateCustomerImpWithSimilar.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0"
        }
        # 组装请求参数
        data = 'name=' + name + '&applyNo=' + applyNo + '&changeId=' + changeId + '&customerId=' + customerId + '&isOnlyBeneRole=Y&gender=' + gender + '&birthday=' + birthday +  '&certiType=' + certiType + '&certiCode=' + certiCode + '&jobCateId=' + jobCateId + '&jobClass=' + jobClass + '&income=' + income + '&familyIncome=' + familyIncome + '&incomeSource=' + incomeSource + '&incomeSourceNote=' + incomeSourceNote + '&companyName=' + companyName + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-查询保费变化接口(客户层)
    def queryMsgChgPrem(self, magicX, interface_ip, applyNo, customerId, name, gender, birthday, certiType, certiCode, jobCateId):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/query/queryMsgChgPrem.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=productId&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=productEftDate&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=payPeriod&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=payYear&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=payFreq&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=amountBfChg&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=premBfChg&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=amountAfChg&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=premAfChg&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=premDif&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&applyNo=' + applyNo + '&customerId=' + customerId + '&name=' + name + '&gender=' + gender + '&birthday=' + birthday +'&certiType=' + certiType + '&certiCode=' + certiCode + '&height=&weight=&jobCateId=' + jobCateId + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-查询客户账号接口(客户层)
    def customerAccQuery(self, magicX, interface_ip, customerId, changeId):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/query/customerAccQuery.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "customerId": customerId,
            "changeId": changeId
		}
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-添加或更改客户账号信息接口(客户层)
    def addOrUpdateCustImpChgAccount(self, magicX, interface_ip, changeId, accountId, accountType, payMode, accountName, bankCode, bankAccount, reserveMobile):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/query/addOrUpdateCustImpChgAccount.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "policyId": '',
            "changeId": changeId,
            "accountId": accountId,
            "accountType": accountType,
            "payerCustId": '',
            "payMode": payMode,
            "accountName": accountName,
            "bankCode": bankCode,
            "bankAccount": bankAccount,
            "reserveMobile": reserveMobile
		}
        data = json.dumps(data)
        
        #data = "{'policyId':'','changeId':'" + changeId + "','accountId':'" + accountId + "','accountType':'" + accountType + "','payerCustId':'','payMode':'" + payMode + "','accountName':'" + accountName + "','bankCode':'" + bankCode + "','bankAccount':'" + bankAccount + "','reserveMobile':'" + reserveMobile + "'}"
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-产品信息详细（犹退、退保、贷款）
    def queryCsPolicyProductBeforeChange(self, magicX, interface_ip, policyNo, changeId, partitionIndi, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=presentPrice&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=investAmount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=loanAccountAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=loanRate&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=loanAccountLimit&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=&serviceCode=' + serviceCode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response




    # 保全录入-产品信息详细（追加保费）  
    def queryCsPolicyProductBeforeChange2(self, magicX, interface_ip, policyNo, changeId, partitionIndi, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange2.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=7&sColumns=%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=presentPrice&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=investAmount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=&serviceCode=' + serviceCode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response





    # 保全录入-产品信息详细（保单复效）   
    def queryCsPolicyProduct(self, magicX, interface_ip, policyNo, changeId, partitionIndi, applyTime):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csRevival/queryCsPolicyProduct.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=16&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=liabilityState&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=chargePeriodType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=chargePeriod&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=coveragePeriod&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=totalPremAf&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=payDueDate&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=revivalAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=revivalInterestAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=loanBalance&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=loanInterest&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=zidianAmount&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=zidianInterestAmount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=extraPrem&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&applyTime=' + applyTime + '&revivalInterestAmount=&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-产品信息详细(减保)
    def queryCsPolicyProduct1(self, magicX, interface_ip, policyNo, changeId, partitionIndi):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/saChange/queryCsPolicyProduct.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=productName&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=coverageEffectDate&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=sumAssured&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=totalPremAf&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=presentPrice&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=adjustAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=changeSumAssured&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=changeTotalPremAf&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=changePresentPrice&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=refundRenewPrem&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response




    # 保全录入-客户受理信息接口   
    def findCsApplylication(self, magicX, interface_ip, env_name, applyTime, applyWay, applyType, serviceCode, policyNo, partitionIndi, changeId, applyNo, subSalesChannel,policyDeliveryMode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/findCsApplylication.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'applyTime=' + applyTime + '&applyWay=' + applyWay + '&applyType=' + applyType + '&serviceCode=' + serviceCode + '&policyNo=' + policyNo + '&partitionIndi=' +partitionIndi + '&changeId=' + changeId + '&applyNo=' + applyNo + '&subSalesChannel=' + subSalesChannel + '&policyDeliveryMode=' + policyDeliveryMode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response
        

    # 保全录入-支付方式及支付账户查询接口   
    def queryAccountList(self, magicX, interface_ip, changeId, partitionIndi, policyId, payMode, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/account/queryAccountList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=7&sColumns=%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=accountId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=accountName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=bankCode&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=bankAccount&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=accountType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=reserveMobile&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=accountId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyId=' + policyId + '&payMode=' + payMode + '&serviceCode=' + serviceCode + '&onbehalf=0&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-增加支付方式及支付账户接口   
    def insertAccount(self, magicX, interface_ip, changeId, partitionIndi, policyId, payMode, serviceCode, accountName, bankCode, bankAccount, reserveMobile):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/account/insertAccount.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyId=' + policyId + '&accountId=&payMode=' + payMode + '&onbehalf=0&serviceCode=' + serviceCode + '&targetMobile=&canOper=&accountName=' + accountName + '&bankCode=' + bankCode + '&bankAccount=' + bankAccount + '&reserveMobile=' + reserveMobile
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-还款账号信息接口   
    def queryCsNeedRepayLoanAccount(self, magicX, interface_ip, policyNo, changeId, partitionIndi, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/loanRepay/queryCsNeedRepayLoanAccount.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=14&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=investId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=false&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=false&mDataProp_2=productName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=false&mDataProp_3=liabilityState&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=false&mDataProp_4=productPresentPrice&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=false&mDataProp_5=productInvestAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=false&mDataProp_6=loanChangeId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=false&mDataProp_7=subFundCode&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=false&mDataProp_8=loanPrincipal&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=false&mDataProp_9=interestCapital&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanInterest&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=rate&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=principalInterest&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=policyPartialRmtAmount&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=&serviceCode=' + serviceCode + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    #保全录入-保单减保计算接口
    def saveProductSaChange(self, magicX, interface_ip,resultData):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/saChange/saveProductSaChange.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = resultData
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response  


    #保全录入-查询产品挂失信息接口
    def querySuspendInfo(self, magicX, interface_ip, applyTime, applyWay, applyType, serviceCode, policyNo, partitionIndi, changeId, applyNo, subSalesChannel, policyDeliveryMode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/suspend/querySuspendInfo.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'applyTime=' + applyTime + '&applyWay=' + applyWay + '&applyType=' + applyType + '&serviceCode=' + serviceCode + '&policyNo=' + policyNo + '&partitionIndi=' + partitionIndi + '&changeId=' + changeId + '&applyNo=' + applyNo + '&subSalesChannel=' + subSalesChannel + '&policyDeliveryMode=' + policyDeliveryMode + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 


    #保全录入-保单挂失信息保存接口
    def saveSuspendInfo(self, magicX, interface_ip, validationFlag, changeId, partitionIndi, policyId, suspendCause):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/suspend/saveSuspendInfo.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'validationFlag=' + validationFlag + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyId=' + policyId + '&suspendCause=' + suspendCause + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response     


    #保全录入-保单挂失信息保存接口
    def saveReissueInfo(self, magicX, interface_ip, validationFlag, changeId, partitionIndi, policyNo, reissueReason, reissueType):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/reissue/saveReissueInfo.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'validationFlag=' + validationFlag + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyNo=' + policyNo + '&reissueTypeSelect=&reissueReason=' + reissueReason + '&reissueType=' + reissueType + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response  


    # 保全录入-保存协议退保金额接口  
    def updateAgreementAmountByProduct(self, magicX, interface_ip, data):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/updateAgreementAmountByProduct.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-保全回退列表接口 
    def csReversalChangeList(self, magicX, interface_ip, policyNo, changeId):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csReversal/csReversalChangeList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=applyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=serviceId&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=validateDate&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=arapDesc&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=totalFeeAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=policyStatus&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=payMode&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=accountName&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=bankCode&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=bankAccount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=taskStatus&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=currTotalFeeAmount&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response




    # 保全录入-减额交清信息查询接口  
    def queryReductionClear(self, magicX, interface_ip, changeId, partitionIndi, policyNo, validateDate):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/reductionClear/queryReductionClear.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=actuaryCode&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=payDueDate&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=presentPrice&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=loanAccountAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=changeSumAssured&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=changePresentPrice&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyNo=' + policyNo + '&validateDate=' + validateDate + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response




    # 保全录入-续保方式变更查询接口
    def queryRenewalInsChg(self, magicX, interface_ip, changeId, partitionIndi, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/renewalChg/queryRenewalInsChg.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=7&sColumns=%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=productName&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=liabilityState&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=sumAssured&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=totalPremAf&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=csIssueDate&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=csExpiryDate&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=renew&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&policyNo=' + policyNo + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-续保方式变更保存接口
    def saveRenewalChg(self, magicX, interface_ip, resultData):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/renewalChg/saveRenewalChg.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }        # 组装请求参数
        data = resultData
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-查询产品满期列表接口
    def queryExpiryProductList(self, magicX, interface_ip, policyNo, changeId, partitionIndi,validateDate):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/drawExpiryPayment/queryExpiryProductList.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=productName&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=liabilityState&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=sumAssured&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=totalPremAf&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=expiryAmount&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=loanAccountAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=cbSbAccount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=bonusAccount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=expiryPayAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=' + validateDate
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-查询产品满期列表接口
    def initBonusMethod(self, magicX, interface_ip, applyTime, applyWay, applyType, serviceCode, policyNo, partitionIndi, changeId, applyNo, subSalesChannel, policyDeliveryMode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/bonus/initBonusMethod.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "applyTime":applyTime,
            "applyWay":applyWay,
            "applyType":applyType,
            "serviceCode":serviceCode,
            "policyNo":policyNo,
            "partitionIndi":partitionIndi,
            "changeId":changeId,
            "applyNo":applyNo,
            "subSalesChannel":subSalesChannel,
            "policyDeliveryMode":policyDeliveryMode,
            "magicX":magicX
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-贷款提交接口   
    def entryConfirm_loan(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, loanApplyAmount, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "loanApplyAmount": loanApplyAmount,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-贷款还款提交接口   
    def entryConfirm_loan_repayment(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, csLoanAccounts, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "csLoanAccounts": csLoanAccounts,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-犹退/退保提交接口   
    def entryConfirm_tb(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, contactNumber, telephone, flag, csProductCals, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "changeDetailModel": {
                "contactNumber": contactNumber,
                "telephone": telephone
            },
            "flag": flag,
            "csProductCals": csProductCals,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-协议退保提交接口   
    def entryConfirm_agreement_tb(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, contactNumber, telephone, sCsArgtSurrender, csProductCals, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "changeDetailModel": {
                "contactNumber": contactNumber,
                "telephone": telephone
            },
            "sCsArgtSurrender":sCsArgtSurrender,
            "csProductCals": csProductCals,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-公司解约提交接口   
    def entryConfirm_company_tb(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, contactNumber, telephone, cancelType, cancelReason, cancelComment, csProductCals, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "changeDetailModel": {
                "contactNumber": contactNumber,
                "telephone": telephone
            },
            "sCsCancelContract": {
                "changeId": changeId,
                "partitionIndi": partitionIndi,
                "cancelType" : cancelType,
                "cancelReason": cancelReason,
                "cancelComment":cancelComment
            },
            "csProductCals": csProductCals,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-客户重要资料变更提交接口   
    def entryConfirm_CD(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, policyProductPremChgs, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "policyProductPremChgs": policyProductPremChgs,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-复效提交接口   
    def entryConfirm_revival(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, csProductCals, scsRevivalInterestFree, changeDetailModel, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "csProductCals": csProductCals,
            "scsRevivalInterestFree": scsRevivalInterestFree,   #是否免息复效。无免息:{}；存在免息:{"interestFreeReason":"1","feeBearWay":"1","feeBearDept":"1"}
            "changeDetailModel": changeDetailModel,   #是否有异常告知。有:{"abnormalTold":"0"}；无:{"abnormalTold":"1"}
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-(追加保费/减保）提交接口   
    def entryConfirm_additional_premium(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, csProductCals, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "csProductCals": csProductCals,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
		}
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-保单挂失/保单解挂/保单补发-提交接口   
    def entryConfirm_suspend(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-保单质押贷款登记/解除保单质押贷款登记-提交接口   
    def entryConfirm_suspend_loan(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, appoint, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "appoint": appoint,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入-保全回退-提交接口   
    def entryConfirm_csReversal(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "payMode": payMode,
            "accountId":accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入-保全通知书寄送标识变更-提交接口   
    def entryConfirm_noticeConfigl(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, noticeConfigList, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "noticeConfigList": noticeConfigList,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入（减额交清）提交接口 
    def entryConfirm_reductionclear(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, csProductCals, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "csProductCals": csProductCals,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        #print data
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入（续保方式变更）提交接口 
    def entryConfirm_RenewalChg(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全录入（保费自垫状态变更）提交接口 
    def entryConfirm_autoPaddingPrem(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, autoPaddingPrem, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "autoPaddingPrem": autoPaddingPrem,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入（满期领取）提交接口 
    def entryConfirm_drawExpiryPayment(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全录入（红利领取方式变更）提交接口 
    def entryConfirm_initBonusMethod(self, magicX, interface_ip, changeId, applyNo, partitionIndi, policyId, policyNo, subSalesChannel, applyTime, serviceCode, bonusPayMode, listId, payMode, accountId, customerInfoList):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEntryConfirm/entryConfirm.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "changeId": changeId,
            "applyNo": applyNo,
            "partitionIndi": partitionIndi,
            "policyId": policyId,
            "policyNo": policyNo,
            "subSalesChannel": subSalesChannel,
            "applyTime": applyTime,
            "serviceCode": serviceCode,
            "bonusMethod":{
                "bonusPayMode":bonusPayMode,
                "relatedPolicyNo":"",
                "listId":listId
            },
            "payMode": payMode,
            "accountId": accountId,
            "customerInfoList": customerInfoList,
            "certiValidEndDate": ''
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 保全复核-保全公共工作池查询接口  
    def queryCsInfoList(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csApply/queryCsInfoList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=changeId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=%23&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=changeStatus&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=serviceCode&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=applyNo&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=policyNo&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=organId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=holdName&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=applyTime&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=finishDate&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=totalAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=totalPremium&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=applyWay&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&fromPage=workPool&policyNo=' + policyNo + '&serviceCode=-1&excludeServiceCodeList=90052%2C30203%2C90058&applyWay=-1&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全复核-工作池领取任务接口  
    def processLockCsChange(self, magicX, interface_ip, changeIds):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csLock/processLockCsChange.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'changeIds=' + changeIds + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全复核提交接口 
    def csEffect(self, magicX, interface_ip, changeId, applyNo, serviceCode, partitionIndi, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEffect/csEffect.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'changeId=' + changeId + '&businessNo=' + applyNo + '&serviceCode=' + serviceCode + '&partitionIndi=' + partitionIndi + '&policyNo=' + policyNo + '&applyNo=' + applyNo + '&sendLetterFlag=Y&sendSmsFlag=Y&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response
 

    # 贷款试算接口
    def trial_loan(self, env_name, magicX, interface_ip, policyNo, calDate):
        if env_name != '预生产':
            # 定义url
            url = "http://" + interface_ip + "/cs-commu/loan/trial.do?magicX=" + magicX
            # 定义headers
            headers = {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            }
            # 组装请求参数
            data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=holderName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=insuredName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=sumAssured&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=totalPremAf&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=presentPrice&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=investAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=loanAccountAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanRate&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=loanAccountLimit&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyCode=' + policyNo + '&calDate=' + calDate
        else:
            # 定义url
            url = "http://" + interface_ip + "/cs-commu/loan/trial.do"
            # 定义headers
            headers = {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            }
            # 组装请求参数
            data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=holderName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=insuredName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=sumAssured&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=totalPremAf&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=presentPrice&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=investAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=loanAccountAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanRate&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=loanAccountLimit&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyCode=' + policyNo + '&calDate=' + calDate  + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 退保试算接口
    def trial_tb(self, env_name, magicX, interface_ip, policyNo, validateDate):
        if env_name != '预生产':
            # 定义url
            url = "http://" + interface_ip + "/cs-commu/surrender/policyTrial.do?magicX=" + magicX
            # 定义headers
            headers = {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            }
            # 组装请求参数
            data = 'sEcho=1&iColumns=17&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=holderName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=insuredName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=sumAssured&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=totalPremAf&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=presentPrice&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=investAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=chargeAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanAccountAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=refundRenewPrem&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=cbSbAccount&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=riskChargeFee&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=bonusAccount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=paidBonus&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&mDataProp_16=surrenderAmount&sSearch_16=&bRegex_16=false&bSearchable_16=true&bSortable_16=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&validateDate=' + validateDate
        else:
            # 定义url
            url = "http://" + interface_ip + "/cs-commu/surrender/policyTrial.do"
            # 定义headers
            headers = {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            }
            # 组装请求参数
            data = 'sEcho=1&iColumns=17&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=holderName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=insuredName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=sumAssured&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=totalPremAf&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=presentPrice&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=investAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=chargeAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanAccountAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=refundRenewPrem&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=cbSbAccount&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=riskChargeFee&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=bonusAccount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=paidBonus&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&mDataProp_16=surrenderAmount&sSearch_16=&bRegex_16=false&bSearchable_16=true&bSortable_16=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&validateDate=' + validateDate + '&magicX=' + magicX        
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 保全撤销查询接口    
    def csCancle_queryCsInfoList(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csApply/queryCsInfoList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=changeStatus&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=serviceCode&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=applyNo&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=holdName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=certiType&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=certiNo&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=applyTime&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=entryTime&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=auditTime&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=applyWay&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&applyNo=&fromPage=revoke&excludeServiceCodeList=90052%2C30203%2C90058&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response
 

    # 保全撤销接口  
    def csCancle(self, magicX, interface_ip, businessNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csEffect/csCancle.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'businessNo=' + businessNo + '&noPassReasonCode=20&noPassReasonComment=&changeStatus=04&operStep=41&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 批处理定义-启用/关闭批处理任务接口  
    def saveBatchJob(self, magicX, interface_ip, batchSystem, jobId, isEnable, jobName):
        # 定义url
        url = "http://" + interface_ip + "/batch-admin/batch/saveBatchJob.do" 
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'batchSystem=' + batchSystem + '&jobId=' + jobId + '&isEnable=' + isEnable + '&jobName=' + jobName + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response
        
 
    # 批处理定义-提交批处理任务接口 
    def saveBatchSubmit(self, magicX, interface_ip, batchSystem, jobId, policyNo, processDate):
        # 定义url
        url = "http://" + interface_ip + "/batch-admin/batch/saveBatchSubmit.do" 
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'jobId=' + jobId + '&batchSystem=' + batchSystem + '&paramCode=policyNo&paramValue=' + policyNo + '&paramCode=processDate&paramValue=' + processDate + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询接口 
    def queryCsInfoList1(self, magicX, interface_ip, applyNo):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/preserve/queryCsInfoList.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=serviceCode&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=applyNo&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=holdName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=certiType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=certiNo&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=applyTime&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=auditTime&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=applyType&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=validateDate&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=finishDate&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=applyWay&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=hasImage&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&applyNo=' + applyNo + '&serviceCode='
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-保单贷款-保单险种信息列表接口
    def queryCsPolicyProductBeforeChange3(self, magicX, interface_ip, policyNo, changeId, partitionIndi, validateDate, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=presentPrice&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=investAmount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=loanAccountAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=loanRate&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=loanAccountLimit&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=' + validateDate + '&serviceCode=' + serviceCode
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-公司解约-查询解约信息接口
    def queryCancelContract(self, magicX, interface_ip, changeId, partitionIndi):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCancelContract.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {"changeId":changeId, "partitionIndi":partitionIndi}
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-公司解约-保单险种信息列表接口
    def queryCsPolicyProductBeforeChange4(self, magicX, interface_ip, policyNo, changeId, partitionIndi, cancelContractType):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=18&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=productId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=productName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=liabilityState&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=sumAssured&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=nbPremium&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=totalPremium&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=presentPrice&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=investAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=chargeAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanAccountAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=refundRenewPrem&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=cbSbAccount&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=riskChargeFee&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=bonusAccount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=paidBonus&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&mDataProp_16=paidAmount&sSearch_16=&bRegex_16=false&bSearchable_16=true&bSortable_16=true&mDataProp_17=surrenderAmount&sSearch_17=&bRegex_17=false&bSearchable_17=true&bSortable_17=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=&cancelContractType=' + cancelContractType + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 


    # 保全查询-协议退保-保单险种信息列表接口
    def queryCsPolicyProductBeforeChange5(self, magicX, interface_ip, policyNo, changeId, partitionIndi, validateDate):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=16&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=productId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=presentPrice&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=investAmount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=chargeAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=loanAccountAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=refundRenewPrem&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=riskChargeFee&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=bonusAccount&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=paidBonus&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=surrenderAmount&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=agreementSurrenderAmount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=agreementSupplymentAmount&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=' + validateDate
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-退保/取消险种-保单险种信息列表接口
    def queryCsPolicyProductBeforeChange6(self, magicX, interface_ip, policyNo, changeId, partitionIndi, validateDate):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/surrender/queryCsPolicyProductBeforeChange.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=15&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=totalPremAf&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=presentPrice&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=investAmount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=chargeAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=loanAccountAmount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=refundRenewPrem&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=cbSbAccount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=riskChargeFee&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=bonusAccount&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=paidBonus&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=surrenderAmount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo +'&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=' + validateDate
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-贷款还款-保单险种信息列表接口
    def queryCsNeedRepayLoanAccount1(self, magicX, interface_ip, policyNo, changeId, partitionIndi, validateDate, serviceCode):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/loanRepay/queryCsNeedRepayLoanAccount.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=14&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=investId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=false&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=false&mDataProp_2=productName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=false&mDataProp_3=liabilityState&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=false&mDataProp_4=productPresentPrice&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=false&mDataProp_5=productInvestAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=false&mDataProp_6=loanChangeId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=false&mDataProp_7=subFundCode&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=false&mDataProp_8=loanPrincipal&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=false&mDataProp_9=interestCapital&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=loanInterest&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=rate&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=principalInterest&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=policyPartialRmtAmount&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&validateDate=' + validateDate + '&serviceCode=' + serviceCode
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response 



    # 保全查询-复效-保单险种信息列表接口   
    def queryCsPolicyProduct2(self, magicX, interface_ip, policyNo, changeId, partitionIndi, validateDate):
        # 定义url
        url = "http://" + interface_ip + "/cs-commu/csRevival/queryCsPolicyProduct.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=16&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=liabilityState&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=liabilityState&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=sumAssured&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=chargePeriodType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=chargePeriod&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=coveragePeriod&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=totalPremAf&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=payDueDate&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=revivalAmount&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=revivalInterestAmount&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=loanBalance&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=loanInterest&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&mDataProp_13=zidianAmount&sSearch_13=&bRegex_13=false&bSearchable_13=true&bSortable_13=true&mDataProp_14=zidianInterestAmount&sSearch_14=&bRegex_14=false&bSearchable_14=true&bSortable_14=true&mDataProp_15=extraPrem&sSearch_15=&bRegex_15=false&bSearchable_15=true&bSortable_15=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&changeId=' + changeId + '&partitionIndi=' + partitionIndi + '&applyTime=' + validateDate + '&revivalInterestAmount='
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



"""收付费接口类"""
class API0000_payment():
    # 柜面收付费查询接口
    def FeeQuery(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/bcp-commu/bcpCash/FeeQuery.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=orgainId&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=accoName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=payOrCollect&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=serviceId&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=amount&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=insertTime&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&orgainId=10000&insertTime=&payOrCollect=&policyNo=' + policyNo + '&serviceId=&userNo=' + magicX + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        for i in range(0,3):
            if isinstance(response,dict) != True or response.has_key('resultData') != True:
                response = API0000_diy().send_post(url, headers, data)
                time.slee(1)
            else:
                break
        resultData = response['resultData']
        len_resultData = len(resultData)
        if len_resultData > 0:
            # 返回是否成功，以及错误信息
            is_success = response['isSuccess']
            if is_success:
                message = ''
            else:
                message = response['message']
        else:
            is_success = False
            message = ''
        return is_success, message, resultData, len_resultData


    # 柜面收付费保存接口
    def CashApproveInfoSave(self, magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, cashOrgId, internalBankName):
        offsetSumId = resultData[0]['offsetSumId']
        accoName = resultData[0]['accoName']
        certiType = resultData[0]['certiType']
        certiCode = resultData[0]['certiCode']
        accoNo = resultData[0]['accoNo']
        bankCode = resultData[0]['bankCode']
        # 定义url
        url = "http://" + interface_ip + "/bcp-commu/bcpCash/CashApproveInfoSave.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "custName": accoName,
            "custCertiType": certiType,
            "custCertiCode": certiCode,
            "accountNo": accoNo,
            "bak3": 'pass',
            "bankEndorse": bankEndorse,
            "feeConfirmDate": feeConfirmDate,
            "internalAccount": internalAccount,
            "internalBankCode": internalBankCode,
            "feeChannel": feeChannel,
            "offsetSumId": offsetSumId,
            "bak2": magicX,
            "bankCode": bankCode,
            "feeStatus": '',
            "cashOrgId": cashOrgId,
            "internalBankName": internalBankName,
            "magicX": magicX
        }

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message

    # 柜面收付费审核接口
    def finishFeeFlow(self, magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, internalBankName):
        offsetSumId = resultData[0]['offsetSumId']
        accoName = resultData[0]['accoName']
        certiType = resultData[0]['certiType']
        certiCode = resultData[0]['certiCode']
        accoNo = resultData[0]['accoNo']
        bankCode = resultData[0]['bankCode']
        # 定义url
        url = "http://" + interface_ip + "/bcp-commu/bcpCash/finishFeeFlow.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "custName": accoName,
            "custCertiType": certiType,
            "custCertiCode": certiCode,
            "accountNo": accoNo,
            "bak2": magicX,
            "bankEndorse": bankEndorse,
            "feeConfirmDate": feeConfirmDate,
            "internalAccount": internalAccount,
            "internalBankCode": internalBankCode,
            "feeChannel": feeChannel,
            "bankCode": bankCode,
            "offsetSumId": offsetSumId,
            "bak3": 'pass',
            "internalBankName": internalBankName,
            "magicX": magicX
        }

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message

    # 资金系统费用查询接口
    def queryBankOffsetSumList(self, magicX, interface_ip, policyNo):
        # 定义url
        url = "http://" + interface_ip + "/bcp-commu/bankOffsetSumVO/queryBankOffsetSumList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=bankCode&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=orgainId&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=accoName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=payOrCollect&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=payMode&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=serviceId&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=amount&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&orgainId=&bankCode=&payOrCollect=&insertTime=&policyNo=' + policyNo + '&submitChannel=&payMode=&serviceId=&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        offsetSumId = response['resultData'][0]['offsetSumId']
        payOrCollect = response['resultData'][0]['payOrCollect']
        # 返回是否成功，以及错误信息
        is_success = response['isSuccess']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message, offsetSumId, payOrCollect

    # 资金系统费用审核确认
    def updateBankOffsetStatus(self, magicX, interface_ip, offsetSumId, payOrCollect):
        # 定义url+data
        if interface_ip == '10.1.6.11':
            url = "http://" + interface_ip + "/bcp-commu/bankOffsetSumVO/updateBankOffsetStatus.do"
            data = '[{"offsetSumId":"' + offsetSumId + ':' + payOrCollect + '"}]&magicX=' + magicX
        elif interface_ip != '10.1.6.11':
            url = "http://" + interface_ip + "/bcp-commu/bankOffsetSumVO/updateBankOffsetStatus.do?magicX=" + magicX
            data = '[{"offsetSumId":"' + offsetSumId + ':' + payOrCollect + '"}]]'
        # 定义headers
        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 获取批次号
        others = response['others']
        payment_no = others[payOrCollect]
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message, payment_no

    # BCP支付方式查询接口（转进账单）
    def queryBankOffsetSum(self, magicX, interface_ip, policyNo):
        # 定义url+data
        url = "http://" + interface_ip + "/bcp-commu/bankChange2Cash/queryBankOffsetSum.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=payMode&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=policyNo&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=bankCode&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=orgainId&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=accoName&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=serviceId&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=amount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=payOrCollect&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=feeStatus&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=insertTime&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=offsetSumId&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&payMode=01&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # BCP支付方式变更转进账单接口
    def doChange2Cash(self, magicX, interface_ip, offsetSumId):
        # 定义url+data
        url = "http://" + interface_ip + "/bcp-commu/bankChange2Cash/doChange2Cash.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'offsetSumId=' + offsetSumId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # BCP支付方式变更转现金交账单
    def doChange2Cash1(self, magicX, interface_ip, offsetSumId):
        # 定义url+data
        url = "http://" + interface_ip + "/bcp-commu/bankChange2Cash/doChange2Cash.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'offsetSumId=' + offsetSumId + '&payMode=12&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # BCP支付方式变更转银行转账接口
    def doChange2Bank(self, magicX, interface_ip, offsetSumId):
        # 定义url+data
        url = "http://" + interface_ip + "/bcp-commu/cashChange2Bank/doChange2Bank.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'offsetSumId=' + offsetSumId + '&payMode=01&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


"""保单综合查询接口类"""
class API0000_queryPolicy():
    # 综合查询-查询按钮接口
    def queryPolicy(self, magicX, interface_ip, policyNo):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/queryPolicy/query.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=proposalNo&sSearch_1=&bRegex_1=false&bSearchable_1=false&bSortable_1=false&mDataProp_2=policyNo&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=name&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=submitChannel&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=effectiveDate&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=expiryDate&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=policyStatus&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=proposalNo&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&proposalNo=&policyNo=' + policyNo +'&queryRole=001'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询-保单信息-日期信息接口
    def policyDetail(self, magicX, interface_ip, policyNo, proposalNo, proposalId, policyId):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/queryPolicy/policyDetail.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&moduleId=policyMainPer'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询-客户信息sheet页_被保人信息接口
    def customerByRole3(self, magicX, interface_ip, policyNo, policyId, proposalId, proposalNo):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/customer/q/customerByRole.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=false&mDataProp_1=name&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=false&mDataProp_2=gender&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=false&mDataProp_3=birthday&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=false&mDataProp_4=certiType&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=false&mDataProp_5=certiCode&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=false&mDataProp_6=benOrder&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=false&mDataProp_7=benePercent&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=false&mDataProp_8=relationship&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=false&mDataProp_9=isLegal&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=false&mDataProp_10=grade&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=false&mDataProp_11=noticeId&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&businessNo=1'
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询-险种信息sheet_险种信息查询接口
    def queryProduct(self, magicX, interface_ip, policyNo, policyId, proposalId, proposalNo):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/queryProduct/queryProduct.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=productCode&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=productName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=insurantId&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=assuredLevelCopies&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=totalPremAF&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=payDueDate&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=effectiveDate&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=expiryDate&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=liabilityState&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询-保全贷款信息sheet页接口
    def loanmain(self, magicX, interface_ip, policyNo, proposalNo, proposalId, policyId):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/loan/q/loanmain.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=12&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=investId&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=loanType&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=productCode&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=productName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=loanPrincipal&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=rate&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=capitalizedInterest&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=loanInterest&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=interestCapital&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=endingOfLoanTerm&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=dealTime&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&magicX=' + magicX  
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询_生存给付信息sheet_账户交易记录查询接口
    def survivalPayAccount(self, magicX, interface_ip, policyNo, proposalNo, proposalId, policyId):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/GRPPolicy/q/list.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=fundCode&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=fundCode&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=dealTime&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=transAmount&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=transAmount&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=rate&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=cashAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=serviceId&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=payOption&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=relatedPolicyNo&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&mDataProp_11=bankCode&sSearch_11=&bRegex_11=false&bSearchable_11=true&bSortable_11=true&mDataProp_12=accountCode&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&moduleId=survivalPayDetial&magicX=' + magicX   
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 综合查询-万能账号信息sheet页_账号记录接口
    def universal_account(self, magicX, interface_ip, policyNo, proposalNo, proposalId, policyId):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/universal/q/account.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=streamId&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=fundCode&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=fundCode&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=subFundCode&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=dealTime&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=recordTime&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=moneyType&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=interestCapital&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=interestCapital&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=interestCapital&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&magicX=' + magicX   
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 综合查询-万能账号信息sheet页_账号记录接口
    def claimlist(self, magicX, interface_ip, policyNo, proposalNo, proposalId, policyId):
        # 定义url+data
        url = "http://" + interface_ip + "/query-commu/claim/q/claimlist.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=5&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=caseNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=arouseName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=applyType&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=productName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=calsignTime&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=claimResult&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=payAmount&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=caseStatus&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=caseNo&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&policyNo=' + policyNo + '&proposalNo=' + proposalNo + '&proposalId=' + proposalId + '&policyId=' + policyId + '&target=1&chlNotTimely=false&businessNo=&moduleId=claimImagePer'  
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 综合查询-理赔基本信息sheet页接口-理赔详情信息
    def claimdetail(self, magicX, interface_ip, r, policyNo, proposalNo, proposalId, policyId, businessNo):
        # 定义url
        url = "http://" + interface_ip + "/query-commu/claim/q/claimdetail.do"

        # 组装请求参数
        data = {
            "r": r,
            "magicX": magicX,
            "policyNo": policyNo,
            "proposalNo": proposalNo,
            "proposalId": proposalId,
            "policyId": policyId,
            "target": '1',
            "chlNotTimely": False,
            "moduleId": 'prem_track_dt',
            "businessNo": businessNo
        }
        # 调用发送get请求函数
        response = API0000_diy().send_get(url, data)
        return response




"""理赔接口类"""
class API0000_clm():
    #理赔报案-保单查询接口
    def Register_queryPolicy(self, magicX, interface_ip, policyNo, caseType):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/Register/queryPolicy.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=11&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=policyNo&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=policyNo&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=grpPolicyNo&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=organName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=productName&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=insurantName&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=customerInfo&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=customerInfo&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=sumAssured&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=liabilityState&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&mDataProp_10=effectiveDate&sSearch_10=&bRegex_10=false&bSearchable_10=true&bSortable_10=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&arouseName=&arouseCertiType=&arouseCertiNo=&arouseSex=&policyNo=' + policyNo + '&caseType=' + caseType + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #理赔报案_保单查询接口
    def Register_registerInfo(self, magicX, interface_ip, arouseName, arouseCertiType, arouseCertiNo, arouseSex, arouseBirthday, telephone, policyNo, productName, organName, grpPolicyNo, productId, organCode, caseType):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/Register/registerInfo.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'arouseName=' + arouseName + '&arouseCertiType=' + arouseCertiType + '&arouseCertiNo=' + arouseCertiNo + '&arouseSex=' + arouseSex + '&arouseBirthday=' + arouseBirthday + '&telephone=' + telephone + '&policyNo=' + policyNo + '&productName=' + productName + '&organName=' + organName + '&grpPolicyNo=' + grpPolicyNo + '&productId=' + productId + '&organCode=' + organCode + '&caseType=' + caseType + '&reportNo=&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #理赔报案-报案事故信息提交接口
    def Register_doFinishSubmit(self, magicX, interface_ip, partitionIndi , accidentDate, claimNature, cityCode, accDesc, claimType, districtCode, reportRemark, hospitalCode2, hospitalCode ,provinceCode, accStreet, reportNo, reportName, telephone, reportChannel, relationShip, reportDate, accCustId, arouseName, arouseCertiType, arouseCertiNo, arouseSex, mobilePhone, arouseBirthday, arouseAge, policyNo, custId):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/Register/doReportSubmit.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        if claimType == '01':
            data = 'grpPolicyNo=&partitionIndi=' + partitionIndi + '&accidentDate=' + accidentDate + '&claimNature=' + claimNature + '&cityCode=' + cityCode + '&accDesc=' + accDesc + '&claimType=' + claimType + '&deathCause' + deathCause + '&districtCode=' + districtCode + '&reportRemark=' + reportRemark + '&hospitalCode2=' + hospitalCode2 + '&hospitalCode=' + hospitalCode + '&provinceCode=' + provinceCode + '&accStreet=' + accStreet + '&reportNo=' + reportNo + '&reportName=' + reportName + '&telephone=' + telephone + '&reportChannel=' + reportChannel + '&relationShip=' + relationShip + '&reportDate=' + reportDate + '&accCustId=' + accCustId + '&arouseName=' + arouseName + '&arouseCertiType=' + arouseCertiType + '&arouseCertiNo=' + arouseCertiNo + '&arouseSex=' + arouseSex + '&mobilePhone=' + mobilePhone + '&arouseBirthday=' + arouseBirthday + '&arouseAge=' + arouseAge + '&policyNo=' + policyNo + '&custId=' + custId + '&magicX=' + magicX
        else:
            data = 'grpPolicyNo=&partitionIndi=' + partitionIndi + '&accidentDate=' + accidentDate + '&claimNature=' + claimNature + '&cityCode=' + cityCode + '&accDesc=' + accDesc + '&claimType=' + claimType + '&districtCode=' + districtCode + '&reportRemark=' + reportRemark + '&hospitalCode2=' + hospitalCode2 + '&hospitalCode=' + hospitalCode + '&provinceCode=' + provinceCode + '&accStreet=' + accStreet + '&reportNo=' + reportNo + '&reportName=' + reportName + '&telephone=' + telephone + '&reportChannel=' + reportChannel + '&relationShip=' + relationShip + '&reportDate=' + reportDate + '&accCustId=' + accCustId + '&arouseName=' + arouseName + '&arouseCertiType=' + arouseCertiType + '&arouseCertiNo=' + arouseCertiNo + '&arouseSex=' + arouseSex + '&mobilePhone=' + mobilePhone + '&arouseBirthday=' + arouseBirthday + '&arouseAge=' + arouseAge + '&policyNo=' + policyNo + '&custId=' + custId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-立案基本信息-申请人信息保存接口
    def updateClaimReporter(self, magicX, interface_ip, partitionIndi, reportNo, applyToRelationship, certiCode, mobilePhone, cityCode, postalCode, applyName, gender, email, districtCode, certiType, birthday, provinceCode, addrDetail):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimReport/updateClaimReporter.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'partitionIndi=' + partitionIndi + '&reportNo=' + reportNo + '&applyToRelationship=' + applyToRelationship + '&certiCode=' + certiCode + '&mobilePhone=' + mobilePhone + '&cityCode=' + cityCode + '&postalCode=' + postalCode + '&applyName=' + applyName + '&gender=' + gender + '&email=' + email + '&districtCode=' + districtCode + '&certiType=' + certiType + '&birthday=' + birthday + '&provinceCode=' + provinceCode + '&addrDetail=' + addrDetail + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-立案基本信息-出险人信息保存接口
    def updateArouseInfo(self, magicX, interface_ip, reportNo, arouseName, arouseCertiEnd, arouseCertiNo, mobilePhone, arouseBirth):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimCase/updateArouseInfo.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'reportNo=' + reportNo + '&arouseName=' + arouseName + '&arouseCertiEnd=' + arouseCertiEnd + '&arouseCertiNo=' + arouseCertiNo + '&mobilePhone=' + mobilePhone + '&arouseBirth=' + arouseBirth + '&arouseAge=' + arouseAge + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-立案基本信息-案件信息保存接口
    def updateClaimCase(self, magicX, interface_ip, caseId, partitionIndi, reportNo, reportChannel, deathDate, provinceCode, materiaApproveTime, useCommonSumAssured, caseNo, accidentDate, diagnoseTime, isgroupFlag, cityCode, claimamout, accDesc, disease, transportType, icdCode2, icdCode, districtCode, accStreet, reportRemark):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimCase/updateClaimCase.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'caseId=' + caseId + '&partitionIndi=' + partitionIndi + '&reportNo=' + reportNo + '&reportChannel=' + reportChannel + '&deathDate=' + deathDate + '&provinceCode=' + provinceCode + '&materiaApproveTime=' + materiaApproveTime + '&useCommonSumAssured=' + useCommonSumAssured + '&caseNo=' +caseNo + '&accidentDate=' + accidentDate + '&hospitalCode2=&hospitalCode=&diagnoseTime=' + diagnoseTime + '&isgroupFlag=' + isgroupFlag + '&cityCode=' + cityCode + '&claimamout=' + claimamout + '&accDesc=' + accDesc + '&disease=' + disease + '&transportType=' + transportType + '&icdCode2=' + icdCode2 + '&icdCode=' +icdCode + '&districtCode=' + districtCode + '&accStreet=' + accStreet + '&reportRemark=' + reportRemark + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-立案基本信息-案件信息有效性校验接口
    def validateOnBasicInfoSubmit(self, magicX, interface_ip, caseId, partitionIndi, reportNo, reportChannel, deathDate, provinceCode, materiaApproveTime, useCommonSumAssured, caseNo, accidentDate, diagnoseTime, isgroupFlag, cityCode, claimamout, accDesc, disease, transportType, icdCode2, icdCode, districtCode, accStreet, reportRemark):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimCase/validateOnBasicInfoSubmit.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'caseId=' + caseId + '&partitionIndi=' + partitionIndi + '&reportNo=' + reportNo + '&reportChannel=' + reportChannel + '&deathDate=' + deathDate + '&provinceCode=' + provinceCode + '&materiaApproveTime=' + materiaApproveTime + '&useCommonSumAssured=' + useCommonSumAssured + '&caseNo=' + caseNo + '&accidentDate=' + accidentDate + '&hospitalCode2=&hospitalCode=&diagnoseTime=' + diagnoseTime + '&isgroupFlag=' + isgroupFlag + '&cityCode=' + cityCode + '&claimamout=' + claimamout + '&accDesc=' + accDesc + '&disease=' + disease + '&transportType=' + transportType + '&icdCode2=' + icdCode2 + '&icdCode=' + icdCode + '&districtCode=' + districtCode + '&accStreet=' + accStreet + '&reportRemark=' + reportRemark + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-影像信息-理赔影像详情查询接口
    def queryClaimChecklistes(self, magicX, interface_ip, reportNo, partitionIndi):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimChecklist/queryClaimChecklistes.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=-1&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=false&bSortable_0=false&mDataProp_1=reportNo&sSearch_1=&bRegex_1=false&bSearchable_1=false&bSortable_1=false&mDataProp_2=checklistItem&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=checklistItem&sSearch_3=&bRegex_3=false&bSearchable_3=false&bSortable_3=false&mDataProp_4=pageTotal&sSearch_4=&bRegex_4=false&bSearchable_4=false&bSortable_4=false&mDataProp_5=status&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=rejectReason&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=isRequired&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=pageTotal&sSearch_8=&bRegex_8=false&bSearchable_8=false&bSortable_8=false&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&reportNo=' + reportNo + '&partitionIndi=' + partitionIndi + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #立案-影像信息-理赔影像详情保存接口
    def updateClaimChecklist(self, magicX, interface_ip, claimChecklistes, sClaimCase):
        # 定义url
        url = "http://" + interface_ip + "/claim-commu/claimChecklist/updateClaimChecklist.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "claimChecklistes": claimChecklistes,
            "sClaimCase": sClaimCase
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



  

"""系统其它接口类"""
class API0000_sys_otherapi():
    # 登录系统接口
    def login(self, env_name, userName, interface_ip):
        #初始化
        magicX = ''
        # 获取密码
        cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
        sql = 'select user_password from t_user where user_name=' + "'" + userName + "'"
        #print sql
        cursor.execute(sql)
        result = cursor.fetchall()
        #判断数据库是否查到用户密码
        if len(result) > 0: 
            userPassword = result[0][0]
            # 定义url
            url = "http://" + interface_ip + "/admin-service/login.do"
            # 定义headers
            headers = {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            }
            # 组装请求参数
            data = {
                "userName": userName,
                "userPassword": userPassword
            }
            # 调用发送post请求函数
            response = API0000_diy().send_post(url, headers, data)
            # 返回是否登录系统成功，以及错误信息
            success = response['success']
            if success == '1':
                is_success = True
            elif success == '0':
                is_success = False
            if is_success:
                message = ''
                magicX = response['magicX']
            else:
                message = '登录失败:' + response['message']
                magicX = ''
        else:
            is_success = False
            message = '数据库查询用户密码失败'
        return is_success, message, magicX


    # 登出系统接口
    def logout(self, magicX, interface_ip):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/logout.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "magicX": magicX
        }
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否登出成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message
        


    # 万能利率配置_插入接口
    def investRateDefAdd(self, magicX, interface_ip, productId, fundCode, rateType, rate, startDate, endDate, publicTime):
        # 定义url
        url = "http://" + interface_ip + "/policy-issue-commu/invest/investRateDefAdd.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'productId=' + productId + '&fundCode=' + fundCode + '&rateType=' + rateType + '&rate=' + rate + '&startDate=' + startDate +'&endDate=' + endDate + '&publicTime=' + publicTime + '&rateId=&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response

 
    # 万能利率配置_查询接口
    def investRateDefQuery(self, magicX, interface_ip, productId, rateType, fundCode, publicTime):
        # 定义url
        url = "http://" + interface_ip + "/policy-issue-commu/invest/investRateDefQuery.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=8&sColumns=%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=2000&mDataProp_0=productId&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=fundCode&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=rateType&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=rate&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=startDate&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=endDate&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=publicTime&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&productId=' + productId + '&rateType=' + rateType + '&fundCode=' + fundCode + '&publicTime=' + publicTime + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 雇员管理保存接口
    def employee_save(self, magicX, interface_ip, employeeCode, gender, certiType, organId, underwritingMark, status, name, birthday, certiCode, title, education):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/employee/save.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0",
            "Accept-Encoding":"gzip, deflate",
            "Accept-Language":"zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
            "Accept":"application/json, text/javascript, */*; q=0.01"
        }
        # 组装请求参数
        data = 'employeeId=&employeeCode=' + employeeCode + '&gender=' + gender + '&certiType=' + certiType + '&organId=' + organId + '&underwritingMark=' + underwritingMark + '&status=' + status + '&name=' + name + '&birthday=' + birthday + '&certiCode=' + certiCode + '&title=' + title + '&education=' + education + '&magicX=' + magicX

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # 雇员管理_自查询接口       
    def employee_query(self, magicX, interface_ip, name, employeeCode):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/employee/query.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=13&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=organId&sSearch_1=&bRegex_1=false&bSearchable_1=false&bSortable_1=false&mDataProp_2=departId&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=employeeCode&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=name&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=gender&sSearch_5=&bRegex_5=false&bSearchable_5=false&bSortable_5=false&mDataProp_6=birthday&sSearch_6=&bRegex_6=false&bSearchable_6=false&bSortable_6=false&mDataProp_7=certiType&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=certiCode&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=education&sSearch_9=&bRegex_9=false&bSearchable_9=false&bSortable_9=false&mDataProp_10=title&sSearch_10=&bRegex_10=false&bSearchable_10=false&bSortable_10=false&mDataProp_11=status&sSearch_11=&bRegex_11=false&bSearchable_11=false&bSortable_11=false&mDataProp_12=employeeId&sSearch_12=&bRegex_12=false&bSearchable_12=true&bSortable_12=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&employeeCode=&name=' + name

        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 雇员管理_用户管理接口       
    def employee_userManage(self, magicX, interface_ip, userName, userPassword, gender, telephone, realName, confirmPassword, userType, email, status, employeeId):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/userManage/empSaveUser.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'userName=' + userName + '&userPassword=' + userPassword + '&gender=' + gender + '&telephone=' + telephone + '&ipAddress=&realName=' + realName + '&confirmPassword=' + confirmPassword + '&userType=' + userType + '&email=' + email + '&status=' + status + '&employeeId=' + str(employeeId) + '&userId=&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # 用户权限管理-查询接口       
    def userManage_queryUser(self, magicX, interface_ip, userName):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/userManage/queryUser.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=10&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=%23&sSearch_0=&bRegex_0=false&bSearchable_0=true&bSortable_0=true&mDataProp_1=abbrName&sSearch_1=&bRegex_1=false&bSearchable_1=true&bSortable_1=true&mDataProp_2=userName&sSearch_2=&bRegex_2=false&bSearchable_2=true&bSortable_2=true&mDataProp_3=realName&sSearch_3=&bRegex_3=false&bSearchable_3=true&bSortable_3=true&mDataProp_4=userTypeDesc&sSearch_4=&bRegex_4=false&bSearchable_4=true&bSortable_4=true&mDataProp_5=userSource&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=telephone&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=email&sSearch_7=&bRegex_7=false&bSearchable_7=true&bSortable_7=true&mDataProp_8=status&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&mDataProp_9=userId&sSearch_9=&bRegex_9=false&bSearchable_9=true&bSortable_9=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&organId=-1&userName=' + userName + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['isSuccess']

        if is_success:
            message = ''
            resultData = response['resultData']   #获取查询结果集，遍历结果集获取对应的employeeId
            userId = response['resultData'][0]['userId']    #获取employeeId
        else:
            message = response['message']
        return is_success, message, userId


    # 用户权限管理-获取可选角色列表       
    def userManage_getRoleList(self, magicX, interface_ip, userId):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/userManage/getRoleList.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'userId=' + userId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
            unAssignedList = response['others']['unAssignedList']
            roleIds = ''
            for i in range(0,len(unAssignedList)):
                roleId = response['others']['unAssignedList'][i]['roleId']
                roleIds = roleIds + str(roleId)
                if i+1 < len(unAssignedList):
                    roleIds = roleIds + '%2C'
        else:
            message = response['message']
            roleIds = ''
        return is_success, message, roleIds


    # 用户权限管理-分配权限       
    def userManage_saveRoles(self, magicX, interface_ip, userId, roleIds):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/userManage/saveRoles.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'userId=' + userId + '&roleIds=' + roleIds + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message
        
 
    # 进账单权限配置       
    def addCashLimitCfg(self, magicX, interface_ip, organId, operatorId, maxAmount, minAmount):
        # 定义url
        url = "http://" + interface_ip + "/bcp-commu/bcpCash/addCashLimitCfg.do?magicX=" + magicX
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'organId=' + organId + '&operatorId=' + operatorId + '&maxAmount=' + maxAmount + '&minAmount=' + minAmount + '&limitId='
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # 修改密码-原密码检查接口
    def checkPassword(self, magicX, interface_ip, old_password_md5):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/password/checkPassword.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'password=' + old_password_md5 + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    # 修改密码接口
    def reset_password(self, magicX, interface_ip, old_password_md5, new_password_md5, customerId):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/password/confirmchange.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'oldPassword=' + old_password_md5 + '&newPassword=' + new_password_md5 + '&customerId=' + customerId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 登录弹出重置密码接口    
    def password_confirmchange(self, magicX, interface_ip, customerId):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/password/confirmchange.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'oldPassword=0659c7992e268962384eb17fafe88364&newPassword=183f3364ed7564a9f5624da2421edeed&customerId=' + customerId + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # 系统用户解锁接口    
    def deblocking(self, magicX, interface_ip, userName):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/deblocking.do"
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'userName=' + userName + '&magicX=' + magicX
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)

        # 返回是否成功，以及错误信息
        is_success = response['success']
        if is_success:
            message = ''
        else:
            message = response['message']
        return is_success, message


    # 产品销售控制配置查询接口
    def queryConditionSalesControl(self, magicX, interface_ip, productId):
        # 定义url
        url = "http://" + interface_ip + "/admin-service/salesControl/queryConditionSalesControl.do" 
        # 定义headers
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = 'sEcho=1&iColumns=9&sColumns=%2C%2C%2C%2C%2C%2C%2C%2C&iDisplayStart=0&iDisplayLength=500&mDataProp_0=productId&sSearch_0=&bRegex_0=false&bSearchable_0=false&bSortable_0=false&mDataProp_1=organId&sSearch_1=&bRegex_1=false&bSearchable_1=false&bSortable_1=false&mDataProp_2=policyChannel&sSearch_2=&bRegex_2=false&bSearchable_2=false&bSortable_2=false&mDataProp_3=partnerId&sSearch_3=&bRegex_3=false&bSearchable_3=false&bSortable_3=false&mDataProp_4=salesChannel&sSearch_4=&bRegex_4=false&bSearchable_4=false&bSortable_4=false&mDataProp_5=startDate&sSearch_5=&bRegex_5=false&bSearchable_5=true&bSortable_5=true&mDataProp_6=endDate&sSearch_6=&bRegex_6=false&bSearchable_6=true&bSortable_6=true&mDataProp_7=status&sSearch_7=&bRegex_7=false&bSearchable_7=false&bSortable_7=false&mDataProp_8=listId&sSearch_8=&bRegex_8=false&bSearchable_8=true&bSortable_8=true&sSearch=&bRegex=false&iSortCol_0=0&sSortDir_0=asc&iSortingCols=1&params=%7B%22productId%22%3A%22' + productId + '%22%2C%22organId%22%3A%22%22%2C%22policyChannel%22%3A%22%22%2C%22salesChannel%22%3A%22%22%2C%22status%22%3A%22%22%2C%22startDate%22%3A%22%22%2C%22endDate%22%3A%22%22%7D&magicX=' + magicX
        #print data
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response


    # 产品销售控制配置查询接口
    def findByBroker(self, agent_code, agent_name, channelType):
        # 定义url
        url = "http://10.8.1.156/gateway/micro-serv-broker-app-uat/openProdSalesCfg/findByBroker" 
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
            "head":{"appid":"JK-TC","appkey":"Jq8wV04lay","sign":"","transDate":"20210415","transTime":"151352","transCode":"BRK-OPEN-506","openId":"OFFLINE"},
            "tenantId": "2017032717500543538",
            "channelType":channelType,
            "brokerCode":agent_code,
            "brokerName":agent_name,
            "applyDate":"2021-04-15",
            "ynLoadProdInfo":"N",
            "issueWay":"1","appkey":"Jq8wV04lay",
            "productDesc":""
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



"""非系统其它接口类"""
class API0000_sys_notherapi():

    # checkdata:执行任务-记录步骤执行信息
    def run_xjob(self, icode, execNo, stepNo, stepTitle, stepParams, stepResult, stepError):
        # 定义url
        url = "http://10.8.1.72/xjob/v1" 
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
                "icode": icode,
                "execNo": execNo,
                "stepNo": stepNo,
                "stepTitle": stepTitle,
                "stepParams": stepParams,
                "stepResult": stepResult,
                "stepError": stepError
            }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response





class myThread3 (threading.Thread):   #继承父类threading.Thread
    def __init__(self, env_name, ip, user_name, password, to_modify_date):
        threading.Thread.__init__(self)
        self.env_name = env_name
        self.ip = ip
        self.user_name = user_name
        self.password = password
        self.to_modify_date = to_modify_date


        
    def run(self):
        try:
            #print '****set_date_' + self.ip + ':begin****'
            #初始化
            service_time_new = ""
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=self.ip, port=22, username=self.user_name, password=self.password, timeout=300,allow_agent=False,look_for_keys=False)
            #修改服务器时间并获取服务器时间
            cmd1 = 'date -s "%s"' % self.to_modify_date   #ssh修复服务器时间命令
            #print cmd1
            stdin, stdout, stderr = ssh.exec_command(cmd1)   #执行ssh命令
            """step2:重启服务器"""
            if self.env_name == '预生产':
                cmd2 = "docker ps | egrep 'policy|base-batch' | awk '{print $1}' | while read line;do docker restart $line ;done"
                #print cmd2
                stdin, stdout, stderr = ssh.exec_command(cmd2)
                result = stdout.read()     #获取进程
                #print result
                cmd3 = "docker ps | egrep 'bcp' | awk '{print $1}' | while read line;do docker restart $line ;done"
                #print cmd3
                stdin, stdout, stderr = ssh.exec_command(cmd3)
                result = stdout.read() 
                #print result
            else:
                cmd2 = "docker ps | egrep 'base|pol' | awk '{print $1}' | while read line;do docker restart $line ;done"
                #print cmd2
                stdin, stdout, stderr = ssh.exec_command(cmd2)
                result = stdout.read()     #获取进程
                #print result
                cmd3 = "docker ps | egrep 'bcp' | awk '{print $1}' | while read line;do docker restart $line ;done"
                #print cmd3
                stdin, stdout, stderr = ssh.exec_command(cmd3)
                result = stdout.read() 
                #print result
            ssh.close()
        except Exception, e:
            #print '---err----'
            print traceback.format_exc()
        finally:
            pass
            #print '****set_date_' + self.ip + ':end****'

"""
多线程：修改服务器时间：
step1:ssh获取服务器1、服务器2的待重启线程列表
step2:再开启多线程，分别对服务器1、服务器2的待重启线程进行依次重启
"""

class myThread (threading.Thread):   #继承父类threading.Thread
    def __init__(self, env_name, ip, user_name, password, to_modify_date):
        threading.Thread.__init__(self)
        self.env_name = env_name
        self.ip = ip
        self.user_name = user_name
        self.password = password
        self.to_modify_date = to_modify_date


        
    def run(self):
        try:
            #print '****set_date_' + self.ip + ':begin****'
            #初始化
            service_time_new = ""
            list = []
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=self.ip, port=22, username=self.user_name, password=self.password, timeout=300,allow_agent=False,look_for_keys=False)
            #修改服务器时间并获取服务器时间
            cmd = 'date -s "%s"' % self.to_modify_date   #ssh修复服务器时间命令
            stdin, stdout, stderr = ssh.exec_command(cmd)   #执行ssh命令
            #print cmd
            """step2:获取待重启服务器进程"""
            if self.env_name == '预生产':
                cmd1 = "docker ps | egrep 'policy|base-batch' | awk '{print $1}'"
                stdin, stdout, stderr = ssh.exec_command(cmd1)
                result1 = stdout.read()     #获取进程
                list1 = result1.split()   #拆分进程存入列表中
                list.extend(list1)
                cmd2 = "docker ps | egrep 'bcp' | awk '{print $1}'"
                stdin, stdout, stderr = ssh.exec_command(cmd2)
                result2 = stdout.read() 
                list2 = result2.split()
                list.extend(list2)
            else:
                cmd1 = "docker ps | egrep 'base|pol' | awk '{print $1}'"
                stdin, stdout, stderr = ssh.exec_command(cmd1)
                result1 = stdout.read()     #获取进程
                list1 = result1.split()   #拆分进程存入列表中
                list.extend(list1)
                cmd2 = "docker ps | egrep 'bcp' | awk '{print $1}'"
                stdin, stdout, stderr = ssh.exec_command(cmd2)
                result2 = stdout.read() 
                list2 = result2.split()
                list.extend(list2)
            ssh.close()
        except Exception, e:
            #print '---err----'
            print traceback.format_exc()
        finally:
            pass
            #print 'list:'
            #print list
            
            threads = []
            for i in range(0, len(list)):
                thread = myThread1(self.env_name, self.ip, self.user_name, self.password, list[i])    # 创建新线程
                threads.append(thread)
                
            for t in threads:
                t.setDaemon(True)
                t.start()

            #join()方法，用于等待线程终止
            for t in threads:
                 t.join()



        



class myThread1 (threading.Thread):   #继承父类threading.Thread
    def __init__(self, env_name, ip, user_name, password, process):
        threading.Thread.__init__(self)
        self.env_name = env_name
        self.ip = ip
        self.user_name = user_name
        self.password = password
        self.process = str(process)


        
    def run(self):
        try:
            #print '****set_date_' + self.ip + ':begin_myThread1****'
            #初始化
            """step1:修复服务器时间"""
            #连接服务器
            ssh = paramiko.SSHClient()
            ssh.load_system_host_keys()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=self.ip, port=22, username=self.user_name, password=self.password, timeout=300,allow_agent=False,look_for_keys=False)

            """step2:获取待重启服务器进程"""
            cmd = "docker restart " + self.process
            #print cmd
            stdin, stdout, stderr = ssh.exec_command(cmd)
            result = stdout.read()     #获取进程
            #print result
            ssh.close()
        except Exception, e:
            #print '---err----'
            print traceback.format_exc()
        finally:
            pass





"""专家系统相关类"""
class API0000_expert_sys():

    # checkdata:执行任务-记录步骤执行信息
    def run_xjob(self, icode, execNo, stepNo, stepTitle, stepParams, stepResult, stepError):
        # 定义url
        url = "http://10.8.1.72/xjob/v1" 
        # 定义headers
        headers = {
            "Content-Type": "application/json;charset=utf-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
        }
        # 组装请求参数
        data = {
                "icode": icode,
                "execNo": execNo,
                "stepNo": stepNo,
                "stepTitle": stepTitle,
                "stepParams": stepParams,
                "stepResult": stepResult,
                "stepError": stepError
            }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = API0000_diy().send_post(url, headers, data)
        return response



    #专家系统-规则解析初始化
    @func_set_timeout(30)#设定函数超执行时间
    def rule_parsing_initialization(self, flow_code, dict):
        """
        规则解析初始化:
        入参：流程编号flow_code
        出参：数据结构list（供脚本使用）flow_list
        1.解析专家系统-数据结构（流程-事件-交易-具象交易-规则）。API0000_expert_sys().query_data_structure()
            入参：流程编号flow_code
            出参：数据结构list（供脚本使用）flow_list
            (1)连接专家系统数据库
            (2)从专家系统数据库获取数据结构json，传给前端展示
                入参：流程编号flow_code
                出参：数据结构dict（供前端展示）flow_dict
            (3)从专家系统数据库获取数据结构(列表)，提供给脚本使用
                入参：流程编号flow_code
                出参：数据结构list（供脚本使用）flow_list
        2.校验数据结构正确性
            入参：流程编号flow_code
            出参：dict
        """
        try:
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''

            #1.解析专家系统-数据结构
            rs = API0000_expert_sys().query_data_structure(flow_code, dict)
            flow_list = rs[0]   #流程列表
            dict = rs[1]
            assert dict['logInfo']['code'] == '1'
            flow_list = dict['flow_list']   #数据结构列表（脚本使用，非前端使用）
            #print flow_list
            #2.校验专家系统-数据结构正确性
            dict = self.verify_data_structure(flow_list, dict)

            #存入dict
            is_success = True
            dict['logInfo']['code'] = '1'
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            pass
            #step3:定义轨迹信息+检查点信息:
            #dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return flow_list, dict




    #从专家系统数据库-查询数据结构
    @func_set_timeout(30)#设定函数超执行时间
    def query_data_structure(self, flow_code, dict):
        """
        解析专家系统-数据结构（流程-事件-交易-具象交易-规则）：
        入参：流程编号flow_code
        出参：数据结构list（供脚本使用）flow_list
        (1)连接专家系统数据库
        (2)从专家系统数据库获取数据结构json，传给前端展示
            入参：流程编号flow_code
            出参：数据结构dict（供前端展示）flow_dict
        (3)从专家系统数据库获取数据结构(列表)，提供给脚本使用
            入参：流程编号flow_code
            出参：数据结构list（供脚本使用）flow_list
        """
        try:
            """step1:连接专家系统数据库"""
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_连接专家系统数据库'  # 节点

            if 'win' in sys.platform:   #windows系统平台
                db_ip = 'localhost'         # 定义数据库ip
                db_user = 'root'      # 定义数据库用户名
                db_password = 'abc123456'   # 定义数据库密码
                db_name = 'test'   #库名
            else:
                db_ip = '10.8.1.157'         # 定义数据库ip
                db_user = 'expert'      # 定义数据库用户名
                db_password = 'expert123'   # 定义数据库密码
                db_name = 'dbexpert'   #库名

            flow_list = []   #数据结构list（供脚本使用）
            # 连接数据库
            try:
                db = pymysql.connect(host=db_ip, user=db_user, passwd=db_password, db=db_name, charset='utf8')
            except Exception, e:
                is_success = False
                message = '专家系统数据库连接失败，请检查用户名、密码、ip、库名是否有误！'
                print traceback.format_exc()
                logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
            assert is_success != False
            #存入dict
            dict['db'] = db


            """step2:从专家系统数据库获取数据结构json，传给前端展示"""
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_从专家系统数据库获取数据结构json，传给前端展示'  # 节点
            rs = self.query_data_structure_step2(db, flow_code, dict)
            flow_dict = rs[0]   #前端展示的数据结构。json类型
            dict = rs[1]
            message = dict['logInfo']['msg']
            assert dict['logInfo']['code'] == '1'

            #传给前端展示
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_记录步骤执行信息'  # 节点

            icode = 'EXEC107'
            execNo = dict['execNo']   #执行编号
            #execNo = '2021090381000000577'
            stepNo = '01'   #步骤编号
            stepTitle = '从专家系统数据库-查询数据结构'   #步骤名称
            stepParams = {}   #步骤执行参数:json字符 ,无则传 “{}”
            stepResult = flow_dict   #步骤执行结果:json字符 ,无则传 “{}”
            stepError = {}   #步骤执行错误信息:json字符 ,无则传 “{}”
            #调用原子脚本
            response = API0000_expert_sys().run_xjob(icode, execNo, stepNo, stepTitle, stepParams, stepResult, stepError)
            if response['header']['istatus'] == '1':
                is_success = True
                message = ''
            else:
                is_success = False
                message = '不成功' 
            assert is_success == True


            #step3:从专家系统数据库获取数据结构(列表)，提供给脚本使用
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_从专家系统数据库获取数据结构(列表)，提供给脚本使用'  # 节点
            rs = self.query_data_structure_step3(db, flow_code, dict)
            flow_list = rs[0]    #数据结构(列表)
            dict = rs[1]
            message = dict['logInfo']['msg']
            assert dict['logInfo']['code'] == '1'

            #存入dict
            is_success = True
            #logging.warning(API0000_diy().text_conversion('&nbsp'*4 + '规则解析初始化_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            #step3:定义轨迹信息+检查点信息:
            dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return flow_list, dict




    #从专家系统数据库获取数据结构(json串)_提供给前端展示
    @func_set_timeout(30)#设定函数超执行时间
    def query_data_structure_step2(self, db, flow_code, dict):
        try:
            #logging.warning(API0000_diy().text_conversion('step2_1_开始'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''

            stage = []   #所有节点（流程、事件。。。规则）
            setArr = []   #专家系统结构列表(所有两个上下节点)
            flow_dict = {'flow':[{'stage':[],'setArr':[]}]}   #初始化数据结构串;stage-所有节点（流程、事件。。。规则）；setArr-所有两个上下节点
            event_list = []   #事件列表
            deal_list = []   #交易列表
            actual_deal_list = []   #具象交易列表
            rule_list = []   #交易列表
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_从专家系统数据库获取数据结构json，传给前端展示'  # 节点

            #根据流程读取专家系统数据结构
            #查询流程下事件
            try:
                cursor = db.cursor()
                sql = "SELECT a.cn_name 流程名称,c.cn_name 事件名称,c.code 事件编码 FROM exp_flow_define a,exp_flow_layout b,exp_event_define c WHERE a.id = b.master_flow_id AND b.event_id = c.id AND b.node_type = '2' AND a.code = '" + flow_code + "'"
                cursor.execute(sql)
                #print sql
                result = cursor.fetchall()
            except Exception, e:
                is_success = False
                message = '专家系统数据库查询规则报错，检查sql（查询流程下事件）是否错误或表中的规则是否存在！'
                print traceback.format_exc()
                logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
            assert is_success != False

            #存入dict
            for i in range(0,len(result)):
                flow_name = ('流程-' + result[i][0])#.encode('utf-8').encode('gbk')   #流程名称
                event_name = ('事件-' +result[i][1])#.encode('utf-8').encode('gbk')   #流程名称
                #存入stage
                stage.append(flow_name)
                stage.append(event_name)
                #存入setArr
                setArr.append([])
                event_list.append({'name':event_name,'code':result[i][2]})
                setArr[-1].append(flow_name)
                setArr[-1].append(event_name)
            #print setArr
            for i in range(0,len(event_list)):
                #print setArr[i][1]
                pass
            #print event_list
            is_success = True


            #查询事件下交易
            for i in range(0,len(event_list)):
                try:
                    cursor = db.cursor()
                    sql1 = "SELECT c.cn_name 事件名称,e.cn_name 交易名称,e.code 交易编码 FROM exp_event_define c,exp_event_layout d,exp_trans_group e WHERE  c.id = d.event_id AND d.node_type = '8' AND d.trans_id = e.id AND c.code = '" + event_list[i]['code'] + "'"
                    cursor.execute(sql1)
                    #print sql
                    result1 = cursor.fetchall()
                except Exception, e:
                    is_success = False
                    message = '专家系统数据库查询规则报错，检查sql（查询事件下交易）是否错误或表中的规则是否存在！'
                    print traceback.format_exc()
                    logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
                assert is_success != False

                #存入dict
                for j in range(0,len(result1)):
                    event_name = ('事件-' + result1[j][0])#.encode('utf-8').encode('gbk')   #事件名称
                    deal_name = ('交易-' + result1[j][1])#.encode('utf-8').encode('gbk')   #交易名称
                    #存入stage
                    stage.append(event_name)
                    stage.append(deal_name)
                    #存入setArr
                    setArr.append([])
                    deal_list.append({'name':deal_name,'code':result1[j][2]})
                    event_list[i]['deal_list'] = deal_list
                    setArr[-1].append(event_name)
                    setArr[-1].append(deal_name)
                #print setArr
                for j in range(0,len(deal_name)):
                    #print setArr[j][1]
                    pass
                #print deal_list
                is_success = True


            #查询交易下具象交易
            for i in range(0,len(deal_list)):
                try:
                    cursor = db.cursor()
                    sql2 = "SELECT e.cn_name 交易名称,g.cn_name 具象交易名称,g.code 具象交易编码 FROM exp_trans_group e,exp_trans_group_member f,exp_trans_define g WHERE e.id = f.group_id AND f.trans_id = g.id AND e.code = '" + deal_list[i]['code'] + "'"
                    cursor.execute(sql2)
                    #print sql
                    result2 = cursor.fetchall()
                except Exception, e:
                    is_success = False
                    message = '专家系统数据库查询规则报错，检查sql（查询交易下具象交易）是否错误或表中的规则是否存在！'
                assert is_success != False

                #存入dict
                for i in range(0,len(result2)):
                    deal_name = ('交易-' + result2[i][0])#.encode('utf-8').encode('gbk')   #交易名称
                    actual_deal_name = ('具象交易-' + result2[i][1])#.encode('utf-8').encode('gbk')   #具象交易名称
                    #存入stage
                    stage.append(deal_name)
                    stage.append(actual_deal_name)
                    #存入setArr
                    setArr.append([])
                    actual_deal_list.append({'name':actual_deal_name,'code':result2[i][2]})
                    setArr[-1].append(deal_name)
                    setArr[-1].append(actual_deal_name)
                #print setArr
                for i in range(0,len(actual_deal_list)):
                    #print setArr[i][1]
                    pass
                #print actual_deal_list
                is_success = True


            #查询具象交易下规则
            for i in range(0,len(actual_deal_list)):
                try:
                    cursor = db.cursor()
                    sql3 = "SELECT g.cn_name 具象交易名称,i.cn_name 规则名称,i.code 规则编码 FROM exp_trans_define g,exp_trans_layout h,exp_rule_define i WHERE g.id = h.trans_id AND h.ref_rule_id = i.id AND g.code = '" + actual_deal_list[i]['code'] + "'"
                    cursor.execute(sql3)
                    #print sql
                    result3 = cursor.fetchall()
                except Exception, e:
                    is_success = False
                    message = '专家系统数据库查询规则报错，检查sql（查询具象交易下规则）是否错误或表中的规则是否存在！'
                assert is_success != False

                #存入dict
                for i in range(0,len(result3)):
                    actual_deal_name = ('具象交易-' + result3[i][0])#.encode('utf-8').encode('gbk')   #具象交易名称
                    rule_name = ('规则-' + result3[i][1])#.encode('utf-8').encode('gbk')   #规则名称
                    #存入stage
                    stage.append(actual_deal_name)
                    stage.append(rule_name)
                    #存入setArr
                    setArr.append([])
                    rule_list.append({'name':rule_name,'code':result3[i][2]})
                    setArr[-1].append(actual_deal_name)
                    setArr[-1].append(rule_name)
            #print len(setArr)
            #print setArr
            """
            for i in range(0,len(setArr)):
                print setArr[i][1]
                pass
            """
            """
            for i in range(0,len(rule_list)):
                print rule_list
            """
            is_success = True

            #存入dict
            dict['logInfo']['code'] = '1' 
            #stage去重
            #print len(stage)
            stage = list(set(stage))
            #print len(stage)
            #print stage
            #print stage
            #定义总json
            flow_dict = {'flow':[{'stage':stage,'setArr':setArr}]}   #初始化流程串;stage-所有节点（流程、事件。。。规则）；setArr-所有两个上下节点

            #存入dict
            dict['flow_dict'] = flow_dict
            #存入dict
            #logging.warning(API0000_diy().text_conversion('step2_1_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            #step3:定义轨迹信息+检查点信息:
            dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return flow_dict, dict




    #从专家系统数据库获取数据结构(json串)_提供给脚本
    @func_set_timeout(30)#设定函数超执行时间
    def query_data_structure_step3(self, db, flow_code, dict):
        try:
            #logging.warning(API0000_diy().text_conversion('step2_2_开始'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''

            #节点初始化 
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-解析专家系统-数据结构_从专家系统数据库获取数据结构(列表)，提供给脚本使用'  # 节点
            flow_list = []   #流程列表

            #查询流程下事件
            try:
                cursor = db.cursor()
                sql11 = "SELECT c.cn_name,c.code FROM exp_flow_define a,exp_flow_layout b,exp_event_define c WHERE    a.id = b.master_flow_id AND b.event_id = c.id AND b.node_type = '2' AND a.code = '" + flow_code + "'"
                cursor.execute(sql11)
                #print sql
                result11 = cursor.fetchall()
            except Exception, e:
                is_success = False
                message = '专家系统数据库查询规则报错，检查sql（查询流程下事件）是否错误或表中的规则是否存在！'
            assert is_success != False

            #遍历事件
            for i in range(0,len(result11)): 
                event_name = result11[i][0]   #事件名称
                event_code = result11[i][1]   #事件编号
                flow_list.append({})
                flow_list[-1]['name'] = event_name 
                flow_list[-1]['code'] = event_code 
                flow_list[-1]['deal'] = []   #交易

                #查询事件下交易
                try:
                    cursor = db.cursor()
                    sql12 = "SELECT e.cn_name,e.code FROM exp_event_define c,exp_event_layout d,exp_trans_group e WHERE c.id = d.event_id AND d.node_type = '8' AND d.trans_id = e.id AND c.code = '" + event_code + "'"
                    cursor.execute(sql12)
                    #print sql
                    result12 = cursor.fetchall()
                except Exception, e:
                    is_success = False
                    message = '专家系统数据库查询规则报错，检查sql（查询事件下交易）是否错误或表中的规则是否存在！'
                assert is_success != False

                #遍历交易
                for i in range(0,len(result12)): 
                    deal_name = result12[i][0]   #交易名称
                    deal_code = result12[i][1]   #交易编号
                    flow_list[-1]['deal'].append({})
                    flow_list[-1]['deal'][-1]['name'] = deal_name 
                    flow_list[-1]['deal'][-1]['code'] = deal_code 
                    flow_list[-1]['deal'][-1]['actual_deal'] = []   #具象交易

                    #查询交易下具象交易
                    try:
                        cursor = db.cursor()
                        sql13 = "SELECT g.cn_name,g.code FROM exp_trans_group e,exp_trans_group_member f,exp_trans_define g WHERE e.id = f.group_id AND f.trans_id = g.id AND e.code = '" + deal_code + "'"
                        cursor.execute(sql13)
                        #print sql
                        result13 = cursor.fetchall()
                    except Exception, e:
                        is_success = False
                        message = '专家系统数据库查询规则报错，检查sql（查询交易下具象交易）是否错误或表中的规则是否存在！'
                    assert is_success != False

                    #遍历具象交易
                    for i in range(0,len(result13)): 
                        actual_deal_name = result13[i][0]   #具象交易名称
                        actual_deal_code = result13[i][1]   #具象交易编号
                        flow_list[-1]['deal'][-1]['actual_deal'].append({})
                        flow_list[-1]['deal'][-1]['actual_deal'][-1]['name'] = actual_deal_name 
                        flow_list[-1]['deal'][-1]['actual_deal'][-1]['code'] = actual_deal_code 
                        flow_list[-1]['deal'][-1]['actual_deal'][-1]['rule'] = []   #规则

                        #查询交易下具象交易
                        try:
                            cursor = db.cursor()
                            sql14 = "SELECT i.cn_name,i.code FROM exp_trans_define g,exp_trans_layout h,exp_rule_define i WHERE g.id = h.trans_id AND h.ref_rule_id = i.id AND g.code = '" + actual_deal_code + "'"
                            cursor.execute(sql14)
                            #print sql
                            result14 = cursor.fetchall()
                        except Exception, e:
                            is_success = False
                            message = '专家系统数据库查询规则报错，检查sql（查询交易下具象交易）是否错误或表中的规则是否存在！'
                        assert is_success != False

                        #遍历具象交易
                        for i in range(0,len(result14)): 
                            #规则名称
                            rule_name = result14[i][0] 
                            #规则编号  
                            rule_code = result14[i][1]   
                            #规则类型。1-数据源规则；2-执行数据源规则；3-结果展示约定规则
                            if 'data_source_rule' in rule_code:
                                rule_type = '1'
                            elif 'execute_data_rule' in rule_code:
                                rule_type = '2'
                            elif 'result_display_rule' in rule_code:
                                rule_type = '3'
                            else:
                                rule_type = 'other'
                            #存入flow_list
                            flow_list[-1]['deal'][-1]['actual_deal'][-1]['rule'].append({})
                            flow_list[-1]['deal'][-1]['actual_deal'][-1]['rule'][-1]['name'] = rule_name 
                            flow_list[-1]['deal'][-1]['actual_deal'][-1]['rule'][-1]['code'] = rule_code 
                            flow_list[-1]['deal'][-1]['actual_deal'][-1]['rule'][-1]['type'] = rule_type 

            #print flow_list
            #存入dict
            dict['flow_list'] = flow_list
            is_success = True
            dict['logInfo']['code'] = '1'
            #logging.warning(API0000_diy().text_conversion('step2_2_结束'))
        except Exception, e: 
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            #step3:定义轨迹信息+检查点信息:
            dict = API0000_diy().result(dict, is_success, message, unit)
        return flow_list, dict


    #校验数据结构正确性
    @func_set_timeout(30)#设定函数超执行时间
    def verify_data_structure(self, flow_list, dict):
        try:
            #初始化
            switch_data_source_rule = ''   #初始化-数据源规则开关。开：True   关：False
            switch_execute_data_source_rules = ''   #初始化-执行数据源规则开关
            switch_execution_result_display_rules = ''   #初始化-结果展示规则开关
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：规则解析初始化-校验数据结构正确性'  # 节点
            #遍历事件
            for i in range(0, len(flow_list)):
                event_name = flow_list[i]['name']   #事件名称
                event_code = flow_list[i]['code']   #事件编号
                #遍历交易
                for j in range(0, len(flow_list[i]['deal'])):
                    deal_name = flow_list[i]['deal'][j]['name']   #交易名称
                    deal_code = flow_list[i]['deal'][j]['code']   #交易编号
                    #遍历交易
                    for x in range(0, len(flow_list[i]['deal'][j]['actual_deal'])):
                        actual_deal_name = flow_list[i]['deal'][j]['actual_deal'][x]['name']   #具象交易名称
                        actual_deal_code = flow_list[i]['deal'][j]['actual_deal'][x]['code']   #具象交易编号
                        actual_deal_list = flow_list[i]['deal'][j]['actual_deal'][x]['rule']   #具象交易下规则列表

                        #转换具象交易下规则列表格式：[{'name':xx,'code':xx}]==>[code,code,code,code]
                        actual_deal_list_new = []   #新具象交易下规则列表[code,code,code,code]
                        for y in range(0,len(actual_deal_list)):
                            actual_deal_list_new.append(actual_deal_list[y]['code'])   #追加

                        #获取具象交易下配置的3种规则列表（数据源规则、执行数据源规则、结果展示规则）
                        rs = self.rule_engine_configuration(actual_deal_code, actual_deal_list, dict)
                        data_source_rule = rs[0]   #数据源规则列表
                        execute_data_source_rules = rs[1]   #执行数据源规则列表
                        execution_result_display_rules = rs[2]   #结果展示规则列表
                        dict = rs[3]
                        #存入dict
                        flow_list[i]['deal'][j]['actual_deal'][x]['data_source_rule'] = data_source_rule    #数据源规则列表
                        flow_list[i]['deal'][j]['actual_deal'][x]['execute_data_source_rules'] = execute_data_source_rules    #执行数据源规则列表
                        flow_list[i]['deal'][j]['actual_deal'][x]['execution_result_display_rules'] = execution_result_display_rules    #结果展示规则列表
                        #规则引擎判断：具象交易下的数据源规则、执行数据源规则、结果展示规则是否满足脚本配置的
                        parent_list = actual_deal_list_new   #专家系统下某具象交易内配置的规则（包含三种规则）
                        sub_list1 = data_source_rule   #数据源规则列表
                        sub_list2 = execute_data_source_rules   #执行数据源规则列表
                        sub_list3 = execution_result_display_rules   #结果展示规则列表
                        #print parent_list
                        #print sub_list1
                        #判断数据源规则列表
                        if API0000_diy().get_sub_list_index(parent_list, sub_list1) == True:
                            switch_data_source_rule = True
                            message = ''
                        else:
                            switch_data_source_rule = False
                            message = event_name + '-' + deal_name + '-' + actual_deal_name + '下‘数据源规则’缺失！'
                        assert switch_data_source_rule == True
                        #执行数据源规则列表
                        if API0000_diy().get_sub_list_index(parent_list, sub_list2) == True:
                            switch_execute_data_source_rules = True
                            message = ''
                        else:
                            switch_execute_data_source_rules = False
                            message = event_name + '-' + deal_name + '-' + actual_deal_name + '下‘执行数据源规则’缺失！'
                        assert switch_execute_data_source_rules == True
                        #结果展示规则列表   
                        if API0000_diy().get_sub_list_index(parent_list, sub_list3) == True:
                            switch_execution_result_display_rules = True
                            message = ''  
                        else:
                            switch_execution_result_display_rules = False
                            message = event_name + '-' + deal_name + '-' + actual_deal_name + '下‘结果展示规则’缺失！'  
                        assert switch_execution_result_display_rules == True 
                        
            #存入dict
            dict['logInfo']['code'] = '1'   #断言是否异常
            is_success = True
            #logging.warning(API0000_diy().text_conversion('解析专家系统规则引擎_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            print message
            #step3:定义轨迹信息+检查点信息:
            dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return dict



    #各具象交易规则引擎配置（数据源规则、执行数据源规则、结果展示规则）
    @func_set_timeout(30)#设定函数超执行时间
    def rule_engine_configuration(self, actual_deal_code, actual_deal_list, dict):
        try:
            #初始化
            is_success = ''
            message = ''
            unit = '各具象交易规则引擎配置'
            data_source_rule = []   #初始化-数据源规则配置列表
            execute_data_source_rules = []   #初始化-执行数据源规则配置列表
            execution_result_display_rules = []   #初始化-执行结果展示规则配置列表

            #配置:各具象交易下的数据源规则、执行数据源规则、执行结果展示规则
            if actual_deal_code == 'actual_deal_code1':   #具象交易：获取jira待分配任务
                #配置-数据源规则配置列表
                data_source_rule = ['data_source_rule1','data_source_rule2']  
                #配置-数据源规则配置列表 
                execute_data_source_rules = ['execute_data_rule1']   
                #配置-数据源规则配置列表
                execution_result_display_rules = ['result_display_rule1']   
            elif actual_deal_code == 'actual_deal_code2':   #具象交易：获取待分配任务的处理人
                #配置-数据源规则配置列表
                data_source_rule = ['data_source_rule1','data_source_rule3']  
                #配置-数据源规则配置列表 
                execute_data_source_rules = ['execute_data_rule3']   
                #配置-数据源规则配置列表
                execution_result_display_rules = ['result_display_rule2']  
            elif actual_deal_code == 'actual_deal_code3':   #具象交易：jira分配任务
                #配置-数据源规则配置列表
                data_source_rule = ['data_source_rule1','data_source_rule4','data_source_rule5']  
                #配置-数据源规则配置列表 
                execute_data_source_rules = ['execute_data_rule5']   
                #配置-数据源规则配置列表
                execution_result_display_rules = ['result_display_rule3'] 
            else:
                #配置-数据源规则配置列表
                data_source_rule = []  
                #配置-数据源规则配置列表 
                execute_data_source_rules = []   
                #配置-数据源规则配置列表
                execution_result_display_rules = [] 
            #存入dict
            dict['logInfo']['code'] = '1'   #断言是否异常
            is_success = True
            #logging.warning(API0000_diy().text_conversion('解析专家系统规则引擎_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            #step3:定义轨迹信息+检查点信息:
            dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return data_source_rule, execute_data_source_rules, execution_result_display_rules, dict



    #读取数据源规则
    @func_set_timeout(30)#设定函数超执行时间
    def read_data_source_rules(self, actual_deal_list, dict):
        try:
            #执行记录日志
            #logging.warning(API0000_diy().text_conversion('&nbsp'*4 + '规则解析初始化' + '_开始'))
            #初始化
            if 'win' in sys.platform:   #windows系统平台
                db_ip = 'localhost'         # 定义数据库ip
                db_user = 'root'      # 定义数据库用户名
                db_password = 'abc123456'   # 定义数据库密码
                db_name = 'test'   #库名
            else:
                db_ip = '10.8.1.157'         # 定义数据库ip
                db_user = 'expert'      # 定义数据库用户名
                db_password = 'expert123'   # 定义数据库密码
                db_name = 'dbexpert'   #库名

            data_rule_list1 = []   #初始化传入的数据规则列表
            data_rule_list2 = []   #初始化传入的执行数据规则列表
            data_rule_list3 = []   #初始化传入的结果约定展示规则列表
            
            resule_list1 = []   #初始化传出的数据源列表
            resule_list2 = []   #初始化传出的执行数据源列表
            resule_list3 = []   #初始化传出的结果约定展示列表
            
            
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：读取数据源规则'  # 节点
            #step1:传入的3类规则存入列表中
            for i in range(0,len(actual_deal_list[0]['rule'])):
                if actual_deal_list[0]['rule'][i]['type'] == '1':   #传入的数据规则列表
                    data_rule_list1.append(actual_deal_list[0]['rule'][i]['name'])
                elif actual_deal_list[0]['rule'][i]['type'] == '2':   #传入的执行数据规则列表
                    data_rule_list2.append(actual_deal_list[0]['rule'][i]['name'])
                elif actual_deal_list[0]['rule'][i]['type'] == '3':   #传入的结果约定展示规则列表
                    data_rule_list3.append(actual_deal_list[0]['rule'][i]['name'])
            #存入list_total列表
            list_total = [data_rule_list1, data_rule_list2, data_rule_list3]   #传入的总列表
            #print list_total

            #step2:解析3类规则录入的内容，依次存入3个列表中，作为出参。
            # 连接数据库
            try:
                db = pymysql.connect(host=db_ip, user=db_user, passwd=db_password, db=db_name, charset='utf8')
            except Exception, e:
                is_success = False
                message = '专家系统数据库连接失败，请检查用户名、密码、ip、库名是否有误！'
            assert is_success != False

            #遍历3类规则
            for aa in range(0,len(list_total)):
                #遍历每类规则的每条规则
                for x in range(0, len(list_total[aa])):
                    #初始化
                    resule_json = {}   #初始化不同交易的数据源结果列表json串
                    rule_name = list_total[aa][x]   #遍历出规则名

                    #读取规则/公式
                    try:
                        cursor = db.cursor()
                        sql = "SELECT body FROM exp_rule_body WHERE descript='" + rule_name + "'"
                        cursor.execute(sql)
                        print '--------'
                        print sql
                        result = cursor.fetchall()
                    except Exception, e:
                        is_success = False
                        message = '专家系统数据库查询规则报错，检查sql是否错误或表中的规则是否存在！'
                    assert is_success != False

                    if result == []:
                        is_success = False
                        message = '专家系统数据库查询规则，查询结果为空！'
                    else:
                        rule = result[0][0]   #获取规则
                        #print rule
                    assert is_success != False

                    #解析规则/公式
                    try:
                        #按行拆分
                        if rule[-4:] == '<br>':
                            rule = rule[0:-4]   #去除末尾的<br>
                        #rule = rule.replace("\\'", "'")
                        rule_list = rule.split("<br>")   #按行拆分
                        #规则解析
                        for i in range(0,len(rule_list)):

                            #解析1：拆分出变量及变量值
                            rule_value_list = rule_list[i].split(" ==")   #按照=拆分为数组

                            #解析2：规则去除html格式
                            pattern = re.compile(r'<[^>]+>',re.S)
                            rule_value_temp = pattern.sub('', rule_value_list[0])
                            rule_value_temp = rule_value_temp.replace("&nbsp;", "")    #&nbsp;转换空
                            #print rule_value_temp
                            #print rule_value_list[1]
                            #print rule_value_list

                            #解析3：存值
                            if rule_value_temp == '数据源-jira用户登录信息 . jira用户名':
                                resule_json['jira_name'] = rule_value_list[1]
                            elif rule_value_temp == '数据源-jira用户登录信息 . jira登录密码':
                                resule_json['jira_password'] = rule_value_list[1]
                            elif rule_value_temp == '数据源-jira用户登录信息 . jira地址':
                                resule_json['jira_url'] = rule_value_list[1]
                            elif rule_value_temp == '数据源-jira查询命令JQL . 查询jira待分配任务JQL':
                                resule_json['JQL1'] = rule_value_list[1]
                            elif rule_value_temp == '执行数据源 . 功能名称':
                                resule_json['function_name'] = rule_value_list[1]             
                            elif rule_value_temp == '结果展示约定 . 展示格式':
                                resule_json['type'] = rule_value_list[1]
                            elif rule_value_temp == 'jira任务信息 . jira任务列表信息':
                                resule_json['jira_list'] = rule_value_list[1]
                            elif rule_value_temp == 'jira任务信息 . 任务编号':
                                resule_json['jira_list'] = rule_value_list[1]
                            elif rule_value_temp == 'jira任务信息 . 待分配人':
                                resule_json['assigned_person'] = rule_value_list[1]

                            #print(rule_value_list[0].decode('utf-8').encode('gb18030'))
                        #print resule_json
                        #3类规则分别存入列表中
                        if aa == 0:
                            resule_list1.append(resule_json)
                        elif aa == 1:
                            resule_list2.append(resule_json)
                        elif aa == 2:
                            resule_list3.append(resule_json)    
                    except Exception, e:
                        is_success = False
                        message = '结息规则/公式报错！'
                        print traceback.format_exc()
                    assert is_success != False
                    #print resule_list1
            #存入dict
            is_success = True
            dict['logInfo']['code'] = '1'
            #执行记录日志
            #logging.warning(API0000_diy().text_conversion('&nbsp'*4 + '规则解析初始化' + '_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            #print '------'
            #print resule_list1
            #print '------'
            #print resule_list2
            #print '------'
            #print resule_list3
            #print '------'
            #print list_total
            pass
            #step3:定义轨迹信息+检查点信息:
            #dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return resule_list1, resule_list2, resule_list3, list_total, dict





    #step2:执行脚本
    @func_set_timeout(30)#设定函数超执行时间
    def execute_script(self, flow_list, dict):
        try:
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            input_parameter = ''   #入参
            out_parameter = []    #存储出参lib

            #遍历事件
            for i in range(0, len(flow_list)): 
                event_name = flow_list[i]['name']   #事件名称
                logging.warning(API0000_diy().text_conversion('&nbsp'*4 + '事件-' + event_name))
                #遍历交易
                for j in range(0, len(flow_list[i]['deal'])):
                    deal_name = flow_list[i]['deal'][j]['name']   #交易名称
                    logging.warning(API0000_diy().text_conversion('&nbsp'*8 + '交易-' + deal_name))
                    #1.读取数据源规则
                    logging.warning(API0000_diy().text_conversion('&nbsp'*16 + '读取数据源规则'))
                    rs = self.read_data_source_rules(flow_list[i]['deal'][0]['actual_deal'], dict)
                    data_source1 = rs[0]   #数据源录入规则列表
                    data_source2 = rs[1]   #执行数据源规则列表
                    data_source3 = rs[2]   #结果约定展示规则列表
                    list_total = rs[3]   #含三类规则列表
                    dict = rs[4]
                    print data_source1

                    #支持入参格式。由[{},{}]转换为{}。
                    input_parameter = data_source1   #入参。格式：[{},{}]
                    dict_initial = {}   #空字典
                    for i in range(0,len(input_parameter)):
                        dict_initial.update(input_parameter[i])   #遍历data_source1，把字典追加空字典中
                    input_parameter = dict_initial
                    #参数化传值
                    for x1 in input_parameter:   #遍历入参列表下每个字典
                        if '【' in input_parameter[x1]:   #如果字典中value值含'【'，启动参数化传值
                            value = input_parameter[x1].encode('utf-8')   #获取字典中含'【'的value值
                            #print value
                            #print type(value)
                            value = value.replace('【', '')
                            value = value.replace('.出参】', '')
                            #print value.decode('utf-8').encode('gb2312')
                            #匹配出参列表中，是否存在交易名与value值相等的
                            for x2 in range(0,len(out_parameter)):
                                if value == out_parameter[x2]['deal_name']:
                                    input_parameter[x1] = out_parameter[x2]['result']            
                    #if i == 1:
                        #exit()
                    #2.执行数据源规则
                    for m in range(0,len(data_source2)):
                        #记录执行日期
                        logging.warning(API0000_diy().text_conversion('&nbsp'*16 + '执行数据源-' + data_source2[m]['function_name']))
                        #匹配功能脚本
                        if data_source2[m]['function_name'] == 'jira查询任务信息':
                            rs = self.jira_query(input_parameter, dict)
                        elif data_source2[m]['function_name'] == '获取jira待分配任务处理人':
                            rs = self.get_assigned_person(input_parameter, dict)
                        elif data_source2[m]['function_name'] == 'jira任务分配':
                            rs = self.jira_assign(input_parameter, dict)
                        #获取出参
                        run_result = rs[0]   #执行结果
                        display_result = rs[1]   #展示结果
                        dict = rs[2]
                        out_parameter.append({'deal_name':deal_name, 'result':run_result})   #执行结果存入data_source_list列表

                    #3.展示结果约定规则
                    logging.warning(API0000_diy().text_conversion('&nbsp'*16 + '展示结果约定规则'))
                    display_type = data_source3[0]['type']   #展示类型
                    dist = API0000_expert_sys().display_result_agreement(event_name, deal_name, display_type, display_result, dict)
                dict['logInfo']['code'] = '1'
            #print data_source
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            pass
            #step3:定义轨迹信息+检查点信息:
            #dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return dict




    #结果展示
    @func_set_timeout(30)#设定函数超执行时间
    def display_result_agreement(self, event_name, deal_name, display_type, display_result, dict):
        try:
            #初始化
            result_list = []    #前端展示值（列表）result_list
            #每个交易第一条展示。（事件名-交易名-执行结果：）
            result_list.append({ 'type':'title', 'titleName':("[事件：" + event_name + "]" + '-' + "[交易：" + deal_name + "]" + '-' + '执行结果展示')})
            #step1:定义展示内容
            if display_type == '文本':
                #每个交易第二条展示：
                result_list.append({'type':'title', 'titleName':display_result['text']})
            elif display_type == '表格':
                #每个交易第二条展示：             
                if deal_name == '获取jira待分配任务':
                    tableHead = ['序号','任务编号','类型','提出人','概要名称']
                elif deal_name == '获取待分配任务的处理人':
                    tableHead = ['序号','任务编号','类型','提出人','概要名称','分配人']
                elif deal_name == 'jira分配任务':
                    tableHead = ['序号','任务编号','分配人','是否分配成功']
                #传前端值
                result_list.append({'type':'table', 'tableHead':tableHead,'trList':display_result['table']})

            #step2:调‘执行任务-记录步骤执行信息’接口"
            icode = 'EXEC107'
            execNo = dict['execNo']   #执行编号
            #execNo = '2021090381000000577'
            stepNo = '02'   #步骤编号
            stepTitle = '展示执行结果'   #步骤名称
            stepParams = {}   #步骤执行参数:json字符 ,无则传 “{}”
            stepResult = result_list   #步骤执行结果:json字符 ,无则传 “{}”
            stepError = {}   #步骤执行错误信息:json字符 ,无则传 “{}”
            response = API0000_expert_sys().run_xjob(icode, execNo, stepNo, stepTitle, stepParams, stepResult, stepError)
            if response['header']['istatus'] == '1':
                is_success = True
                message = ''
            else:
                is_success = False
                message = '不成功' 
            assert is_success == True
            #存入dict
            dict['logInfo']['code'] = '1'
            #执行记录日志
            #logging.warning(API0000_diy().text_conversion('&nbsp'*4 + '规则解析初始化' + '_结束'))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            print traceback.format_exc()
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            is_success = False
            dict['logInfo']['code'] = '0'
            message = '执行时间超时退出'
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            pass
            #step3:定义轨迹信息+检查点信息:
            #dict = API0000_diy().result(dict, is_success, message, unit)
            #print dict['logInfo']['msg']
        return dict




    #查询jira任务信息
    def jira_query(self, input_parameter, dict):
        try:
            #初始化
            run_result = []   #执行结果
            display_result = {'text':'','table':[]}   #展示结果

            #获取数据源规则内容
            jira_url =  input_parameter['jira_url']   #jira地址
            jira_name = input_parameter['jira_name']   #jira登录用户名
            jira_password = input_parameter['jira_password']   #jira登录密码
            JQL = input_parameter['JQL1']   #jira登录密码
            print jira_name

            #step1：连接jira
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：查询jira任务信息-连接jira'  # 节点
            try:
                jira = JIRA(auth=(jira_name, jira_password), options={'server': jira_url})
            except Exception, e:
                is_success = False
                message = '连接jira失败，请检查用户名、密码、jira地址是否正确！'
            assert is_success != False

            #step2：获取查询到的任务对象
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：查询jira任务信息-获取查询到的任务对象'  # 节点
            #查询任务
            try:
                issues = jira.search_issues(JQL, maxResults=-1)
            except Exception, e:
                is_success = False
                message = 'JQL查询任务列表对象报错！'
            assert is_success != False

            #step3：遍历issues_list列表（测试组未完成的'待测试'任务），并展示
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：查询jira任务信息-遍历issues_list列表'  # 节点
            for j in range(0,len(issues)):
                #遍历获取问题对象
                try:
                    issue = jira.issue(issues[j])
                except Exception, e:
                    is_success = False
                    message = 'jira获取问题对象报错！'
                assert is_success != False

                #问题key
                if issue.key == None:
                    issue_key = ''
                else:
                    issue_key = issue.key

                #项目名称
                if jira.project(issue.fields.project).name == None:
                    project_name = ''
                else:
                    project_name = jira.project(issue.fields.project).name  

                #任务类型
                if issue.fields.issuetype.name == None:
                    issue_issuetype = ''
                else:
                    issue_issuetype = issue.fields.issuetype.name

                #提出人
                if str(issue.fields.assignee) == 'None':
                    issue_assignee = ''
                else:
                    issue_assignee = str(issue.fields.assignee)

                #概要名称
                if issue.fields.summary == None:
                    issue_summary = ''
                else:
                    issue_summary = issue.fields.summary.encode('utf-8').replace(u'\xa0', u' ')

                #任务状态
                if issue.fields.status == None:
                    issue_status = ''
                else:
                    issue_status = str(issue.fields.status)

                #存入执行结果run_result
                run_result.append({})
                run_result[-1]['issue_key'] = issue_key 
                run_result[-1]['project_name'] = project_name 
                run_result[-1]['issue_issuetype'] = issue_issuetype
                run_result[-1]['issue_assignee'] = issue_assignee
                run_result[-1]['issue_summary'] = issue_summary
                run_result[-1]['issue_status'] = issue_status   

                #存入展示结果display_result
                #1.表格格式
                display_result['table'].append([])
                display_result['table'][-1].append(str(j + 1))
                display_result['table'][-1].append(issue_key)
                display_result['table'][-1].append(issue_issuetype)
                display_result['table'][-1].append(issue_assignee)
                display_result['table'][-1].append(issue_summary)
                #2.文本格式
            display_result['text'] = '成功获取到' + str(j+1) + '条任务。'
            print '-------'
            print display_result
            print '-------'
            dict['logInfo']['code'] = '1'   #断言是否异常
            is_success = True
            #logging.warning(API0000_diy().text_conversion('结束'))
            #print dict
        except Exception, e:
            dict['logInfo']['code'] = '0'
            is_success = False
            print traceback.format_exc()
            #记录异常日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            dict['logInfo']['code'] = '0'
            is_success = False
            message = '执行时间超时退出'
            #记录执行时间超时日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            dict = API0000_diy().result(dict, is_success, message, unit)
        return run_result,display_result, dict



    #获取测试任务的待分配人
    def get_assigned_person(self, input_parameter, dict):
        try:
            #初始化
            run_result = []   #执行结果
            display_result = {'text':'','table':[]}   #展示结果

            #获取数据源规则内容
            jira_url =  input_parameter['jira_url']   #jira地址
            jira_name = input_parameter['jira_name']   #jira登录用户名
            jira_password = input_parameter['jira_password']   #jira登录密码
            jira_list = input_parameter['jira_list']   #jira任务lib

            #分析数据源中jira_list格式，若非参数化传值，进行split拆分
            if isinstance(jira_list, unicode):
                jira_list = jira_list.split(',')
            else:
                list_temp = []
                #组装列表
                for i in range(0, len(jira_list)):
                    list_temp.append(jira_list[i]['issue_key'])
                jira_list = list_temp

            #遍历任务
            for i in range(0,len(jira_list)):
                #遍历任务key
                issue_key = jira_list[i]

                #step1：查询jira任务信息
                #节点初始化
                is_success = ''
                message = ''
                unit = '节点：获取测试任务的待分配人-查询jira任务信息'  # 节点
                #连接jira
                try:
                    jira = JIRA(auth=(jira_name, jira_password), options={'server': jira_url})
                except Exception, e:
                    is_success = False
                    message = '连接jira失败，请检查用户名、密码、jira地址是否正确！'
                assert is_success != False

                #查询任务信息
                try:
                    issue = jira.issue(issue_key)
                except Exception, e:
                    is_success = False
                    message = 'jira获取问题对象报错！'
                assert is_success != False

                #问题key
                if issue.key == None:
                    issue_key = ''
                else:
                    issue_key = issue.key

                #项目名称
                if jira.project(issue.fields.project).name == None:
                    project_name = ''
                else:
                    project_name = jira.project(issue.fields.project).name

                #任务类型
                if issue.fields.issuetype.name == None:
                    issue_issuetype = ''
                else:
                    issue_issuetype = issue.fields.issuetype.name 

                #提出人
                if str(issue.fields.assignee) == 'None':
                    issue_assignee = ''
                else:
                    issue_assignee = str(issue.fields.assignee)

                #概要名称
                if issue.fields.summary == None:
                    issue_summary = ''
                else:
                    issue_summary = issue.fields.summary.encode('utf-8').replace(u'\xa0', u' ')

                #任务状态
                if issue.fields.status == None:
                    issue_status = ''
                else:
                    issue_status = str(issue.fields.status)

                #任务概述
                if issue.fields.description == None:
                    issue_description = ''
                else:
                    issue_description = issue.fields.description.encode('utf-8').replace(u'\xa0', u' ')

                if issue_key in ('JK001-18706', 'JK001-18707'):
                    assigned_person = 'xingchunyu'
                else:
                    assigned_person = 'chunqing.zhang'

                #存入执行结果run_result
                run_result.append({})
                run_result[-1]['issue_key'] = issue_key 
                run_result[-1]['project_name'] = project_name 
                run_result[-1]['issue_issuetype'] = issue_issuetype
                run_result[-1]['issue_assignee'] = issue_assignee
                run_result[-1]['issue_summary'] = issue_summary
                run_result[-1]['issue_status'] = issue_status  
                run_result[-1]['assigned_person'] = assigned_person 

                #存入展示结果display_result
                #1.表格格式
                display_result['table'].append([])
                display_result['table'][-1].append(i + 1)
                display_result['table'][-1].append(issue_key)
                display_result['table'][-1].append(issue_issuetype)
                display_result['table'][-1].append(issue_assignee)
                display_result['table'][-1].append(issue_summary) 
                display_result['table'][-1].append(assigned_person)
                #2.文本格式
            display_result['text'] = '成功获取到' + str(i + 1) + '条任务的待分配人。'
            dict['logInfo']['code'] = '1'   #断言是否异常
            is_success = True
            #logging.warning(API0000_diy().text_conversion('结束'))
            #print dict
        except Exception, e:
            dict['logInfo']['code'] = '0'
            is_success = False
            print traceback.format_exc()
            #记录异常日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            dict['logInfo']['code'] = '0'
            is_success = False
            message = '执行时间超时退出'
            #记录执行时间超时日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            dict = API0000_diy().result(dict, is_success, message, unit)
        return run_result, display_result, dict



    #jira任务分配
    @func_set_timeout(30)#设定函数超执行时间
    def jira_assign(self, input_parameter, dict):
        try:
            #初始化
            run_result = []   #执行结果
            display_result = {'text':'','table':[]}   #展示结果

            #获取数据源规则内容
            jira_url =  input_parameter['jira_url']   #jira地址
            jira_name = input_parameter['jira_name']   #jira登录用户名
            jira_password = input_parameter['jira_password']   #jira登录密码
            assigned_person_list = input_parameter['assigned_person']   #待分配人
            jira_list = input_parameter['jira_list']   #jira任务

            #分析数据源中任务jira_list格式，若非参数化传值，进行split拆分
            if isinstance(jira_list, unicode):
                jira_list = jira_list.split(',')
            else:
                list_temp = []
                #组装列表
                for i in range(0, len(jira_list)):
                    list_temp.append(jira_list[i]['issue_key'])
                jira_list = list_temp

            #分析数据源中待分配人issue_assignee格式，若非参数化传值，进行split拆分
            if isinstance(assigned_person_list, unicode):
                assigned_person_list = assigned_person_list.split(',')
            else:
                list_temp = []
                #组装列表
                for i in range(0, len(assigned_person_list)):
                    list_temp.append(assigned_person_list[i]['assigned_person'])
                assigned_person_list = list_temp

            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：jira任务分配'  # 节点

            #连接jira
            try:
                jira = JIRA(auth=(jira_name, jira_password), options={'server': jira_url})
            except Exception, e:
                is_success = False
                message = '连接jira失败，请检查用户名、密码、jira地址是否正确！'
            assert is_success != False

            #jira任务分配
            for i in range(0, len(jira_list)):
                if len(assigned_person_list) == 1:
                    assigned_person = assigned_person_list[0]
                elif len(assigned_person_list) == len(jira_list):
                    assigned_person = assigned_person_list[i]
                #获取任务对象    
                issue = jira.issue(jira_list[i])
                issue.update(
                        {
                            "assignee": {"name": assigned_person}
                        }
                    )
                #存入展示结果display_result
                #1.表格格式
                display_result['table'].append([])
                display_result['table'][-1].append(i+1)
                display_result['table'][-1].append(jira_list[i])
                display_result['table'][-1].append(assigned_person)
                display_result['table'][-1].append('成功')
                #2.文本格式
            display_result['text'] = '成功分配' + str(1) + '条任务。'

            dict['logInfo']['code'] = '1'   #断言是否异常
            is_success = True
            #logging.warning(API0000_diy().text_conversion('结束'))
            #print dict
        except Exception, e:
            dict['logInfo']['code'] = '0'
            is_success = False
            print traceback.format_exc()
            #记录异常日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp异常位置:\n')  + traceback.format_exc())
        except func_timeout.exceptions.FunctionTimedOut:
            dict['logInfo']['code'] = '0'
            is_success = False
            message = '执行时间超时退出'
            #记录执行时间超时日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        finally:
            dict = API0000_diy().result(dict, is_success, message, unit)
        return run_result, display_result, dict








