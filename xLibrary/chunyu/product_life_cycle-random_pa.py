# coding=utf-8

import os
import sys
import time
import json
import random
import openpyxl
import traceback
import webbrowser
reload(sys)
sys.setdefaultencoding("utf8")
from openpyxl import load_workbook
from xLibrary.chunyu.API0000 import *

"""
一、获取excel数据
二、组装生命周期列表:   get_excel_date(filename)-->出参business_list, sum_business, product_life_cycle_list
    step1:获取去重复后全部业务场景(business_list)
          1.初始化业务列表
          2.追加续期业务
          3.追加保全项目
    step2:组装生命周期列表
          1.获取生命周期年数years
          2.组装。   product_life_cycle_list格式[[{'business':'','apply_date':''},{},{}],[{},{},{}]],product_life_cycle_list[i]为每一年;product_life_cycle_list[i][j]每一年内每个业务)
              (1)列表第一层级:生命周期年度
              (2)列表第二层级:每个年度内业务（业务名、申请日期）
                 1)随机获取保全项目次数:退保的保全次数1次、贷款还款小于等于贷款次数、其他保全项目小于excel中定义
                 2)得到随机保全项的随机申请日期:退保申请日期在最后业务申请日期与满期之间、贷款还款申请日期随机在两个贷款申请日期之间、其他保全项目随机申请日期
                 注:贷款还款申请日期:第一次还款申请日期处于第一次贷款申请日期与满期之间、非第一次还款:1.贷款还款列表中最后一次业务为贷款还款，不生成还款；2.贷款还款列表中最后一次业务为非贷款还款，申请日期处于最后一次还款贷款申请日期之后的贷款申请日期与满期之间
              (3)判断随机保全项的随机申请日期属于保单的第几年度
              (4)随机保全项业务存入生命周期列表。
                 1)根据业务申请日期年度存入生命周期列表第一层级对应的年度末尾
                 2)每一年度内业务按申请日期排序
    step3:获取业务数总和数(sum_business)
三、创建并组装场景脚本.py文件:    create_process_file(filename, business_list, product_life_cycle_list) 
    step1:excel获取业务参数
    step2:创建场景脚本.py文件
    step3:组装场景脚本.py文件
    step4:执行场景脚本.py文件
四、创建并组装测试报告.html文件
    step1:生成测试报告
    step2:执行脚本
"""
class random_pa_test:

    def __init__(self):
        pass

    def run(self, py_file_path, excle_file_path, html_file_path, execNo, input_dict):
        try:
            #初始化
            list_cycle = []   #生命周期（业务数总和+生命周期列表）列表
            env_name = '-'
            run_type = '-'
            product_life_cycle_times = '1'
            #判断是否传入命令行第二个的参数-json串。未传入取excel表格中数据；传入则获取命令行第二个的参数
            if input_dict != {}:
                temp1_input_dict = input_dict
                temp1_input_dict = temp1_input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
                if 'win' in sys.platform:
                    temp1_input_dict = temp1_input_dict.decode('gb2312').encode('utf-8')   #把命令行参数的gb2312编码格式转换为utf-8格式
                temp1_input_dict = json.loads(temp1_input_dict)   #json转dict格式
                filename =  temp1_input_dict['filename']   #生命周期循环次数
            #一、获取excel数据
            wb = load_workbook(filename, data_only=True)
            sh = wb['random_pa']   #产品生命周期&随机保全项sheet
            product_life_cycle_times = int(sh['B7'].value)   #生命周期循环次数次数  
            #判断是否传入命令行第二个的参数-json串。未传入取excel表格中数据；传入则获取命令行第二个的参数
            if input_dict != {}:
                temp_input_dict = input_dict
                temp_input_dict = temp_input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
                if 'win' in sys.platform:
                    temp_input_dict = temp_input_dict.decode('gb2312').encode('utf-8')   #把命令行参数的gb2312编码格式转换为utf-8格式
                temp_input_dict = json.loads(temp_input_dict)   #json转dict格式
                product_life_cycle_times = int(temp_input_dict['product_life_cycle_times'])   #生命周期循环次数
            wb.close()
            
            for times in range(0, product_life_cycle_times): 
                #二、组装生命周期列表
                rs = random_pa_test().package_product_life_cycle_list(filename, times, input_dict)    
                business_list = rs[0]   #每次生命周期中业务去重列表
                sum_business = rs[1]   #每次生命周期中业务数总和
                product_life_cycle_list = rs[2]   #每次生命周期中生命周期列表
                years = rs[3]
                list_cycle.append({'sum_business': sum_business, 'product_life_cycle_list': product_life_cycle_list})
                #三、创建并组装场景脚本.py文件
                rs = random_pa_test().create_process_file(filename, py_file_path, business_list, product_life_cycle_list, years, times, sum_business, execNo, input_dict)
                run_type = rs[0]
                env_name = rs[1]
                #四、创建并组装测试报告.html文件
            random_pa_test().create_testreport_file(env_name, product_life_cycle_times, list_cycle, run_type, filename, excle_file_path, html_file_path, execNo) 
        except Exception, e:
            print '---err---------'
            traceback.print_exc()
            random_pa_test().except_report(env_name, product_life_cycle_times, list_cycle, run_type, filename, excle_file_path, html_file_path, execNo)  
        finally:
            #删除多余sheet页，只保留'record'sheet页
            wb = load_workbook(excle_file_path)
            sheets = wb.get_sheet_names()
            for i in range(0, len(sheets)):
                if sheets[i] != 'record':
                    sh = wb[sheets[i]]
                    wb.remove(sh)
            wb.save(excle_file_path)
            wb.close()


    
    #二、组装生命周期列表      
    def package_product_life_cycle_list(self, filename, times, input_dict):
        """step1:获取去重复后全部业务场景"""
        #1.初始化业务列表
        wb = load_workbook(filename, data_only=True)
        sh = wb['random_pa']   #产品生命周期&随机保全项sheet
        sh1 = wb['nb_param']   #nb_param-sheet页
        applicationDate = sh.cell(row=2, column=2).value   #新契约投保日期
        product_life_cycle_times = int(sh['B7'].value)   #生命周期循环次数
        #判断是否传入命令行第二个的参数-json串。未传入取excel表格中数据；传入则获取命令行第二个的参数
        if input_dict != {}:
            input_dict = input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
            if 'win' in sys.platform:
                input_dict = input_dict.decode('gb2312').encode('utf-8')   #把命令行参数的gb2312编码格式转换为utf-8格式
            input_dict = json.loads(input_dict)   #json转dict格式
            applicationDate = input_dict['applicationDate']   #投保日期
            product_life_cycle_times = int(input_dict['product_life_cycle_times'])   #生命周期循环次数
        productCode = sh1.cell(row=28, column=2).value  #产品代码
        business_list = ['新契约']
        num_business = 0   #初始化固定业务类型数
        num_pa = 0   #初始化保全项类型数
        #2.追加固定业务
        for i in range(2, 42):   #遍历列
            value = sh.cell(row=3, column=i).value   #获取每个单元格值
            if value != None:    #如果单元格值不为空，追加到业务列表
                business_list.append(value)
                num_business = num_business + 1
        #3.追加保全项目    
        for i in range(2, 42):   #遍历列
            value = sh.cell(row=5, column=i).value   #获取每个单元格值
            if value != None:    #如果单元格值不为空，追加到业务列表
                business_list.append(value)
                num_pa = num_pa + 1   #保全项类型数
        #print num_pa

        """step2:组装生命周期列表"""
        #1.获取生命周期年数years
        if str(sh1['B31'].value) == '999':
            years = 100
        else:
            years = int(sh1['B31'].value)
        #2.组装
            #(1)列表第一层级:全部年度
        product_life_cycle_list = []   #初始化生命周期列表
        for i in range(0, years):
            product_life_cycle_list.append([])
            #(2)列表第二层级:每个年度内业务
            #新契约
        app_date = applicationDate   #新契约投保日期
        product_life_cycle_list[0].append({'business':sh.cell(row=60, column=1).value, 'apply_date': app_date})  
            #固定业务
        for i in range(2, num_business+2):
            if sh.cell(row=3, column=i).value == '续期': 
                count_business = int(sh.cell(row=4, column=i).value)   #固定业务次数
                for j in range(1, count_business+1):
                    if API0000_diy().isvaild_date(str(int(applicationDate[0:4])+j) + applicationDate[4:10]):#判断组装的日期是否有效
                        apply_date = str(int(applicationDate[0:4])+j) + applicationDate[4:10]   #若拼接的日期有效。日期不变
                    else:
                        apply_date = str(int(applicationDate[0:4])+j) + '-02-28'   
                    product_life_cycle_list[j].append({'business':sh.cell(row=3, column=i).value, 'apply_date': apply_date}) 
            elif sh.cell(row=3, column=i).value == '生存金派发': 
                count_business = int(sh.cell(row=4, column=i).value)   #固定业务次数
                if productCode == '3237':
                    for j in range(1, count_business+1):
                        if API0000_diy().isvaild_date(str(int(applicationDate[0:4])+j+3) + applicationDate[4:10]):#判断组装的日期是否有效
                            apply_date = API0000_diy().date_add_subtract(str(int(applicationDate[0:4])+j+3) + applicationDate[4:10], 1)   #若拼接的日期有效。日期不变
                        else:
                            apply_date = API0000_diy().date_add_subtract((str(int(applicationDate[0:4])+j+3) + '-02-28'), 1) 
                        product_life_cycle_list[j+3].append({'business':sh.cell(row=3, column=i).value, 'apply_date': apply_date}) 
                else:
                    a = 1
            elif sh.cell(row=3, column=i).value == '持续奖金派发':
                count_business = int(sh.cell(row=4, column=i).value)   #固定业务次数
                if productCode in ('1205','1206','3208','3211','5207','5206','3209','5201','5202','5211','5242','5225','5241','8204','5213','3213','8211','8212','8213','8214','5214','8216','8217','3245','3249','3247','3248','3243','3247B','3255','3256','3257CA','3258','3257','3264','3265','3267','3267CO','3265CB','3270','3267CB','8233','8237','8233CA1','3257CA1','8237CA1'):
                    for j in range(1, count_business+1):
                        if API0000_diy().isvaild_date(str(int(applicationDate[0:4])+j+5) + applicationDate[4:10]):#判断组装的日期是否有效
                            apply_date = API0000_diy().date_add_subtract((str(int(applicationDate[0:4])+j+5) + applicationDate[4:10]), 1)   #若拼接的日期有效。日期不变
                        else:
                            apply_date = API0000_diy().date_add_subtract((str(int(applicationDate[0:4])+j+5) + '-02-28'), 1) 
                        product_life_cycle_list[j+5].append({'business':sh.cell(row=3, column=i).value, 'apply_date': apply_date}) 
            #保全项目
        for i in range(2, num_pa+2):
            max_count_pa = int(sh.cell(row=6, column=i).value) 
            count_pa = random.randint(1, max_count_pa) 
                ##随机获取保全项目次数（#退保的保全次数1次、贷款还款小于等于贷款次数、其他保全项目小于excel中定义）
            """
            if sh.cell(row=5, column=i).value == '退保':   #退保的保全次数1次
                count_pa = 1   
            elif sh.cell(row=5, column=i).value == '贷款还款':
                #print 'i:' + str(i)
                if '贷款' not in business_list:
                    count_pa = 0
                elif '贷款' in business_list:              #贷款还款小于等于贷款次数
                    index = business_list.index('贷款')
                    max_loan = sh.cell(row=6, column=index-(len(business_list)-num_pa)+2).value
                    max_loan = int(max_loan)
                    max_count_pa = int(sh.cell(row=6, column=i).value)
                    if max_count_pa <= max_loan:
                        count_pa = random.randint(1, max_count_pa)
                    elif max_count_pa > max_loan:
                        count_pa = random.randint(1, max_loan)
                #print 'count_pa:' + str(count_pa)
            else:                                           #其他保全项目小于excel中定义
                max_count_pa = int(sh.cell(row=6, column=i).value) 
                count_pa = random.randint(1, max_count_pa) 
                #print 'loan_count_pa:' + str(count_pa)
            """
            list_loan = []
            for j in range(0, count_pa):
                    ##得到随机保全项的随机申请日期（退保申请日期在最后业务申请日期与满期之间、贷款还款申请日期随机在两个贷款申请日期之间、其他保全项目随机申请日期）
                """
                if sh.cell(row=5, column=i).value == '退保':        #退保申请日期在最后一个业务申请日期与满期之间
                    for dgc in range(len(product_life_cycle_list), 0, -1):
                        if len(product_life_cycle_list[dgc-1]) > 0:
                            last_list = dgc
                            break
                    start_date = product_life_cycle_list[last_list-1][-1]['apply_date']
                    end_date = API0000_diy().date_add_subtract(sh.cell(row=2, column=2).value, 365*years)
                    pa_apply_date = API0000_diy().random_date(start_date, end_date)
                elif sh.cell(row=5, column=i).value == '贷款还款':   #贷款还款申请日期随机在两个贷款申请日期之间（第一次还款申请日期处于第一次贷款申请日期与满期之间）
                    if j == 0:   #第一次还款前组装全部贷款还款列表list_loan。[{'business':'loan', 'apply_date': product_life_cycle_list[ixx][jxx]['apply_date']},{},{}]
                        for ixx in range(0,len(product_life_cycle_list)):
                            for jxx in range(0,len(product_life_cycle_list[ixx])):
                                index = -1
                                if product_life_cycle_list[ixx][jxx]['business'] == '贷款':
                                    list_loan.append({'business':'loan', 'apply_date': product_life_cycle_list[ixx][jxx]['apply_date']})
  
                    if j == 0:   #第一次还款申请日期处于第一次贷款申请日期与满期之间
                        start_date = list_loan[0]['apply_date']
                        end_date = API0000_diy().date_add_subtract(sh.cell(row=2, column=2).value, 365*years)
                        pa_apply_date = API0000_diy().random_date(start_date, end_date)
                        list_loan.append({'business':'loan_repayment', 'apply_date':pa_apply_date})   
                    elif j != 0: #非第一次还款:1.贷款还款列表中最后一次业务为贷款还款，不生成还款；2.贷款还款列表中最后一次业务为非贷款还款，申请日期处于最后一次还款贷款申请日期之后的贷款申请日期与满期之间
                        if list_loan[-1]['business'] == 'loan_repayment':   #1.贷款还款列表中最后一次业务为贷款还款，不生成还款
                            break
                        else:   #2.贷款还款列表中最后一次业务为非贷款还款，申请日期处于最后一次还款贷款申请日期之后的贷款申请日期与满期之间
                            for slo in range(len(list_loan), 0, -1):
                                if list_loan[slo-1]['business'] == 'loan_repayment':
                                    index1 = slo-1
                                    #print 'index1:' + str(index1)
                                    break
                            start_date = list_loan[index1+1]['apply_date']
                            end_date = API0000_diy().date_add_subtract(sh.cell(row=2, column=2).value, 365*years)
                            pa_apply_date = API0000_diy().random_date(start_date, end_date) 
                            list_loan.append({'business':'loan_repayment', 'apply_date':pa_apply_date})
                    #print 'pa_apply_date:' + pa_apply_date
                    temp = ''
                    for slac in range(0,len(list_loan)):  #贷款还款列表根据申请日期升序排序
                         for scec in range(0,slac):
                            if list_loan[scec]['apply_date'] > list_loan[slac]['apply_date']:
                                temp = list_loan[scec]
                                list_loan[scec] = list_loan[slac]
                                list_loan[slac] = temp
                else:   #非退保、贷款还款。申请日期随机生成
                    start_date = API0000_diy().date_add_subtract(app_date, 30)   #设定随机生成保全项目的初始端日期
                    end_date = API0000_diy().date_add_subtract(app_date, 365*years-100)   #设定随机生成保全项目的结束端日期
                    pa_apply_date = API0000_diy().random_date(start_date, end_date)   #得到随机保全项的随机日期
                """
                start_date = API0000_diy().date_add_subtract(app_date, 30)   #设定随机生成保全项目的初始端日期
                end_date = API0000_diy().date_add_subtract(app_date, 365*years-100)   #设定随机生成保全项目的结束端日期
                pa_apply_date = API0000_diy().random_date(start_date, end_date)   #得到随机保全项的随机日期
                #(3)判断随机保全项的随机申请日期属于保单的第几年度
                eff_date = API0000_diy().date_add_subtract(app_date, 1)   #获取新契约生效日期
                if API0000_diy().isvaild_date(pa_apply_date[0:4] + eff_date[4:10]) == True:
                    anniversary_date = pa_apply_date[0:4] + eff_date[4:10]   #获取随机申请日期所在年度的保单周年日
                else:
                    anniversary_date = pa_apply_date[0:4] + '-03-01'
                if pa_apply_date >= anniversary_date:
                    pa_years = int(pa_apply_date[0:4]) - int(eff_date[0:4]) + 1
                else:
                    pa_years = int(pa_apply_date[0:4]) - int(eff_date[0:4]) 
                #(4)随机保全项业务存入生命周期列表
                product_life_cycle_list[pa_years-1].append({'business':sh.cell(row=5, column=i).value, 'apply_date': pa_apply_date})   #追加到生命周期列表中所在年的最后一列
                temp = 0   #对生命周期列表中所在年业务按申请日期重新排序
                for z in range(0, len(product_life_cycle_list[pa_years-1])):
                    for k in range(0, z):
                        if product_life_cycle_list[pa_years-1][k]['apply_date'] > product_life_cycle_list[pa_years-1][z]['apply_date']:
                            temp = product_life_cycle_list[pa_years-1][k]['apply_date']
                            product_life_cycle_list[pa_years-1][k]['apply_date'] = product_life_cycle_list[pa_years-1][z]['apply_date']
                            product_life_cycle_list[pa_years-1][z]['apply_date'] = temp   
        """step4:获取业务数总和"""
        sum_business= 0  
        for i in range(0, len(product_life_cycle_list)):
            for j in range(0,len(product_life_cycle_list[i])):
                sum_business = sum_business + 1   #全部业务数和
        wb.close()
        return business_list, sum_business, product_life_cycle_list, years   #business_list:去重复后全部业务场景;sum_business:业务数总和数;product_life_cycle_list:组装的产品生命周期列表

        
    #三、创建并组装场景脚本.py文件
    def create_process_file(self, filename, py_file_path, business_list, product_life_cycle_list, years, times, sum_business, execNo, input_dict):
        """step1:excel获取业务参数"""
        file_type = 'product_life_cycle-random_pa'    #初始化文件类型为随机保全项目
        rs = product_life_cycle_test().get_business_param(filename, file_type, input_dict, business_list)
        run_type = rs[0]    #脚本运行类型；手工计算运行；系统组件运行；手工计算+系统运行
        env_name = rs[1]    #环境名称
        loanApplyAmount = rs[2]    #贷款申请金额
        addInvestAmount = rs[3]    #追加保费申请金额
        modify_info_dict = rs[4]    #客户重要资料变更欲修改信息（角色类型、性别、出生日期、证件类型、证件号码、职业代码/名称、个人年收入、家庭年收入、收入来源、收入来源备注、工作单位）
        
        """step2:创建场景脚本.py文件"""   
        #1.创建.py文件
        file = open(py_file_path, 'w')
        
        """step3:组装场景脚本.py文件"""  
        #1.写入导入模块脚本
        message = """#!/usr/bin/python
# coding=utf-8

import sys
import time
import json
import logging
import datetime
import openpyxl
import traceback
from openpyxl import load_workbook
reload(sys)
sys.setdefaultencoding("utf8")"""
        file.write(message + "\n")

        file.write("from xLibrary.chunyu.API0000 import API0000_diy\n")
        #（三种：手工计算类、系统计运行类、手工计算+系统运行类）
        #手工计算类
        if run_type == '手工计算运行':
            #1.import文件
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.calc_script.calc_nb.calc_API0001 import calc_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.calc_script.calc_renew.calc_API0002 import calc_API0002\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0003 import calc_API0003\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_survivalFee.calc_API0004 import calc_API0004\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0005 import calc_API0005\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan.calc_API0006 import calc_API0006\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan_repayment.calc_API0007 import calc_API0007\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0008 import calc_API0008\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.calc_script.calc_CD.calc_API0009 import calc_API0009\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0010 import calc_API0010\n")
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_continuous_bonus.calc_API0011 import calc_API0011\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.calc_script.calc_additional_premium.calc_API0012 import calc_API0012\n")                     
            #2.写入脚本正文    
            file.write("\n\n\n" + "class test:\n")
            file.write("\n\n" + "    def test(self):\n")
                #(1)try内容。
            file.write("        try:\n")
            file.write("            logging.basicConfig(level=logging.WARNING , format='%(message)s  %(asctime)s')\n")
            file.write("            dict = {'policy_info':{'policyNo':''}, 'public':{'execNo':'" + str(execNo) + "'}, 'logInfo':{'msg':''}}\n")
            file.write("            current_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')\n")
            file.write("            times = " + str(times) + "\n")
            file.write("            sum_business = " + str(sum_business) + "\n")
            file.write("            result = -1\n")
            file.write("            execNo = '" + str(execNo) + "'\n")
            file.write("            env_name = '" + env_name + "'\n")
            file.write("            env_name = env_name.encode('utf-8')\n")
            file.write("            applicationDate = '" + str(product_life_cycle_list[0][0]['apply_date']) + "'\n")
            message = """            print '---dict_manual---'
            dict_manual = API0000_diy().define_dict()
            dict_manual['public']['execNo'] = '%s'
            test_type = 'product_life_cycle'
            dict_manual = API0000_diy().store_nbdata(test_type, applicationDate, dict_manual)"""% str(execNo)
            file.write(message + "\n")
            file.write("            dict_manual['public']['filename'] = r'" + str(filename) + "'\n")
            for i in range(0, years):
                file.write("            print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("            result = result + 1\n")
                        file.write("            dict_manual = calc_API0001().calc_nb(dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("            pay_due_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0002().calc_renew(pay_due_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0003().calc_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0004().calc_survivalFee(sendDate, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0005().calc_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("            dict_manual = calc_API0006().calc_loan(apply_date, loanApplyAmount, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0007().calc_loan_repayment(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0008().calc_trial_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0009().calc_CD(apply_date, " + modify_info_dict + ", dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0010().calc_trial_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0011().calc_continuous_bonus(sendDate, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("            dict_manual = calc_API0012().calc_additional_premium(apply_date, addInvestAmount, dict_manual)\n")
                        file.write("            result = result + 1\n")
            file.write("            print '=================end=================================='\n")
            file.write("            print '---dict_manual:---'\n")
            file.write("            print dict_manual\n")
                #(2)except内容 
            message = """        except Exception, e:
            dict_manual['logInfo']['err'] = str(e)
            print '---err-----'
            traceback.print_exc()
            exit()"""
            file.write(message + "\n")
                #(3)finally内容 
            file.write("        finally :\n")
                #结果写入excel
            #filename = 'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'
            if 'win' not in sys.platform:
                file.write("            filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                if int(times) == 0:
                    message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                    file.write(message + "\n")
                    os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("            filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """            wb = load_workbook(filename)
            sheet_name = 'result' + str(times+1)
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n")   
            message = """            if result == sum_business:
                result1 = sum_business
            else:
                result1 = result + 1
            for i in range(0, result1):
                business_apply_date = dict_manual['track_info'][i]['trackTime']   #业务申请日期
                if dict_manual['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。新契约收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount'])                  
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。续期收费金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。犹豫期退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。犹豫期退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice'])               
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。贷款总金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'  
                    if dict_manual['track_info'][i]['msg'] == '': 
                        business_check_result = '验证正确。贷款还款金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) 
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。生存金派发金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。持续奖金派发金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。追加保费收付费:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result
                sh['A' + str(i+6)] = ''
                sh['B' + str(i+6)] = ''
                sh['C' + str(i+6)] = ''
                sh['D' + str(i+6)] = ''"""
            file.write(message + "\n")
            file.write("            sh2 = wb['record']" + "\n")
            file.write("            rows = sh2.max_row" + "\n")
            file.write("            sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("            sh2['C' + str(rows+1)] = json.dumps(dict_manual)" +  "\n")
            file.write("            wb.save(filename)" + "\n")
            file.write("            wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")
        elif run_type == '系统组件运行':
            #1.import文件
            file.write("from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002 import sys_API0002\n")
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.sys_script.nb.sys_API0001 import sys_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.sys_script.renew.sys_API0004 import sys_API0004\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0009 import sys_API0009\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_survivalFee.sys_API0011 import sys_API0011\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.sys_script.tb.sys_API0010 import sys_API0010\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.sys_script.loan.sys_API0007 import sys_API0007\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.sys_script.loan_repayment.sys_API0008 import sys_API0008\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.sys_script.trial_tb.sys_API0013 import sys_API0013\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.sys_script.CD.sys_API0016 import sys_API0016\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0015 import sys_API0015\n")    
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_continuous_bonus.sys_API0020 import sys_API0020\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.sys_script.additional_premium.sys_API0021 import sys_API0021\n")
            #2.写入脚本正文    
            file.write("\n\n\n" + "class test:\n")
            file.write("\n\n" + "    def test(self):\n")
            file.write("        try:\n")
            file.write("            logging.basicConfig(level=logging.WARNING , format='%(message)s  %(asctime)s')\n")
            file.write("            dict = {'policy_info':{'policyNo':''}, 'public':{'execNo':'" + str(execNo) + "'}, 'logInfo':{'msg':''}}\n")
            file.write("            current_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')\n")
            file.write("            times = " + str(times) + "\n")
            file.write("            sum_business = " + str(sum_business) + "\n")
            file.write("            result = -1\n")
            file.write("            execNo = '" + str(execNo) + "'\n")
            file.write("            env_name = '" + env_name + "'\n")
            file.write("            env_name = env_name.encode('utf-8')\n")
            file.write("            applicationDate = '" + str(product_life_cycle_list[0][0]['apply_date']) + "'\n")
                #(1)try内容
            file.write("            print '---dict---'\n")
            file.write("            dict = API0000_diy().define_dict()\n")
            file.write("            dict['public']['filename'] = r'" + str(filename) + "'\n")
            file.write("            dict['public']['execNo'] = '" + str(execNo) + "'\n")
            for i in range(0, years):
                file.write("            print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("            result = result + 1\n")
                        file.write("            dict = sys_API0001().nb(env_name, applicationDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("            times_renew = 1\n")
                        file.write("            dict = sys_API0004().renew(times_renew, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0011().send_survivalFee(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0010().tb(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("            dict = sys_API0007().loan(apply_date, loanApplyAmount, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0008().loan_repayment(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("            validateDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0013().trial_tb(validateDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0016().CD(apply_date, " + modify_info_dict + ", dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0020().send_continuous_bonus(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")   
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("            dict = sys_API0021().additional_premium(apply_date, addInvestAmount, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")   
            file.write("            print '=================end=================================='\n")
            #(2)except内容
            file.write("            print '---dict:---'\n")
            file.write("            print dict\n")
            file.write("            print dict['policy_info']['policyNo']\n")
            message = """        except Exception, e:
                dict['logInfo']['err'] = str(e)
                print '---err------'
                traceback.print_exc()
                print dict
                exit()""" 
            file.write(message + "\n") 
            #(3)finally内容
            file.write("        finally:" + "\n")  
            file.write("            print dict" + "\n")                
            if 'win' not in sys.platform:
                file.write("            filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                if int(times) == 0:
                    message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                    file.write(message + "\n")
                    os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("            filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """            wb = load_workbook(filename)
            sheet_name = 'result' + str(times+1)
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n")
            message = """            for i in range(0,result):
                business_apply_date = dict['track_info'][i]['trackTime']   #业务申请日期
                if dict['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    business_check_result = '新契约收付费金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][i]['trackData']['product'][0]['amount'])                  
                elif dict['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    business_check_result = '续期收费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict_manual['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])               
                elif dict['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    business_check_result = '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                elif dict['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'    
                    business_check_result = '贷款还款金额:' + str(dict['track_info'][i]['trackData']['payment']) 
                elif dict['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                elif dict['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    business_check_result = '追加保费收付费:' + str(dict['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result

            #存在报错场景    
            if result < sum_business:
                if dict['logInfo']['msg'] == None or dict['logInfo']['msg'] == '':
                    msg = str(dict['logInfo']['err'])
                else:
                    msg = str(dict['logInfo']['msg'])
                if result == 0:
                    sh['A' + str(5)] = '新契约'  
                    sh['B' + str(5)] = applicationDate
                    sh['C' + str(5)] = '新契约收付费金额、产品保费、保额' 
                    sh['D' + str(5)] = '验证不通过。' + msg
                else:
                    if dict['track_info'][result]['trackType'] == 'renew':
                        business_name = '续期'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '续期收费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                        business_name = '犹豫期退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'hesitation_tb':
                        business_name = '犹豫期退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result   
                    elif dict['track_info'][result]['trackType'] == 'trial_tb':
                        business_name = '退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、产品现价、其他'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan':
                        business_name = '贷款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan_repayment':
                        business_name = '贷款还款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款还款金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'survivalFee':
                        business_name = '生存金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '生存金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'CD':
                        business_name = '客户重要资料变更'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '客户重要资料变更收付费、保额、保费'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'tb':
                        business_name = '退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、现价'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'continuous_bonus':
                        business_name = '持续奖金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '持续奖金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'additional_premium':
                        business_name = '追加保费'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result"""
            file.write(message + "\n")
            file.write("            sh2 = wb['record']" + "\n")
            file.write("            rows = sh2.max_row" + "\n")
            file.write("            sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("            sh2['B' + str(rows+1)] = json.dumps(dict)" +  "\n")
            file.write("            wb.save(filename)" + "\n")
            file.write("            wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")
        elif run_type == '手工计算+系统运行':
            #1.import文件
                #(1)import手工文件
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.calc_script.calc_nb.calc_API0001 import calc_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.calc_script.calc_renew.calc_API0002 import calc_API0002\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0003 import calc_API0003\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_survivalFee.calc_API0004 import calc_API0004\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0005 import calc_API0005\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan.calc_API0006 import calc_API0006\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan_repayment.calc_API0007 import calc_API0007\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0008 import calc_API0008\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.calc_script.calc_CD.calc_API0009 import calc_API0009\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0010 import calc_API0010\n")
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_continuous_bonus.calc_API0011 import calc_API0011\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.calc_script.calc_additional_premium.calc_API0012 import calc_API0012\n")
                #(2)import系统文件
            file.write("from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002 import sys_API0002\n")
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.sys_script.nb.sys_API0001 import sys_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.sys_script.renew.sys_API0004 import sys_API0004\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0009 import sys_API0009\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_survivalFee.sys_API0011 import sys_API0011\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.sys_script.tb.sys_API0010 import sys_API0010\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.sys_script.loan.sys_API0007 import sys_API0007\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.sys_script.loan_repayment.sys_API0008 import sys_API0008\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.sys_script.trial_tb.sys_API0013 import sys_API0013\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.sys_script.CD.sys_API0016 import sys_API0016\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0015 import sys_API0015\n")    
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_continuous_bonus.sys_API0020 import sys_API0020\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.sys_script.additional_premium.sys_API0021 import sys_API0021\n")
            #2.写入脚本正文    
            file.write("\n\n\n" + "class test:\n")
            file.write("\n\n" + "    def test(self):\n")
            file.write("        try:\n")
            file.write("            logging.basicConfig(level=logging.WARNING , format='%(message)s  %(asctime)s')\n")
            file.write("            dict = {'policy_info':{'policyNo':''}, 'public':{'execNo':'" + str(execNo) + "'}, 'logInfo':{'msg':''}}\n")
            file.write("            current_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')\n")
            file.write("            times = " + str(times) + "\n")
            file.write("            sum_business = " + str(sum_business) + "\n")
            file.write("            result_manual = -1\n")
            file.write("            result_sys = -1\n")
            file.write("            execNo = '" + str(execNo) + "'\n")
            file.write("            env_name = '" + env_name + "'\n")
            file.write("            env_name = env_name.encode('utf-8')\n")
            file.write("            applicationDate = '" + str(product_life_cycle_list[0][0]['apply_date']) + "'\n")
                #(1)try内容
            message = """            try:
                print '---dict_manual---'
                dict_manual = API0000_diy().define_dict()
                dict_manual['public']['execNo'] = '%s'
                test_type = 'product_life_cycle'
                dict_manual = API0000_diy().store_nbdata(test_type, applicationDate, dict_manual)"""% str(execNo)
            file.write(message + "\n")
            file.write("                dict_manual['public']['filename'] = r'" + str(filename) + "'\n")
            for i in range(0, years):
                file.write("                print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                result_manual = result_manual + 1\n")
                        file.write("                dict_manual = calc_API0001().calc_nb(dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                pay_due_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0002().calc_renew(pay_due_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0003().calc_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0004().calc_survivalFee(sendDate, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0005().calc_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("                dict_manual = calc_API0006().calc_loan(apply_date, loanApplyAmount, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0007().calc_loan_repayment(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0008().calc_trial_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0009().calc_CD(apply_date, " + modify_info_dict + ", dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0010().calc_trial_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0011().calc_continuous_bonus(sendDate, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("                dict_manual = calc_API0012().calc_additional_premium(apply_date, addInvestAmount, dict_manual)\n") 
                        file.write("                result_manual = result_manual + 1\n")
            file.write("            finally:\n")            
            file.write("                print '---dict---'\n")
            file.write("                dict = API0000_diy().define_dict()\n")
            file.write("                dict['public']['filename'] = r'" + str(filename) + "'\n")
            file.write("                dict['public']['execNo'] = '" + str(execNo) + "'\n")
            for i in range(0, years):
                file.write("                print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                result_sys = result_sys + 1\n")
                        file.write("                dict = sys_API0001().nb(env_name, applicationDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                times_renew = 1\n")
                        file.write("                dict = sys_API0004().renew(times_renew, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0011().send_survivalFee(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0010().tb(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("                dict = sys_API0007().loan(apply_date, loanApplyAmount, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0008().loan_repayment(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                validateDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0013().trial_tb(validateDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0016().CD(apply_date, " + modify_info_dict + ", dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0020().send_continuous_bonus(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n") 
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("                dict = sys_API0021().additional_premium(apply_date, addInvestAmount, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")                          
                #(2)except内容                     
            file.write("            print '---dict_manual:---'\n")
            file.write("            print dict_manual\n")
            file.write("            print '---dict:---'\n")
            file.write("            print dict\n")
            file.write("            print dict['policy_info']['policyNo']\n")
            message = """        except Exception, e:
            dict['logInfo']['err'] = str(e)
            print '---err------'
            traceback.print_exc()
            print dict
            exit()"""
            file.write(message + "\n")    
                #(3)finally内容
            file.write("        finally:" + "\n") 
            file.write("            try:" + "\n")            
            #filename = 'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'
            if 'win' not in sys.platform:
                file.write("                filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                if int(times) == 0:
                    message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                    file.write(message + "\n")
                    os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("                filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """                wb = load_workbook(filename)
                sheet_name = 'result' + str(times+1)
                wb.create_sheet(title=sheet_name)
                sh = wb[sheet_name]
                result = min(int(result_manual), int(result_sys))
                sh['A1'] = str(result)
                if dict['policy_info']['policyNo'] != '':
                    sh['A2'] = str(dict['policy_info']['policyNo'])
                else:
                    sh['A2'] = '-'
                sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n")
            
            message = """                #两种场景：1.无报错场景时，对全部结果比对。2.存在报错场景时，获取手工计算与系统存有运行的组件数量的最小值，并比对非最后一个组件的结果（最小值-1）
                for i in range(0,result):
                    business_apply_date = dict_manual['track_info'][i]['trackTime']   #业务申请日期
                    if dict_manual['track_info'][i]['trackType'] == 'nb':
                        business_name = '新契约'    #业务名
                        business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '新契约收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][i]['trackData']['product'][0]['amount'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '新契约收付费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['premium'] != dict_manual['track_info'][i]['trackData']['product'][0]['premium']:
                                business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['amount'] != dict_manual['track_info'][i]['trackData']['product'][0]['amount']:
                                business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount']) + ';'                         
                    if dict_manual['track_info'][i]['trackType'] == 'renew':
                        business_name = '续期'
                        business_check_info = '续期收费金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '续期收费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '续期收费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                        business_name = '犹豫期退保试算'
                        business_check_info = '犹豫期退保金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][i]['trackType'] == 'hesitation_tb':
                        business_name = '犹豫期退保'
                        business_check_info = '犹豫期退保金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][i]['trackType'] == 'trial_tb':
                        business_name = '退保试算'
                        business_check_info = '退保金额、产品现价、其他'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                                business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']:
                                business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']:
                                business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem'])   + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']:
                                business_check_result = business_check_result + '累计生息账户, ' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']:
                                business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']:
                                business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']:
                                business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']) + ';'                  
                    if dict_manual['track_info'][i]['trackType'] == 'loan':
                        business_name = '贷款'
                        business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'                        
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '贷款总金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                            if dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']:
                                business_check_result = business_check_result + '贷款额度,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['productCode'] not in ('1205','1206','3208','3209','3211','3245','3247','3247B','3255','3256','3257','3257CA','3264','3265','3265CB','3267','3267CB','3267CO','3270','5201','5202','5206','5207','5211','5213','5214','8214','8216','8217','8233','8237'):
                                if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                                    business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';' 
                            else:
                                if business_check_result == '验证不通过。':
                                    business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:0' + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                    if dict_manual['track_info'][i]['trackType'] == 'loan_repayment':
                        business_name = '贷款还款'
                        business_check_info = '贷款还款金额'    
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '贷款还款金额:' + str(dict['track_info'][i]['trackData']['payment']) 
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '贷款还款金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][i]['trackType'] == 'survivalFee':
                        business_name = '生存金派发'
                        business_check_info = '生存金派发金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '生存金派发金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][i]['trackType'] == 'CD':
                        business_name = '客户重要资料变更'
                        business_check_info = '客户重要资料变更收付费、保额、保费'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过  
                            business_check_result = '验证不通过。'
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '客户重要资料变更收付费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['premium'] != dict_manual['track_info'][i]['trackData']['product'][0]['premium']:
                                business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['amount'] != dict_manual['track_info'][i]['trackData']['product'][0]['amount']:
                                business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount']) + ';' 
                    if dict_manual['track_info'][i]['trackType'] == 'tb':
                        business_name = '退保'
                        business_check_info = '退保金额、现价'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                                business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']:
                                business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']:
                                business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']:
                                business_check_result = business_check_result + '累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']:
                                business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']:
                                business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']:
                                business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']) + ';' 
                    if dict_manual['track_info'][i]['trackType'] == 'continuous_bonus':
                        business_name = '持续奖金派发'
                        business_check_info = '持续奖金派发金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '持续奖金派发金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][i]['trackType'] == 'additional_premium':
                        business_name = '追加保费'
                        business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                        if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '追加保费收付费:' + str(dict['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                        elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                                business_check_result = business_check_result + '追加保费收付费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['addInvestAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']:
                                business_check_result = business_check_result + '追加保费申请金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '追加保费扣费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';' 
                            if dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount']:
                                business_check_result = business_check_result + '实际追加保费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount']) + ';' 
                    sh['A' + str(i+5)] = business_name  
                    sh['B' + str(i+5)] = business_apply_date
                    sh['C' + str(i+5)] = business_check_info
                    sh['D' + str(i+5)] = business_check_result"""
            file.write(message + "\n")
            
            message = """                #两种场景：1.若无报错。不运行此部分。2.若存在报错，获取手工计算与系统计算运行的组件数量的最小值，并比对最后一个组件的结果  
                if result != sum_business:
                    business_apply_date = dict_manual['track_info'][result]['trackTime']   #业务申请日期
                    if dict_manual['track_info'][result]['trackType'] == 'nb':
                        business_name = '新契约'    #业务名
                        business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容           
                    elif dict_manual['track_info'][result]['trackType'] == 'renew':
                        business_name = '续期'
                        business_check_info = '续期收费金额'
                    elif dict_manual['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                        business_name = '犹豫期退保试算'
                        business_check_info = '犹豫期退保金额'
                    elif dict_manual['track_info'][result]['trackType'] == 'hesitation_tb':
                        business_name = '犹豫期退保'
                        business_check_info = '犹豫期退保金额'
                    elif dict_manual['track_info'][result]['trackType'] == 'trial_tb':
                        business_name = '退保试算'
                        business_check_info = '退保金额、产品现价、其他'                 
                    elif dict_manual['track_info'][result]['trackType'] == 'loan':
                        business_name = '贷款'
                        business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    elif dict_manual['track_info'][result]['trackType'] == 'loan_repayment':
                        business_name = '贷款还款'
                        business_check_info = '贷款还款金额'    
                    elif dict_manual['track_info'][result]['trackType'] == 'survivalFee':
                        business_name = '生存金派发'
                        business_check_info = '生存金派发金额'
                    elif dict_manual['track_info'][result]['trackType'] == 'CD':
                        business_name = '客户重要资料变更'
                        business_check_info = '客户重要资料变更收付费、保额、保费'
                    elif dict_manual['track_info'][result]['trackType'] == 'tb':
                        business_name = '退保'
                        business_check_info = '退保金额、现价'
                    elif dict_manual['track_info'][result]['trackType'] == 'continuous_bonus':
                        business_name = '持续奖金派发'
                        business_check_info = '持续奖金派发金额'
                    elif dict_manual['track_info'][result]['trackType'] == 'additional_premium':
                        business_name = '追加保费'
                        business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    if dict['track_info'][result]['msg'] == '' and dict_manual['track_info'][result]['msg'] == '':
                        if dict_manual['track_info'][result]['trackType'] == 'nb':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '新契约收付费金额:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][result]['trackData']['product'][0]['amount'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '新契约收付费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['premium'] != dict_manual['track_info'][result]['trackData']['product'][0]['premium']:
                                    business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['premium']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['amount'] != dict_manual['track_info'][result]['trackData']['product'][0]['amount']:
                                    business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['amount']) + ';'                         
                        if dict_manual['track_info'][result]['trackType'] == 'renew':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '续期收费金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '续期收费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                        if dict_manual['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                        if dict_manual['track_info'][result]['trackType'] == 'hesitation_tb':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                        if dict_manual['track_info'][result]['trackType'] == 'trial_tb':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                    business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']:
                                    business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                    business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                    business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']:
                                    business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem'])   + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']:
                                    business_check_result = business_check_result + '累计生息账户, ' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']:
                                    business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']:
                                    business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']:
                                    business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']) + ';'                  
                        if dict_manual['track_info'][result]['trackType'] == 'loan':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'                        
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '贷款总金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                                if dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountLimit']:
                                    business_check_result = business_check_result + '贷款额度,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                    business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['productCode'] not in ('1205','1206','3208','3209','3211','3245','3247','3247B','3255','3256','3257','3257CA','3264','3265','3265CB','3267','3267CB','3267CO','3270','5201','5202','5206','5207','5211','5213','5214','8214','8216','8217','8233','8237'):
                                    if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                        business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';' 
                                else:
                                    if business_check_result == '验证不通过。':
                                        business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:0' + ';贷款额度:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'])
                        if dict_manual['track_info'][result]['trackType'] == 'loan_repayment':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '贷款还款金额:' + str(dict['track_info'][result]['trackData']['payment']) 
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '贷款还款金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                        if dict_manual['track_info'][result]['trackType'] == 'survivalFee':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '生存金派发金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                        if dict_manual['track_info'][result]['trackType'] == 'CD':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过  
                                business_check_result = '验证不通过。'
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '客户重要资料变更收付费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['premium'] != dict_manual['track_info'][result]['trackData']['product'][0]['premium']:
                                    business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['premium']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['amount'] != dict_manual['track_info'][result]['trackData']['product'][0]['amount']:
                                    business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['amount']) + ';' 
                        if dict_manual['track_info'][result]['trackType'] == 'tb':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                    business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']:
                                    business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                    business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                    business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']:
                                    business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']:
                                    business_check_result = business_check_result + '累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']:
                                    business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']:
                                    business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']:
                                    business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']) + ';' 
                        if dict_manual['track_info'][result]['trackType'] == 'continuous_bonus':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][result]['trackData']['payment'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                business_check_result = business_check_result + '持续奖金派发金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                        if dict_manual['track_info'][result]['trackType'] == 'additional_premium':
                            if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                                business_check_result = '验证正确。' + '追加保费收付费:' + str(dict['track_info'][result]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][result]['trackData']['product'][0]['surrenderAmount'])
                            elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                                business_check_result = '验证不通过。'
                                if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                    business_check_result = business_check_result + '追加保费收付费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['addInvestAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['addInvestAmount']:
                                    business_check_result = business_check_result + '追加保费申请金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                    business_check_result = business_check_result + '追加保费扣费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';' 
                                if dict['track_info'][result]['trackData']['product'][0]['surrenderAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['surrenderAmount']:
                                    business_check_result = business_check_result + '实际追加保费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['surrenderAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['surrenderAmount']) + ';' 
                    elif dict['track_info'][result]['msg'] != '' and dict_manual['track_info'][result]['msg'] != '':
                        if dict_manual['track_info'][result]['msg'] in dict['track_info'][result]['msg']:    #验证通过
                            business_check_result = '验证正确。' + '阻断提示为:' + dict_manual['track_info'][result]['msg']
                        else:
                            business_check_result = '验证不通过。' + '手工计算阻断提示为:' + dict_manual['track_info'][result]['msg'] + ';系统计算阻断提示为:' + dict['track_info'][result]['msg']
                    elif dict['track_info'][result]['msg'] != '' and dict_manual['track_info'][result]['msg'] == '':
                        business_check_result = '验证不通过。' + '手工计算未阻断' + ';系统计算阻断提示为:' + dict['track_info'][result]['msg']
                    elif dict['track_info'][result]['msg'] == '' and dict_manual['track_info'][result]['msg'] != '':
                        business_check_result = '验证不通过。' + '手工计算未阻断提示为:' + dict_manual['track_info'][result]['msg'] + ';系统计算未阻断。'
                    sh['A' + str(result+5)] = business_name  
                    sh['B' + str(result+5)] = business_apply_date
                    sh['C' + str(result+5)] = business_check_info
                    sh['D' + str(result+5)] = business_check_result"""
            file.write(message + "\n")
            file.write("            except Exception, e:" + "\n")
            if 'win' not in sys.platform:
                file.write("                filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                if int(times) == 0:
                    message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                    file.write(message + "\n")
                    os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("                filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """                wb = load_workbook(filename)
                sheet_name = 'result' + str(times+1)
                wb.create_sheet(title=sheet_name)
                sh = wb[sheet_name]
                result = -1
                sh['A1'] = str(result)
                if dict['policy_info']['policyNo'] != '':
                    sh['A2'] = str(dict['policy_info']['policyNo'])
                else:
                    sh['A2'] = '-'
                sh['A3'] = '请仔细检查生命周期excel或者组件代码！'"""
            file.write(message + "\n")
            file.write("            finally:" + "\n")
            file.write("                sh2 = wb['record']" + "\n")
            file.write("                rows = sh2.max_row" + "\n")
            file.write("                sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("                sh2['B' + str(rows+1)] = json.dumps(dict)" +  "\n")
            file.write("                sh2['C' + str(rows+1)] = json.dumps(dict_manual)" +  "\n")
            file.write("                wb.save(filename)" + "\n")
            file.write("                wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")
            
        """step4:执行场景脚本.py文件"""
        file.close()
        os.system(py_file_path)
        return run_type, env_name


    #四、创建并组装测试报告.html文件
    def create_testreport_file(self, env_name, product_life_cycle_times, list_cycle, run_type, filename, excle_file_path, html_file_path, execNo):
        product_life_cycle_times = int(product_life_cycle_times)
        """step1:生成测试报告"""
        wb1 = load_workbook(filename)
        sh1 = wb1['nb_param']   #新契约参数sheet
        product = str(sh1['B28'].value)
        wb1.close()
        #1获取生命周期结果
        wb = load_workbook(excle_file_path)
            #统计执行成功的生命周期数
        num_right_result = 0
        list_result = []
        for ss in range(0, len(list_cycle)):
            product_life_cycle_list = list_cycle[ss]['product_life_cycle_list']
            #print product_life_cycle_list
            sh = wb['result' + str(ss+1)]
            result = str(sh['A1'].value)   #业务编号1~x
            list_result.append(result)
            if result != '-2':
                num_right_result = num_right_result + 1
        #2.创建.html文件
        file = open(html_file_path, 'w')
        
        #3.写入脚本
            #(1)写入测试报告脚本-标题、结果汇总
        message = """<html>
    <head>
        <meta charset='utf-8'>
        <title>产品生命周期</title>
    </head>
    <body>
        <h3>%s产品生命周期测试报告-%s</h3>
        <h6>测试编号:%s</h6>    
        <h4>结果汇总(%s):</h4>	
        <table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='30'>
                <td>生命周期总次数</td>
                <td>生命周期成功次数</td>
            </tr>
            <tr width='30'>"""%(product, run_type, str(execNo), env_name)
        file.write(message + "\n")
        file.write("                <td>" + str(product_life_cycle_times) + "</td>" + "\n")
        file.write("                <td>" + str(num_right_result) + "</td>" + "\n")
        file.write("            </tr>" + "\n")
        file.write("        </table>" + "\n")

            #(2)写入测试报告脚本-生命周期详细
            #每个生命周期的第一行
        max_sum_business = 0    #获取所有生命周期中最多业务数
        for i in range(0, len(list_cycle)):
            if list_cycle[i]['sum_business'] > max_sum_business:
                max_sum_business = list_cycle[i]['sum_business']
        message = """		<h4>生命周期详细:</h4>	
        <table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
            <tr align='center'>
                <td width='300'>生命周期序号</td>"""
        file.write(message + "\n")
            #遍历生命周的每一年，得到每个生命周期中每一年业务数最多的值。
        list_sum = []   #初始化所有生命周期中每年的业务最多数列表
        for xx in range(0,len(list_cycle[0]['product_life_cycle_list'])):
            sum = 0
            for yy in range(0, len(list_cycle)):
                if len(list_cycle[yy]['product_life_cycle_list'][xx]) > sum:
                    sum = len(list_cycle[yy]['product_life_cycle_list'][xx])
            list_sum.append(sum)
            colspan = sum
            file.write("				<td colspan='" + str(colspan) + "'  width='300'>第" + str(xx+1) + "年</td>" + "\n")
        file.write("			</tr>" + "\n")
            #每个生命周期的业务名
        for ss in range(0, len(list_cycle)):
            product_life_cycle_list = list_cycle[ss]['product_life_cycle_list']
            sum_business = int(list_cycle[ss]['sum_business'])
            sh = wb['result' + str(ss+1)]
            result = str(sh['A1'].value)   #业务编号1~x
            msg = sh['A2'].value   #报错信息
            if msg == None:
                msg = ''
            file.write("			<tr align='center'>" + "\n")
            file.write("				<td rowspan='2'>" + str(ss+1) + "</td>" + "\n")
            for i in range(0, len(product_life_cycle_list)):
                if len(product_life_cycle_list[i]) == 0:
                    file.write("				<td width='200'></td>" + "\n")
                    for xcy in range(1, list_sum[i]-len(product_life_cycle_list[i])):
                        file.write("				<td width='200'></td>" + "\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("				<td width='200'>新契约</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("				<td width='200'>续期</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("				<td width='200'>犹豫期退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("				<td width='200'>生存金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("				<td width='200'>退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("				<td width='200'>贷款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("				<td width='200'>贷款还款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("				<td width='200'>退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("				<td width='200'>客户重要资料变更</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("				<td width='200'>犹豫期退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("				<td width='200'>持续奖金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("				<td width='200'>追加保费</td>" + "\n")
                #若果某年的业务数小于其他生命周期某年最少业务数，进行补齐表格        
                if len(product_life_cycle_list[i]) != 0 and j == (len(product_life_cycle_list[i])-1) and len(product_life_cycle_list[i]) < list_sum[i]:
                    for xcy in range(0, list_sum[i]-len(product_life_cycle_list[i])):
                        file.write("				<td width='200'></td>" + "\n")  
            file.write("			</tr>" + "\n")
            
                #是否成功标识
            file.write("			<tr align='center'>" + "\n")  
            if int(result) == -1:   #-1无成功执行业务-全部业务均未执行场景。
                #遍历每年
                for i in range(0, len(product_life_cycle_list)):
                    #若这一年的全部生命周期的业务数都为0时，补充html中表格的列
                    if list_sum[i] == 0:
                        file.write("                <td></td>" + "\n")
                    #对每年的全部生命周期遍历
                    for j in range(0,len(product_life_cycle_list[i])):
                        file.write("                <td>-</td>" + "\n") 
                    #与其他生命周期这一年最大业务数比较，补充html中表格的列
                    for zz in range(0, list_sum[i] - len(product_life_cycle_list[i])):
                        file.write("                <td></td>" + "\n") 
            elif int(result) == 0:   #无成功执行业务-新契约报错场景。
                #遍历每年
                for i in range(0, len(product_life_cycle_list)):
                    #若这一年的全部生命周期的业务数都为0时，补充html中表格的列
                    if list_sum[i] == 0:
                        file.write("                <td></td>" + "\n")
                    #对每年的全部生命周期遍历
                    for j in range(0,len(product_life_cycle_list[i])):
                        #html表格中第一个业务-新契约的单元格录入×
                        if i == 0 and j == 0:
                            file.write("                <td><font size='3' color='red'>×</font></td>" + "\n") 
                        #其他业务单元格录入-
                        else:    
                            file.write("                <td>-</td>" + "\n") 
                    #与其他生命周期这一年最大业务数比较，补充html中表格的列
                    for zz in range(0, list_sum[i] - len(product_life_cycle_list[i])):
                        file.write("                <td></td>" + "\n") 
            elif int(result) > 0 and sum_business > int(result):  #存在成功执行业务-成功业务数值小于总业务数
                z = 0   #初始化数据，每个业务数+1，为了与成功执行业务数result比较，判断哪些业务正确了，哪些失败了，哪些业务未被执行
                #遍历每年
                for i in range(0, len(product_life_cycle_list)): 
                    #若这一年的全部生命周期的业务数都为0时，补充html中表格的列
                    if list_sum[i] == 0:
                        file.write("                <td></td>" + "\n")
                    #对每年的全部生命周期遍历
                    for j in range(0,len(product_life_cycle_list[i])):
                        z = z + 1   #遍历每个业务后+1
                        if z <= int(result)+1:   #判断成功执行的业务。
                            if sh['D' + str(z+4)].value[0:5] != '验证不通过':
                                file.write("                <td>√</td>" + "\n")
                            elif sh['D' + str(z+4)].value[0:5] == '验证不通过':
                                file.write("                <td><font size='3' color='red'>×</font></td>" + "\n")
                        elif z > int(result)+1:  #判断其他未被执行的业务。
                            file.write("                <td>-</td>" + "\n") 
                    #与其他生命周期这一年最大业务数比较，补充html中表格的列
                    for zz in range(0, list_sum[i] - len(product_life_cycle_list[i])):
                        file.write("                <td></td>" + "\n")            
            elif int(result) > 0 and sum_business == int(result):  #存在成功执行业务-成功业务数值等于总业务数
                z = 0  
                for i in range(0, len(product_life_cycle_list)):
                    #若这一年的全部生命周期的业务数都为0时，补充html中表格的列
                    if list_sum[i] == 0:
                        file.write("                <td></td>" + "\n")
                    #对每年的全部生命周期遍历
                    for j in range(0,len(product_life_cycle_list[i])):
                        z = z + 1
                        if sh['D' + str(z+4)].value[0:5] == '验证正确。':
                            file.write("				<td>√</td>" + "\n")
                        elif sh['D' + str(z+4)].value[0:5] == '验证不通过':
                            file.write("				<td><font size='3' color='red'>×</font></td>" + "\n")
                    #与其他生命周期这一年最大业务数比较，补充html中表格的列
                    for zz in range(0, list_sum[i] - len(product_life_cycle_list[i])):
                        file.write("                <td></td>" + "\n")    
            file.write("            </tr>" + "\n")    
        file.write("        </table>" + "\n\n\n")
        
            #(3)写入测试报告脚本-报告详细
        file.write("		<h4>报告详细:</h4>" + "\n")
        for gg in range(0, len(list_cycle)):
            sh = wb['result' + str(gg+1)]
            result = str(sh['A1'].value)
            product_life_cycle_list = list_cycle[gg]['product_life_cycle_list']
            if gg != 0:   #每个生命周期直接加入一行空格
                file.write("		<h1></h1>" + "\n")
            if list_result[gg] == '-2':
                file.write("        <h4>生命周期报错原因:</h4>" + "\n") 
                message = """       <table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='50'>
                <td>接口返回报错信息</td>
                <td>%s;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;请仔细检查生命周期excel！</td>
            </tr>
            <tr width='50'>
                <td>日志返回报错信息</td>
                <td>%s</td>
            </tr>
            </table>"""%(msg, err)
                file.write(message + "\n") 
            else:
                file.write("		<table table border='2' bordercolor='black' width='800' cellspacing='0' cellpadding='5'>" + "\n")
                file.write("			<tr width='30'>" + "\n")
                file.write("				<td colspan='2'>生命周期序号:" + str(gg+1) + "</td>" + "\n")
                file.write("			</tr>" + "\n")
                
                file.write("			<tr width='30'>" + "\n")
                file.write("				<td colspan='2'>保单号:" + sh['A2'].value + "</td>" + "\n")
                file.write("			</tr>" + "\n")

                z = 0
                for i in range(0, len(product_life_cycle_list)):  #每一年
                    file.write("			<tr width='30'>" + "\n")
                    file.write("				<td colspan='2'>第" + str(i+1) + "年</td>" + "\n")
                    file.write("			</tr>" + "\n")
                    for j in range(0, len(product_life_cycle_list[i])): 
                        z = z + 1
                        if len(product_life_cycle_list[i]) > 0:
                            #每一年内每个业务名
                            file.write("			<tr width='30'>" + "\n")
                            if z <= int(result)+1:
                                file.write("				<td rowspan='3'>" + sh['A' + str(z+4)].value + "</td>" + "\n")
                                file.write("				<td>申请日期:" + sh['B' + str(z+4)].value + "</td>" + "\n")
                            else:
                                file.write("				<td rowspan='3'>-</td>" + "\n")
                                file.write("				<td>申请日期:-</td>" + "\n")
                            file.write("			</tr>" + "\n")
                            
                            #验证内容
                            file.write("			<tr width='30'>" + "\n")
                            if z <= int(result)+1:
                                file.write("				<td>验证内容:" + sh['C' + str(z+4)].value + "</td>" + "\n")
                            else:
                                file.write("				<td>-</td>" + "\n")
                            file.write("			</tr>" + "\n")

                            #每一年内每个业务验证结果
                            file.write("			<tr width='30'>" + "\n") 
                            if z <= int(result)+1:
                                file.write("				<td>验证结果:" + sh['D' + str(z+4)].value + "</td>" + "\n")
                            else:
                                file.write("				<td>-</td>" + "\n")
                            file.write("			</tr>" + "\n")
                file.write("		</table>" + "\n")        
        file.write("	</body>" + "\n")
        file.write("</html>" + "\n")

        wb.close()     
        file.close()
        #执行html文件(本地打开，服务端不打开html文件)
        if 'win' in sys.platform:
            os.system(html_file_path)



    def except_report(self, env_name, product_life_cycle_times, list_cycle, run_type, filename, excle_file_path, html_file_path, execNo):
        product_life_cycle_times = int(product_life_cycle_times)
        """step1:生成测试报告"""
        wb1 = load_workbook(filename)
        sh1 = wb1['nb_param']   #新契约参数sheet
        product = str(sh1['B28'].value)
        wb1.close()
        #1获取生命周期结果
        wb = load_workbook(excle_file_path)

        #2.创建.html文件
        file = open(html_file_path, 'w')
        if list_cycle != []:
        #3.写入脚本
            #(1)写入测试报告脚本-标题、结果汇总
            message = """<html>
    <head>
        <meta charset='utf-8'>
        <title>产品生命周期</title>
    </head>
    <body>
        <h3>%s产品生命周期测试报告-%s</h3>
        <h6>测试编号:%s</h6>    
        <h4>结果汇总(%s):</h4>  
        <table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='30'>
                <td>生命周期总次数</td>
                <td>生命周期成功次数</td>
            </tr>
            <tr width='30'>"""%(product, run_type, str(execNo), env_name)
            file.write(message + "\n")
            file.write("                <td>" + str(product_life_cycle_times) + "</td>" + "\n")
            file.write("                <td>-</td>" + "\n")
            file.write("            </tr>" + "\n")
            file.write("        </table>" + "\n")

                #(2)写入测试报告脚本-生命周期详细
                #每个生命周期的第一行
            max_sum_business = 0    #获取所有生命周期中最多业务数
            for i in range(0, len(list_cycle)):
                if list_cycle[i]['sum_business'] > max_sum_business:
                    max_sum_business = list_cycle[i]['sum_business']
            message = """       <h4>生命周期详细:</h4>    
        <table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
            <tr align='center'>
                <td width='300'>生命周期序号</td>"""
            file.write(message + "\n")
                #遍历生命周的每一年，得到每个生命周期中每一年业务数最多的值。
            list_sum = []   #初始化所有生命周期中每年的业务最多数列表

            for xx in range(0,len(list_cycle[0]['product_life_cycle_list'])):
                sum = 0
                for yy in range(0, len(list_cycle)):
                    if len(list_cycle[yy]['product_life_cycle_list'][xx]) > sum:
                        sum = len(list_cycle[yy]['product_life_cycle_list'][xx])
                list_sum.append(sum)
                colspan = sum
                file.write("                <td colspan='" + str(colspan) + "'  width='300'>第" + str(xx+1) + "年</td>" + "\n")
            file.write("            </tr>" + "\n")
                #每个生命周期的业务名
            for ss in range(0, len(list_cycle)):
                product_life_cycle_list = list_cycle[ss]['product_life_cycle_list']
                sum_business = int(list_cycle[ss]['sum_business'])
                file.write("            <tr align='center'>" + "\n")
                file.write("                <td rowspan='2'>" + str(ss+1) + "</td>" + "\n")
                for i in range(0, len(product_life_cycle_list)):
                    if len(product_life_cycle_list[i]) == 0:
                        file.write("                <td width='200'></td>" + "\n")
                        for xcy in range(1, list_sum[i]-len(product_life_cycle_list[i])):
                            file.write("                <td width='200'></td>" + "\n")
                    for j in range(0, len(product_life_cycle_list[i])):
                        if product_life_cycle_list[i][j]['business'] == '新契约':
                            file.write("                <td width='200'>新契约</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '续期':
                            file.write("                <td width='200'>续期</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                            file.write("                <td width='200'>犹豫期退保</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                            file.write("                <td width='200'>生存金派发</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '退保':
                            file.write("                <td width='200'>退保</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '贷款':
                            file.write("                <td width='200'>贷款</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                            file.write("                <td width='200'>贷款还款</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '退保试算':
                            file.write("                <td width='200'>退保试算</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                            file.write("                <td width='200'>客户重要资料变更</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                            file.write("                <td width='200'>犹豫期退保试算</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                            file.write("                <td width='200'>持续奖金派发</td>" + "\n")
                        elif product_life_cycle_list[i][j]['business'] == '追加保费':
                            file.write("                <td width='200'>追加保费</td>" + "\n")
                    #若果某年的业务数小于其他生命周期某年最少业务数，进行补齐表格        
                    if len(product_life_cycle_list[i]) != 0 and j == (len(product_life_cycle_list[i])-1) and len(product_life_cycle_list[i]) < list_sum[i]:
                        for xcy in range(0, list_sum[i]-len(product_life_cycle_list[i])):
                            file.write("                <td width='200'></td>" + "\n")  
                file.write("            </tr>" + "\n")
                
                    #是否成功标识
                file.write("            <tr align='center'>" + "\n")  
                #遍历每年
                for i in range(0, len(product_life_cycle_list)):
                    #若这一年的全部生命周期的业务数都为0时，补充html中表格的列
                    if list_sum[i] == 0:
                        file.write("                <td></td>" + "\n")
                    #对每年的全部生命周期遍历
                    for j in range(0,len(product_life_cycle_list[i])):
                        file.write("                <td>-</td>" + "\n") 
                    #与其他生命周期这一年最大业务数比较，补充html中表格的列
                    for zz in range(0, list_sum[i] - len(product_life_cycle_list[i])):
                        file.write("                <td></td>" + "\n")          
            file.write("            </tr>" + "\n")    
        file.write("        </table>" + "\n\n\n")
        
            #(3)写入测试报告脚本-报告详细
        file.write("        <h4>生命周期报错:</h4>" + "\n") 
        message = """<table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
        <tr>
            <td>组件检查未成功：</td>
            <td>请仔细检查生命周期excel或者组件代码！</td>
        </tr>
    </table>"""
        file.write(message + "\n")         
        file.write("    </body>" + "\n")
        file.write("</html>" + "\n")

        wb.close()     
        file.close()
        #执行html文件(本地打开，服务端不打开html文件)
        if 'win' in sys.platform:
            os.system(html_file_path)




if __name__ == "__main__":
    #初始化测试编号
    execNo = 'test001'
    input_dict = {}
    #初始化测试编号
    #接收参数：第一位为测试编号
    if len(sys.argv) == 1:
        execNo = sys.argv[1]   #测试编号     
    elif len(sys.argv) > 1:
        execNo = sys.argv[1]   #测试编号 
        input_dict = sys.argv[2]   #输入的json串 

    #初始化
    if 'win' in sys.platform:
        py_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.py'  #组装生成的py文件觉得路径
        excle_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.xlsx'   #执行中临时存储数据的excel文件路径
        html_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.html'   #html测试报告文件路径
    else:
        py_file_path = '/data/xServer/xRunner/' + str(execNo) + '.py'   #组装生成的py文件觉得路径
        excle_file_path = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'   #执行中临时存储数据的excel文件路径
        html_file_path = '/data/xServer/xReport/' + str(execNo) + '.html'   #html测试报告文件路径  



    #调用函数
    random_pa_test().run(py_file_path, excle_file_path, html_file_path, execNo, input_dict)
    #python D:\xLibrary\chunyu\product_life_cycle-random_pa.py test0002 "{'run_type':'手工计算+系统运行','env_name':'预生产','applicationDate':'2016-11-20','product_life_cycle_times':'2','filename':'D:\\xLibrary\\chunyu\\doc\\product_life_cycle\\product_life_cycle_data.xlsx'}"

