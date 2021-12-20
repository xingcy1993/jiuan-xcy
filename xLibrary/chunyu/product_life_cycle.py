# coding=utf-8

import os
import sys
import time
import json
import logging
import openpyxl
import traceback
import webbrowser
reload(sys)
sys.setdefaultencoding("utf8")
from openpyxl import load_workbook
from xLibrary.chunyu.API0000 import *

"""
一、获取excel数据+组装生命周期列表:   get_excel_date(filename)-->出参business_list, sum_business, product_life_cycle_list
step1:获取excel数据
step2:获取去重复后全部业务场景(business_list)
step3:获取业务数总和数(sum_business)
step4:组装生命周期列表(product_life_cycle_list,格式[[{'business':'','apply_date':''},{},{}],[{},{},{}]],product_life_cycle_list[i]为每一年;product_life_cycle_list[i][j]每一年内每个业务)

二、创建并组装场景脚本.py文件:    create_process_file(filename, business_list, product_life_cycle_list) 
step1:excel获取业务参数
step2:创建场景脚本.py文件
step3:组装场景脚本.py文件
step4:执行场景脚本.py文件

三、创建并组装测试报告.html文件
step1:生成测试报告
step2:执行脚本
"""
class product_life_cycle_test:

    def __init__(self):
        pass


    def run(self, py_file_path, excle_file_path, html_file_path, execNo, filename):
        try:
            #一、获取excel数据+组装生命周期列表 
            rs = product_life_cycle_test().get_excel_date(filename)   #一、获取excel数据+组装生命周期列表  
            business_list = rs[0]   #业务列表
            sum_business = rs[1]   #全部业务数和
            product_life_cycle_list = rs[2]   #生命周期列表
            years = rs[3]   #生命周期年数
            #print business_list
            #print sum_business
            #print product_life_cycle_list
            #print years
            #二、创建并组装场景脚本.py文件
            rs = product_life_cycle_test().create_process_file(filename, py_file_path, business_list, product_life_cycle_list, years, sum_business, execNo, input_dict) 
            run_type = rs[0]
            env_name = rs[1]
            #三、创建并组装测试报告.html文件
            product_life_cycle_test().create_testreport_file(env_name, filename, excle_file_path, run_type, sum_business, product_life_cycle_list, html_file_path, execNo) 
        except Exception, e:
            print '---err-------------'
            traceback.print_exc()
            product_life_cycle_test().except_report(excle_file_path, filename, html_file_path, run_type, execNo, env_name, product_life_cycle_list)
        finally:
            #删除多余sheet页，只保留'record'sheet页
            wb = load_workbook(excle_file_path)
            sheets = wb.get_sheet_names()
            for i in range(0, len(sheets)):
                if sheets[i] != 'record':
                    sh = wb[sheets[i]]
                    wb.remove(sh)
            wb.save(excle_file_path)
            wb.close()
            

            
    """一、获取excel数据+组装生命周期列表 
    step1:获取excel数据
    step2:获取去重复后全部业务场景(business_list)
    step3:获取业务数总和数(sum_business)
    step4:组装生命周期列表(product_life_cycle_list,格式[[{'business':'','apply_date':''},{},{}],[{},{},{}]],product_life_cycle_list[i]为每一年;product_life_cycle_list[i][j]每一年内每个业务)
    """
    def get_excel_date(self, filename):
        """step1:获取excel数据"""
        wb = load_workbook(filename)
        sh = wb['product_life_cycle']   #产品生命周期sheet
        columns = sh.max_column   #总列数
        rows = sh.max_row   #总行数
        wb.close() 

        """step2:获取去重复后全部业务场景"""
        business_list = []   #初始化业务列表
        for i in range(3, rows + 1, 2):   #遍历行
            for j in range(3, columns + 1):   #遍历列
                value = sh.cell(row=i, column=j).value   #获取每个单元格值
                if value != None:    #如果单元格值不为空，追加到业务列表
                    business_list.append(value)
        business_list = list(set(business_list))   #业务列表去重(根据set无重复原理)
                        
        """step3:组装生命周期列表"""
        years = 0   #初始化生命周期年数
        #1.获取生命周期年数years
        for i in range(3, rows + 1):  #
            value = sh.cell(row=i, column=1).value
            if value != None:    
                business_list.append(value)
                years = years + 1
        #2.组装
            #(1)列表第一层级:全部年度
        product_life_cycle_list = []   #初始化生命周期列表
        for i in range(0, years):
            product_life_cycle_list.append([])
            #(2)列表第二层级:每个年度内业务
        k = 0
        for i in range(3, rows + 1, 2):
            for j in range(3, columns + 1):
                if sh.cell(row=i, column=j).value != None:
                    product_life_cycle_list[k].append({})
                    product_life_cycle_list[k][j-3]['business'] = sh.cell(row=i, column=j).value   #业务名称
                    product_life_cycle_list[k][j-3]['apply_date'] = sh.cell(row=i+1, column=j).value   #申请日期
            k = k+1

        """step4:获取业务数总和"""
        sum_business= 0  
        for i in range(0, len(product_life_cycle_list)):
            for j in range(0,len(product_life_cycle_list[i])):
                sum_business = sum_business + 1   #全部业务数和
        return business_list, sum_business, product_life_cycle_list, years  #business_list:去重复后全部业务场景;sum_business:业务数总和数;product_life_cycle_list:组装的产品生命周期列表



    """二、创建并组装场景脚本.py文件
    step1:excel获取业务参数
    step2:创建场景脚本.py文件
    step3:组装场景脚本.py文件
    step4:执行场景脚本.py文件
    """
    def create_process_file(self, filename, py_file_path, business_list, product_life_cycle_list, years, sum_business, execNo, input_dict):
        """step1:excel获取业务参数"""
        file_type = 'product_life_cycle'    #初始化文件类型为产品生命周期
        rs = product_life_cycle_test().get_business_param(filename, file_type, input_dict, business_list)
        run_type = rs[0]    #脚本运行类型；手工计算运行；系统组件运行；手工计算+系统运行
        env_name = rs[1]    #环境名称
        loanApplyAmount = rs[2]    #贷款申请金额
        addInvestAmount = rs[3]    #追加保费申请金额
        modify_info_dict = rs[4]    #客户重要资料变更欲修改信息（角色类型、性别、出生日期、证件类型、证件号码、职业代码/名称、个人年收入、家庭年收入、收入来源、收入来源备注、工作单位）
        
        """step2:创建场景脚本.py文件"""   
        #1.创建.py文件
        file = open(py_file_path, 'w')
        
        """step3:组装场景脚本.py文件"""  
        #1.import文件
        product_life_cycle_test().py_file_import(file, run_type, business_list)
        #2.写入脚本正文：
            #(1)定义类+定义函数+写入try内容    
        product_life_cycle_test().py_file_try(file, run_type, execNo, sum_business, env_name, product_life_cycle_list, filename, years, loanApplyAmount, modify_info_dict, addInvestAmount)
            #(2)except内容 
        product_life_cycle_test().py_file_except(file, run_type)
            #(3)finally内容
        product_life_cycle_test().py_file_finally(file, run_type)
            
        """step4:执行场景脚本.py文件"""
        file.close()
        os.system(py_file_path)
        return run_type, env_name




    """三、创建并组装测试报告.html文件
    step1:生成测试报告
    step2:执行脚本
    """
    def create_testreport_file(self, env_name, filename, excle_file_path, run_type, sum_business, product_life_cycle_list, html_file_path, execNo):
        """step1:生成测试报告"""
        #1获取生命周期结果
        wb = load_workbook(excle_file_path)
        sh = wb['result1']
        result = str(sh['A1'].value)   #结果：-2失败；非-2时，为业务编号0~x
        msg = sh['A2'].value   #接口报错返回信息
        if msg == None:
            msg = ''
        err = sh['A3'].value   #日志报错返回信息
        wb1 = load_workbook(filename)
        sh1 = wb1['nb_param']   #新契约参数sheet
        product = str(sh1['B28'].value)
        wb1.close()
        #2.创建.html文件
        file = open(html_file_path, 'w')
         #3.写入脚本
            #(1)写入测试报告脚本-标题、结果汇总
        message = """<html>
	<head>
		<meta charset='utf-8'>
		<title>产品生命周期</title>
	</head>
	<body>
		<h3>%s产品生命周期测试报告-%s:</h3>
        <h6>测试编号：%s</h6>    
		<h4>结果汇总(%s环境)：</h4>	
		<table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
			<tr width='30'>
				<td>生命周期总次数</td>
				<td>生命周期成功次数</td>
			</tr>
			<tr width='30'>
				<td>1</td>
				<td>1</td>
			</tr>
		</table>"""%(product, run_type, str(execNo), env_name)
        file.write(message + "\n")
            #(2)写入测试报告脚本-生命周期详细
            #每个生命周期的第一行
        message = """		<h4>生命周期详细:</h4>	
		<table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
			<tr align='center'>
				<td width='150'>生命周期序号</td>"""
        file.write(message + "\n")
        for i in range(0, len(product_life_cycle_list)):
            colspan = len(product_life_cycle_list[i])
            file.write("				<td colspan='" + str(colspan) + "'  width='200'>第" + str(i+1) + "年</td>" + "\n")
        file.write("			</tr>" + "\n")
            #每个生命周期的第二行
        message = """			<tr align='center'>
				<td rowspan='2'>1</td>"""
        file.write(message + "\n")
        for i in range(0, len(product_life_cycle_list)):
            if len(product_life_cycle_list[i]) == 0:
                file.write("				<td width='200'></td>" + "\n")
            for j in range(0, len(product_life_cycle_list[i])):
                if product_life_cycle_list[i][j]['business'] == '新契约':
                    file.write("				<td width='200'>新契约</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '续期':
                    file.write("				<td width='200'>续期</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                    file.write("				<td width='200'>犹豫期退保</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                    file.write("				<td width='200'>生存金派发</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '退保':
                    file.write("				<td width='200'>退保</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '贷款':
                    file.write("				<td width='200'>贷款</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                    file.write("				<td width='200'>贷款还款</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '退保试算':
                    file.write("				<td width='200'>退保试算</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                    file.write("				<td width='200'>客户重要资料变更</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                    file.write("				<td width='200'>犹豫期退保试算</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                    file.write("				<td width='200'>持续奖金派发</td>" + "\n")
                elif product_life_cycle_list[i][j]['business'] == '追加保费':
                    file.write("				<td width='200'>追加保费</td>" + "\n")
        file.write("			</tr>" + "\n")
            #是否成功标识
        file.write("			<tr align='center'>" + "\n")  

        ##无成功执行业务-全部业务均未执行场景。
        if int(result) == -1:
            #遍历每年
            for i in range(0, len(product_life_cycle_list)):
                #若这一年的业务数都为0时，补充html中表格的列
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td></td>" + "\n")
                #对每年的全部生命周期遍历
                for j in range(0,len(product_life_cycle_list[i])):
                    file.write("                <td>-</td>" + "\n") 
        elif int(result) == 0:    #无成功执行业务-新契约报错场景。
            #遍历每年
            for i in range(0, len(product_life_cycle_list)):
                #若这一年的业务数都为0时，补充html中表格的列
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td></td>" + "\n")
                #对每年的全部生命周期遍历
                for j in range(0,len(product_life_cycle_list[i])):
                    #html表格中第一个业务-新契约的单元格录入×
                    if i ==0 and j == 0:
                        file.write("                <td><font size='3' color='red'>×</font></td>" + "\n") 
                    #其他业务单元格录入-
                    else:    
                        file.write("                <td>-</td>" + "\n") 
        elif int(result) > 0 and sum_business > int(result):   #存在成功执行业务-成功业务数值小于总业务数
            z = 0  
            for i in range(0, len(product_life_cycle_list)):
                #若这一年的业务数都为0时，补充html中表格的列
                if len(product_life_cycle_list[i]) == 0:
                    file.write("				<td></td>" + "\n")
                #对每年的全部生命周期遍历
                for j in range(0,len(product_life_cycle_list[i])):
                    z = z + 1   #遍历每个业务后+1
                    if z <= int(result)+1:   #判断成功执行的业务。
                        if sh['D' + str(z+4)].value[0:5] != '验证不通过':
                            file.write("                <td>√</td>" + "\n")
                        elif sh['D' + str(z+4)].value[0:5] == '验证不通过':
                            file.write("                <td><font size='3' color='red'>×</font></td>" + "\n")
                    elif z > int(result)+1:  #判断其他未被执行的业务。
                        file.write("                <td>-</td>" + "\n") 
        elif int(result) > 0 and sum_business == int(result):    #存在成功执行业务-成功业务数值等于总业务数
            z = 0
            #遍历每年
            for i in range(0, len(product_life_cycle_list)):
                #若这一年的业务数都为0时，补充html中表格的列
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td></td>" + "\n")
                #对每年的全部生命周期遍历
                for j in range(0,len(product_life_cycle_list[i])):
                    z = z + 1
                    if sh['D' + str(z+4)].value[0:5] != '验证不通过':
                        file.write("                <td>√</td>" + "\n")
                    elif sh['D' + str(z+4)].value[0:5] == '验证不通过':
                        file.write("                <td><font size='3' color='red'>×</font></td>" + "\n")
            file.write("			</tr>" + "\n")    
        file.write("		</table>" + "\n\n\n")
        
            #(3)写入测试报告脚本-报告详细
        z = 0
        file.write("		<h4>报告详细:</h4>" + "\n")
        file.write("		<table table border='2' bordercolor='black' width='800' cellspacing='0' cellpadding='5'>" + "\n")
        
        file.write("			<tr width='30'>" + "\n")
        file.write("				<td colspan='2'>保单号：" + sh['A2'].value + "</td>" + "\n")
        file.write("			</tr>" + "\n")
        
        for i in range(0, len(product_life_cycle_list)):  #每一年
            file.write("			<tr width='30'>" + "\n")
            file.write("				<td colspan='2'>第" + str(i+1) + "年</td>" + "\n")
            file.write("			</tr>" + "\n")
            for j in range(0, len(product_life_cycle_list[i])): 
                z = z + 1
                #每一年内每个业务名
                file.write("			<tr width='30'>" + "\n")  
                if z <= int(result)+1:
                    file.write("				<td rowspan='3'>" + sh['A' + str(z+4)].value + "</td>" + "\n")
                    file.write("				<td>申请日期：" + sh['B' + str(z+4)].value + "</td>" + "\n")
                else:
                    file.write("				<td rowspan='3'>-</td>" + "\n")
                    file.write("				<td>申请日期:-</td>" + "\n")
                file.write("			</tr>" + "\n")
                
                #验证内容
                file.write("			<tr width='30'>" + "\n")
                if z <= int(result)+1:
                    file.write("				<td>验证内容：" + sh['C' + str(z+4)].value + "</td>" + "\n")
                else:
                    file.write("				<td>-</td>" + "\n")
                file.write("			</tr>" + "\n")

                #每一年内每个业务验证结果
                file.write("			<tr width='30'>" + "\n") 
                if z <= int(result)+1:
                    file.write("				<td>验证结果：" + sh['D' + str(z+4)].value + "</td>" + "\n")
                else:
                    file.write("				<td>-</td>" + "\n")
                file.write("			</tr>" + "\n")
        file.write("		</table>" + "\n")        
        file.write("	</body>" + "\n")
        file.write("</html>" + "\n")
        wb.close()     
        file.close()
        #执行html文件(本地打开，服务端不打开html文件)
        if 'win' in sys.platform:
            os.system(html_file_path)



    #异常场景=测试报告
    def except_report(self, excle_file_path, filename, html_file_path, run_type, execNo, env_name, product_life_cycle_list):
        """step1:生成测试报告"""
        #1获取生命周期结果
        wb = load_workbook(excle_file_path)
        #不存在result1sheet页时：
        if 'result1' not in wb.sheetnames:
            wb1 = load_workbook(filename)
            sh1 = wb1['nb_param']   #新契约参数sheet
            product = str(sh1['B28'].value)
            wb1.close()
            #2.创建.html文件
            file = open(html_file_path, 'w')
             #3.写入脚本
                #(1)写入测试报告脚本-标题、结果汇总
            message = """<html>
    <head>
        <meta charset='utf-8'>
        <title>产品生命周期</title>
    </head>
    <body>
        <h3>%s产品生命周期测试报告-%s:</h3>
        <h6>测试编号：%s</h6>    
        <h4>结果汇总(%s环境)：</h4>    
        <table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='30'>
                <td>生命周期总次数</td>
                <td>生命周期成功次数</td>
            </tr>
            <tr width='30'>
                <td>1</td>
                <td>0</td>
            </tr>
        </table>"""%(product, run_type, str(execNo), env_name)
            file.write(message + "\n") 
            #(2)写入测试报告脚本-生命周期详细
            #每个生命周期的第一行
            message = """       <h4>生命周期详细:</h4>    
        <table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
            <tr align='center'>
                <td width='150'>生命周期序号</td>"""
            file.write(message + "\n")
            for i in range(0, len(product_life_cycle_list)):
                colspan = len(product_life_cycle_list[i])
                file.write("                <td colspan='" + str(colspan) + "'  width='200'>第" + str(i+1) + "年</td>" + "\n")
            file.write("            </tr>" + "\n")
                #每个生命周期的第二行
            message = """           <tr align='center'>
                    <td rowspan='2'>1</td>"""
            file.write(message + "\n")
            for i in range(0, len(product_life_cycle_list)):
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td width='200'></td>" + "\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                <td width='200'>新契约</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                <td width='200'>续期</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                <td width='200'>犹豫期退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                <td width='200'>生存金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                <td width='200'>退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                <td width='200'>贷款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                <td width='200'>贷款还款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                <td width='200'>退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                <td width='200'>客户重要资料变更</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                <td width='200'>犹豫期退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                <td width='200'>持续奖金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                <td width='200'>追加保费</td>" + "\n")
            file.write("            </tr>" + "\n")
                #是否成功标识
            file.write("            <tr align='center'>" + "\n")  
            for i in range(0, len(product_life_cycle_list)):
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td></td>" + "\n")
                else:
                    for j in range(0,len(product_life_cycle_list[i])):
                        file.write("                <td>-</td>" + "\n")  
            file.write("        </table>" + "\n") 

            #(3)写入测试报告脚本-标题、结果汇总
            file.write("        <h4>生命周期报错原因:</h4>" + "\n") 
            message = """<table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
            <tr>
                <td>组件检查未成功：</td>
                <td>请仔细检查生命周期excel或者组件代码！</td>
            </tr>
        </table>"""
            file.write(message + "\n") 
            file.write("    </body>" + "\n")
            file.write("</html>" + "\n")
            wb.close()     
            file.close()
            #执行html文件(本地打开，服务端不打开html文件)
            if 'win' in sys.platform:
                os.system(html_file_path)
        #存在result1sheet页时：
        else:
            sh = wb['result1']
            result = str(sh['A1'].value)   #结果：-2失败；非-2时，为业务编号0~x
            msg = sh['A2'].value   #接口报错返回信息
            if msg == None:
                msg = ''
            err = sh['A3'].value   #日志报错返回信息
            wb1 = load_workbook(filename)
            sh1 = wb1['nb_param']   #新契约参数sheet
            product = str(sh1['B28'].value)
            wb1.close()
            #2.创建.html文件
            file = open(html_file_path, 'w')
             #3.写入脚本
                #(1)写入测试报告脚本-标题、结果汇总
            message = """<html>
    <head>
        <meta charset='utf-8'>
        <title>产品生命周期</title>
    </head>
    <body>
        <h3>%s产品生命周期测试报告-%s:</h3>
        <h6>测试编号：%s</h6>    
        <h4>结果汇总(%s环境)：</h4>    
        <table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='30'>
                <td>生命周期总次数</td>
                <td>生命周期成功次数</td>
            </tr>
            <tr width='30'>
                <td>1</td>
                <td>0</td>
            </tr>
        </table>"""%(product, run_type, str(execNo), env_name)
            file.write(message + "\n") 
            #(2)写入测试报告脚本-生命周期详细
            #每个生命周期的第一行
            message = """       <h4>生命周期详细:</h4>    
        <table table border='2' bordercolor='black' cellspacing='0' cellpadding='5'>
            <tr align='center'>
                <td width='150'>生命周期序号</td>"""
            file.write(message + "\n")
            for i in range(0, len(product_life_cycle_list)):
                colspan = len(product_life_cycle_list[i])
                file.write("                <td colspan='" + str(colspan) + "'  width='200'>第" + str(i+1) + "年</td>" + "\n")
            file.write("            </tr>" + "\n")
                #每个生命周期的第二行
            message = """           <tr align='center'>
                    <td rowspan='2'>1</td>"""
            file.write(message + "\n")
            for i in range(0, len(product_life_cycle_list)):
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td width='200'></td>" + "\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                <td width='200'>新契约</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                <td width='200'>续期</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                <td width='200'>犹豫期退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                <td width='200'>生存金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                <td width='200'>退保</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                <td width='200'>贷款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                <td width='200'>贷款还款</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                <td width='200'>退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                <td width='200'>客户重要资料变更</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                <td width='200'>犹豫期退保试算</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                <td width='200'>持续奖金派发</td>" + "\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                <td width='200'>追加保费</td>" + "\n")
            file.write("            </tr>" + "\n")
                #是否成功标识
            file.write("            <tr align='center'>" + "\n")  
            for i in range(0, len(product_life_cycle_list)):
                if len(product_life_cycle_list[i]) == 0:
                    file.write("                <td></td>" + "\n")
                else:
                    for j in range(0,len(product_life_cycle_list[i])):
                        file.write("                <td>-</td>" + "\n")  
            file.write("        </table>" + "\n") 

            #(3)写入测试报告脚本-标题、结果汇总
            file.write("        <h4>生命周期报错原因:</h4>" + "\n") 
            message = """<table table border='2' bordercolor='black' width='400' cellspacing='0' cellpadding='5'>
            <tr width='50'>
                <td>接口返回报错信息</td>
                <td>%s;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;请仔细检查生命周期excel！</td>
            </tr>
            <tr width='50'>
                <td>日志返回报错信息</td>
                <td>%s</td>
            </tr>
        </table>"""%(msg, err)
            file.write(message + "\n") 
            file.write("    </body>" + "\n")
            file.write("</html>" + "\n")
            wb.close()     
            file.close()
            #执行html文件(本地打开，服务端不打开html文件)
            if 'win' in sys.platform:
                os.system(html_file_path)


    #excel获取业务参数
    def get_business_param(self, filename, file_type, input_dict, business_list):
        #初始化
        run_type = ''    #脚本运行类型
        env_name = ''    #环境类型
        loanApplyAmount = 100   #贷款申请金额初始化
        addInvestAmount = 1000    #追加保费申请金额
        modify_info_dict = ''   #客户重要资料变更-欲修改信息初始化
        #1.excel获取业务参数
        wb = load_workbook(filename)
        #判断文件类型是产品生命周期or随机保全项目
        if file_type == 'product_life_cycle':
            sh = wb['product_life_cycle']   #产品生命周期sheet
            sh1 = wb['other_param']   #其他业务参数（除新契约业务外）sheet
            env_name = sh['C1'].value   #环境名称
            run_type = sh['A1'].value   #脚本运行类型；手工计算运行；系统组件运行；手工计算+系统运行
        elif file_type == 'product_life_cycle-random_pa':
            sh = wb['random_pa']   #产品生命周期sheet
            sh1 = wb['other_param']   #其他业务参数（除新契约业务外）sheet
            env_name = sh['B1'].value   #环境名称
            run_type = '手工计算+系统运行'   #脚本运行类型；手工计算运行；系统组件运行；手工计算+系统运行
        #判断是否传入命令行第二个的参数-json串。未传入取excel表格中数据；传入则获取命令行第二个的参数
        if input_dict != {}:
            input_dict = input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
            if 'win' in sys.platform:
                input_dict = input_dict.decode('gb2312').encode('utf-8')   #把命令行参数的gb2312编码格式转换为utf-8格式
            input_dict = json.loads(input_dict)   #json转dict格式
            run_type =  input_dict['run_type']   #运行类型
            env_name =  input_dict['env_name']   #环境类型
        #2.获取贷款参数-贷款申请金额（默认为1000）
        if '贷款' in business_list:
            #获取贷款参数:贷款申请金额
            if sh1['C2'].value == None:
                loanApplyAmount = 100   #若excle未录入贷款申请金额，默认为100
            else:
                loanApplyAmount = sh1['C2'].value
        if '追加保费' in business_list:
            #获取追加保费参数:追加保费申请金额
            if sh1['C13'].value == None:
                addInvestAmount = 1000   #若excle未录入追加保费申请金额，默认为1000
            else:
                addInvestAmount = sh1['C13'].value
                
        #3.获取客户重要资料变更参数-欲修改信息（角色类型、性别、出生日期、证件类型、证件号码、职业代码/名称、个人年收入、家庭年收入、收入来源、收入来源备注、工作单位）        
        if '客户重要资料变更' in business_list:
            #角色类型
            if sh1['C3'].value == None:
                roleType = '2'
            else:
                roleType = str(sh1['C3'].value)
            #客户原证件号
            oldcertiCode = str(sh1['E3'].value)
            #性别
            if sh1['G3'].value == None:
                gender = {'is_modify':'N', 'new':''}
            else:
                gender = {'is_modify':'Y', 'new':str(sh1['G3'].value)}
            #出生日期
            if sh1['G4'].value == None:
                birthday = {'is_modify':'N', 'new':''}
            else:
                birthday = {'is_modify':'Y', 'new':str(sh1['G4'].value)}
            #证件类型
            if sh1['G5'].value == None:
                certiType = {'is_modify':'N', 'new':''}
            else:
                certiType = {'is_modify':'Y', 'new':str(sh1['G5'].value)}   
            #证件号码
            if sh1['G6'].value == None:
                certiCode = {'is_modify':'N', 'new':''}
            else:
                certiCode = {'is_modify':'Y', 'new':str(sh1['G6'].value)}     
            #职业代码/名称
            if sh1['G7'].value == None:
                jobCateId = {'is_modify':'N', 'new':''}
            else:
                jobCateId = {'is_modify':'Y', 'new':str(sh1['G7'].value)}       
            #个人年收入（万元）
            if sh1['G8'].value == None:
                income = {'is_modify':'N', 'new':''}
            else:
                income = {'is_modify':'Y', 'new':str(sh1['G8'].value)}      
            #家庭年收入（万元）
            if sh1['G9'].value == None:
                familyIncome = {'is_modify':'N', 'new':''}
            else:
                familyIncome = {'is_modify':'Y', 'new':str(sh1['G9'].value)}     
            #收入来源
            if sh1['G10'].value == None:
                incomeSource = {'is_modify':'N', 'new':''}
            else:
                incomeSource = {'is_modify':'Y', 'new':str(sh1['G10'].value)}      
            #收入来源备注
            if sh1['G11'].value == None:
                incomeSourceNote = {'is_modify':'N', 'new':''}
            else:
                incomeSourceNote = {'is_modify':'Y', 'new':str(sh1['G11'].value)}   
            #工作单位
            if sh1['G12'].value == None:
                companyName = {'is_modify':'N', 'new':''}
            else:
                companyName = {'is_modify':'Y', 'new':str(sh1['G12'].value)}   
            modify_info_dict = {
                        'roleType': roleType,   #角色类型:1-投保人；2-被保人；3-受益人；4-代理人（保单无对应角色，默认为被保人）
                        'oldcertiCode': oldcertiCode,  #客户原证件号
                        'gender': gender,    #性别:0-男，1-女
                        'birthday': birthday,     #出生日期
                        'certiType': certiType,     #证件类型
                        'certiCode': certiCode,     #证件号码
                        'jobCateId': jobCateId,     #职业代码/名称
                        'income': income,     #个人年收入（万元）
                        'familyIncome': familyIncome,    #家庭年收入（万元）
                        'incomeSource': incomeSource,    #收入来源
                        'incomeSourceNote': incomeSourceNote,    #收入来源备注
                        'companyName': companyName    #工作单位
                        }
            modify_info_dict = json.dumps(modify_info_dict)
        else:
            pass
        wb.close()
        return run_type, env_name, loanApplyAmount, addInvestAmount, modify_info_dict



    #手工计算import内容
    def py_file_import(self, file, run_type, business_list):
        message = """#!/usr/bin/python
# coding=utf-8
import sys
import time
import json
import logging
import openpyxl
import datetime
import traceback
from openpyxl import load_workbook
reload(sys)
sys.setdefaultencoding("utf8")"""
        file.write(message + "\n")
        file.write("from xLibrary.chunyu.API0000 import API0000_diy\n")
        if run_type == '手工计算运行' or run_type == '手工计算+系统运行':
            #1.import文件
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.calc_script.calc_nb.calc_API0001 import calc_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.calc_script.calc_renew.calc_API0002 import calc_API0002\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0003 import calc_API0003\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_survivalFee.calc_API0004 import calc_API0004\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0005 import calc_API0005\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan.calc_API0006 import calc_API0006\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.calc_script.calc_loan_repayment.calc_API0007 import calc_API0007\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_tb.calc_API0008 import calc_API0008\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.calc_script.calc_CD.calc_API0009 import calc_API0009\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.calc_script.calc_hesitation_tb.calc_API0010 import calc_API0010\n")
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.calc_script.calc_continuous_bonus.calc_API0011 import calc_API0011\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.calc_script.calc_additional_premium.calc_API0012 import calc_API0012\n")                     
        if run_type == '系统组件运行' or run_type == '手工计算+系统运行':
            file.write("from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002 import sys_API0002\n")
            for i in range(0, len(business_list)):
                if business_list[i] == '新契约':
                    file.write("from xLibrary.chunyu.sys_script.nb.sys_API0001 import sys_API0001\n")
                elif business_list[i] == '续期':
                    file.write("from xLibrary.chunyu.sys_script.renew.sys_API0004 import sys_API0004\n")
                elif business_list[i] == '犹豫期退保':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0009 import sys_API0009\n")
                elif business_list[i] == '生存金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_survivalFee.sys_API0011 import sys_API0011\n")
                elif business_list[i] == '退保':
                    file.write("from xLibrary.chunyu.sys_script.tb.sys_API0010 import sys_API0010\n")
                elif business_list[i] == '贷款':
                    file.write("from xLibrary.chunyu.sys_script.loan.sys_API0007 import sys_API0007\n")
                elif business_list[i] == '贷款还款':
                    file.write("from xLibrary.chunyu.sys_script.loan_repayment.sys_API0008 import sys_API0008\n")
                elif business_list[i] == '退保试算':
                    file.write("from xLibrary.chunyu.sys_script.trial_tb.sys_API0013 import sys_API0013\n")
                elif business_list[i] == '客户重要资料变更':
                    file.write("from xLibrary.chunyu.sys_script.CD.sys_API0016 import sys_API0016\n")
                elif business_list[i] == '犹豫期退保试算':
                    file.write("from xLibrary.chunyu.sys_script.hesitation_tb.sys_API0015 import sys_API0015\n")    
                elif business_list[i] == '持续奖金派发':
                    file.write("from xLibrary.chunyu.sys_script.send_continuous_bonus.sys_API0020 import sys_API0020\n")
                elif business_list[i] == '追加保费':
                    file.write("from xLibrary.chunyu.sys_script.additional_premium.sys_API0021 import sys_API0021\n")


    #lianxi.py文件写入脚本正文+try内容
    def py_file_try(self, file, run_type, execNo, sum_business, env_name, product_life_cycle_list, filename, years, loanApplyAmount, modify_info_dict, addInvestAmount):
        #定义类
        file.write("\n\n\n" + "class test:\n")
        #定义函数
        file.write("\n\n" + "    def test(self):\n")
        #写入try内容
        file.write("        try:\n")
        file.write("            logging.basicConfig(level=logging.WARNING , format='%(message)s  %(asctime)s')\n")
        file.write("            dict = {'policy_info':{'policyNo':''}, 'public':{'execNo':'" + str(execNo) + "'}, 'logInfo':{'msg':''}}\n")
        file.write("            current_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')\n")
        file.write("            sum_business = " + str(sum_business) + "\n")
        file.write("            execNo = '" + str(execNo) + "'\n")
        file.write("            env_name = '" + env_name + "'\n")
        file.write("            env_name = env_name.encode('utf-8')\n")
        file.write("            applicationDate = '" + str(product_life_cycle_list[0][0]['apply_date']) + "'\n")
        if run_type == '手工计算运行':
            file.write("            result = -1\n")
            message = """            print '---dict_manual---'
            dict_manual = API0000_diy().define_dict()
            dict_manual['public']['execNo'] = '%s'
            test_type = 'product_life_cycle'
            dict_manual = API0000_diy().store_nbdata(test_type, applicationDate, dict_manual)"""% str(execNo)
            file.write(message + "\n")
            file.write("            dict_manual['public']['filename'] = r'" + str(filename) + "'\n")
            for i in range(0, years):
                file.write("            print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("            result = result + 1\n")
                        file.write("            dict_manual = calc_API0001().calc_nb(dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("            pay_due_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0002().calc_renew(pay_due_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0003().calc_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0004().calc_survivalFee(sendDate, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0005().calc_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("            dict_manual = calc_API0006().calc_loan(apply_date, loanApplyAmount, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0007().calc_loan_repayment(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0008().calc_trial_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0009().calc_CD(apply_date, " + modify_info_dict + ", dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0010().calc_trial_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict_manual = calc_API0011().calc_continuous_bonus(sendDate, dict_manual)\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("            dict_manual = calc_API0012().calc_additional_premium(apply_date, addInvestAmount, dict_manual)\n")
            file.write("            result1 = result\n")
            file.write("            print '=================end=================================='\n")
            file.write("            print '---dict_manual:---'\n")
            file.write("            print dict_manual\n")
        elif run_type == '系统组件运行':
            file.write("            result = -1\n") 
            file.write("            print '---dict---'\n")
            file.write("            dict = API0000_diy().define_dict()\n")
            file.write("            dict['public']['filename'] = r'" + str(filename) + "'\n")
            file.write("            dict['public']['execNo'] = '" + str(execNo) + "'\n")
            for i in range(0, years):
                file.write("            print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("            result = result + 1\n")
                        file.write("            dict = sys_API0001().nb(env_name, applicationDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("            times_renew = 1\n")
                        file.write("            dict = sys_API0004().renew(times_renew, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0011().send_survivalFee(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0010().tb(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("            dict = sys_API0007().loan(apply_date, loanApplyAmount, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0008().loan_repayment(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("            validateDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0013().trial_tb(validateDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0016().CD(apply_date, " + modify_info_dict + ", dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("            sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            dict = sys_API0020().send_continuous_bonus(sendDate, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")   
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("            apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("            dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("            dict = sys_API0021().additional_premium(apply_date, addInvestAmount, dict)\n")
                        file.write("            assert dict['logInfo']['code'] == '1'\n")
                        file.write("            result = result + 1\n")   
            file.write("            print '=================end=================================='\n")
        elif run_type == '手工计算+系统运行':
            file.write("            result_manual = -1\n")
            file.write("            result_sys = -1\n")
            message = """            try:
                print '---dict_manual---'
                dict_manual = API0000_diy().define_dict()
                dict_manual['public']['execNo'] = '%s'
                test_type = 'product_life_cycle'
                dict_manual = API0000_diy().store_nbdata(test_type, applicationDate, dict_manual)"""% str(execNo)
            file.write(message + "\n")
            file.write("                dict_manual['public']['filename'] = r'" + str(filename) + "'\n")
            for i in range(0, years):
                file.write("                print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                result_manual = result_manual + 1\n")
                        file.write("                dict_manual = calc_API0001().calc_nb(dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                pay_due_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0002().calc_renew(pay_due_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0003().calc_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0004().calc_survivalFee(sendDate, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0005().calc_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("                dict_manual = calc_API0006().calc_loan(apply_date, loanApplyAmount, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0007().calc_loan_repayment(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0008().calc_trial_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0009().calc_CD(apply_date, " + modify_info_dict + ", dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0010().calc_trial_hesitation_tb(apply_date, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict_manual = calc_API0011().calc_continuous_bonus(sendDate, dict_manual)\n")
                        file.write("                result_manual = result_manual + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("                dict_manual = calc_API0012().calc_additional_premium(apply_date, addInvestAmount, dict_manual)\n") 
                        file.write("                result_manual = result_manual + 1\n")
            file.write("            finally:\n")            
            file.write("                print '---dict---'\n")
            file.write("                dict = API0000_diy().define_dict()\n")
            file.write("                dict['public']['filename'] = r'" + str(filename) + "'\n")
            file.write("                dict['public']['execNo'] = '" + str(execNo) + "'\n")
            for i in range(0, years):
                file.write("                print '=================years:" + str(i+1) + "=================================='\n")
                for j in range(0, len(product_life_cycle_list[i])):
                    if product_life_cycle_list[i][j]['business'] == '新契约':
                        file.write("                result_sys = result_sys + 1\n")
                        file.write("                dict = sys_API0001().nb(env_name, applicationDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '续期':
                        file.write("                times_renew = 1\n")
                        file.write("                dict = sys_API0004().renew(times_renew, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '生存金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0011().send_survivalFee(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0010().tb(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                loanApplyAmount = '" + str(loanApplyAmount) + "'\n")
                        file.write("                dict = sys_API0007().loan(apply_date, loanApplyAmount, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '贷款还款':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0008().loan_repayment(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '退保试算':
                        file.write("                validateDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0013().trial_tb(validateDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '客户重要资料变更':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0016().CD(apply_date, " + modify_info_dict + ", dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '犹豫期退保试算':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0015().hesitation_tb_csCancle(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n")
                    elif product_life_cycle_list[i][j]['business'] == '持续奖金派发':
                        file.write("                sendDate = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                dict = sys_API0020().send_continuous_bonus(sendDate, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n") 
                    elif product_life_cycle_list[i][j]['business'] == '追加保费':
                        file.write("                apply_date = '" + str(product_life_cycle_list[i][j]['apply_date']) + "'\n")
                        file.write("                dict = sys_API0002().Modify_servertime(apply_date, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                addInvestAmount = '" + str(addInvestAmount) + "'\n")
                        file.write("                dict = sys_API0021().additional_premium(apply_date, addInvestAmount, dict)\n")
                        file.write("                assert dict['logInfo']['code'] == '1'\n")
                        file.write("                result_sys = result_sys + 1\n") 


    #lianxi.py文件写入脚本正文中except内容
    def py_file_except(self, file, run_type):
        if run_type == '手工计算运行':
            message = """        except Exception, e:
            dict_manual['logInfo']['err'] = str(e)
            print '---err----------'
            traceback.print_exc()
            exit()"""
            file.write(message + "\n")
        elif run_type == '系统组件运行':
            file.write("            print '---dict:---'\n")
            file.write("            print dict\n")
            file.write("            print dict['policy_info']['policyNo']\n")
            message = """        except Exception, e:
            dict['logInfo']['err'] = str(e)
            print '---err------'
            traceback.print_exc()
            print dict
            exit()""" 
            file.write(message + "\n") 
        elif run_type == '手工计算+系统运行':
            file.write("            print '---dict_manual:---'\n")
            file.write("            print dict_manual\n")
            file.write("            print '---dict:---'\n")
            file.write("            print dict\n")
            file.write("            print dict['policy_info']['policyNo']\n")
            message = """        except Exception, e:
            dict['logInfo']['err'] = str(e)
            print '---err------'
            traceback.print_exc()
            print dict
            exit()"""
            file.write(message + "\n") 


    #lianxi.py文件写入脚本正文中finally内容
    def py_file_finally(self, file, run_type):
        if run_type == '手工计算运行':
            file.write("        finally :\n")
            #结果写入excel
            if 'win' not in sys.platform:
                file.write("            filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                file.write(message + "\n")
                os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("            filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """            wb = load_workbook(filename)
            sheet_name = 'result1'
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n") 

            message = """            if result == sum_business:
                result1 = sum_business
            else:
                result1 = result + 1
            for i in range(0, result1):
                business_apply_date = dict_manual['track_info'][i]['trackTime']   #业务申请日期
                if dict_manual['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。新契约收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount'])                  
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。续期收费金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。犹豫期退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。犹豫期退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice'])               
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。贷款总金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'  
                    if dict_manual['track_info'][i]['msg'] == '': 
                        business_check_result = '验证正确。贷款还款金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) 
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。生存金派发金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。退保金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。持续奖金派发金额:' + str(dict_manual['track_info'][i]['trackData']['payment'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                elif dict_manual['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    if dict_manual['track_info'][i]['msg'] == '':
                        business_check_result = '验证正确。追加保费收付费:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                    else:
                        business_check_result = '验证不通过:' + dict_manual['track_info'][i]['msg']
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result
                sh['A' + str(i+6)] = ''
                sh['B' + str(i+6)] = ''
                sh['C' + str(i+6)] = ''
                sh['D' + str(i+6)] = ''"""
            file.write(message + "\n")
            file.write("            sh2 = wb['record']" + "\n")
            file.write("            rows = sh2.max_row" + "\n")
            file.write("            sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("            sh2['C' + str(rows+1)] = json.dumps(dict_manual)" +  "\n")
            file.write("            wb.save(filename)" + "\n")
            file.write("            wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")
        elif run_type == '系统组件运行':
            file.write("        finally:" + "\n")  
            file.write("            print dict" + "\n")                
            if 'win' not in sys.platform:
                file.write("            filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                file.write(message + "\n")
                os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("            filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """            wb = load_workbook(filename)
            sheet_name = 'result1'
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n")
            message = """            for i in range(0,result):
                business_apply_date = dict['track_info'][i]['trackTime']   #业务申请日期
                if dict['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    business_check_result = '新契约收付费金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][i]['trackData']['product'][0]['amount'])                  
                elif dict['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    business_check_result = '续期收费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])               
                elif dict['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    business_check_result = '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                elif dict['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'    
                    business_check_result = '贷款还款金额:' + str(dict['track_info'][i]['trackData']['payment']) 
                elif dict['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                elif dict['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    business_check_result = '追加保费收付费:' + str(dict['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result

            #存在报错场景    
            if result < sum_business:
                if dict['logInfo']['msg'] == None or dict['logInfo']['msg'] == '':
                    msg = str(dict['logInfo']['err'])
                else:
                    msg = str(dict['logInfo']['msg'])
                if result == 0:
                    sh['A' + str(5)] = '新契约'  
                    sh['B' + str(5)] = applicationDate
                    sh['C' + str(5)] = '新契约收付费金额、产品保费、保额' 
                    sh['D' + str(5)] = '验证不通过。' + msg
                else:
                    if dict['track_info'][result]['trackType'] == 'renew':
                        business_name = '续期'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '续期收费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                        business_name = '犹豫期退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'hesitation_tb':
                        business_name = '犹豫期退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'trial_tb':
                        business_name = '退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、产品现价、其他'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan':
                        business_name = '贷款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan_repayment':
                        business_name = '贷款还款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款还款金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'survivalFee':
                        business_name = '生存金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '生存金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'CD':
                        business_name = '客户重要资料变更'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '客户重要资料变更收付费、保额、保费'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'tb':
                        business_name = '退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、现价'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'continuous_bonus':
                        business_name = '持续奖金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '持续奖金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'additional_premium':
                        business_name = '追加保费'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result"""
            file.write(message + "\n")
            file.write("            sh2 = wb['record']" + "\n")
            file.write("            rows = sh2.max_row" + "\n")
            file.write("            sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("            sh2['B' + str(rows+1)] = json.dumps(dict)" +  "\n")
            file.write("            wb.save(filename)" + "\n")
            file.write("            wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")
        elif run_type == '手工计算+系统运行':
            file.write("        finally:" + "\n")             
            if 'win' not in sys.platform:
                file.write("            filename = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'" + "\n")
                message = """            wb = openpyxl.Workbook()
            ws = wb.active  # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws.title = 'record'  #更改工作表ws的title
            ws['A1'] = '时间'
            ws['B1'] = '系统dict'
            ws['C1'] = '手工计算dict'
            wb.save(filename)"""
                file.write(message + "\n")
                os.system('chmod -R 777 /data/xServer/xRunner')
            else:   
                file.write("            filename = r'D:\\xLibrary\\chunyu\\product_life_cyle_test\\lianxi.xlsx'")
            file.write("\n")
            message = """            wb = load_workbook(filename)
            sheet_name = 'result1'
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            result = min(int(result_manual), int(result_sys))
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])"""
            file.write(message + "\n")
            
            message = """            #两种场景：1.无报错场景时，对全部结果比对。2.存在报错场景时，获取手工计算与系统存有运行的组件数量的最小值，并比对非最后一个组件的结果（最小值-1）
            for i in range(0,result):
                business_apply_date = dict_manual['track_info'][i]['trackTime']   #业务申请日期
                if dict_manual['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '新契约收付费金额:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][i]['trackData']['product'][0]['amount'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '新契约收付费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['premium'] != dict_manual['track_info'][i]['trackData']['product'][0]['premium']:
                            business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['amount'] != dict_manual['track_info'][i]['trackData']['product'][0]['amount']:
                            business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount']) + ';'                         
                if dict_manual['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '续期收费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '续期收费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                if dict_manual['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                if dict_manual['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                if dict_manual['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                            business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']:
                            business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                            business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                            business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']:
                            business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem'])   + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']:
                            business_check_result = business_check_result + '累计生息账户, ' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']:
                            business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']:
                            business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']:
                            business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']) + ';'                  
                if dict_manual['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'                        
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '贷款总金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';'
                        if dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']:
                            business_check_result = business_check_result + '贷款额度,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                            business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['productCode'] not in ('1205','1206','3208','3209','3211','3245','3247','3247B','3255','3256','3257','3257CA','3264','3265','3265CB','3267','3267CB','3267CO','3270','5201','5202','5206','5207','5211','5213','5214','8214','8216','8217','8233','8237'):
                            if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                                business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';' 
                        else:
                            if business_check_result == '验证不通过。':
                                business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:0' + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                if dict_manual['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'    
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '贷款还款金额:' + str(dict['track_info'][i]['trackData']['payment']) 
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '贷款还款金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                if dict_manual['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '生存金派发金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                if dict_manual['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过  
                        business_check_result = '验证不通过。'
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '客户重要资料变更收付费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['premium'] != dict_manual['track_info'][i]['trackData']['product'][0]['premium']:
                            business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['premium']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['amount'] != dict_manual['track_info'][i]['trackData']['product'][0]['amount']:
                            business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['amount']) + ';' 
                if dict_manual['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']:
                            business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']:
                            business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                            business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']:
                            business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']:
                            business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['refundRenewPrem']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']:
                            business_check_result = business_check_result + '累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['cbSbAccount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']:
                            business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['riskChargeFee']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']:
                            business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['bonusAccount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']:
                            business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['paidBonus']) + ';' 
                if dict_manual['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        business_check_result = business_check_result + '持续奖金派发金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                if dict_manual['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    if dict['track_info'][i]['trackData'] == dict_manual['track_info'][i]['trackData']:    #验证通过
                        business_check_result = '验证正确。' + '追加保费收付费:' + str(dict['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                    elif dict['track_info'][i]['trackData'] != dict_manual['track_info'][i]['trackData']:  #验证不通过 
                        business_check_result = '验证不通过。'
                        if dict['track_info'][i]['trackData']['payment'] != dict_manual['track_info'][i]['trackData']['payment']:
                            business_check_result = business_check_result + '追加保费收付费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['payment']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['addInvestAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']:
                            business_check_result = business_check_result + '追加保费申请金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['chargeAmount']:
                            business_check_result = business_check_result + '追加保费扣费,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['investAmount']) + ';' 
                        if dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'] != dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount']:
                            business_check_result = business_check_result + '实际追加保费金额,' + '系统计算:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount']) + ',手工计算:' + str(dict_manual['track_info'][i]['trackData']['product'][0]['surrenderAmount']) + ';' 
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result"""
            file.write(message + "\n")
            
            message = """            #两种场景：1.若无报错。不运行此部分。2.若存在报错，获取手工计算与系统计算运行的组件数量的最小值，并比对最后一个组件的结果
            if result != sum_business:
                business_apply_date = dict_manual['track_info'][result]['trackTime']   #业务申请日期
                if dict_manual['track_info'][result]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容           
                elif dict_manual['track_info'][result]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                elif dict_manual['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                elif dict_manual['track_info'][result]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                elif dict_manual['track_info'][result]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'                 
                elif dict_manual['track_info'][result]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                elif dict_manual['track_info'][result]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'    
                elif dict_manual['track_info'][result]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                elif dict_manual['track_info'][result]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                elif dict_manual['track_info'][result]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                elif dict_manual['track_info'][result]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                elif dict_manual['track_info'][result]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                if dict['track_info'][result]['msg'] == '' and dict_manual['track_info'][result]['msg'] == '':
                    if dict_manual['track_info'][result]['trackType'] == 'nb':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '新契约收付费金额:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][result]['trackData']['product'][0]['amount'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '新契约收付费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['premium'] != dict_manual['track_info'][result]['trackData']['product'][0]['premium']:
                                business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['premium']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['amount'] != dict_manual['track_info'][result]['trackData']['product'][0]['amount']:
                                business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['amount']) + ';'                         
                    if dict_manual['track_info'][result]['trackType'] == 'renew':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '续期收费金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '续期收费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][result]['trackType'] == 'hesitation_tb':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '犹豫期退保金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '犹豫期退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                    if dict_manual['track_info'][result]['trackType'] == 'trial_tb':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']:
                                business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']:
                                business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem'])   + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']:
                                business_check_result = business_check_result + '累计生息账户, ' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']:
                                business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']:
                                business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']:
                                business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']) + ';'                  
                    if dict_manual['track_info'][result]['trackType'] == 'loan':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'                        
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '贷款总金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';'
                            if dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountLimit']:
                                business_check_result = business_check_result + '贷款额度,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['productCode'] not in ('1205','1206','3208','3209','3211','3245','3247','3247B','3255','3256','3257','3257CA','3264','3265','3265CB','3267','3267CB','3267CO','3270','5201','5202','5206','5207','5211','5213','5214','8214','8216','8217','8233','8237'):
                                if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                    business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';' 
                            else:
                                if business_check_result == '验证不通过。':
                                    business_check_result = '验证正确。' + '贷款总金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:0' + ';贷款额度:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'])
                    if dict_manual['track_info'][result]['trackType'] == 'loan_repayment':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '贷款还款金额:' + str(dict['track_info'][result]['trackData']['payment']) 
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '贷款还款金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][result]['trackType'] == 'survivalFee':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '生存金派发金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][result]['trackType'] == 'CD':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过  
                            business_check_result = '验证不通过。'
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '客户重要资料变更收付费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['premium'] != dict_manual['track_info'][result]['trackData']['product'][0]['premium']:
                                business_check_result = business_check_result + '产品保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['premium']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['premium']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['amount'] != dict_manual['track_info'][result]['trackData']['product'][0]['amount']:
                                business_check_result = business_check_result + '产品保额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['amount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['amount']) + ';' 
                    if dict_manual['track_info'][result]['trackType'] == 'tb':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '退保金额:' + str(dict['track_info'][result]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '退保金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['presentPrice'] != dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']:
                                business_check_result = business_check_result + '产品现价,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['presentPrice']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['presentPrice']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['investAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']:
                                business_check_result = business_check_result + '万能账户价值,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '手续费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']:
                                business_check_result = business_check_result + '欠款余额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['loanAccountAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem'] != dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']:
                                business_check_result = business_check_result + '退还续期保费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['refundRenewPrem']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['cbSbAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']:
                                business_check_result = business_check_result + '累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['cbSbAccount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['riskChargeFee'] != dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']:
                                business_check_result = business_check_result + '退还未到期风险保险费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['riskChargeFee']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['bonusAccount'] != dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']:
                                business_check_result = business_check_result + '红利累计生息账户,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['bonusAccount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['paidBonus'] != dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']:
                                business_check_result = business_check_result + '已分红金额扣回,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['paidBonus']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['paidBonus']) + ';' 
                    if dict_manual['track_info'][result]['trackType'] == 'continuous_bonus':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '生存金派发金额:' + str(dict['track_info'][result]['trackData']['payment'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            business_check_result = business_check_result + '持续奖金派发金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                    if dict_manual['track_info'][result]['trackType'] == 'additional_premium':
                        if dict['track_info'][result]['trackData'] == dict_manual['track_info'][result]['trackData']:    #验证通过
                            business_check_result = '验证正确。' + '追加保费收付费:' + str(dict['track_info'][result]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][result]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][result]['trackData']['product'][0]['surrenderAmount'])
                        elif dict['track_info'][result]['trackData'] != dict_manual['track_info'][result]['trackData']:  #验证不通过 
                            business_check_result = '验证不通过。'
                            if dict['track_info'][result]['trackData']['payment'] != dict_manual['track_info'][result]['trackData']['payment']:
                                business_check_result = business_check_result + '追加保费收付费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['payment']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['payment']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['addInvestAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['addInvestAmount']:
                                business_check_result = business_check_result + '追加保费申请金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['addInvestAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['chargeAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['chargeAmount']:
                                business_check_result = business_check_result + '追加保费扣费,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['investAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['investAmount']) + ';' 
                            if dict['track_info'][result]['trackData']['product'][0]['surrenderAmount'] != dict_manual['track_info'][result]['trackData']['product'][0]['surrenderAmount']:
                                business_check_result = business_check_result + '实际追加保费金额,' + '系统计算:' + str(dict['track_info'][result]['trackData']['product'][0]['surrenderAmount']) + ',手工计算:' + str(dict_manual['track_info'][result]['trackData']['product'][0]['surrenderAmount']) + ';' 
                elif dict['track_info'][result]['msg'] != '' and dict_manual['track_info'][result]['msg'] != '':
                    if dict_manual['track_info'][result]['msg'] in dict['track_info'][result]['msg']:    #验证通过
                        business_check_result = '验证正确。' + '阻断提示为:' + dict_manual['track_info'][result]['msg']
                    else:
                        business_check_result = '验证不通过。' + '手工计算阻断提示为:' + dict_manual['track_info'][result]['msg'] + ';系统计算阻断提示为:' + dict['track_info'][result]['msg']
                elif dict['track_info'][result]['msg'] != '' and dict_manual['track_info'][result]['msg'] == '':
                    business_check_result = '验证不通过。' + '手工计算未阻断' + ';系统计算阻断提示为:' + dict['track_info'][result]['msg']
                elif dict['track_info'][result]['msg'] == '' and dict_manual['track_info'][result]['msg'] != '':
                    business_check_result = '验证不通过。' + '手工计算阻断提示为:' + dict_manual['track_info'][result]['msg'] + ';系统计算未阻断。'
                sh['A' + str(result+5)] = business_name  
                sh['B' + str(result+5)] = business_apply_date
                sh['C' + str(result+5)] = business_check_info
                sh['D' + str(result+5)] = business_check_result"""
            file.write(message + "\n")
            file.write("            sh2 = wb['record']" + "\n")
            file.write("            rows = sh2.max_row" + "\n")
            file.write("            sh2['A' + str(rows+1)] = current_time " + "\n")
            file.write("            sh2['B' + str(rows+1)] = json.dumps(dict)" +  "\n")
            file.write("            sh2['C' + str(rows+1)] = json.dumps(dict_manual)" +  "\n")
            file.write("            wb.save(filename)" + "\n")
            file.write("            wb.close()" + "\n")
            file.write("\n\n\n" + "test().test()\n")




if __name__ == "__main__":
    #初始化
    execNo = 'test0001'    #测试编号  
    input_dict = {}     #输入的json串
    if 'win' in sys.platform:
        py_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.py'  #组装生成的py文件觉得路径
        excle_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.xlsx'   #执行中临时存储数据的excel文件路径
        html_file_path = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.html'   #html测试报告文件路径
    else:
        py_file_path = '/data/xServer/xRunner/' + str(execNo) + '.py'   #组装生成的py文件觉得路径
        excle_file_path = '/data/xServer/xRunner/' + str(execNo) + '.xlsx'   #执行中临时存储数据的excel文件路径
        html_file_path = '/data/xServer/xReport/' + str(execNo) + '.html'   #html测试报告文件路径 
    #接收参数：第一位为测试编号，第二位json串
    if len(sys.argv) == 2:
        execNo = sys.argv[1]   #测试编号     
    elif len(sys.argv) > 2:
        execNo = sys.argv[1]   #测试编号 
        input_dict = sys.argv[2]   #输入的json串 
    #接收参数：第二位json串中的测试数据excel文件路径
    if input_dict != {}:
        temp_input_dict = input_dict
        temp_input_dict = temp_input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
        if 'win' in sys.platform:
            temp_input_dict = temp_input_dict.decode('gb2312').encode('utf-8')   #把命令行参数的gb2312编码格式转换为utf-8格式
        temp_input_dict = json.loads(temp_input_dict)   #json转dict格式
        filename =  temp_input_dict['filename']   #获取测试数据excel文件路径

    #执行脚本
    product_life_cycle_test().run(py_file_path, excle_file_path, html_file_path, execNo, filename)

    #python D:\xLibrary\chunyu\product_life_cycle.py test0002 "{'run_type':'系统运行','env_name':'预生产','filename':'D:\\xLibrary\\chunyu\\doc\\product_life_cycle\\product_life_cycle_data.xlsx'}"