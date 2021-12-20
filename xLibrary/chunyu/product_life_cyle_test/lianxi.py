#!/usr/bin/python
# coding=utf-8
import sys
import time
import json
import logging
import openpyxl
import datetime
import traceback
from openpyxl import load_workbook
reload(sys)
sys.setdefaultencoding("utf8")
from xLibrary.chunyu.API0000 import API0000_diy
from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002 import sys_API0002
from xLibrary.chunyu.sys_script.nb.sys_API0001 import sys_API0001



class test:


    def test(self):
        try:
            logging.basicConfig(level=logging.WARNING , format='%(message)s  %(asctime)s')
            dict = {'policy_info':{'policyNo':''}, 'public':{'execNo':'test0001'}, 'logInfo':{'msg':''}}
            current_time = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')
            sum_business = 1
            execNo = 'test0001'
            env_name = '预生产'
            env_name = env_name.encode('utf-8')
            applicationDate = '2017-03-10'
            result = -1
            print '---dict---'
            dict = API0000_diy().define_dict()
            dict['public']['filename'] = r'D:\xLibrary\chunyu\doc\product_life_cycle\product_life_cycle_data.xlsx'
            dict['public']['execNo'] = 'test0001'
            print '=================years:1=================================='
            result = result + 1
            dict = sys_API0001().nb(env_name, applicationDate, dict)
            assert dict['logInfo']['code'] == '1'
            result = result + 1
            print '=================years:2=================================='
            print '=================years:3=================================='
            print '=================years:4=================================='
            print '=================years:5=================================='
            print '=================years:6=================================='
            print '=================years:7=================================='
            print '=================years:8=================================='
            print '=================years:9=================================='
            print '=================years:10=================================='
            print '=================end=================================='
            print '---dict:---'
            print dict
            print dict['policy_info']['policyNo']
        except Exception, e:
            dict['logInfo']['err'] = str(e)
            print '---err------'
            traceback.print_exc()
            print dict
            exit()
        finally:
            print dict
            filename = r'D:\xLibrary\chunyu\product_life_cyle_test\lianxi.xlsx'
            wb = load_workbook(filename)
            sheet_name = 'result1'
            wb.create_sheet(title=sheet_name)
            sh = wb[sheet_name]
            sh['A1'] = str(result)
            if dict['policy_info']['policyNo'] != '':
                sh['A2'] = str(dict['policy_info']['policyNo'])
            else:
                sh['A2'] = '-'
            sh['A3'] = str(dict['logInfo']['msg'])
            for i in range(0,result):
                business_apply_date = dict['track_info'][i]['trackTime']   #业务申请日期
                if dict['track_info'][i]['trackType'] == 'nb':
                    business_name = '新契约'    #业务名
                    business_check_info = '新契约收付费金额、产品保费、保额'   #验证内容
                    business_check_result = '新契约收付费金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品保费:' + str(dict['track_info'][i]['trackData']['product'][0]['premium']) + ';产品保额:' + str(dict['track_info'][i]['trackData']['product'][0]['amount'])                  
                elif dict['track_info'][i]['trackType'] == 'renew':
                    business_name = '续期'
                    business_check_info = '续期收费金额'
                    business_check_result = '续期收费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_hesitation_tb':
                    business_name = '犹豫期退保试算'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'hesitation_tb':
                    business_name = '犹豫期退保'
                    business_check_info = '犹豫期退保金额'
                    business_check_result = '犹豫期退保金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'trial_tb':
                    business_name = '退保试算'
                    business_check_info = '退保金额、产品现价、其他'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])               
                elif dict['track_info'][i]['trackType'] == 'loan':
                    business_name = '贷款'
                    business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                    business_check_result = '贷款总金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice']) + ';贷款额度:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountLimit']) + ';欠款余额:' + str(dict['track_info'][i]['trackData']['product'][0]['loanAccountAmount'])
                elif dict['track_info'][i]['trackType'] == 'loan_repayment':
                    business_name = '贷款还款'
                    business_check_info = '贷款还款金额'    
                    business_check_result = '贷款还款金额:' + str(dict['track_info'][i]['trackData']['payment']) 
                elif dict['track_info'][i]['trackType'] == 'survivalFee':
                    business_name = '生存金派发'
                    business_check_info = '生存金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'CD':
                    business_name = '客户重要资料变更'
                    business_check_info = '客户重要资料变更收付费、保额、保费'
                    business_check_result = '验证正确。' + '客户重要资料变更收付费金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'tb':
                    business_name = '退保'
                    business_check_info = '退保金额、现价'
                    business_check_result = '退保金额:' + str(dict['track_info'][i]['trackData']['payment']) + ';产品现价:' + str(dict['track_info'][i]['trackData']['product'][0]['presentPrice'])
                elif dict['track_info'][i]['trackType'] == 'continuous_bonus':
                    business_name = '持续奖金派发'
                    business_check_info = '持续奖金派发金额'
                    business_check_result = '生存金派发金额:' + str(dict['track_info'][i]['trackData']['payment'])
                elif dict['track_info'][i]['trackType'] == 'additional_premium':
                    business_name = '追加保费'
                    business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                    business_check_result = '追加保费收付费:' + str(dict['track_info'][i]['trackData']['payment']) + ';追加保费申请金额:' + str(dict['track_info'][i]['trackData']['product'][0]['addInvestAmount']) + ';追加保费扣费:' + str(dict['track_info'][i]['trackData']['product'][0]['chargeAmount']) + ';实际追加保费金额:' + str(dict['track_info'][i]['trackData']['product'][0]['surrenderAmount'])
                sh['A' + str(i+5)] = business_name  
                sh['B' + str(i+5)] = business_apply_date
                sh['C' + str(i+5)] = business_check_info
                sh['D' + str(i+5)] = business_check_result

            #存在报错场景    
            if result < sum_business:
                if dict['logInfo']['msg'] == None or dict['logInfo']['msg'] == '':
                    msg = str(dict['logInfo']['err'])
                else:
                    msg = str(dict['logInfo']['msg'])
                if result == 0:
                    sh['A' + str(5)] = '新契约'  
                    sh['B' + str(5)] = applicationDate
                    sh['C' + str(5)] = '新契约收付费金额、产品保费、保额' 
                    sh['D' + str(5)] = '验证不通过。' + msg
                else:
                    if dict['track_info'][result]['trackType'] == 'renew':
                        business_name = '续期'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '续期收费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'trial_hesitation_tb':
                        business_name = '犹豫期退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'hesitation_tb':
                        business_name = '犹豫期退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '犹豫期退保金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'trial_tb':
                        business_name = '退保试算'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、产品现价、其他'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan':
                        business_name = '贷款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款总金额、产品现价、贷款额度、欠款余额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'loan_repayment':
                        business_name = '贷款还款'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '贷款还款金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'survivalFee':
                        business_name = '生存金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '生存金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'CD':
                        business_name = '客户重要资料变更'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '客户重要资料变更收付费、保额、保费'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'tb':
                        business_name = '退保'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '退保金额、现价'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'continuous_bonus':
                        business_name = '持续奖金派发'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '持续奖金派发金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
                    elif dict['track_info'][result]['trackType'] == 'additional_premium':
                        business_name = '追加保费'
                        business_apply_date = dict['track_info'][result]['trackTime']
                        business_check_info = '追加保费收付费、追加保费申请金额、追加保费扣费、实际追加保费金额'
                        business_check_result = '验证不通过。' + msg
                        sh['A' + str(i+6)] = business_name  
                        sh['B' + str(i+6)] = business_apply_date
                        sh['C' + str(i+6)] = business_check_info
                        sh['D' + str(i+6)] = business_check_result
            sh2 = wb['record']
            rows = sh2.max_row
            sh2['A' + str(rows+1)] = current_time 
            sh2['B' + str(rows+1)] = json.dumps(dict)
            wb.save(filename)
            wb.close()



test().test()
