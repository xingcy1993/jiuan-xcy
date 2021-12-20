# coding=utf-8

import sys
import time
import logging
import pymysql
import requests
import traceback
from openpyxl import load_workbook
from func_timeout import func_set_timeout
import func_timeout
reload(sys)
sys.setdefaultencoding("utf8")
from xLibrary.chunyu.API0000 import *


"""产品万能结息利率配置：
    录入字段：险种（下拉选择项）、类型（单选框：清空后插入数据、不清空插入数据）
    step1:读取excel录入的产品万能结息利率配置内容。
    step2:判断录入险种是否存在 + excel录入数据规范。
          1.判断录入险种是否存在
          2.excel录入数据规范1：不允许存在空或多个空格
          3.excel录入数据规范2：判断'账户类型/利率类型/结息利率/起始日期/截止日期/公布时间'下录入的内容是否有误
          4.excel录入数据规范3：录入的保底利率，最多不超过1条
          5.excel录入数据规范4：录入的公布利率对应的公布日期不能存在重复
    step3:判断（清空后插入数据、不清空插入数据）不同类似对应的规则。
          1.选择‘清空后插入数据’，调用查询接口，判断数据是否为空。不为空提示：请咨询相关开发老师清空xxx产品万能结息利率配置内容。
          2.选择‘不清空插入数据’，调用查询接口，判断系统与录入的保底利率是否重复+系统公布利率对应的公布日期与录入的是否存在重复。
    step4:调用插入接口，插入产品万能结息利率配置内容。
"""
class sys_API0022:

    def __init__(self):
        pass


    @func_set_timeout(30)#设定函数超执行时间
    def investRateDefAdd_announcement_rate(self, dict):
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('产品万能结息利率配置开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            #dict中获取
            filename = dict['public']['filename']   
            """step1:读取excel录入的产品万能结息利率配置内容"""
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*读取excel录入的产品万能结息利率配置内容开始'))
            #初始化
            is_success = ''
            message = ''
            unit = '读取excel录入的产品万能结息利率配置内容'
            #获取excel文件对象
            wb = load_workbook(filename)
            #获取sheet页对象
            sheet = wb['Sheet1']
            #读取最大行
            rows = sheet.max_row
            list_excel_value = []
            for i in ['A', 'B', 'C', 'D', 'E', 'F']:
                list_excel_value.append([])
                for j in range(1,rows):
                    list_excel_value[-1].append(str(sheet[i + str(j+1)].value).strip())
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*读取excel录入的产品万能结息利率配置内容结束'))


            """step2:判断excel录入数据规范。"""
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*判断excel录入数据规范开始'))
            #初始化
            is_success = ''
            message = ''
            unit = '判断excel录入数据规范'
            #1.判断录入险种是否存在
            risk_name = dict['public']['risk_name']   #险种。格式：8237-君康安享金账户终身寿险（万能型）
            productCode = risk_name.split('-')[0]   #险种code
            productId = str(API0000_diy().query_productId(env_name, productCode))   #productId
            dict['public']['productId'] = productId
            #判断数据库是否可查询到录入的险种是否
            if productId == '':
                is_success = False
                message = '数据库未查询到录入的险种！'
            assert is_success != False

            #2.excel录入数据规范1：不允许存在空或多个空格
            pattern1 = r'^(\s)+$'   #正则公式：一个或多个空格
            for i in range(0, len(list_excel_value)):
                for j in range(0, len(list_excel_value[i])):
                    if list_excel_value[i][j] == '' or re.match(pattern1, list_excel_value[i][j]) != None or list_excel_value[i][j] == 'None':
                        is_success = False
                        message = 'excel录入数据,第' + str(j+2) + '行第' + str(i+1) + '列存在空！'
                        break
                if is_success == False:
                    break
            assert is_success != False

            #3.excel录入数据规范2：判断'账户类型/利率类型/结息利率/起始日期/截止日期/公布时间'下录入的内容是否有误
            for i in range(1, rows):
                #判断'账户类型'列
                if list_excel_value[0][i-1] not in ['万能活期账户', '万能定期账户']:
                    is_success = False
                    message = 'excel录入数据的账户类型，第' +str(i+1) + '行录入内容有误！'
                    break
                #判断'利率类型'列
                if list_excel_value[1][i-1] not in ['保底利率', '公布利率']:
                    is_success = False
                    message = 'excel录入数据的利率类型，第' +str(i+1) + '行录入内容有误！'
                    break
                #判断'结息利率'列
                pattern2 = r'^(\d)+\.(\d)+$'   #正则公式：小数点
                if re.match(pattern2, list_excel_value[2][i-1]) == None:
                    is_success = False
                    message = 'excel录入数据的结息利率，第' +str(i+1) + '行录入内容有误！'
                    break
                #判断'起始日期'列
                pattern3 = r'^\d{4}-\d{2}-\d{2}$'   #正则公式：日期xxxx-xx-xx
                if re.match(pattern3, list_excel_value[3][i-1]) == None:
                    is_success = False
                    message = 'excel录入数据的起始日期，第' +str(i+1) + '行录入内容有误！'
                    break
                #判断'截止日期'列
                if re.match(pattern3, list_excel_value[4][i-1]) == None:
                    is_success = False
                    message = 'excel录入数据的截止日期，第' +str(i+1) + '行录入内容有误！'
                    break
                #判断'公布时间'列
                if re.match(pattern3, list_excel_value[5][i-1]) == None:
                    is_success = False
                    message = 'excel录入数据的公布时间，第' +str(i+1) + '行录入内容有误！'
                    break
            assert is_success != False

            #4.excel录入数据规范3：录入的保底利率，最多不超过1条
            guaranteed_interest_rate_num = 0   #初始化保底利率录入的个数
            guaranteed_interest_rate_i = ''   #初始化保底利率对应的条数下标
            for i in range(0, len(list_excel_value[1])):
                if list_excel_value[1][i] == '保底利率':
                    guaranteed_interest_rate_i = i   #获取保底利率对应的条数下标
                    guaranteed_interest_rate_num = guaranteed_interest_rate_num + 1   #保底利率录入的个数+1
            if guaranteed_interest_rate_num > 1:
                is_success = False
                message = '保底利率录入重复！'
            assert is_success != False

            #5.excel录入数据规范4：录入的公布利率对应的公布日期不能存在重复
            list1 = list_excel_value[5]
            if guaranteed_interest_rate_i != '':
                del list1[guaranteed_interest_rate_i]
            set_list1 = set(list_excel_value[5])   #存储公布利率的列表去重
            if len(list1) != len(set_list1):
                is_success = False
                message = '公布利率录入存在重复！'
            assert is_success != False
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*判断excel录入数据规范开始结束'))


            """step3:判断（清空后插入数据、不清空插入数据）不同类似对应的规则。"""
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*判断系统内数据与excel录入数据是否冲突开始'))
            #登录系统
                #初始化
            is_success = ''
            message = ''
            unit = '登录系统'
            interface_ip = API0000_diy().ip_conf(env_name)   #环境ip
            rs = API0000_sys_otherapi().login(env_name, username_entry, interface_ip)
            is_success = rs[0]
            message = rs[1]
            magicX = rs[2]
            dict['public']['magicX'] = magicX
            assert is_success == True
            
            #万能利率配置查询接口
                #初始化
            is_success = ''
            message = ''
            unit = '万能利率配置查询接口'
            rateType = ''
            fundCode = ''
            publicTime = ''
            productId = dict['public']['productId']
            #产品万能利率配置查询接口
            response = API0000_sys_otherapi().investRateDefQuery(magicX, interface_ip, productId, rateType, fundCode, publicTime)
            is_success = response['isSuccess']
            if is_success:
                message = ''
                resultData = response['resultData']
            else:
                message = response['message']
                resultData = ''
            #断言
            assert is_success == True

            #判断系统内数据与excel录入数据是否冲突
            is_success = ''
            message = ''
            unit = '判断系统内数据与excel录入数据是否冲突'
            #1.选择‘清空后插入数据’，调用查询接口，判断数据是否为空。不为空提示：请咨询相关开发老师，清空xxx产品万能结息利率配置内容。
            if dict['public']['insert_type'] == '清空后插入数据':
                #调用查询接口，判断数据是否为空
                if len(resultData) != 0:
                    is_success = False
                    message = '请咨询相关开发老师，提前清空' + risk_name + '产品万能结息利率配置内容。本组件不提供删除时数据库操作！'
                #断言
                assert is_success == True
            #2.选择‘不清空插入数据’，调用查询接口，判断系统与录入的保底利率是否重复+系统公布利率对应的公布日期与录入的是否存在重复。
            elif dict['public']['insert_type'] == '不清空插入数据':
                if len(resultData) > 0:
                    #保底利率：系统与录入的是否重复
                    flag_sys_guaranteed_interest_rate = ''   #初始化系统是否存在保底利率
                    for i in range(0, len(resultData)):
                        if str(resultData[i]['rateType']) == '1':
                            flag_sys_guaranteed_interest_rate = '1'
                    if flag_sys_guaranteed_interest_rate == '1' and guaranteed_interest_rate_num == 1:
                        is_success = False
                        message = '系统已存在保底利率，不可再录入保底利率！'
                    assert is_success != False
                    #公布利率：系统存在的公布日期与录入的是否存在重复
                    num_repeat = 0   #初始化系统与录入的公布利率对应的公布日期重复个数为0
                    resultData_temp = resultData   #去除系统保底利率的记录列表
                    for i in range(0, len(resultData_temp)):
                        if str(resultData_temp[i]['rateType']) == '1':
                            del resultData_temp[i]
                            break    
                    for i in range(0, len(resultData_temp)):   #获取重复个数
                        for j in range(0, len(list1)):
                            if resultData_temp[i]['publicTime'] == list1[j]:
                                num_repeat = num_repeat + 1
                    if num_repeat > 0:
                        is_success = False
                        message = '系统中公布利率对应的公布日期与录入的存在重复！'
                    assert is_success != False
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*判断系统内数据与excel录入数据是否冲突结束'))


            #step4:调用插入接口，插入产品万能结息利率配置内容。
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*插入excel录入的产品万能结息利率配置内容开始'))
                #初始化
            is_success = ''
            message = ''
            unit = '万能利率配置插入接口'
            for i in range(0, len(list_excel_value[0])):
                #账号类型
                if list_excel_value[0][i] == '万能活期账户':
                    fundCode = '01'
                elif list_excel_value[0][i] == '万能定期账户':
                    fundCode = '02'
                #利率类型
                if list_excel_value[1][i] == '保底利率':
                    rateType = '1'
                elif list_excel_value[1][i] == '公布利率':
                    rateType = '2'
                rate = str(list_excel_value[2][i])   #结息利率
                startDate = list_excel_value[3][i]   #起始日期
                endDate = list_excel_value[4][i]   #起始日期
                publicTime = list_excel_value[5][i]   #公布时间
                #调用接口
                response = API0000_sys_otherapi().investRateDefAdd(magicX, interface_ip, productId, fundCode, rateType, rate, startDate, endDate, publicTime)
                #获取响应值
                if response['success'] == True and response['message'] == '处理成功':
                    is_success = True
                    message = ''
                else:
                    is_success = False
                    message = response['message']      
                assert is_success == True

            #登出系统
            rs = API0000_sys_otherapi().logout(magicX, interface_ip)
            is_success = rs[0]
            message = rs[1]
            unit = '登出系统'
            assert is_success == True
            magicX = ''
            dict['public']['magicX'] = ''
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*插入excel录入的产品万能结息利率配置内容结束'))
            dict['logInfo']['code'] = '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('产品万能结息利率配置结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            dict['logInfo']['code'] = '0'    #记录异常标识
            dict['logInfo']['err'] = unit + ':' + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
                logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
            #记录异常日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*异常位置:\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict




if __name__ == "__main__":   
    try:
        begin_time = datetime.datetime.now()   #记录开始时间
        #1.自定义录入+初始化
        execNo = 'test001'   #测试编号
        env_name = 'uat4'   #环境名称
        env_name = env_name.encode('utf-8')
        username_entry = 'xcy1'    #登录用户
        risk_name = '8237-君康安享金账户终身寿险（万能型）'   #险种。格式：8237-君康安享金账户终身寿险（万能型）
        insert_type = '不清空插入数据'   #插入类型。清空后插入数据or不清空插入数据
        filename = r'D:\xLibrary\chunyu\doc\investRateDefAdd_announcement_rate\investRateDefAdd_announcement_rate.xlsx'
        #2.接收参数：第一位为测试编号，第二位json串
        if len(sys.argv) == 2:
            execNo = sys.argv[1]   #测试编号     
        elif len(sys.argv) > 2:
            execNo = sys.argv[1]   #测试编号 
            input_dict = sys.argv[2]   #输入的json串
            input_dict = input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
            if 'win' in sys.platform:
                input_dict = input_dict.decode('gb2312').encode('utf-8') 
            input_dict = json.loads(input_dict)   #json转dict格式
            env_name = input_dict['env_name']    #环境
            username_entry = input_dict['username_entry']    #新契约录入用户
            risk_name = input_dict['risk_name']    #险种。格式：8237-君康安享金账户终身寿险（万能型）
            insert_type = input_dict['insert_type']    #插入类型。清空后插入数据or不清空插入数据
            filename = input_dict['filename']   #excel文件路径
        interface_ip = API0000_diy().ip_conf(env_name)
        #3.定义dict
        dict = {'execNo': execNo, 'public': {'env_name':env_name, 'username_entry': username_entry, 'interface_ip': interface_ip, 'insert_type': insert_type, 'risk_name': risk_name, 'filename':filename, 'magicX':''}, 'logInfo': {'code':'0', 'msg':'', 'err':'', 'result':'','unit':''}}
        #4.1定义logging文件路径
        if 'win' in sys.platform:
            logging_filename = 'D:\\xLibrary\\chunyu\\%s.html' % str(execNo)
            #判断windows是否存在某文件，存在则删除文件
            if os.path.exists(logging_filename):
                os.remove(logging_filename)
        else:
            logging_filename = '/data/xServer/xReport/%s.html' % str(execNo)
        #4.2定义logging格式
        logging.basicConfig(level=logging.WARNING , format='%(message)s  &nbsp&nbsp&nbsp&nbsp%(asctime)s <br/><br/>', filename=logging_filename, filemode='a')
        #5.执行脚本
        dict = sys_API0022().investRateDefAdd_announcement_rate(dict)
    except Exception, e:   #常见异常的捕捉
        dict['logInfo']['code'] = '0'
        #记录异常日志
        logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*异常位置:\n')  + traceback.format_exc())
        #异常场景-登出系统（防止与用户被挂起）
        if dict['public']['magicX'] != '':
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
            dict['public']['magicX'] = ''
    except func_timeout.exceptions.FunctionTimedOut:    #函数超时异常的捕捉
        #记录执行时间超时日志
        logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        dict['logInfo']['code'] = '0'
        dict['logInfo']['msg'] = '执行时间超时退出。' + dict['logInfo']['msg']
        #异常场景-登出系统（防止与用户被挂起）
        if dict['public']['magicX'] != '':
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
            dict['public']['magicX'] = ''
    finally:
        try:
            end_time = datetime.datetime.now()   #记录结束时间
            #打印终端开始符
            print '\n\n' + '#'*3 + 'RESULT' + '#'*3
            if 'win' in sys.platform:   #windows系统平台
                if dict['logInfo']['code'] == '1':
                    print '万能结息利率是否配置成功：成功'.decode('utf-8').encode('gb2312')
                    print ('环境：' + env_name).decode('utf-8').encode('gb2312')
                    print ('险种：' + risk_name).decode('utf-8').encode('gb2312')
                    print ('耗时：' + API0000_diy().logger_count_time(begin_time, end_time)).decode('utf-8').encode('gbk')
                else:
                    print '万能结息利率是否配置成功：不成功'.decode('utf-8').encode('gb2312')
                    print '不成功原因：'.decode('utf-8').encode('gb2312'),;API0000_diy().solve_terminal_garbled(dict['logInfo']['msg'])
                    print ('环境：' + env_name).decode('utf-8').encode('gb2312')
                    print ('险种：' + risk_name).decode('utf-8').encode('gb2312')
                    print ('耗时：' + API0000_diy().logger_count_time(begin_time, end_time)).decode('utf-8').encode('gbk')
                    #print 'dict:'
                    #print dict
            else:    #非windows系统平台
                if dict['logInfo']['code'] == '1':
                    result_str = "万能结息利率是否配置成功：成功,,环境：%s,,险种：%s,,耗时：%s" % (env_name, risk_name, API0000_diy().logger_count_time(begin_time, end_time))
                else:
                    result_str = "万能结息利率是否配置成功：不成功,,不成功原因：%s,,环境：%s,,险种：%s,,耗时：%s,,dict:%s" % (dict['logInfo']['msg'], risk_name, API0000_diy().logger_count_time(begin_time, end_time), json.dumps(dict))
                print result_str
            #打印终端结束符
            print '#'*3 + 'RESULT' + '#'*3
        except Exception, e:
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*组件结果展示部分脚本报错:\n')  + traceback.format_exc())







