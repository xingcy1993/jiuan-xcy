#coding=utf-8

import os
import sys
import time
import json
import random
import logging
import datetime
import traceback
from openpyxl import load_workbook
from func_timeout import func_set_timeout
import func_timeout
reload(sys)
sys.setdefaultencoding("utf8")
from xLibrary.chunyu.API0000 import *
from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002_1 import sys_API0002


class sys_API0001:

    def __init__(self):
        pass

    @func_set_timeout(300)#设定函数超执行时间 
    def nb (self,env_name, applicationDate, dict):
        """
            step1:存储新契约数据(包含两部分：定义dict、excel获取数据存入dict)
            step2:判断是否需要修改服务器时间
            step3:新契约录入
            step4:新契约复核
            step5:新契约核保
            step6:新契约收费
            step7:新契约回执
            step8:定义轨迹信息+检查点信息
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('新契约出单_开始(预计耗时:不需修改系统时间60s;需修改系统时间1分50s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            job = dict['logInfo']['job']   #场景名
            #step1:存储新契约数据+验证契约录入用户、复核用户是否可正常登录系统
            dict = sys_API0001().store_nbdata(env_name, applicationDate, dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常

            #step2:判断是否需要修改服务器时间
            dict = sys_API0002().Modify_servertime(applicationDate, job, dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常
            #time.sleep(20)

            #step3:新契约录入
            rs = sys_API0001().nb_entry(dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常
            proposalId = rs[0]
            dict = rs[1]

            #step4:新契约复核
            rs = sys_API0001().nb_review(proposalId, dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常
            magicX = rs[0]
            others = rs[1]
            dict = rs[2]
            if others == '2':   #是否进入人工核保
                #step5:新契约核保
                dict = sys_API0001().nb_uw(magicX, others, dict)
                assert dict['logInfo']['code'] == '1'   #断言是否异常
            #step6:新契约收费
            dict = sys_API0001().nb_payment(magicX, dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常
            #step7:新契约回执
            dict = sys_API0001().nb_receipt(magicX, dict)
            assert dict['logInfo']['code'] == '1'   #断言是否异常
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('新契约出单_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            dict['logInfo']['code'] = '0'   #dict['logInfo']['code']记录异常标识
            #记录异常日志
            logging.warning(API0000_diy().text_conversion('异常位置:\n') + traceback.format_exc())
            #异常场景-登出系统（防止与用户被挂起）
            if dict['public']['magicX'] != '':
                API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
                API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_review'])
                logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
                dict['public']['magicX'] = ''
        except func_timeout.exceptions.FunctionTimedOut:
            #记录执行时间超时日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = '执行时间超时退出。' + dict['logInfo']['msg']
            #异常场景-登出系统（防止与用户被挂起）
            if dict['public']['magicX'] != '':
                API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
                API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_review'])
                logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
                dict['public']['magicX'] = ''
        finally:
            dict = sys_API0001().save_check_info(dict)
        return dict


    """存储新契约数据+验证契约录入用户、复核用户是否可正常登录系统"""
    @func_set_timeout(30)#设定函数超执行时间
    def store_nbdata(self,env_name, applicationDate, dict):
        """
           step1:获取excel值+存入dict中
           step2:验证契约录入用户、复核用户是否可正常登录系统
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*出单前准备：数据存储+用户验证_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''
            message = ''
            magicX = ''
            unit = ''
            """step1:获取excel值存入dict中"""
            unit = '系统-新契约_获取excel值存入dict中'  # 节点
            #根据脚本执行平台获取excel文件路径（win系统为自测，linux为自动化网页平台）
            filename = dict['public']['filename']   #dict中获取
            #获取excel文件对象
            wb = load_workbook(filename)
            #获取sheet页对象
            sh = wb['nb_param']
            # 读取公共信息
            dict['public']['env_name'] = env_name  # 环境名称
            interface_ip = dict['public']['interface_ip']   #环境ip
            #dict['public']['username_entry'] = sh['B3'].value  # 契约录入用户
            #dict['public']['username_review'] = sh['B4'].value  # 契约复核用户
            dict['policy_info']['policyNo'] = ''  #保单号
            if sh['D5'].value == '北京分公司':   #所属机构
                dict['policy_info']['organId'] = '10002'
            elif sh['D5'].value == '黑龙江分公司':
                dict['policy_info']['organId'] = '10003'
            elif sh['D5'].value == '上海分公司':
                dict['policy_info']['organId'] = '10004'
            elif sh['D5'].value == '江苏分公司':
                dict['policy_info']['organId'] = '10005'
            elif sh['D5'].value == '浙江分公司':
                dict['policy_info']['organId'] = '10006'
            elif sh['D5'].value == '北京分公司营业本部':
                dict['policy_info']['organId'] = '10007'
            elif sh['D5'].value == '上海分公司营业本部':
                dict['policy_info']['organId'] = '10008'
            elif sh['D5'].value == '黑龙江分公司营业本部':
                dict['policy_info']['organId'] = '10009'
            elif sh['D5'].value == '':
                dict['policy_info']['organId'] = ''
            dict['public']['is_HolderExemption'] = sh['B7'].value  # 是否投保人豁免
            dict['public']['is_2risk'] = sh['B25'].value  #是否双主险
            # 读取保单信息：
            dict['policy_info']['policyChannel'] = sh['B5'].value  # 公司渠道
            dict['policy_info']['applicationDate'] = applicationDate  # 投保日期
            dict['policy_info']['payment_nb']['payMode'] = str(sh['B24'].value)  # 缴费方式
            # 读取保单信息-投保人信息
            dict['policy_info']['holder_info']['sameASInsurd'] = sh['B8'].value  # 投被保人是否同一人
            dict['policy_info']['holder_info']['name'] = sh['B10'].value  # 投保人姓名
            dict['policy_info']['holder_info']['gender'] = sh['B11'].value  # 投保人性别
            dict['policy_info']['holder_info']['pbHoldBirth'] = sh['B12'].value  # 投保人出生日期
            dict['policy_info']['holder_info']['certiType'] = sh['B13'].value  # 投保人证件类型
            dict['policy_info']['holder_info']['certiCode'] = sh['B14'].value  # 投保人证件号码
            dict['policy_info']['holder_info']['pbCertiValidEndDate'] = sh['B15'].value  # 投保人证件止期
            dict['policy_info']['holder_info']['mobilePhone'] = sh['B16'].value  # 投保人手机号
            # 读取保单信息-被保人信息
                #非投保人豁免，投被保人不为同一人时：
            if (dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0'):
                insured1 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                insured1['insuName'] = sh['B17'].value  # 被保人姓名
                insured1['insuGender'] = sh['B18'].value  # 被保人性别
                insured1['insuBirth'] = sh['B19'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B20'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B21'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B22'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B23'].value  # 被保人手机号
                insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                insured1 = {}
                insured1['insuRelation'] = '00'  # 被保人与投保人关系
                insured1['insuName'] = sh['B10'].value  # 被保人姓名
                insured1['insuGender'] = sh['B11'].value  # 被保人性别
                insured1['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B16'].value  # 被保人手机号
                insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #投保人豁免时：
            elif dict['public']['is_HolderExemption'] == '1':
                insured1 = {}
                insured2 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                insured1['insuName'] = sh['B17'].value  # 被保人姓名
                insured1['insuGender'] = sh['B18'].value  # 被保人性别
                insured1['insuBirth'] = sh['B19'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B20'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B21'].value  # 被保人证件号码
                insured1['no'] = '1'  # 第1被保人
                insured1['insuCertiValidEndDate'] = sh['B22'].value  # 被保人证件止期
                insured1['insuMobile'] = sh['B23'].value  # 被保人手机号
                insured2['insuRelation'] = '00'  # 被保人与投保人关系
                insured2['insuName'] = sh['B10'].value  # 被保人姓名
                insured2['insuGender'] = sh['B11'].value  # 被保人性别
                insured2['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured2['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured2['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured2['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured2['insuMobile'] = sh['B16'].value  # 被保人手机号
                insured2['no'] = '2'  # 第2被保人
                dict['policy_info']['insured_info'].append(insured1)
                dict['policy_info']['insured_info'].append(insured2)
                
            # 读取保单信息-主险信息
                #判断是否双主险。得到主险个数
            if sh['B25'].value == '0':
                main_risk_num = 1   #主险个数
            elif sh['B25'].value == '1':
                main_risk_num = 2   #主险个数
            
                #根据主险个数，组装主险信息列表。
            for i in range(0, main_risk_num):
                unit = '系统-新契约_获取excel值存入dict中'
                main_risk_info = {}
                if i == 0:
                    main_risk_column = 2   #excel中主险所在的列
                elif i ==1:
                    main_risk_column = 5   #excel中主险所在的列
                main_risk_info['main_risk_num'] = str(i+1) # 第x主险
                main_risk_info['Main_Rider'] = 'M' # 主附险标识
                main_risk_info['productCode'] = str(sh.cell(row=28, column=main_risk_column).value)   #险种代码
                unit = '系统-新契约_获取excel值存入dict中_连接数据库'
                try:
                    cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
                except:
                    is_success = False
                    message = '连接数据库不成功！'
                finally:
                    assert is_success != False
                sql = 'select * from t_product where PRODUCT_CODE=' + "'" + main_risk_info['productCode'] + "'"
                unit = '系统-新契约_获取excel值存入dict中_数据库查询主险种信息'
                cursor.execute(sql)
                result = cursor.fetchall()
                unit = '系统-新契约_获取excel值存入dict中'
                try:
                    main_risk_info['productId'] = str(result[0][0])  # 险种id
                except:
                    is_success = False
                    message = 'excel录入的险种' + str(main_risk_info['productCode']) + ',数据库中不存在！'
                finally:
                    assert is_success != False
                main_risk_info['liability_state'] = '01'   #险种状态
                main_risk_info['chargeMode'] = sh.cell(row=29, column=main_risk_column).value  # 缴费频率
                main_risk_info['coverPeriodType'] = sh.cell(row=30, column=main_risk_column).value  # 保障期间类型
                main_risk_info['coveragePeriod'] = sh.cell(row=31, column=main_risk_column).value  # 保障期间
                main_risk_info['chargePeriodType'] = sh.cell(row=32, column=main_risk_column).value  # 缴费期间类型
                main_risk_info['chargePeriod'] = sh.cell(row=33, column=main_risk_column).value  # 缴费期间
                main_risk_info['sumAssured'] = sh.cell(row=34, column=main_risk_column).value  # 保额
                main_risk_info['stdPremBf'] = sh.cell(row=35, column=main_risk_column).value  # 保费
                main_risk_info['units'] = sh.cell(row=36, column=main_risk_column).value  # 份数
                main_risk_info['planFreq'] = sh.cell(row=37, column=main_risk_column).value  # 年金领取方式 或 年金/生存金领取方式
                main_risk_info['payOption'] = sh.cell(row=38, column=main_risk_column).value  # 年金使用方式
                main_risk_info['bonusYearOrAge'] = sh.cell(row=39, column=main_risk_column).value  # 领取年期或年龄
                main_risk_info['renew'] = sh.cell(row=40, column=main_risk_column).value  # 是否可续保
                # 存入dict
                dict['policy_info']['main_risk_info'].append(main_risk_info)

            # 读取保单信息-附加险信息
            for i in range(3, int(sh['B26'].value)+3):
                unit = '系统-新契约_获取excel值存入dict中'
                sub_risk_info = {} 
                sub_risk_info['belong_mainrisk_productId'] = sh.cell(row=28, column=2).value  # 所属主险的productId
                sub_risk_info['Main_Rider'] = 'R' # 主附险标识
                sub_risk_info['productCode'] = str(sh.cell(row=28, column=i).value)   #险种代码
                unit = '系统-新契约_获取excel值存入dict中_连接数据库'
                cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
                sql = 'select * from t_product where PRODUCT_CODE=' + "'" + sub_risk_info['productCode'] + "'"
                unit = '系统-新契约_获取excel值存入dict中_数据库查询附加险种信息'
                cursor.execute(sql)
                result = cursor.fetchall()
                unit = '系统-新契约_获取excel值存入dict中'
                try:
                    sub_risk_info['productId'] = str(result[0][0])  # 险种id
                except:
                    is_success = False
                    message = 'excel录入的险种' + str(sub_risk_info['productCode']) + ',数据库中不存在！'
                finally:
                    assert is_success != False
                sub_risk_info['liability_state'] = '01'   #险种状态
                sub_risk_info['chargeMode'] = sh.cell(row=29, column=i).value  # 缴费频率
                sub_risk_info['coverPeriodType'] = sh.cell(row=30, column=i).value  # 保障期间类型
                sub_risk_info['coveragePeriod'] = sh.cell(row=31, column=i).value  # 保障期间
                sub_risk_info['chargePeriodType'] = sh.cell(row=32, column=i).value  # 缴费期间类型
                sub_risk_info['chargePeriod'] = sh.cell(row=33, column=i).value  # 缴费期间
                sub_risk_info['sumAssured'] = sh.cell(row=34, column=i).value  # 保额
                sub_risk_info['stdPremBf'] = sh.cell(row=35, column=i).value  # 保费
                sub_risk_info['units'] = sh.cell(row=36, column=i).value  # 份数
                sub_risk_info['planFreq'] = sh.cell(row=37, column=i).value  # 年金领取方式 或 年金/生存金领取方式
                sub_risk_info['payOption'] = sh.cell(row=38, column=i).value  # 年金使用方式
                sub_risk_info['bonusYearOrAge'] = sh.cell(row=39, column=i).value  # 领取年期或年龄
                sub_risk_info['renew'] = sh.cell(row=40, column=i).value  # 是否可续保
                # 存入dict
                dict['policy_info']['sub_risk_info'].append(sub_risk_info) 
            is_success = True
            # 加入断言
            assert is_success == True
            """step2:验证契约录入用户、复核用户是否可正常登录系统"""
            unit = '系统-验证契约录入用户、复核用户是否可正常登录系统:'   #节点标识
            #登录用户
            rs = API0000_diy().verify_userlogin(env_name, dict['public']['username_entry'], dict['public']['interface_ip'])
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            # 加入断言
            assert is_success == True
            #登出用户
            rs = API0000_diy().verify_userlogin(env_name, dict['public']['username_review'], dict['public']['interface_ip'])
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息

            #3.校验录入险种，脚本是否支持出单
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：校验录入险种，脚本是否支持出单：'  # 节点
            list_risk = ['10001','10002','10003','10004','10006','10007','10008','10009','10011','10012','10014','10017','10020','10021','10035','10036','10037','10038','10042','10044','10046','10047','10048','10049','10050','10053','10054','10063','10065','10071','10073','10076','10077','10078','10079','10080','10081','10082','10083','10097','10098','10099','10100','10106','10107','10108','10109','10110','10111','10112','10113','10118','10130','10131','10132','10133','10135','10138','10139','10140','10141','10142','10148','10159','10160','10161','10162','10163','10164','10166','10168','10169','10170','10171','10173','10174','10175','10177','10184','10189','10193','10194','10200','10204','10207','10211','10214','10223','10224','10225','10226','10129','10231','10240','10241','10143','10144','10203','10205','10206','10232','10128','10202','10185', '10186', '10187','10188','10227','10228','10229','10230','10119','10209','10222','10180','10181','10127','10172','10233','10120','10121','10122','10216']
            dict_risk = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']
            for i in range(0, len(dict_risk)):
                productId = dict_risk[i]['productId']
                if productId not in list_risk:
                    is_success = False
                    message = 'excel录入的险种' + str(dict_risk[i]['productCode']) + '暂不支持出单，请联系管理员进行配置！'
                else:
                    is_success = True
                    message = ''
                assert is_success == True
            #结束
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*出单前准备：数据存储+用户验证_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
            # 加入断言
            assert is_success == True
            dict['logInfo']['code'] = '1'   #记录无异常标识
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'   #记录异常标识
            dict['logInfo']['err'] = unit + ':' + str(e)   #存入脚本日志报错信息
            if message == '':
                message = str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n') + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入接口报错信息
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict



    """新契约录入"""
    @func_set_timeout(180)#设定函数超执行时间
    def nb_entry(self, dict):
        """
           step1:登录系统-新契约录入用户
           step2:投保单号提交接口
           step3:保单基本信息录入接口
           step4:投保人信息录入接口
           step5:投保人告知接口
           step6:被保人信息录入接口
           step7:被保人查询接口
           step8:被保人告知接口
           step9:收费方式信息录入接口
           step10:险种信息录入接口
           step11:新契约录入提交接口
           step12:登出系统-新契约录入用户
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约录入_开始(预计耗时:20s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            magicX = ''
            """step1:登录系统-新契约录入用户"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '系统-新契约录入_登录系统'  # 节点标识
            message = ''   #信息标识
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   #环境名称
            interface_ip = dict['public']['interface_ip']   #接口ip
            username = dict['public']['username_entry']    #新契约录入用户
            #调用原子脚本
            rs = API0000_sys_otherapi().login(env_name,username, interface_ip)
            #获取所需响应值
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            magicX = rs[2]   #会话id
            #断言
            assert is_success == True
            dict['public']['magicX'] = magicX
            
            """step2:投保单号提交接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_投保单号提交接口'  # 节点
            #dict获取数据+定义数据
            proposalNo = API0000_diy().random_proposalNo()  # 投保单号（随机生成保单号）
            policyType = 1  # 保单类型：1-个险
            #调用原子脚本
            response = API0000_nb().entry_newProposal(magicX, interface_ip, policyType, proposalNo)
            #获取所需响应值
            is_success = response['success']
            if is_success == True:
                others = response['others']
                # 获取生成的保单号
                policyNo = others['policyNo']
                # 获取服务器时间
                registDate = others['registDate']
                # 获取proposalId
                proposalId = others['proposalId']
                message = ''
            else:
                policyNo = ''
                registDate = ''
                proposalId = ''
                message = response['message']
            #断言
            assert is_success == True
            dict['policy_info']['policyNo'] = policyNo  # 保单号
            
            """step3:保单基本信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_保单基本信息录入接口'  # 节点
            #dict获取数据+定义数据      
                #公司渠道
            policyChannel = dict['policy_info']['policyChannel']  
                #投保日期
            applicationDate = dict['policy_info']['applicationDate'] 
                #定义生效日期 effectiveDate
            date1 = datetime.datetime.strptime(applicationDate, '%Y-%m-%d')
            date2 = date1 + datetime.timedelta(days=1)
            effectiveDate = date2.strftime('%Y-%m-%d')
                #定义所属机构 organId；定义销售组织 channelId；定义合作机构 headPartner；定义合作网点 partnerId;定义代理人 agentId
            if env_name == 'uat4' and policyChannel == '03':
                organId = '10008'
                channelId = '100533'
                headPartner = '100000'
                partnerId = '100004'
                agentId = '5097183066'
            elif env_name == 'uat4' and policyChannel == '09':
                organId = '10007'
                channelId = '100389'
                headPartner = '120000'
                partnerId = '20010554'
                agentId = '107702'
            elif env_name == 'uat6' and policyChannel == '03':
                organId = '10008'
                channelId = '100004'
                headPartner = '100000'
                partnerId = '100004'
                agentId = '100342'
            elif env_name == 'uat6' and policyChannel == '06':
                organId = '10008'
                channelId = '800314'
                headPartner = '29002493'
                partnerId = '29002494'
                agentId = '5097184964'
            elif env_name == 'uat6' and policyChannel == '09':
                organId = '10007'
                channelId = '100389'
                headPartner = '120000'
                partnerId = '120001'
                agentId = '107702'
            elif env_name == 'uat7' and policyChannel == '03':
                organId = '10008'
                channelId = '100608'
                headPartner = '1100000'
                partnerId = '1970378'
                agentId = '100334'
            elif env_name == 'uat7' and policyChannel == '06':
                organId = '10007'
                channelId = '800226'
                headPartner = '20003358'
                partnerId = '20003359'
                agentId = '5097188336'
            elif env_name == 'uat7' and policyChannel == '09':
                organId = '10007'
                channelId = '100389'
                headPartner = '120000'
                partnerId = '120000'
                agentId = '107702'
            elif env_name == 'uat8' and policyChannel == '03':
                organId = '10008'
                channelId = '100533'
                headPartner = '100000'
                partnerId = '100004'
                agentId = '5097183066'
            elif env_name == 'uat8' and policyChannel == '06':
                organId = '10007'
                channelId = '800227'
                headPartner = '20003358'
                partnerId = '20003359'
                agentId = '5097183136'
            elif env_name == 'uat8' and policyChannel == '09':
                organId = '10007'
                channelId = '100389'
                headPartner = '120000'
                partnerId = '120001'
                agentId = '107702'
            elif env_name == '预生产' and policyChannel == '03':
                organId = '10030'
                channelId = '800280'
                headPartner = '100000'
                partnerId = '100004'
                agentId = '100598'
            elif env_name == '预生产' and policyChannel == '06':
                organId = '10007'
                channelId = '800296'
                headPartner = '20005380'
                partnerId = '20005381'
                agentId = '1216962'
            elif env_name == '预生产' and policyChannel == '09':
                organId = '10007'
                channelId = '100389'
                headPartner = '120000'
                partnerId = '120001'
                agentId = '107702'
                #定义销售渠道 salesChannelText；定义 salesChannel;定义柜员代码 tellerCode;定义销售人员代码 sellerCode;定义销售人员姓名 sellerName；定义提交渠道 submitChannel
            if policyChannel == '03':
                salesChannelText = '银邮代理'
                salesChannel = '0200'
                tellerCode = '1'
                sellerCode = '1'
                sellerName = '1'
                submitChannel = '03'
            elif policyChannel == '06':
                salesChannelText = '个人代理'
                salesChannel = '0100'
                tellerCode = ''
                sellerCode = ''
                sellerName = ''
                submitChannel = '01'
            elif policyChannel == '09':
                salesChannelText = '保险专业代理'
                salesChannel = '0300'
                tellerCode = ''
                sellerCode = ''
                sellerName = ''
                submitChannel = '05'
                #定义出单方式 subSalesChannel ;定义是否自垫 autoPaddingPrem;定义争议方式 disputedType;定义递送方式 policyDeliveryMode
            subSalesChannel = 'NG-01'
            autoPaddingPrem = '0'
            disputedType = ''
            policyDeliveryMode = '3' 
            if dict['policy_info']['organId'] != '':   #判断excel是否录入了所属机构
                organId = dict['policy_info']['organId']  
            #调用原子脚本
            response = API0000_nb().proposal_updateProposalInfo(magicX, interface_ip, proposalNo, policyNo, proposalId, organId, policyChannel, channelId, salesChannelText, salesChannel, headPartner, partnerId, tellerCode, sellerCode, sellerName, agentId, submitChannel, subSalesChannel, applicationDate, effectiveDate, autoPaddingPrem, disputedType, policyDeliveryMode)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True           

            """step4:投保人信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_投保人信息录入接口'  # 节点
            #dict获取数据+定义数据
            is_HolderExemption = dict['public']['is_HolderExemption']
            if dict['policy_info']['holder_info']['sameASInsurd'] == '1' or is_HolderExemption == '1': #投被保人为同一人，和投保人豁免，sameASInsurd均录入1
                sameASInsurd = '1'
            else:
                sameASInsurd = ''
            Holder_name = dict['policy_info']['holder_info']['name']   # 投保人姓名
            Holder_gender = dict['policy_info']['holder_info']['gender']   # 投保人性别
            Holder_Birth = dict['policy_info']['holder_info']['pbHoldBirth']   # 投保人出生日期
            Holder_certiType = dict['policy_info']['holder_info']['certiType']   # 投保人证件类型
            Holder_certiCode = dict['policy_info']['holder_info']['certiCode']   # 投保人证件号码
            Holder_CertiValidEndDate = dict['policy_info']['holder_info']['pbCertiValidEndDate']   # 投保人证件止期
            Holder_mobilePhone = dict['policy_info']['holder_info']['mobilePhone']   # 投保人手机号
            Holder_isMarried = '20'  # 婚姻状况
            Holder_height = '167'  # 身高/CM
            Holder_weight = '67'  # 体重 /KG
            Holder_officeTel = '010-98343543'  # 固定电话
            Holder_homeTel = '010-98343543'  # 家庭电话
            Holder_industry = '03'  # 工作行业
            Holder_workCompany = '东方小清华'  # 工作单位/学校名称
            Holder_email = API0000_diy().random_email()  # 电子邮箱
            Holder_jobCode = '0001001'  # 职业代码/名称
            Holder_jobClass = '1'  # 职业等级
            Holder_driverLicenseType = '06'  # 驾照类型
            Holder_nationality = 'CHN'  # 国籍/地区
            Holder_nationnality1 = '01'  # 民族
            Holder_educationId = '40'  # 学位
            Holder_medicalInsType = '04'  # 医保类型
            Holder_incomeSource = '01'  # 收入来源
            Holder_incomeSourceNote = 'wu'  # 收入来源备注
            Holder_annualIncome = '5000000'  # 年收入/万元（系统自动转换为30万元）
            Holder_familyIncome = '5000000'  # 家庭年收入/万元（系统自动转换为30万元）
            Holder_premBudget = '300000'  # 保费预算/元
            Holder_residentType = '1'  # 居民类型
            Holder_taxPayerType = '01'  # 税收居民类型
            Holder_postalCode = '121200'  # 邮编
            Holder_provinceCode = '110000'  # 省份
            Holder_cityCode = '110100'  # 市区
            Holder_districtCode = '110101'  # 地区
            Holder_addrDetail = '建国门东大街22号'  # 详细地址
            #调用原子脚本
            response = API0000_nb().savePersonalHolder(magicX, interface_ip, proposalNo, policyChannel, sameASInsurd, Holder_name, Holder_gender, Holder_Birth, Holder_certiType, Holder_certiCode, Holder_CertiValidEndDate, Holder_isMarried, Holder_height, Holder_weight, Holder_mobilePhone, Holder_officeTel, Holder_homeTel, Holder_industry, Holder_workCompany, Holder_email, Holder_jobCode, Holder_jobClass, Holder_driverLicenseType, Holder_nationality, Holder_nationnality1, Holder_educationId, Holder_medicalInsType, Holder_incomeSource, Holder_incomeSourceNote, Holder_annualIncome, Holder_familyIncome, Holder_premBudget, Holder_residentType, Holder_taxPayerType, Holder_postalCode, Holder_provinceCode, Holder_cityCode, Holder_districtCode, Holder_addrDetail)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']
                customerId = others['customerId']
                message = ''
            else:
                customerId = ''
                message = response['message']
            #断言
            assert is_success == True
            
            """step5:投保人告知接口""" 
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_投保人告知接口'  # 节点
            #dict获取数据+定义数据
            result = '[{"id":"505004","value":"Y"},{"id":"5050045001","value":""}]'   #告知内容
            #调用原子脚本
            response = API0000_nb().Holder_saveNotifyInfo(magicX, interface_ip, proposalId, result, customerId)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step6:被保人信息录入接口"""            
            if (is_HolderExemption == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0') or is_HolderExemption == '1':
                #节点初始化
                is_success = ''
                message = ''
                unit = '系统-新契约录入_被保人信息录入接口'  # 节点
                #dict获取数据+定义数据
                insuRelation = dict['policy_info']['insured_info'][0]['insuRelation']   # 被保人与投保人关系
                insuName = dict['policy_info']['insured_info'][0]['insuName']  # 被保人姓名
                insuGender = dict['policy_info']['insured_info'][0]['insuGender']  # 被保人性别
                insuBirth = dict['policy_info']['insured_info'][0]['insuBirth']  # 被保人出生日期
                insuCertiType = dict['policy_info']['insured_info'][0]['insuCertiType']  # 被保人证件类型
                insuCertiCode = dict['policy_info']['insured_info'][0]['insuCertiCode']  # 被保人证件号码
                insuCertiValidEndDate =dict['policy_info']['insured_info'][0]['insuCertiValidEndDate']  # 被保人证件止期
                insuMobile = dict['policy_info']['insured_info'][0]['insuMobile']  # 被保人手机号
                insuHeight = '178'  # 身高/CM
                insuWeight = '78'  # 体重/KG
                insuMarriage = '20'  # 婚姻状况
                if API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth) < 18:
                    insuWorkCode = '0000001'# 职业代码/名称-儿童、18岁前儿童
                else:
                    insuWorkCode = '0001001'  # 职业代码/名称-内勤
                insujobClass = '1'  # 职业等级
                insuEmail = API0000_diy().random_email()  # 电子邮箱
                insuWorkType = '02'  # 工作行业
                insuCompany = 'apple'  # 工作单位/学校名称
                insuNationatiy = 'CHN'  # 国籍/地区
                insunationnality1 = '01'  # 民族
                insuofficeTel = '010-93456444'  # 固定电话
                insumedicalInsType = '04'  # 医保类型
                insudriverLicenseType = '06'  # 驾照类型
                insuhomeTel = '010-993433344'  # 家庭电话
                insuincomeSource = '01'  # 收入来源
                insuincomeSourceNote = 'wu'  # 收入来源备注
                insuannualIncome = '8000000'  # 年收入/万元
                insuresidentType = '1'  # 居民类型
                insutaxPayerType = '01'  # 税收居民类型
                insueducationId = '40'  # 学位
                insuPostalCode = '120000'  # 邮编
                insuProvinceCode = '110000'  # 省
                insuCityCode = '110100'  # 市
                insuDistrictCode = '110101'  # 区
                insuAddrDetail = '建国门东大街22号'  # 详细地址
                #调用原子脚本
                response = API0000_nb().savePersonalInsured(magicX, interface_ip, proposalId, policyChannel, insuRelation, insuName, insuGender, insuBirth, insuHeight, insuWeight, insuMarriage, insuCertiType, insuCertiCode, insuCertiValidEndDate, insuWorkCode, insujobClass, insuEmail, insuWorkType, insuCompany, insuMobile, insuNationatiy, insunationnality1, insuofficeTel, insumedicalInsType, insudriverLicenseType, insuhomeTel, insuincomeSource, insuincomeSourceNote, insuannualIncome, insuresidentType, insutaxPayerType, insueducationId, insuPostalCode, insuProvinceCode, insuCityCode, insuDistrictCode, insuAddrDetail)

                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True
            
            """step7:被保人查询接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_被保人查询接口'  # 节点
            #调用原子脚本
            response = API0000_nb().queryPersonalInsured(magicX, interface_ip, proposalNo)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']
                message = ''
            else:
                others = ''
                message = response['message']
            # 根据判断是否豁免，确定出被保人insuId、被保人姓名、被保人insucustomerId、豁免投保人insuId、豁免投保人姓名、豁免投保人insucustomerId
            if len(others) == 1:
                insuId = others[0]['insuId']  # 被保人insuId
                insuName = others[0]['insuName']  # 被保人姓名
                insucustomerId = others[0]['insucustomerId']  # 被保人insucustomerId
                insuId1 = ''  # 豁免投保人insuId
                insuName1 = ''  # 豁免投保人姓名
                insucustomerId1 = ''  # 豁免投保人insucustomerId
            elif len(others) == 2:
                # 判断豁免投保人处于others中的第几位
                for i in range(0, 2):
                    if others[i]['insuName'] == dict['policy_info']['holder_info']['name']:
                        j = i  # 得到豁免投保人处于others中的位置
                        # 判断录入的被保人处于others中的第几位
                        if j == 0:
                            k = 1
                        elif j == 1:
                            k = 0;
                insuId = others[k]['insuId']  # 被保人insuId
                insuName = others[k]['insuName']  # 被保人姓名
                insucustomerId = others[k]['insucustomerId']  # 被保人insucustomerId
                insuId1 = others[j]['insuId']  # 豁免投保人insuId
                insuName1 = others[j]['insuName']  # 豁免投保人姓名
                insucustomerId1 = others[j]['insucustomerId']  # 豁免投保人insucustomerId
            assert is_success == True 
                       
            """step8:被保人告知接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_被保人告知接口'  # 节点
            #dict获取数据+定义数据
            result = '[{"id":"505004","value":"Y"},{"id":"5050045001","value":""}]'   #告知内容
            # 1.投保人非豁免、投被保人关系不为同一人时录入;2.投保人豁免时录入
            #调用原子脚本
            if (is_HolderExemption == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0') or is_HolderExemption == '1':
                response = API0000_nb().Insured_saveNotifyInfo(magicX, interface_ip, proposalId, result, insucustomerId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True
            
            """step9:收费方式信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_收费方式信息录入接口'  # 节点
            #dict获取数据+定义数据
            payMode = dict['policy_info']['payment_nb']['payMode']
            name = dict['policy_info']['holder_info']['name']  # 账户名(投保人姓名)
            bankCode = '01'  # 开户银行
            bankAccount = '64556' + str(random.randint(100000000000,999999999999)) # 银行卡号
            reserveMobile = '13556546366'  # 银行预留手机号
            #调用原子脚本
            response = API0000_nb().payModeSave(magicX, interface_ip, proposalId, payMode, name, bankCode, bankAccount, reserveMobile)
            #获取所需响应值
            # 返回是否成功，以及错误信息
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step10:险种信息录入接口"""
            is_success = ''
            message = ''
            unit = '系统-新契约录入_险种信息录入接口-主险1'  # 节点
            is_2risk = dict['public']['is_2risk']  #是否双主险
            ##先录入主险1
                #dict获取数据+定义数据
            productId = dict['policy_info']['main_risk_info'][0]['productId']  # 险种id
            chargeMode = dict['policy_info']['main_risk_info'][0]['chargeMode']  # 缴费频率
            coverPeriodType = dict['policy_info']['main_risk_info'][0]['coverPeriodType']  # 保障期间类型
            coveragePeriod = dict['policy_info']['main_risk_info'][0]['coveragePeriod']  # 保障期间
            chargePeriodType = dict['policy_info']['main_risk_info'][0]['chargePeriodType']  # 缴费期间类型
            chargePeriod = dict['policy_info']['main_risk_info'][0]['chargePeriod']  # 缴费期间
            sumAssured = dict['policy_info']['main_risk_info'][0]['sumAssured']  # 保额
            stdPremBf = dict['policy_info']['main_risk_info'][0]['stdPremBf']  # 保费
            units = dict['policy_info']['main_risk_info'][0]['units']  # 份数
            planFreq = dict['policy_info']['main_risk_info'][0]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
            payOption = dict['policy_info']['main_risk_info'][0]['payOption']  # 年金使用方式
            bonusYearOrAge = dict['policy_info']['main_risk_info'][0]['bonusYearOrAge']  # 领取年期或年龄
            renew = dict['policy_info']['main_risk_info'][0]['renew']  # 是否可续保
            mainItemId = ''
                #调用原子脚本
            response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId, mainItemId, insuId, chargeMode, coverPeriodType, coveragePeriod, chargePeriodType, chargePeriod, units, sumAssured, stdPremBf, planFreq, payOption, bonusYearOrAge, headPartner, organId, renew)
                #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
                others = response['others']
                mainItemId1 = others['itemId']
                sys_stdPremBf = others['stdPremBf']   #保费
                sys_sumAssured = others['sumAssured']   #保额
            else:
                message = response['message']
                mainItemId1 = ''
                sys_stdPremBf = ''
                sys_sumAssured = ''

                #存入dict
            dict['policy_info']['main_risk_info'][0]['stdPremBf'] = sys_stdPremBf  # 保费
            dict['policy_info']['main_risk_info'][0]['sumAssured'] = sys_sumAssured  # 保额
            ##录入主险1-豁免附加险
            if is_HolderExemption == '1':
                is_success = ''
                message = ''
                unit = '系统-新契约录入_险种信息录入接口-主险1_豁免附加险'  # 节点
                #dict获取数据+定义数据
                productId1 = dict['policy_info']['sub_risk_info'][0]['productId']  # 险种id
                chargeMode1 = dict['policy_info']['sub_risk_info'][0]['chargeMode']  # 缴费频率
                coverPeriodType1 = dict['policy_info']['sub_risk_info'][0]['coverPeriodType']  # 保障期间类型
                coveragePeriod1 = dict['policy_info']['sub_risk_info'][0]['coveragePeriod']  # 保障期间
                chargePeriodType1 = dict['policy_info']['sub_risk_info'][0]['chargePeriodType']  # 缴费期间类型
                chargePeriod1 = dict['policy_info']['sub_risk_info'][0]['chargePeriod']  # 缴费期间
                sumAssured1 = dict['policy_info']['sub_risk_info'][0]['sumAssured']  # 保额
                stdPremBf1 = dict['policy_info']['sub_risk_info'][0]['stdPremBf']  # 保费
                units1 = dict['policy_info']['sub_risk_info'][0]['units']  # 份数
                planFreq1 = dict['policy_info']['sub_risk_info'][0]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                payOption1 = dict['policy_info']['sub_risk_info'][0]['payOption']  # 年金使用方式
                bonusYearOrAge1 = dict['policy_info']['sub_risk_info'][0]['bonusYearOrAge']  # 领取年期或年龄
                renew = dict['policy_info']['sub_risk_info'][0]['renew']  # 是否可续保
                #调用原子脚本
                response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew)   
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                    others = response['others']
                    sys_stdPremBf = others['stdPremBf']   #保费
                    sys_sumAssured = others['sumAssured']   #保额
                else:
                    message = response['message']
                    sys_stdPremBf = ''
                    sys_sumAssured = ''

                    #存入dict
                dict['policy_info']['sub_risk_info'][0]['stdPremBf'] = sys_stdPremBf  # 保费
                dict['policy_info']['sub_risk_info'][0]['sumAssured'] = sys_sumAssured  # 保额
            ##判断存在附加险不存在豁免险时
            if is_HolderExemption == '0' and len(dict['policy_info']['sub_risk_info']) > 0:
                is_success = ''
                message = ''
                unit = '系统-新契约录入_险种信息录入接口-非豁免的附加险'  # 节点
                for i in range(0, len(dict['policy_info']['sub_risk_info'])):
                    #dict获取数据+定义数据
                    productId1 = dict['policy_info']['sub_risk_info'][i]['productId']  # 险种id
                    chargeMode1 = dict['policy_info']['sub_risk_info'][i]['chargeMode']  # 缴费频率
                    coverPeriodType1 = dict['policy_info']['sub_risk_info'][i]['coverPeriodType']  # 保障期间类型
                    coveragePeriod1 = dict['policy_info']['sub_risk_info'][i]['coveragePeriod']  # 保障期间
                    chargePeriodType1 = dict['policy_info']['sub_risk_info'][i]['chargePeriodType']  # 缴费期间类型
                    chargePeriod1 = dict['policy_info']['sub_risk_info'][i]['chargePeriod']  # 缴费期间
                    sumAssured1 = dict['policy_info']['sub_risk_info'][i]['sumAssured']  # 保额
                    stdPremBf1 = dict['policy_info']['sub_risk_info'][i]['stdPremBf']  # 保费
                    units1 = dict['policy_info']['sub_risk_info'][i]['units']  # 份数
                    planFreq1 = dict['policy_info']['sub_risk_info'][i]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                    payOption1 = dict['policy_info']['sub_risk_info'][i]['payOption']  # 年金使用方式
                    bonusYearOrAge1 = dict['policy_info']['sub_risk_info'][i]['bonusYearOrAge']  # 领取年期或年龄
                    renew1 = dict['policy_info']['sub_risk_info'][i]['renew']  # 是否可续保
                    # 调用原子化脚本
                    response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew1) 
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        message = ''
                        others = response['others']
                        sys_stdPremBf = others['stdPremBf']   #保费
                        sys_sumAssured = others['sumAssured']   #保额
                    else:
                        message = response['message']
                        sys_stdPremBf = ''
                        sys_sumAssured = ''

                        #存入dict
                    dict['policy_info']['sub_risk_info'][i]['stdPremBf'] = sys_stdPremBf  # 保费
                    dict['policy_info']['sub_risk_info'][i]['sumAssured'] = sys_sumAssured  # 保额
            ##判断存在附加险也存在豁免险时
            elif is_HolderExemption == '1' and len(dict['policy_info']['sub_risk_info']) > 1:
                is_success = ''
                message = ''
                unit = '系统-新契约录入_险种信息录入接口-非豁免的附加险'  # 节点
                for i in range(1, len(dict['policy_info']['sub_risk_info'])):
                    productId1 = dict['policy_info']['sub_risk_info'][i]['productId']  # 险种id
                    chargeMode1 = dict['policy_info']['sub_risk_info'][i]['chargeMode']  # 缴费频率
                    coverPeriodType1 = dict['policy_info']['sub_risk_info'][i]['coverPeriodType']  # 保障期间类型
                    coveragePeriod1 = dict['policy_info']['sub_risk_info'][i]['coveragePeriod']  # 保障期间
                    chargePeriodType1 = dict['policy_info']['sub_risk_info'][i]['chargePeriodType']  # 缴费期间类型
                    chargePeriod1 = dict['policy_info']['sub_risk_info'][i]['chargePeriod']  # 缴费期间
                    sumAssured1 = dict['policy_info']['sub_risk_info'][i]['sumAssured']  # 保额
                    stdPremBf1 = dict['policy_info']['sub_risk_info'][i]['stdPremBf']  # 保费
                    units1 = dict['policy_info']['sub_risk_info'][i]['units']  # 份数
                    planFreq1 = dict['policy_info']['sub_risk_info'][i]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                    payOption1 = dict['policy_info']['sub_risk_info'][i]['payOption']  # 年金使用方式
                    bonusYearOrAge1 = dict['policy_info']['sub_risk_info'][i]['bonusYearOrAge']  # 领取年期或年龄
                    renew1 = dict['policy_info']['sub_risk_info'][i]['renew']  # 是否可续保
                    # 调用原子化脚本
                    response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew1) 
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        message = ''
                        others = response['others']
                        sys_stdPremBf = others['stdPremBf']   #保费
                        sys_sumAssured = others['sumAssured']   #保额
                    else:
                        message = response['message']
                        sys_stdPremBf = ''
                        sys_sumAssured = ''

                        #存入dict
                    dict['policy_info']['sub_risk_info'][i]['stdPremBf'] = sys_stdPremBf  # 保费
                    dict['policy_info']['sub_risk_info'][i]['sumAssured'] = sys_sumAssured  # 保额
            ##存在双主险时-录入第二个主险 
            if is_2risk == '1':
                is_success = ''
                message = ''
                unit = '系统-新契约录入_险种信息录入接口-录入第二个主险'  # 节点
                productId2 = dict['policy_info']['main_risk_info'][-1]['productId']  # 险种id
                chargeMode2 = dict['policy_info']['main_risk_info'][-1]['chargeMode']  # 缴费频率
                coverPeriodType2 = dict['policy_info']['main_risk_info'][-1]['coverPeriodType']  # 保障期间类型
                coveragePeriod2 = dict['policy_info']['main_risk_info'][-1]['coveragePeriod']  # 保障期间
                chargePeriodType2 = dict['policy_info']['main_risk_info'][-1]['chargePeriodType']  # 缴费期间类型
                chargePeriod2 = dict['policy_info']['main_risk_info'][-1]['chargePeriod']  # 缴费期间
                sumAssured2 = dict['policy_info']['main_risk_info'][-1]['sumAssured']  # 保额
                stdPremBf2 = dict['policy_info']['main_risk_info'][-1]['stdPremBf']  # 保费
                units2 = dict['policy_info']['main_risk_info'][-1]['units']  # 份数
                planFreq2 = dict['policy_info']['main_risk_info'][-1]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                payOption2 = dict['policy_info']['main_risk_info'][-1]['payOption']  # 年金使用方式
                bonusYearOrAge2 = dict['policy_info']['main_risk_info'][-1]['bonusYearOrAge']  # 领取年期或年龄
                renew2 = dict['policy_info']['main_risk_info'][-1]['renew']  # 是否可续保
                mainItemId2 = ''  # 主险标识
                # 调用原子化脚本      
                response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId2, mainItemId2, insuId, chargeMode2, coverPeriodType2, coveragePeriod2, chargePeriodType2, chargePeriod2, units2, sumAssured2, stdPremBf2, planFreq2, payOption2, bonusYearOrAge2, headPartner, organId, renew2) 
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                    others = response['others']
                    sys_stdPremBf = others['stdPremBf']   #保费
                    sys_sumAssured = others['sumAssured']   #保额
                else:
                    message = response['message']
                    sys_stdPremBf = ''
                    sys_sumAssured = ''

                    #存入dict
                dict['policy_info']['main_risk_info'][-1]['stdPremBf'] = sys_stdPremBf  # 保费
                dict['policy_info']['main_risk_info'][-1]['sumAssured'] = sys_sumAssured  # 保额
            assert is_success == True

            """step11:新契约录入提交前规则校验接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_提交前规则校验接口'  # 节点
            #调用原子脚本
            response = API0000_nb().valiDateRule(magicX, interface_ip, proposalId)
            #获取所需响应值
            if response['success'] == True:
                is_success = True
                message = ''
            else:
                is_success = False
                message = response['message']
            #断言
            assert is_success == True  
            
            """step12:新契约录入提交接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_提交接口'  # 节点
            #调用原子脚本
            response = API0000_nb().reviewProposal(magicX, interface_ip, proposalId)
            #获取所需响应值
            if (response['success'] == True) and (response['message'] == ''):
                is_success = True
                others = ''
                message = ''
            else:
                is_success = False
                others = response['others']
                if others != None:
                    for i in range(0, len(others)):
                        if i == 0:
                            message = response['message'] + ':1.' + others[i]['issueContent']
                        else:
                            message = message + str(i+1) + '.' + others[i]['issueContent']
            #断言
            assert is_success == True       
            
            """step13:登出系统-新契约录入用户"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-新契约录入_登出系统'  # 节点
            #调用原子脚本
            rs = API0000_sys_otherapi().logout(magicX, interface_ip)
            #获取所需响应值
            magicX = ''
            is_success = rs[0]
            message = rs[1]
            assert is_success == True
            dict['public']['magicX'] = ''

            """存入dict"""
                #保单信息
            dict['policy_info']['proposalNo'] = proposalNo  # 投保单号
            dict['policy_info']['policyType'] = policyType  # 保单号类型
            dict['policy_info']['effectiveDate'] = effectiveDate  # 生效日期
            dict['policy_info']['organId'] = organId  # 所属机构
            dict['policy_info']['policyChannel'] = policyChannel  # 公司渠道
                #投保人信息
            dict['policy_info']['holder_info']['isMarried'] = Holder_isMarried  # 婚姻状况
            dict['policy_info']['holder_info']['height'] = Holder_height  # 身高/CM
            dict['policy_info']['holder_info']['weight'] = Holder_weight  # 体重 /KG
            dict['policy_info']['holder_info']['officeTel'] = Holder_officeTel  # 固定电话
            dict['policy_info']['holder_info']['homeTel'] = Holder_homeTel  # 家庭电话
            dict['policy_info']['holder_info']['industry'] = Holder_industry  # 工作行业
            dict['policy_info']['holder_info']['workCompany'] = Holder_workCompany  # 工作单位/学校名称
            dict['policy_info']['holder_info']['email'] = Holder_email  # 电子邮箱
            dict['policy_info']['holder_info']['jobCode'] = Holder_jobCode  # 职业代码/名称
            dict['policy_info']['holder_info']['jobClass'] = Holder_jobClass  # 职业等级
            dict['policy_info']['holder_info']['driverLicenseType'] = Holder_driverLicenseType  # 驾照类型
            dict['policy_info']['holder_info']['nationality'] = Holder_nationality  # 国籍/地区
            dict['policy_info']['holder_info']['nationnality1'] = Holder_nationnality1  # 民族
            dict['policy_info']['holder_info']['educationId'] = Holder_educationId  # 学位
            dict['policy_info']['holder_info']['medicalInsType'] = Holder_medicalInsType  # 医保类型
            dict['policy_info']['holder_info']['incomeSource'] = Holder_incomeSource  # 收入来源
            dict['policy_info']['holder_info']['incomeSourceNote'] = Holder_incomeSourceNote  # 收入来源备注
            dict['policy_info']['holder_info']['annualIncome'] = Holder_annualIncome  # 年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['familyIncome'] = Holder_familyIncome  # 家庭年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['premBudget'] = Holder_premBudget  # 保费预算/元
            dict['policy_info']['holder_info']['residentType'] = Holder_residentType  # 居民类型
            dict['policy_info']['holder_info']['taxPayerType'] = Holder_taxPayerType  # 税收居民类型
            dict['policy_info']['holder_info']['postalCode'] = Holder_postalCode  # 邮编
            dict['policy_info']['holder_info']['provinceCode'] = Holder_provinceCode  # 省份
            dict['policy_info']['holder_info']['cityCode'] = Holder_cityCode  # 市区
            dict['policy_info']['holder_info']['districtCode'] = Holder_districtCode  # 地区
            dict['policy_info']['holder_info']['addrDetail'] = Holder_addrDetail   # 详细地址
                #被保人信息
                #①非投保人豁免，投被保人不为同一人时
            if dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = insuofficeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                #②非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = Holder_officeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
                #③非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '1':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['officeTel'] = insuofficeTel   # 固定电话
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId1   #客户id
                dict['policy_info']['insured_info'][1]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][1]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][1]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][1]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][1]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][1]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][1]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][1]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][1]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][1]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][1]['insuofficeTel'] = Holder_officeTel   # 固定电话
                dict['policy_info']['insured_info'][1]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][1]['insudriverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][1]['insuhomeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][1]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][1]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][1]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][1]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][1]['insutaxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][1]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][1]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][1]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][1]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][1]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][1]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
            dict['logInfo']['code'] = '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约录入_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            dict['logInfo']['err'] = unit + ':' + str(e)
            if message == '':
                message = str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return proposalId,dict


    """新契约复核"""
    @func_set_timeout(60)#设定函数超执行时间
    def nb_review(self,proposalId, dict):
        """
           step1:复核用户登录
           step2:新契约复核updatePropStatus接口
           step3:新契约复核提交接口
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约复核_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            magicX = ''   #初始化会话id
            """step1:登录系统-新契约复核用户"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '系统-新契约复核_登录系统'  # 节点
            message = ''   #信息
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   #环境名称
            interface_ip = dict['public']['interface_ip']   #接口ip
            username = dict['public']['username_review']    #新契约复核用户
            #调用原子脚本
            rs = API0000_sys_otherapi().login(env_name,username, interface_ip)
            #获取所需响应值
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            magicX = rs[2]   #会话id
            #断言
            assert is_success == True
            dict['public']['magicX'] = magicX

            """step2:登录系统-新契约复核updatePropStatus接口"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '系统-新契约复核_新契约复核updatePropStatus接口'  # 节点
            message = ''   #信息
            #调用原子脚本
            response = API0000_nb().updatePropStatus(magicX, interface_ip, proposalId)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step3:登录系统-新契约复核提交接口"""
            #节点初始化
            is_success = ''
            unit = '系统-新契约复核_新契约复核提交接口'  # 节点
            message = ''
            #dict获取数据+定义数据
            personUW = 'false'   #是否强制人工核保
            #调用原子脚本
            response = API0000_nb().confirmProposal(magicX, interface_ip, proposalId, personUW)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']   # others为1时自动核保通过；为2时人工核保
                message = ''
            else:
                others = ''
                message = response['message']
            #断言
            assert is_success == True 
            dict['logInfo']['code'] = '1' 
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约复核_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0' 
            dict['logInfo']['err'] = unit + ':' + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return magicX, others, dict


    """新契约核保"""
    @func_set_timeout(30)#设定函数超执行时间
    def nb_uw(self, magicX, others, dict):
        """
           step1:判断自动核保或人工核保
           step2:新契约人工核保_核保共享池查询接口
           step3:新契约人工核保_锁定任务接口
           step4:新契约人工核保_获取保单信息接口
           step5:新契约人工核保_险种层提交核保决定接口
           step6:新契约人工核保_契约核保结论提交接口
        """
        
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约核保_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            """step1:判断自动核保或人工核保"""
            if others == '1':  #新契约自核通过
                pass
            elif others == '2':   #进入新契约人工核保
                """step2:新契约人工核保_核保共享池查询接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约人工核保_核保共享池查询接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                interface_ip = dict['public']['interface_ip']   #接口ip
                policyNo = dict['policy_info']['policyNo']  # 保单号
                #调用原子脚本
                response = API0000_uw().uwPolicyQuery_nb(magicX, interface_ip, policyNo)
                #获取所需响应值
                is_success = response['isSuccess']
                if is_success:
                    resultData = response['resultData']
                    uwId = resultData[0]['uwId']
                    mainId = resultData[0]['mainId']
                    message = ''
                else:
                    uwId = ''
                    mainId = ''
                    message = ''
                    message = response['message']
                #断言
                assert is_success == True
                
                """step3:新契约人工核保_锁定任务接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约人工核保_锁定任务接口'  # 节点
                message = ''
                #调用原子脚本
                response = API0000_uw().uwLock_nb(magicX, interface_ip, mainId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True

                """step4:新契约人工核保_获取保单信息接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约人工核保_获取保单信息接口'  # 节点
                message = ''
                #调用原子脚本
                response = API0000_uw().uw_queryPersonalPolicy(magicX, interface_ip, uwId, mainId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    others = response['others']
                    uwProducts = others['uwProducts']   # 核保险种信息
                    message = ''
                else:
                    uwProducts = ''
                    message = response['message']
                #断言
                assert is_success == True
                
                """step5:新契约人工核保_险种层提交核保决定接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约人工核保_险种层提交核保决定接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                uwDecision = '01'  # 核保决定 01-核保通过
                decisionComment = '核保通过'  # 核保决定描述
                uwExceptTxt = ''  # 除外责任

                num = 0
                lenght = len(uwProducts)   # 获取险种个数
                for a in range(0, lenght):
                    uwItemId = uwProducts[a]['uwItemId']
                    #调用原子脚本
                    response = API0000_uw().updateUwDecision_nb(magicX, interface_ip, uwItemId, uwId, uwDecision, decisionComment, uwExceptTxt, policyNo)
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        num = num + 1
                if num == lenght:
                    is_success = True
                    message = ''
                else:
                    is_success = False
                    message = '险种层提交核保决定未成功'
                #断言
                assert is_success == True

                """step6:新契约人工核保_契约核保结论提交接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约人工核保_契约核保结论提交接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                uwPolicyDecision = '01'  # 保单核保结论 01-标准通过
                #调用原子脚本
                response = API0000_uw().commitUwComplete(magicX, interface_ip, uwId, uwPolicyDecision)
                #获取所需响应值
                status = response['status']
                if status == '0':
                    is_success = True
                    message = ''
                else:
                    is_success = False
                    message = response['message']
                #断言
                assert is_success == True
                dict['logInfo']['code'] == '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约核保_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + ':' + str(e)
            if message == '':
                message = str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict


    """新契约收费"""
    @func_set_timeout(60)#设定函数超执行时间
    def nb_payment(self, magicX, dict):
        """
           step1:柜面收付费查询接口
           step2:柜面收付费保存接口
           step3:柜面收付费审核接口
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约收费_开始(预计耗时:30s)'))
            """step1:柜面收付费查询接口"""
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''   #初始化是否成功标识
            unit = '系统-新契约收费_柜面收付费查询接口'  # 节点
            message = ''   #初始化信息
            payment = 0   #初始化收付费金额 
            #dict获取数据+定义数据
            interface_ip = dict['public']['interface_ip']   #接口ip
            policyNo = dict['policy_info']['policyNo']  # 保单号

            for i in range(0, 300):  #收付费查询(每0.5秒查询一次，最大200次，直到查到结果便结束查询)
                #调用原子脚本
                rs = API0000_payment().FeeQuery(magicX, interface_ip, policyNo)
                is_success = rs[0]
                message = rs[1]
                resultData = rs[2]
                len_resultData = rs[3]
                if len_resultData > 0:
                    break
                else:
                    time.sleep(1)
            #存入dict
            payment = float(resultData[0]['amount'])   #收付费金额
            dict['policy_info']['payment_nb']['payment'] = payment  # 新契约收费金额                    
            assert is_success == True
            """step2:柜面收付费保存接口"""    
            #节点初始化
            is_success = ''
            unit = '系统-新契约收费_柜面收付费保存接口'  # 节点
            message = ''
            #dict获取数据+定义数据
            bankEndorse = '123123'  # 银行凭证号码
            feeConfirmDate = dict['policy_info']['applicationDate']  # 费用确认日期
            internalAccount = '08-400301040007336'  # 公司收费账号
            internalBankCode = '05'  # 收费机构
            feeChannel = 'CASH-10017-AR'  # 支付通道
            cashOrgId = '10017'  # 机构名称
            internalBankName = '中国农业银行绥化市大有支行'  # 所属银行
            #调用原子脚本
            rs = API0000_payment().CashApproveInfoSave(magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, cashOrgId, internalBankName)
            #获取所需响应值
            is_success = rs[0]
            message = rs[1]
            assert is_success == True

            """step3:柜面收付费审核接口"""    
            #节点初始化
            is_success = ''
            unit = '系统-新契约收费_柜面收付费审核接口'  # 节点
            message = ''
            #调用原子脚本
            rs = API0000_payment().finishFeeFlow(magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, internalBankName)
            #获取所需响应值
            is_success = rs[0]
            message = rs[1]
            assert is_success == True 
            dict['logInfo']['code'] == '1' 
            end_time = datetime.datetime.now()   #结束时间 
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约收费_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))   
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + ':' + str(e)
            if message == '':
                message = str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict
            

    """新契约回执"""
    @func_set_timeout(60)#设定函数超执行时间
    def nb_receipt(self, magicX, dict):
        """
           step1:保单确认回执查询接口
           step2:保单确认回执接口
           step3:登出系统-新契约复核用户
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约回执_开始(预计耗时:30s)'))
           #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''    #初始化是否成功标识
            unit = '系统-新契约回执_保单确认回执'  # 节点
            message = ''    #初始化信息
            #dict获取数据+定义数据
            interface_ip = dict['public']['interface_ip']  # 接口ip
            policyNo = dict['policy_info']['policyNo']  # 保单号
            callBackDate = dict['policy_info']['applicationDate']  # 回执日期
            if dict['policy_info']['main_risk_info'][0]['productId'] not in ('10120','10121','10122'):
                """step1:保单确认回执查询接口"""
                #节点初始化
                is_success = ''    #初始化是否成功标识
                unit = '系统-新契约回执_保单确认回执查询接口'  # 节点
                message = ''    #初始化信息
                for i in range(0, 200):
                    #调用原子脚本
                    response = API0000_nb().query_policyReceiptConfirm(magicX, interface_ip, policyNo)
                    #获取所需响应值
                    is_success = response['isSuccess']
                    if is_success:
                        message = ''
                        resultData = response['resultData']
                    else:
                        message = response['message']
                        resultData = ''
                    #重复发送接口请求
                    if resultData != '' and resultData != None:
                        break
                    else:
                        time.sleep(0.5)
                assert is_success == True 
                
                """step2:保单确认回执接口"""
                #节点初始化
                is_success = ''
                unit = '系统-新契约回执_保单确认回执接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                proposalNo = dict['policy_info']['proposalNo']  # 投保单号
                callBackDate = dict['policy_info']['applicationDate']  # 回执日期
                callBackWay = '01'  # 回执方式
                appointVisitStartTime = ''  # 预约回访起始日期
                appointVisitEndTime = ''  # 预约回访终止日期
                #调用原子脚本
                response = API0000_nb().policyReceiptConfirm(magicX, interface_ip, policyNo, proposalNo, callBackDate, callBackWay, appointVisitStartTime, appointVisitEndTime)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True    
            else:
                time.sleep(1)
            """step2:登出系统-新契约复核用户"""
            #节点初始化
            is_success = ''
            unit = '系统-新契约复核_登出系统'  # 节点
            message = ''
            #调用原子脚本
            rs = API0000_sys_otherapi().logout(magicX, interface_ip)
            #获取所需响应值
            magicX = ''
            is_success = rs[0]
            message = rs[1]
            assert is_success == True
            dict['public']['magicX'] = ''
            #存入dict:
            dict['policy_info']['policy_status'] = '01'   #保单状态。01有效；02失效
            dict['policy_info']['callBackDate'] = callBackDate   #回执日期
            dict['logInfo']['code'] == '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约回执_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + ':' + str(e)
            if message == '':
                message = str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict
            

    """定义轨迹信息+检查点信息"""
    @func_set_timeout(30)#设定函数超执行时间
    def save_check_info(self, dict):
        """
           step1:定义轨迹信息
           step2:定义检查点信息
        """
        try:
            #初始化
            unit = '定义轨迹信息+检查点信息'    #节点标识
            #获取dict数据
            """step1:定义检查点信息"""
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   # 环境名称
            #dict获取数据+定义数据
            policyNo = dict['policy_info']['policyNo']  # 保单号
            applicationDate = dict['policy_info']['applicationDate']
            #获取产品列表
            product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #保单产品列表
            product = []
            if dict['logInfo']['code'] == '1':
                #产品信息product_info追加到dict['check']['calc_nb']['product']中
                for i in range (0,len(product_list)):
                    productId = str(product_list[i]['productId'])   #产品id
                    productCode = API0000_diy().query_productCode(env_name, productId)   #获取产品code
                    premium = float(product_list[i]['stdPremBf'])   #获取产品保费
                    amount = float(product_list[i]['sumAssured'])   #获取产品保额
                    #产品列表中每个产品生成一个产品信息product_info
                    product_info = {'productCode':productCode,'premium':premium,'amount':amount}
                    #存入dict['check']['calc_nb']中
                    product.append(product_info)
                dict['check']['calc_nb']['product'] = product
                #新契约收费金额存入dict['check']['calc_nb']中
                dict['check']['calc_nb']['payment'] = float(dict['policy_info']['payment_nb']['payment'])   #新契约收费金额
                dict['check']['calc_nb']['apply_date'] = applicationDate   #投保日期
                dict['check']['calc_nb']['msg'] = ''   #投保日期
                """step2:定义轨迹信息"""
                #定义轨迹信息
                track_info = {'trackType':'nb', 'trackTime':applicationDate, 'trackData':{'product':product, 'payment':float(dict['policy_info']['payment_nb']['payment'])}, 'msg': ''}
                dict['track_info'].append(track_info)   #存入轨迹信息
            else:
                message = dict['logInfo']['msg']
                err = dict['logInfo']['err']
                if dict['logInfo']['msg'] == '' or dict['logInfo']['msg'] == None:
                    message = str(dict['logInfo']['err'])
                dict['check']['calc_nb']['product'] = ''
                #新契约收费金额存入dict['check']['calc_nb']中
                dict['check']['calc_nb']['payment'] = dict['policy_info']['payment_nb']['payment']   #新契约收费金额
                dict['check']['calc_nb']['apply_date'] = applicationDate   #投保日期
                dict['check']['calc_nb']['msg'] = message 
                #定义轨迹信息
                track_info = {'trackType':'nb', 'trackTime':applicationDate, 'trackData':{'product':'', 'payment':dict['policy_info']['payment_nb']['payment']}, 'msg': message}
                dict['track_info'].append(track_info)   #存入轨迹信息
        except Exception, e:
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*定义轨迹信息+检查点信息部分脚本报错:\n')  + traceback.format_exc())
        finally:
            pass 
        return dict



        
if __name__ == "__main__":   
    try:
        begin_time = datetime.datetime.now()   #开始时间
        #1.初始化
        is_run = True   #是否执行python脚本 
        msg = ''   #执行python脚本前存在的错误信息
        job = 'nb'   #场景名
        #2.本地定义/服务端获取-脚本执行数据：(1)win系统（本地）:自定义数据(2)服务器端:接收并校验参数数据（第一位为测试编号，第二位json串）
        if 'win' in sys.platform:   #(1)win系统（本地）
            execNo = 'test001'   #测试编号
            env_name = '预生产'
            env_name = env_name.encode('utf-8')
            applicationDate = '2018-02-01'   #投保日期
            username_entry = 'xcy1'   #保全录入用户名
            username_review = 'xcy2'   #保全复核用户名
            filename = r'D:\xLibrary\chunyu\doc\product_life_cycle\product_life_cycle_data.xlsx'   #excel测试数据路径 
        else:   #(2)服务器端:接收并校验参数数据（第一位为测试编号，第二位json串）
            if len(sys.argv) != 3:   #判读后台出参数量（3位：脚本路径+测试编号+json串）
                is_run = False
                msg = '后台传值数有误！'
                execNo = ''
                env_name = ''
                applicationDate = ''   
                username_entry = ''  
                username_review = '' 
                filename = ''
            elif len(sys.argv) == 3:
                execNo = sys.argv[1]   #测试编号
                input_dict = sys.argv[2]   #输入的json串
                input_dict = input_dict.replace("'", '"')   #json转dict，需要把‘单引号’转换为‘双引号’
                #后台传值json串中字段列表
                key_list1 = ['env_name','applicationDate','username_entry','username_review','filename']
                #checkdata界面录入字段-后台传值json-脚本，字段是否存在及读取字段值。不存在时为''
                rs = API0000_diy().checkdata_json_check(input_dict,key_list1)
                is_run = rs[0] 
                msg = rs[1]
                env_name = rs[2][0]   #环境类型
                applicationDate = rs[2][1]   #保全申请日期
                username_entry = rs[2][2]   #保全录入用户
                username_review = rs[2][3]   #保全复核用户
                filename = rs[2][4]
       #3.校验录入字段是否符合规范
        key_dict = {'env_name':env_name,'applicationDate':applicationDate}   #录入字段中待校验字典
        if msg == '':
            rs = API0000_diy().checkdata_json_check1(key_dict)
            is_run = rs[0]
            msg = rs[1]
        #4.定义dict
        if is_run == True:
            msg = ''
        dict = API0000_diy().define_dict() 
        dict['logInfo']['job'] = job
        dict['logInfo']['msg'] = msg
        dict['public']['env_name'] =  env_name    #环境类型
        dict['public']['filename'] =  filename    #excel测试数据路径
        dict['public']['username_entry'] = username_entry
        dict['public']['username_review'] = username_review
        dict['public']['interface_ip'] = API0000_diy().ip_conf(env_name)   #环境ip
        #5.1定义logging文件路径
        if 'win' in sys.platform:
            logging_filename = 'D:\\xLibrary\\chunyu\\%s.html' % str(execNo)
            #判断windows是否存在某文件，存在则删除文件
            if os.path.exists(logging_filename):
                os.remove(logging_filename)
        else:
            logging_filename = '/data/xServer/xReport/%s.html' % str(execNo) 
        #5.2定义logging格式
        logging.basicConfig(level=logging.WARNING , format='%(message)s  &nbsp&nbsp&nbsp&nbsp%(asctime)s <br/><br/>', filename=logging_filename, filemode='a')
        #6.执行脚本
        if is_run == True:
            dict = sys_API0001().nb(env_name, applicationDate, dict)
    except Exception, e:   #常见异常的捕捉
        dict['logInfo']['code'] = '0'
        #记录异常日志
        logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*异常位置:\n')  + traceback.format_exc())
        #异常场景-登出系统（防止与用户被挂起）
        if dict['public']['magicX'] != '':
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_review'])
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
            dict['public']['magicX'] = ''
    except func_timeout.exceptions.FunctionTimedOut:    #函数超时异常的捕捉
        #记录执行时间超时日志
        logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*执行时间超时退出'))
        dict['logInfo']['code'] = '0'
        dict['logInfo']['msg'] = '执行时间超时退出。' + dict['logInfo']['msg']
        #异常场景-登出系统（防止与用户被挂起）
        if dict['public']['magicX'] != '':
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_entry'])
            API0000_sys_otherapi().deblocking(dict['public']['magicX'], dict['public']['interface_ip'], dict['public']['username_review'])
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*运行异常-用户解锁'))
            dict['public']['magicX'] = ''
    finally:
        try:
            end_time = datetime.datetime.now()   #记录结束时间
            #打印终端开始符
            print '\n\n' + '#'*3 + 'RESULT' + '#'*3
            if 'win' in sys.platform:   #windows系统平台
                if dict['logInfo']['code'] == '1':
                    print '是否出单成功：成功'.decode('utf-8').encode('gb2312')
                    print ('环境：' + env_name).decode('utf-8').encode('gb2312')
                    print ('投保日期：' + applicationDate).decode('utf-8').encode('gb2312')
                    print ('保单号：' + dict['policy_info']['policyNo']).decode('utf-8').encode('gb2312')
                    print ('是否修改时间重启应用：' + dict['track_change'][-1]['trackData']['is_Modify_flag'] + '(耗时:' + dict['track_change'][-1]['trackData']['consume_time'] + ')').decode('utf-8').encode('gb2312')
                    print ('新契约出单耗时：' + API0000_diy().logger_count_time1(API0000_diy().logger_count_time(begin_time, end_time), dict['track_change'][-1]['trackData']['consume_time'])).decode('utf-8').encode('gbk')
                    print ('总耗时：' + API0000_diy().logger_count_time(begin_time, end_time)).decode('utf-8').encode('gbk')
                else:
                    print '是否出单成功：不成功'.decode('utf-8').encode('gb2312')
                    print '不成功原因：'.decode('utf-8').encode('gb2312'),;API0000_diy().solve_terminal_garbled(dict['logInfo']['msg'])
                    print ('环境：' + env_name).decode('utf-8').encode('gb2312')
                    print ('投保日期：' + applicationDate).decode('utf-8').encode('gb2312')
                    print ('保单号：' + dict['policy_info']['policyNo']).decode('utf-8').encode('gb2312')
                    print ('耗时：' + API0000_diy().logger_count_time(begin_time, end_time)).decode('utf-8').encode('gbk')
            else:    #非windows系统平台
                if dict['logInfo']['code'] == '1':
                    result_str = "是否出单成功：成功,,环境：%s,,投保日期：%s,,保单号：%s,,是否修改时间重启应用：%s(耗时：%s):,,新契约出单耗时：%s,,总耗时：%s" % (env_name, applicationDate, dict['policy_info']['policyNo'], dict['track_change'][-1]['trackData']['is_Modify_flag'], dict['track_change'][-1]['trackData']['consume_time'], API0000_diy().logger_count_time1(API0000_diy().logger_count_time(begin_time, end_time), dict['track_change'][-1]['trackData']['consume_time']), API0000_diy().logger_count_time(begin_time, end_time))
                else:
                    result_str = "是否出单成功：不成功,,不成功原因：%s,,环境：%s,,投保日期：%s,,保单号：%s,,总耗时：%s,,dict:%s" % (dict['logInfo']['msg'], env_name, applicationDate, dict['policy_info']['policyNo'], API0000_diy().logger_count_time(begin_time, end_time), json.dumps(dict))
                print result_str
            #打印终端结束符
            print '#'*3 + 'RESULT' + '#'*3
        except Exception, e:
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*组件结果展示部分脚本报错:\n')  + traceback.format_exc())







