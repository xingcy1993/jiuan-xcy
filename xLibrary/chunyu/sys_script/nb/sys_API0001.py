#coding=utf-8

import os
import sys
import time
import json
import random
import logging
import datetime
import traceback
from openpyxl import load_workbook
from func_timeout import func_set_timeout
import func_timeout
reload(sys)
sys.setdefaultencoding("utf8")
from xLibrary.chunyu.API0000 import *
from xLibrary.chunyu.sys_script.Modify_servertime.sys_API0002_1 import sys_API0002


class sys_API0001:

    def __init__(self):
        pass

    """存储新契约数据+验证契约录入用户、复核用户是否可正常登录系统"""
    @func_set_timeout(30)#设定函数超执行时间
    def store_nbdata(self, env_name, dict):
        """
           step1:获取excel值+存入dict中
           step2:验证契约录入用户、复核用户是否可正常登录系统
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*出单前准备：数据存储+用户验证_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''
            message = ''
            magicX = ''
            unit = ''
            # 获取dict值
            env_name = dict['public']['env_name']
            interface_ip = dict['public']['interface_ip']
            """step1:验证录入的代理人是否有误"""
            #节点初始化
            is_success = ''
            message = ''
            #连接数据库
            cursor = API0000_diy().db_conf(env_name,'wift_iiws')   #连接数据库
            unit = '系统-验证录入的代理人/合作网点信息是否有误'   #节点标识
            #获取dict值
            agent = dict['public']['agent']   #录入代理人信息
            agent_code = agent.split('-')[0]   #代理人code
            agent_name = agent.split('-')[1]   #代理人姓名
            #代理人查询出“所属机构”、“公司渠道”、“销售组织”、“销售渠道”
            sql = "SELECT b.ORGAN_ID,b.policy_channel,a.CHANNEL_id,d.name,b.SALES_CHANNEL,a.AGENT_ID FROM  t_agent a,t_chl_org_insurance b,d_sales_channel d WHERE a.CHANNEL_id=b.CHANNEL_id AND b.SALES_CHANNEL=d.code AND a.agent_code='" + agent_code + "' AND a.agent_name='" + agent_name + "'"
            cursor.execute(sql)
            result = cursor.fetchall()
            if len(result) == 0:
                is_success = False
                message = '录入的代理人信息，数据库中不存在！'
            else:
                policyChannel = result[0][1]   #公司渠道
                if env_name in ('uat7','uat8','预生产') and policyChannel in ('06','09'):
                    if policyChannel == '06':
                        channelType = '1'
                    elif policyChannel == '09':
                        channelType = '2'
                    #调用业管销控查询接口
                    response = API0000_sys_otherapi().findByBroker(agent_code, agent_name, channelType)
                    if response['t']['resultCode'] == '10':
                        is_success = True
                        message = ''
                    else:
                        is_success = False
                        message = '请检查业管销控查询接口：http://10.8.1.156/gateway/micro-serv-broker-app-uat/openProdSalesCfg/findByBroker是否可正常申请！'
                else:
                    is_success = True
                    message = ''
                assert is_success == True
                unit = '系统-验证录入的代理人/合作网点信息是否有误'
                if is_success == True:
                    #获取数据库查询结果
                    organId = result[0][0]   #所属机构
                    policyChannel = result[0][1]   #公司渠道
                    channelId = result[0][2]   #销售组织
                    salesChannelText = result[0][4]   #销售渠道名
                    salesChannel = result[0][4]   #销售渠道id
                    agentId = result[0][5]   #代理人
                    #合作网点查询出合作机构、提交渠道
                    partner = dict['public']['partner']   #录入代理人信息
                    partner_code = partner.split('-')[0]   #合作机构code
                    partner_name = partner.split('-')[1]   #合作机构名称
                    sql1 = "select if(parent_id is NULL,partner_id,parent_id),PARTNER_ID,SUBMIT_CHANNEL from t_chl_org_partner where partner_code='" + partner_code + "' AND partner_name='" + partner_name + "'"
                    cursor.execute(sql1)
                    result1 = cursor.fetchall()
                    if len(result1) == 0:
                        headPartner = ''
                        partnerId = ''
                        submitChannel = ''
                        is_success = False
                        message = '录入的合作网点信息，数据库不存在'
                    else:
                        is_success = True
                        message = ''
                        headPartner = result1[0][0]   #合作机构
                        partnerId = result1[0][1]   #合作网点
                        submitChannel = result1[0][2]   #提交渠道
            assert is_success == True

            """step2:验证契约录入用户、复核用户是否可正常登录系统"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '系统-验证契约录入用户、复核用户是否可正常登录系统'   #节点标识
            #登录用户
            rs = API0000_diy().verify_userlogin(env_name, dict['public']['username_entry'], dict['public']['interface_ip'])
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            # 加入断言
            assert is_success == True
            #登出用户
            rs = API0000_diy().verify_userlogin(env_name, dict['public']['username_review'], dict['public']['interface_ip'])
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            # 加入断言
            assert is_success == True


            """step3:获取excel值存入dict中"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：读取excel值存入dict中'  # 节点
            #根据脚本执行平台获取excel文件路径（win系统为自测，linux为自动化网页平台）
            filename = dict['public']['filename']   #dict中获取
            #获取excel文件对象
            wb = load_workbook(filename)
            #获取sheet页对象
            sh = wb['nb_param']
            #区分“非法定受益人”与“法定受益人”的excel内单元格获取位置
            if dict['nb_type'] == 'sys_API0001.4':    #受益人非法定
                count = 13   #相差9个位置（受影响字段为：“缴费方式”~“是否可续保”字段）
            else:
                count = 0
            # 读取公共信息
            dict['policy_info']['payment_nb']['payMode'] = str(sh['B' + str(28+count)].value)  # 缴费方式
            dict['policy_info']['payment_nb']['bankCode'] = str(sh['B' + str(29+count)].value)  # 开户银行
            dict['public']['is_2risk'] = sh['B' + str(30+count)].value  #是否双主险
            dict['public']['is_HolderExemption'] = sh['B7'].value  # 是否投保人豁免
            # 保单基本信息
            dict['policy_info']['organId'] = organId   #所属机构
            dict['policy_info']['policyChannel'] = policyChannel   #公司渠道
            dict['policy_info']['channelId'] = channelId   #销售组织
            dict['policy_info']['salesChannelText'] = salesChannelText   #销售渠道名
            dict['policy_info']['salesChannel'] = salesChannel   #销售渠道id
            dict['policy_info']['headPartner'] = headPartner   #合作机构
            dict['policy_info']['partnerId'] = partnerId   #合作网点
            dict['policy_info']['agentId'] = agentId   #代理人
            dict['policy_info']['submitChannel'] = submitChannel   #提交渠道
            dict['policy_info']['subSalesChannel'] = 'NG-01'   #出单方式
            dict['policy_info']['tellerCode'] = '1'   #柜员代码
            dict['policy_info']['sellerCode'] = '1'   #销售人员代码
            dict['policy_info']['sellerName'] = '1'   #销售人员姓名
            dict['policy_info']['autoPaddingPrem'] = '0'   #是否自垫
            dict['policy_info']['disputedType'] = ''   #争议方式
            dict['policy_info']['policyDeliveryMode'] = '3'   #递送方式
            # 读取保单信息-投保人信息
            dict['policy_info']['holder_info']['sameASInsurd'] = sh['B8'].value  # 投被保人是否同一人
            dict['policy_info']['holder_info']['name'] = sh['B10'].value  # 投保人姓名
            dict['policy_info']['holder_info']['gender'] = sh['B11'].value  # 投保人性别
            dict['policy_info']['holder_info']['pbHoldBirth'] = sh['B12'].value  # 投保人出生日期
            dict['policy_info']['holder_info']['certiType'] = sh['B13'].value  # 投保人证件类型
            dict['policy_info']['holder_info']['certiCode'] = sh['B14'].value  # 投保人证件号码
            dict['policy_info']['holder_info']['pbCertiValidEndDate'] = sh['B15'].value  # 投保人证件止期
            dict['policy_info']['holder_info']['Holder_nationality'] = sh['B16'].value  # 投保人国籍
            dict['policy_info']['holder_info']['mobilePhone'] = sh['B17'].value  # 投保人手机号
            dict['policy_info']['holder_info']['officeTel'] = sh['B18'].value  # 投保人固定电话
            #校验录入的'投被保人是否同一人'与'被保人与投保人关系'是否有冲突
            insuRelation = sh['B9'].value  # 被保人与投保人关系
            if dict['policy_info']['holder_info']['sameASInsurd'] == '0':
                if insuRelation == '00':
                    is_success = False
                    message = '“投被保人是否同一人”录入不为同一人时，“被保人与投保人关系”字段不可以录入“00-本人！”'
            elif dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                if insuRelation != '00':
                    is_success = False
                    message = '“投被保人是否同一人”录入同一人时，“被保人与投保人关系”字段不可以录入非“00-本人！”'
            assert is_success != False

            # 读取保单信息-被保人信息
                #非投保人豁免，投被保人不为同一人时：
            if (dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0'):
                insured1 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                if insured1['insuRelation'] != '00':   #被保人与投保人关系不为本人
                    insured1['insuName'] = sh['B19'].value  # 被保人姓名
                    insured1['insuGender'] = sh['B20'].value  # 被保人性别
                    insured1['insuBirth'] = sh['B21'].value  # 被保人出生日期
                    insured1['insuCertiType'] = sh['B22'].value  # 被保人证件类型
                    insured1['insuCertiCode'] = sh['B23'].value  # 被保人证件号码
                    insured1['insuCertiValidEndDate'] = sh['B24'].value  # 被保人证件止期
                    insured1['insuNationatiy'] = sh['B25'].value  # 被保人国籍
                    insured1['insuMobile'] = sh['B26'].value  # 被保人手机号
                    insured1['officeTel'] = sh['B27'].value  # 被保人固定电话
                    insured1['no'] = '1'  # 第1被保人
                elif insured1['insuRelation'] == '00':   #被保人与投保人关系为本人
                    insured1['insuRelation'] = '00'  # 被保人与投保人关系
                    insured1['insuName'] = sh['B10'].value  # 被保人姓名
                    insured1['insuGender'] = sh['B11'].value  # 被保人性别
                    insured1['insuBirth'] = sh['B12'].value  # 被保人出生日期
                    insured1['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                    insured1['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                    insured1['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                    insured1['insuNationatiy'] = sh['B16'].value  # 被保人国籍
                    insured1['insuMobile'] = sh['B17'].value  # 被保人手机号
                    insured1['officeTel'] = sh['B18'].value  # 被保人固定电话
                    insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                insured1 = {}
                insured1['insuRelation'] = '00'  # 被保人与投保人关系
                insured1['insuName'] = sh['B10'].value  # 被保人姓名
                insured1['insuGender'] = sh['B11'].value  # 被保人性别
                insured1['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured1['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured1['insuNationatiy'] = sh['B16'].value  # 被保人国籍
                insured1['insuMobile'] = sh['B17'].value  # 被保人手机号
                insured1['officeTel'] = sh['B18'].value  # 被保人固定电话
                insured1['no'] = '1'  # 第1被保人
                dict['policy_info']['insured_info'].append(insured1)
                #投保人豁免时：
            elif dict['public']['is_HolderExemption'] == '1':
                insured1 = {}
                insured2 = {}
                insured1['insuRelation'] = sh['B9'].value  # 被保人与投保人关系
                #投保人豁免时，“被保人与投保人关系”字段不可以录入“00-本人”。
                if insured1['insuRelation'] == '00':
                    is_success = False
                    message = '投保人豁免时，“被保人与投保人关系”字段不可以录入“00-本人！”'
                    assert is_success == True
                insured1['insuName'] = sh['B19'].value  # 被保人姓名
                insured1['insuGender'] = sh['B20'].value  # 被保人性别
                insured1['insuBirth'] = sh['B21'].value  # 被保人出生日期
                insured1['insuCertiType'] = sh['B22'].value  # 被保人证件类型
                insured1['insuCertiCode'] = sh['B23'].value  # 被保人证件号码
                insured1['no'] = '1'  # 第1被保人
                insured1['insuCertiValidEndDate'] = sh['B24'].value  # 被保人证件止期
                insured1['insuNationatiy'] = sh['B25'].value  # 被保人国籍
                insured1['insuMobile'] = sh['B26'].value  # 被保人手机号
                insured1['officeTel'] = sh['B27'].value  # 被保人固定电话
                insured2['insuRelation'] = '00'  # 被保人与投保人关系
                insured2['insuName'] = sh['B10'].value  # 被保人姓名
                insured2['insuGender'] = sh['B11'].value  # 被保人性别
                insured2['insuBirth'] = sh['B12'].value  # 被保人出生日期
                insured2['insuCertiType'] = sh['B13'].value  # 被保人证件类型
                insured2['insuCertiCode'] = sh['B14'].value  # 被保人证件号码
                insured2['insuCertiValidEndDate'] = sh['B15'].value  # 被保人证件止期
                insured2['insuNationatiy'] = sh['B16'].value  # 被保人国籍
                insured2['insuMobile'] = sh['B17'].value  # 被保人手机号
                insured1['officeTel'] = sh['B18'].value  # 被保人固定电话
                insured2['no'] = '2'  # 第2被保人
                dict['policy_info']['insured_info'].append(insured1)
                dict['policy_info']['insured_info'].append(insured2) 

            # 读取保单信息-受益人信息
            if dict['nb_type'] == 'sys_API0001.4':    #受益人非法定
                dict['policy_info']['ben_count'] = sh['B28'].value    #受益人个数
                for i in range(0,int(dict['policy_info']['ben_count'])):
                    #判断字段是否录入
                    for j in range(30, 41):
                        if sh.cell(row=j, column=i+2).value == None:
                            is_success = False
                            message = '受益人信息部分存在字段未录入情况。未录入字段为第' + str(j) + '行，第' + str(i+2) + '列！'
                            assert is_success == True
                    dict['policy_info']['benf_info'].append({})
                    dict['policy_info']['benf_info'][-1]['benIsHolder'] = sh.cell(row=30, column=i+2).value    #受益人与投保人是否为同一人
                    dict['policy_info']['benf_info'][-1]['benfRelation'] = sh.cell(row=31, column=i+2).value    #受益人与被保人关系
                    dict['policy_info']['benf_info'][-1]['benfType'] = sh.cell(row=32, column=i+2).value    #受益人类型
                    dict['policy_info']['benf_info'][-1]['benfOrder'] = sh.cell(row=33, column=i+2).value    #受益人顺序
                    dict['policy_info']['benf_info'][-1]['benfPercent'] = sh.cell(row=34, column=i+2).value    #受益人比例

                    #投被保人不为豁免
                    if dict['public']['is_HolderExemption'] == '0':
                        #1.投、被保人同一人：
                        if  dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                            #1.1受益人不为投保人，受益人不为被保人时
                            if dict['policy_info']['benf_info'][-1]['benIsHolder'] == '0' and dict['policy_info']['benf_info'][-1]['benfRelation'] != '00':
                                dict['policy_info']['benf_info'][-1]['benfName'] = sh.cell(row=35, column=i+2).value    #受益人姓名
                                dict['policy_info']['benf_info'][-1]['benfGender'] = sh.cell(row=36, column=i+2).value    #受益人性别
                                dict['policy_info']['benf_info'][-1]['benfBirth'] = sh.cell(row=37, column=i+2).value    #受益人出生日期
                                dict['policy_info']['benf_info'][-1]['benfCertiType'] = sh.cell(row=38, column=i+2).value    #受益人证件类型
                                dict['policy_info']['benf_info'][-1]['benfCertiCode'] = sh.cell(row=39, column=i+2).value    #受益人证件号码
                                dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = sh.cell(row=40, column=i+2).value    #受益人证件止期
                            #1.2受益人同时为投、被保人时
                            elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '1' and dict['policy_info']['benf_info'][-1]['benfRelation'] == '00':
                                dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['holder_info']['name']    #受益人姓名
                                dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['holder_info']['gender']     #受益人性别
                                dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['holder_info']['pbHoldBirth']     #受益人出生日期
                                dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['holder_info']['certiType']     #受益人证件类型
                                dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['holder_info']['certiCode']     #受益人证件号码
                                dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['holder_info']['pbCertiValidEndDate']     #受益人证件止期
                            #1.3受益人为投保人但不为被保人，或受益人不为投保人但为被保人时
                            else:
                                is_success = False
                                message = '投、被保人为同一人时，受益人不可“为投保人而不为被保人”，或“为被保人不为投被人”！'
                                assert is_success == True  
                        #投、被保人不为同一人
                        elif dict['policy_info']['holder_info']['sameASInsurd'] == '0':
                            #受益人同时为投、被保人
                            if dict['policy_info']['benf_info'][-1]['benIsHolder'] == '1' and dict['policy_info']['benf_info'][-1]['benfRelation'] == '00':
                                is_success = False
                                message = '“投、被保人不为同一人”，受益人不可与投、被保人同时为同一人！”'
                                assert is_success == True 
                            #受益人为投保人，不为被保人时
                            elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '1' and dict['policy_info']['benf_info'][-1]['benfRelation'] != '00':
                                dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['holder_info']['name']    #受益人姓名
                                dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['holder_info']['gender']     #受益人性别
                                dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['holder_info']['pbHoldBirth']     #受益人出生日期
                                dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['holder_info']['certiType']     #受益人证件类型
                                dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['holder_info']['certiCode']     #受益人证件号码
                                dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['holder_info']['pbCertiValidEndDate']     #受益人证件止期
                            #受益人为被投保人，不为投保人时
                            elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '0' and dict['policy_info']['benf_info'][-1]['benfRelation'] == '00':
                                dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['insured_info'][-1]['insuName']    #受益人姓名
                                dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['insured_info'][-1]['insuGender']     #受益人性别
                                dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['insured_info'][-1]['insuBirth']     #受益人出生日期
                                dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['insured_info'][-1]['insuCertiType']     #受益人证件类型
                                dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['insured_info'][-1]['insuCertiCode']     #受益人证件号码
                                dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['insured_info'][-1]['insuCertiValidEndDate']     #受益人证件止期
                            #受益人不为被投保人，不为投保人时
                            elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '0' and dict['policy_info']['benf_info'][-1]['benfRelation'] != '00':
                                dict['policy_info']['benf_info'][-1]['benfName'] = sh.cell(row=35, column=i+2).value   #受益人姓名
                                dict['policy_info']['benf_info'][-1]['benfGender'] = sh.cell(row=36, column=i+2).value     #受益人性别
                                dict['policy_info']['benf_info'][-1]['benfBirth'] = sh.cell(row=37, column=i+2).value   #受益人出生日期
                                dict['policy_info']['benf_info'][-1]['benfCertiType'] = sh.cell(row=38, column=i+2).value     #受益人证件类型
                                dict['policy_info']['benf_info'][-1]['benfCertiCode'] = sh.cell(row=39, column=i+2).value     #受益人证件号码
                                dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = sh.cell(row=40, column=i+2).value     #受益人证件止期
                    #投被保人为豁免
                    elif dict['public']['is_HolderExemption'] == '1':
                        #受益人同时为投、被保人
                        if dict['policy_info']['benf_info'][-1]['benIsHolder'] == '1' and dict['policy_info']['benf_info'][-1]['benfRelation'] == '00':
                            dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['holder_info']['name']    #受益人姓名
                            dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['holder_info']['gender']     #受益人性别
                            dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['holder_info']['pbHoldBirth']     #受益人出生日期
                            dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['holder_info']['certiType']     #受益人证件类型
                            dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['holder_info']['certiCode']     #受益人证件号码
                            dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['holder_info']['pbCertiValidEndDate']     #受益人证件止期
                        #受益人为投保人，不为被保人时
                        elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '1' and dict['policy_info']['benf_info'][i]['benfRelation'] != '00':
                            dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['holder_info']['name']    #受益人姓名
                            dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['holder_info']['gender']     #受益人性别
                            dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['holder_info']['pbHoldBirth']     #受益人出生日期
                            dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['holder_info']['certiType']     #受益人证件类型
                            dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['holder_info']['certiCode']     #受益人证件号码
                            dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['holder_info']['pbCertiValidEndDate']     #受益人证件止期
                        #受益人为被投保人，不为投保人时
                        elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '0' and dict['policy_info']['benf_info'][-1]['benfRelation'] == '00':
                            dict['policy_info']['benf_info'][-1]['benfName'] = dict['policy_info']['insured_info'][-1]['insuName']    #受益人姓名
                            dict['policy_info']['benf_info'][-1]['benfGender'] = dict['policy_info']['insured_info'][-1]['insuGender']     #受益人性别
                            dict['policy_info']['benf_info'][-1]['benfBirth'] = dict['policy_info']['insured_info'][-1]['insuBirth']     #受益人出生日期
                            dict['policy_info']['benf_info'][-1]['benfCertiType'] = dict['policy_info']['insured_info'][-1]['insuCertiType']     #受益人证件类型
                            dict['policy_info']['benf_info'][-1]['benfCertiCode'] = dict['policy_info']['insured_info'][-1]['insuCertiCode']     #受益人证件号码
                            dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = dict['policy_info']['insured_info'][-1]['insuCertiValidEndDate']     #受益人证件止期
                        #受益人不为被投保人，不为投保人时
                        elif dict['policy_info']['benf_info'][-1]['benIsHolder'] == '0' and dict['policy_info']['benf_info'][i]['benfRelation'] != '00':
                            dict['policy_info']['benf_info'][-1]['benfName'] = sh.cell(row=35, column=i+2).value  #受益人姓名
                            dict['policy_info']['benf_info'][-1]['benfGender'] = sh.cell(row=36, column=i+2).value     #受益人性别
                            dict['policy_info']['benf_info'][-1]['benfBirth'] = sh.cell(row=37, column=i+2).value    #受益人出生日期
                            dict['policy_info']['benf_info'][-1]['benfCertiType'] = sh.cell(row=38, column=i+2).value     #受益人证件类型
                            dict['policy_info']['benf_info'][-1]['benfCertiCode'] = sh.cell(row=39, column=i+2).value     #受益人证件号码
                            dict['policy_info']['benf_info'][-1]['benfCertiValidDate'] = sh.cell(row=40, column=i+2).value     #受益人证件止期


            # 读取保单信息-主险信息
                #判断是否双主险。得到主险个数
            if dict['public']['is_2risk'] == '0':
                main_risk_num = 1   #主险个数
            elif dict['public']['is_2risk'] == '1':
                main_risk_num = 2   #主险个数
                #根据主险个数，组装主险信息列表。
            for i in range(0, main_risk_num):
                main_risk_info = {}
                if i == 0:
                    main_risk_column = 2   #excel中主险所在的列
                elif i ==1:
                    main_risk_column = 5   #excel中主险所在的列
                main_risk_info['main_risk_num'] = str(i+1) # 第x主险
                main_risk_info['Main_Rider'] = 'M' # 主附险标识
                main_risk_info['productCode'] = str(sh.cell(row=34+count, column=main_risk_column).value)   #险种代码
                try:
                    cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
                except:
                    is_success = False
                    message = '连接数据库不成功！'
                finally:
                    assert is_success != False
                sql = 'select * from t_product where PRODUCT_CODE=' + "'" + main_risk_info['productCode'] + "'"
                cursor.execute(sql)
                result = cursor.fetchall()
                try:
                    main_risk_info['productId'] = str(result[0][0])  # 险种id
                except:
                    is_success = False
                    if main_risk_info['productCode'] == 'None':
                        message = "excel中存在主险的'险种代码'字段未录入，请检查！"
                    else:
                        message = 'excel中录入的主险' + str(main_risk_info['productCode']) + ',数据库中不存在,请检查是否录入有误！'
                finally:
                    assert is_success != False

                main_risk_info['liability_state'] = '01'   #险种状态
                main_risk_info['chargeMode'] = sh.cell(row=35+count, column=main_risk_column).value  # 缴费频率
                main_risk_info['coverPeriodType'] = sh.cell(row=36+count, column=main_risk_column).value  # 保障期间类型
                main_risk_info['coveragePeriod'] = sh.cell(row=37+count, column=main_risk_column).value  # 保障期间
                main_risk_info['chargePeriodType'] = sh.cell(row=38+count, column=main_risk_column).value  # 缴费期间类型
                main_risk_info['chargePeriod'] = sh.cell(row=39+count, column=main_risk_column).value  # 缴费期间
                main_risk_info['sumAssured'] = sh.cell(row=40+count, column=main_risk_column).value  # 保额
                main_risk_info['stdPremBf'] = sh.cell(row=41+count, column=main_risk_column).value  # 保费
                main_risk_info['units'] = sh.cell(row=42+count, column=main_risk_column).value  # 份数
                main_risk_info['planFreq'] = sh.cell(row=43+count, column=main_risk_column).value  # 年金领取方式 或 年金/生存金领取方式
                main_risk_info['payOption'] = sh.cell(row=44+count, column=main_risk_column).value  # 年金使用方式
                main_risk_info['bonusYearOrAge'] = sh.cell(row=45+count, column=main_risk_column).value  # 领取年期或年龄
                main_risk_info['renew'] = sh.cell(row=46+count, column=main_risk_column).value  # 是否续保 
                #判断主险的保额保费是否录入
                if main_risk_info['sumAssured'] == None and main_risk_info['stdPremBf'] == None:
                    is_success = False
                    message = 'excel中存在主险-保额、保费未录入！'
                assert is_success != False
                #存在保额保费为未录入，进行补充为0
                if main_risk_info['sumAssured'] == None:
                    main_risk_info['sumAssured'] = '0'
                if main_risk_info['stdPremBf'] == None:
                    main_risk_info['stdPremBf'] = '0'
                #存在保额保费为0，进行补充
                """
                if main_risk_info['sumAssured'] == '0' and main_risk_info['stdPremBf'] != '0':
                    main_risk_info['sumAssured'] = main_risk_info['stdPremBf']
                elif main_risk_info['sumAssured'] != '0' and main_risk_info['stdPremBf'] == '0':
                    main_risk_info['stdPremBf'] = main_risk_info['sumAssured']
                """
                # 存入dict
                dict['policy_info']['main_risk_info'].append(main_risk_info)
            # 读取保单信息-主险1附加险信息
            for i in range(3, int(sh['B' + str(31+count)].value)+3):
                sub_risk_info = {} 
                sub_risk_info['belong_mainrisk_productId'] = 'main_risk_1'  # 所属第几主险
                sub_risk_info['Main_Rider'] = 'R' # 主附险标识
                sub_risk_info['productCode'] = str(sh.cell(row=34+count, column=i).value)   #险种代码
                cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
                sql = 'select * from t_product where PRODUCT_CODE=' + "'" + sub_risk_info['productCode'] + "'"
                cursor.execute(sql)
                result = cursor.fetchall()
                try:
                    sub_risk_info['productId'] = str(result[0][0])  # 险种id
                except:
                    is_success = False
                    if sub_risk_info['productCode'] == 'None':
                        message = "excel中存在附加险的'险种代码'字段未录入，请检查！"
                    else:
                        message = 'excel中录入的附加险' + str(sub_risk_info['productCode']) + ',数据库中不存在,请检查是否录入有误！'
                finally:
                    assert is_success != False
                sub_risk_info['liability_state'] = '01'   #险种状态
                sub_risk_info['chargeMode'] = sh.cell(row=35+count, column=i).value  # 缴费频率
                sub_risk_info['coverPeriodType'] = sh.cell(row=36+count, column=i).value  # 保障期间类型
                sub_risk_info['coveragePeriod'] = sh.cell(row=37+count, column=i).value  # 保障期间
                sub_risk_info['chargePeriodType'] = sh.cell(row=38+count, column=i).value  # 缴费期间类型
                sub_risk_info['chargePeriod'] = sh.cell(row=39+count, column=i).value  # 缴费期间
                sub_risk_info['sumAssured'] = sh.cell(row=40+count, column=i).value  # 保额
                sub_risk_info['stdPremBf'] = sh.cell(row=41+count, column=i).value  # 保费
                sub_risk_info['units'] = sh.cell(row=42+count, column=i).value  # 份数
                sub_risk_info['planFreq'] = sh.cell(row=43+count, column=i).value  # 年金领取方式 或 年金/生存金领取方式
                sub_risk_info['payOption'] = sh.cell(row=44+count, column=i).value  # 年金使用方式
                sub_risk_info['bonusYearOrAge'] = sh.cell(row=45+count, column=i).value  # 领取年期或年龄
                sub_risk_info['renew'] = sh.cell(row=46+count, column=i).value  # 是否续保 
                #判断附加险的保额保费是否录入
                if sub_risk_info['sumAssured'] == None and sub_risk_info['stdPremBf'] == None:
                    is_success = False
                    message = 'excel中存在附加险-保额、保费未录入！'
                assert is_success != False
                #存在保额保费为未录入，进行补充为0
                if sub_risk_info['sumAssured'] == None:
                    sub_risk_info['sumAssured'] = '0'
                if sub_risk_info['stdPremBf'] == None:
                    sub_risk_info['stdPremBf'] = '0'
                #存在保额保费为0，进行补充
                if sub_risk_info['sumAssured'] == '0' and sub_risk_info['stdPremBf'] != '0':
                    sub_risk_info['sumAssured'] = sub_risk_info['stdPremBf']
                elif sub_risk_info['sumAssured'] != '0' and sub_risk_info['stdPremBf'] == '0':
                    sub_risk_info['stdPremBf'] = sub_risk_info['sumAssured']
                # 存入dict
                dict['policy_info']['sub_risk_info'].append(sub_risk_info)

            # 读取保单信息-主险2附加险信息
            if dict['public']['is_2risk'] == '1':
                for i in range(0, int(sh['B' + str(32+count)].value)):
                    sub_risk_info = {} 
                    sub_risk_info['belong_mainrisk_productId'] = 'main_risk_2'  # 所属第几主险
                    sub_risk_info['Main_Rider'] = 'R' # 主附险标识
                    sub_risk_info['productCode'] = str(sh.cell(row=34+count, column=i+6).value)   #险种代码
                    cursor = API0000_diy().db_conf(env_name, 'wift_iiws')
                    sql = 'select * from t_product where PRODUCT_CODE=' + "'" + sub_risk_info['productCode'] + "'"
                    cursor.execute(sql)
                    result = cursor.fetchall()
                    try:
                        sub_risk_info['productId'] = str(result[0][0])  # 险种id
                    except:
                        is_success = False
                        if sub_risk_info['productCode'] == 'None':
                            message = "excel中存在附加险的'险种代码'字段未录入，请检查！"
                        else:
                            message = 'excel中录入的附加险' + str(sub_risk_info['productCode']) + ',数据库中不存在,请检查是否录入有误！'
                    finally:
                        assert is_success != False
                    sub_risk_info['liability_state'] = '01'   #险种状态
                    sub_risk_info['chargeMode'] = sh.cell(row=35+count, column=i+6).value  # 缴费频率
                    sub_risk_info['coverPeriodType'] = sh.cell(row=36+count, column=i+6).value  # 保障期间类型
                    sub_risk_info['coveragePeriod'] = sh.cell(row=37+count, column=i+6).value  # 保障期间
                    sub_risk_info['chargePeriodType'] = sh.cell(row=38+count, column=i+6).value  # 缴费期间类型
                    sub_risk_info['chargePeriod'] = sh.cell(row=39+count, column=i+6).value  # 缴费期间
                    sub_risk_info['sumAssured'] = sh.cell(row=40+count, column=i+6).value  # 保额
                    sub_risk_info['stdPremBf'] = sh.cell(row=41+count, column=i+6).value  # 保费
                    sub_risk_info['units'] = sh.cell(row=42+count, column=i+6).value  # 份数
                    sub_risk_info['planFreq'] = sh.cell(row=43+count, column=i+6).value  # 年金领取方式 或 年金/生存金领取方式
                    sub_risk_info['payOption'] = sh.cell(row=44+count, column=i+6).value  # 年金使用方式
                    sub_risk_info['bonusYearOrAge'] = sh.cell(row=45+count, column=i+6).value  # 领取年期或年龄
                    sub_risk_info['renew'] = sh.cell(row=46+count, column=i+6).value  # 是否续保 
                    #判断附加险的保额保费是否录入
                    if sub_risk_info['sumAssured'] == None and sub_risk_info['stdPremBf'] == None:
                        is_success = False
                        message = 'excel中存在附加险-保额、保费未录入！'
                    assert is_success != False
                    #存在保额保费为未录入，进行补充为0
                    if sub_risk_info['sumAssured'] == None:
                        sub_risk_info['sumAssured'] = '0'
                    if sub_risk_info['stdPremBf'] == None:
                        sub_risk_info['stdPremBf'] = '0'
                    #存在保额保费为0，进行补充
                    if sub_risk_info['sumAssured'] == '0' and sub_risk_info['stdPremBf'] != '0':
                        sub_risk_info['sumAssured'] = sub_risk_info['stdPremBf']
                    elif sub_risk_info['sumAssured'] != '0' and sub_risk_info['stdPremBf'] == '0':
                        sub_risk_info['stdPremBf'] = sub_risk_info['sumAssured']
                    # 存入dict
                    dict['policy_info']['sub_risk_info'].append(sub_risk_info) 
            is_success = True

            #3.校验录入险种，脚本是否支持出单
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：出单前判断-校验录入险种，脚本是否支持出单'  # 节点
            list_risk = ['10001','10002','10003','10004','10006','10007','10008','10009','10011','10012','10014','10017','10020','10021','10035','10036','10037','10038','10042','10044','10046','10047','10048','10049','10050','10053','10054','10063','10065','10071','10073','10076','10077','10078','10079','10080','10081','10082','10083','10097','10098','10099','10100','10106','10107','10108','10109','10110','10111','10112','10113','10118','10130','10131','10132','10133','10134','10135','10136','10138','10139','10140','10141','10142','10148','10159','10160','10161','10162','10163','10164','10166','10168','10169','10170','10171','10173','10174','10175','10177','10184','10189','10193','10194','10200','10204','10207','10211','10214','10223','10224','10225','10226','10129','10231','10240','10241','10245','10248','10250','10251','10143','10144','10203','10205','10206','10232','10128','10202','10185', '10186', '10187','10188','10227','10228','10229','10230','10119','10209','10222','10180','10181','10127','10172','10233','10120','10121','10122','10216','10246','10252','10253']
            dict_risk = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']
            for i in range(0, len(dict_risk)):
                productId = dict_risk[i]['productId']
                if productId not in list_risk:
                    is_success = False
                    message = 'excel录入的险种' + str(dict_risk[i]['productCode']) + '暂不支持出单，请联系管理员进行配置！'
                else:
                    is_success = True
                    message = ''
                assert is_success == True

            #4.验证录入的投保人手机号是否被其他投保人使用
            is_success = ''
            message = ''
            unit = '节点：出单前判断-校验录入的投保人手机号是否被其他投保人使用'  # 节点
            holderer_list = []   #投保人列表
            try:
                cursor = API0000_diy().db_conf(env_name, 'wift_pa')
            except:
                is_success = False
                message = '连接数据库不成功！'
            finally:
                assert is_success != False
            sql = "select holder_cust_id from t_pa_policy_holder where mobile='" + str(dict['policy_info']['holder_info']['mobilePhone']) + "'"
            cursor.execute(sql)
            result = cursor.fetchall()
            for i in range(0, len(result)):
                holderer_list.append(result[i][0])
            holderer_list = set(holderer_list)  #列表去重
            holderer_list = list(holderer_list)

            if len(holderer_list) > 0:   #数据库查询投保人表，录入的投保人手机号查询结果不为空
                for i in range(0,len(holderer_list)):
                    sql1 = "select * from t_pa_customer where customer_id='" + str(holderer_list[i]) + "' and name='" + str(dict['policy_info']['holder_info']['name']) + "' and gender='" + str(dict['policy_info']['holder_info']['gender']) + "' and certi_type='" + str(dict['policy_info']['holder_info']['certiType']) + "' and certi_code='" + str(dict['policy_info']['holder_info']['certiCode']) + "' and mobile='" + str(dict['policy_info']['holder_info']['mobilePhone']) + "'"
                    cursor.execute(sql1)
                    result1 = cursor.fetchall()
                    if len(result1) == 0:   #数据库查询客户表，录入的投保人手机号存在其他投保人。
                        is_success = False
                        message = '经数据库查询，录入的投保人手机号被其他投保人使用！'
                        break
                    elif len(result1) > 0:   #数据库查询客户表，录入的投保人手机号为本人。
                        is_success = True
                        message = ''
            elif len(holderer_list) == 0:  #数据库查询投保人表，录入的投保人手机号查询结果为空
                is_success = True
                message = ''
            assert is_success == True

            #5.验证录入的投被保人关系

            #结束时间
            end_time = datetime.datetime.now()   
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*出单前准备：数据存储+用户验证_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
            # 加入断言
            assert is_success == True
            dict['logInfo']['code'] = '1'   #记录无异常标识
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'   #记录异常标识
            dict['logInfo']['err'] = unit + str(e)   #存入脚本日志报错信息
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n') + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入接口报错信息
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict




    """新契约录入"""
    @func_set_timeout(120)#设定函数超执行时间
    def nb_entry(self, dict):
        """
           step1:登录系统-新契约录入用户
           step2:投保单号提交接口
           step3:保单基本信息录入接口
           step4:投保人信息录入接口
           step5:投保人告知接口
           step6:被保人信息录入接口
           step7:被保人查询接口
           step8:被保人告知接口
           step9:收费方式信息录入接口
           step10:险种信息录入接口
           step11:新契约录入提交接口
           step12:登出系统-新契约录入用户
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约录入_开始(预计耗时:20s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            magicX = ''
            """step1:登录系统-新契约录入用户"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '节点：系统-新契约录入_登录系统'  # 节点标识
            message = ''   #信息标识
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   #环境名称
            interface_ip = dict['public']['interface_ip']   #接口ip
            username = dict['public']['username_entry']    #新契约录入用户
            #调用原子脚本
            rs = API0000_sys_otherapi().login(env_name,username, interface_ip)
            #获取所需响应值
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            magicX = rs[2]   #会话id
            #断言
            assert is_success == True
            dict['public']['magicX'] = magicX
            
            """step2:投保单号提交接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_投保单号提交接口'  # 节点
            #dict获取数据+定义数据
            proposalNo = API0000_diy().random_proposalNo()  # 投保单号（随机生成保单号）
            dict['policy_info']['proposalNo'] = proposalNo
            policyType = 1  # 保单类型：1-个险
            #调用原子脚本
            response = API0000_nb().entry_newProposal(magicX, interface_ip, policyType, proposalNo)
            #获取所需响应值
            is_success = response['success']
            if is_success == True:
                others = response['others']
                # 获取生成的保单号
                policyNo = others['policyNo']
                # 获取服务器时间
                registDate = others['registDate']
                # 获取proposalId
                proposalId = others['proposalId']
                message = ''
            else:
                policyNo = ''
                registDate = ''
                proposalId = ''
                message = response['message']
            #断言
            assert is_success == True
            dict['policy_info']['policyNo'] = policyNo  # 保单号
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*投保单号：' + str(policyNo)))
            """step3:保单基本信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_保单基本信息录入接口'  # 节点
            #dict获取数据+定义数据     
                #投保日期 
            applicationDate = dict['policy_info']['applicationDate']   
                #定义生效日期 effectiveDate
            date1 = datetime.datetime.strptime(applicationDate, '%Y-%m-%d')
            date2 = date1 + datetime.timedelta(days=1)
            effectiveDate = date2.strftime('%Y-%m-%d')
                #保单基本信息
            organId = dict['policy_info']['organId']   #所属机构
            policyChannel = dict['policy_info']['policyChannel']   #公司渠道
            channelId = dict['policy_info']['channelId']   #销售组织
            salesChannelText = dict['policy_info']['salesChannelText']   #销售渠道名
            salesChannel = dict['policy_info']['salesChannel']   #销售渠道id
            headPartner = dict['policy_info']['headPartner']   #合作机构
            partnerId = dict['policy_info']['partnerId']   #合作网点
            agentId = dict['policy_info']['agentId']   #代理人
            submitChannel = dict['policy_info']['submitChannel']   #提交渠道
            subSalesChannel = dict['policy_info']['subSalesChannel']
            tellerCode = dict['policy_info']['tellerCode']   #柜员代码
            sellerCode = dict['policy_info']['sellerCode']   #销售人员代码
            sellerName = dict['policy_info']['sellerName']   #销售人员姓名
            autoPaddingPrem = dict['policy_info']['autoPaddingPrem']   #是否自垫
            disputedType = dict['policy_info']['disputedType']   #争议方式
            policyDeliveryMode = dict['policy_info']['policyDeliveryMode']   #递送方式
            #调用原子脚本
            response = API0000_nb().proposal_updateProposalInfo(magicX, interface_ip, proposalNo, policyNo, proposalId, organId, policyChannel, channelId, salesChannelText, salesChannel, headPartner, partnerId, tellerCode, sellerCode, sellerName, agentId, submitChannel, subSalesChannel, applicationDate, effectiveDate, autoPaddingPrem, disputedType, policyDeliveryMode)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True           

            """step4:投保人信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_投保人信息录入接口'  # 节点
            #dict获取数据+定义数据
            is_HolderExemption = dict['public']['is_HolderExemption']
            if dict['policy_info']['holder_info']['sameASInsurd'] == '1' or is_HolderExemption == '1': #投被保人为同一人，和投保人豁免，sameASInsurd均录入1
                sameASInsurd = '1'
            else:
                sameASInsurd = ''
            Holder_name = dict['policy_info']['holder_info']['name']   # 投保人姓名
            Holder_gender = dict['policy_info']['holder_info']['gender']   # 投保人性别
            Holder_Birth = dict['policy_info']['holder_info']['pbHoldBirth']   # 投保人出生日期
            Holder_certiType = dict['policy_info']['holder_info']['certiType']   # 投保人证件类型
            Holder_certiCode = dict['policy_info']['holder_info']['certiCode']   # 投保人证件号码
            Holder_CertiValidEndDate = dict['policy_info']['holder_info']['pbCertiValidEndDate']   # 投保人证件止期
            Holder_nationality = dict['policy_info']['holder_info']['Holder_nationality']   # 投保人国籍
            Holder_mobilePhone = dict['policy_info']['holder_info']['mobilePhone']   # 投保人手机号
            Holder_officeTel = dict['policy_info']['holder_info']['officeTel']  # 投保人固定电话
            Holder_isMarried = '20'  # 婚姻状况
            Holder_height = '167'  # 身高/CM
            Holder_weight = '67'  # 体重 /KG
            Holder_homeTel = '010-98343543'  # 家庭电话
            Holder_industry = '03'  # 工作行业
            Holder_workCompany = '东方小清华'  # 工作单位/学校名称
            Holder_email = API0000_diy().random_email()  # 电子邮箱
            Holder_jobCode = '0001001'  # 职业代码/名称
            Holder_jobClass = '1'  # 职业等级
            Holder_driverLicenseType = '06'  # 驾照类型
            Holder_nationnality1 = '01'  # 民族
            Holder_educationId = '40'  # 学位
            Holder_medicalInsType = '04'  # 医保类型
            Holder_incomeSource = '01'  # 收入来源
            Holder_incomeSourceNote = 'wu'  # 收入来源备注
            Holder_annualIncome = '5000000'  # 年收入/万元（系统自动转换为30万元）
            Holder_familyIncome = '5000000'  # 家庭年收入/万元（系统自动转换为30万元）
            Holder_premBudget = '30000'  # 保费预算/元
            Holder_residentType = '1'  # 居民类型
            Holder_taxPayerType = '01'  # 税收居民类型
            Holder_postalCode = '121200'  # 邮编
            Holder_provinceCode = '110000'  # 省份
            Holder_cityCode = '110100'  # 市区
            Holder_districtCode = '110101'  # 地区
            Holder_addrDetail = '建国门东大街22号'  # 详细地址
            #调用原子脚本
            response = API0000_nb().savePersonalHolder(magicX, interface_ip, proposalNo, policyChannel, sameASInsurd, Holder_name, Holder_gender, Holder_Birth, Holder_certiType, Holder_certiCode, Holder_CertiValidEndDate, Holder_isMarried, Holder_height, Holder_weight, Holder_mobilePhone, Holder_officeTel, Holder_homeTel, Holder_industry, Holder_workCompany, Holder_email, Holder_jobCode, Holder_jobClass, Holder_driverLicenseType, Holder_nationality, Holder_nationnality1, Holder_educationId, Holder_medicalInsType, Holder_incomeSource, Holder_incomeSourceNote, Holder_annualIncome, Holder_familyIncome, Holder_premBudget, Holder_residentType, Holder_taxPayerType, Holder_postalCode, Holder_provinceCode, Holder_cityCode, Holder_districtCode, Holder_addrDetail)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']
                customerId = others['customerId']
                message = ''
            else:
                customerId = ''
                message = response['message']
            #断言
            assert is_success == True
            
            """step5:投保人告知接口""" 
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_投保人告知接口'  # 节点
            #dict获取数据+定义数据
            result = '[{"id":"505004","value":"Y"},{"id":"5050045001","value":""}]'   #告知内容
            #调用原子脚本
            response = API0000_nb().Holder_saveNotifyInfo(magicX, interface_ip, proposalId, result, customerId)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step6:被保人信息录入接口"""            
            if (is_HolderExemption == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0') or is_HolderExemption == '1':
                #节点初始化
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_被保人信息录入接口'  # 节点
                #dict获取数据+定义数据
                insuRelation = dict['policy_info']['insured_info'][0]['insuRelation']   # 被保人与投保人关系
                if insuRelation == '00':
                    insucustomerId = customerId
                else:
                    insucustomerId = ''
                insuName = dict['policy_info']['insured_info'][0]['insuName']  # 被保人姓名
                insuGender = dict['policy_info']['insured_info'][0]['insuGender']  # 被保人性别
                insuBirth = dict['policy_info']['insured_info'][0]['insuBirth']  # 被保人出生日期
                insuCertiType = dict['policy_info']['insured_info'][0]['insuCertiType']  # 被保人证件类型
                insuCertiCode = dict['policy_info']['insured_info'][0]['insuCertiCode']  # 被保人证件号码
                insuCertiValidEndDate =dict['policy_info']['insured_info'][0]['insuCertiValidEndDate']  # 被保人证件止期
                insuNationatiy = dict['policy_info']['insured_info'][0]['insuNationatiy']  # 被保人国籍
                insuMobile = dict['policy_info']['insured_info'][0]['insuMobile']  # 被保人手机号
                insuofficeTel = dict['policy_info']['insured_info'][0]['officeTel']  # 被保人固定电话
                insuHeight = '178'  # 身高/CM
                insuWeight = '78'  # 体重/KG
                insuMarriage = '20'  # 婚姻状况
                if API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth) <= 6:
                    insuWorkCode = '2147001'   #职业代码/名称-学龄前儿童
                elif (API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth) > 6 and API0000_diy().calc_policy_insured_age(effectiveDate, insuBirth) < 18):
                    insuWorkCode = '0000001'   #职业代码/名称-儿童、18岁前儿童
                else:
                    insuWorkCode = '0001001'  #职业代码/名称-内勤
                insujobClass = '1'  # 职业等级
                insuEmail = API0000_diy().random_email()  # 电子邮箱
                insuWorkType = '02'  # 工作行业
                insuCompany = 'apple'  # 工作单位/学校名称
                insunationnality1 = '01'  # 民族
                insumedicalInsType = '04'  # 医保类型
                insudriverLicenseType = '06'  # 驾照类型
                insuhomeTel = '010-993433344'  # 家庭电话
                insuincomeSource = '01'  # 收入来源
                insuincomeSourceNote = 'wu'  # 收入来源备注
                insuannualIncome = '5000000'  # 年收入/万元
                insuresidentType = '1'  # 居民类型
                insutaxPayerType = '01'  # 税收居民类型
                insueducationId = '40'  # 学位
                insuPostalCode = '120000'  # 邮编
                insuProvinceCode = '110000'  # 省
                insuCityCode = '110100'  # 市
                insuDistrictCode = '110101'  # 区
                insuAddrDetail = '建国门东大街22号'  # 详细地址
                #调用原子脚本
                response = API0000_nb().savePersonalInsured(magicX, interface_ip, proposalId, insucustomerId, policyChannel, insuRelation, insuName, insuGender, insuBirth, insuHeight, insuWeight, insuMarriage, insuCertiType, insuCertiCode, insuCertiValidEndDate, insuWorkCode, insujobClass, insuEmail, insuWorkType, insuCompany, insuMobile, insuNationatiy, insunationnality1, insuofficeTel, insumedicalInsType, insudriverLicenseType, insuhomeTel, insuincomeSource, insuincomeSourceNote, insuannualIncome, insuresidentType, insutaxPayerType, insueducationId, insuPostalCode, insuProvinceCode, insuCityCode, insuDistrictCode, insuAddrDetail)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True
            
            """step7:被保人查询接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_被保人查询接口'  # 节点
            #调用原子脚本
            response = API0000_nb().queryPersonalInsured(magicX, interface_ip, proposalNo)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']
                message = ''
            else:
                others = ''
                message = response['message']
            # 根据判断是否豁免，确定出被保人insuId、被保人姓名、被保人insucustomerId、豁免投保人insuId、豁免投保人姓名、豁免投保人insucustomerId
            if len(others) == 1:
                insuId = others[0]['insuId']  # 被保人insuId
                insuName = others[0]['insuName']  # 被保人姓名
                insucustomerId = others[0]['insucustomerId']  # 被保人insucustomerId
                insuId1 = ''  # 豁免投保人insuId
                insuName1 = ''  # 豁免投保人姓名
                insucustomerId1 = ''  # 豁免投保人insucustomerId
            elif len(others) == 2:
                # 判断豁免投保人处于others中的第几位
                for i in range(0, 2):
                    if others[i]['insuName'] == dict['policy_info']['holder_info']['name']:
                        j = i  # 得到豁免投保人处于others中的位置
                        # 判断录入的被保人处于others中的第几位
                        if j == 0:
                            k = 1
                        elif j == 1:
                            k = 0;
                insuId = others[k]['insuId']  # 被保人insuId
                insuName = others[k]['insuName']  # 被保人姓名
                insucustomerId = others[k]['insucustomerId']  # 被保人insucustomerId
                insuId1 = others[j]['insuId']  # 豁免投保人insuId
                insuName1 = others[j]['insuName']  # 豁免投保人姓名
                insucustomerId1 = others[j]['insucustomerId']  # 豁免投保人insucustomerId
            assert is_success == True 
                       
            """step8:被保人告知接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_被保人告知接口'  # 节点
            #dict获取数据+定义数据
            result = '[{"id":"505004","value":"Y"},{"id":"5050045001","value":""}]'   #告知内容
            # 1.投保人非豁免、投被保人关系不为同一人时录入;2.投保人豁免时录入
            #调用原子脚本
            if (is_HolderExemption == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0') or is_HolderExemption == '1':
                response = API0000_nb().Insured_saveNotifyInfo(magicX, interface_ip, proposalId, result, insucustomerId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True
            
            """step~：受益人信息录入接口"""
            if dict['nb_type'] == 'sys_API0001.4':
                #节点初始化
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_受益人信息录入接口'  # 节点
                #dict获取数据+定义数据
                isLegal = ''    #是否法定
                ben_count = dict['policy_info']['ben_count']   #受益人个数
                for i in range(0, int(ben_count)):
                    benIsHolder = dict['policy_info']['benf_info'][i]['benIsHolder']   #受益人与投保人是否为同一人
                    benfRelation = dict['policy_info']['benf_info'][i]['benfRelation']   #受益人与被保人关系
                    if benIsHolder == '1' and benfRelation != '00':   #受益人为投保人不为被保人
                        benfcustomerId = customerId
                        holderCustomerId = customerId
                        benIsHolder = 'Y'
                    elif  benIsHolder == '0' and benfRelation == '00':   #受益人为被保人不为投被人
                        benfcustomerId = insucustomerId
                        holderCustomerId = insucustomerId
                        benIsHolder = ''
                    elif benIsHolder == '0' and benfRelation != '00':   #受益人不为投保人、被保人
                        benfcustomerId = ''
                        holderCustomerId = ''
                        benIsHolder = ''
                    elif benIsHolder == '1' and benfRelation == '00':   #受益人为投保人、被保人
                        benfcustomerId = customerId
                        holderCustomerId = customerId
                        benIsHolder = 'Y'
                    insured = insuId   #
                    benfType = dict['policy_info']['benf_info'][i]['benfType']   #受益人类型;1-身故受益人；2-生存受益人
                    benfOrder = dict['policy_info']['benf_info'][i]['benfOrder']   #受益顺序
                    benfPercent = dict['policy_info']['benf_info'][i]['benfPercent']   #受益比例(%)
                    benfName = dict['policy_info']['benf_info'][i]['benfName']   #受益人姓名
                    benfGender = dict['policy_info']['benf_info'][i]['benfGender']   #受益人性别
                    benfBirth = dict['policy_info']['benf_info'][i]['benfBirth']   #受益人出生日期
                    benfCertiType = dict['policy_info']['benf_info'][i]['benfCertiType']    #受益人证件类型
                    benfCertiCode = dict['policy_info']['benf_info'][i]['benfCertiCode']    #受益人证件号码
                    benfCertiValidDate = dict['policy_info']['benf_info'][i]['benfCertiValidDate']    #受益人证件止期
                    benfMobile = '18903456754'   #受益人移动电话
                    officeTel = '010-93435343'   #受益人固定电话
                    homeTel = ''   #受益人家庭电话
                    benfNationatiy = 'CHN'    #受益人国籍
                    taxPayerType = '01'   #受益人税收居民类型
                    benfEmail = '453589sdfDF1@qq.com'   #受益人电子邮箱
                    if API0000_diy().calc_policy_insured_age(effectiveDate, benfBirth) <= 6:
                        benfWorkCode = '2147001'   #职业代码/名称-学龄前儿童
                    elif (API0000_diy().calc_policy_insured_age(effectiveDate, benfBirth) > 6 and API0000_diy().calc_policy_insured_age(effectiveDate, benfBirth) < 18):
                        benfWorkCode = '0000001'   #职业代码/名称-儿童、18岁前儿童
                    else:
                        benfWorkCode = '0001001'  #职业代码/名称-内勤
                    benfPostalCode = '121200'   #受益人邮编
                    benfProvinceCode = '110000'   #受益人省份
                    benfCityCode = '110100'   #受益人市区
                    benfDistrictCode = '110101'   #受益人地区
                    benfAddrDetail = '建国门东大街108号'   #受益人详细地址
                    #调用原子脚本
                    response = API0000_nb().savePersonalBenf(magicX, interface_ip, proposalId, benfcustomerId, holderCustomerId, isLegal, benIsHolder, insured, benfRelation, benfType, benfOrder, benfPercent, benfName, benfGender, benfBirth, benfCertiType, benfCertiCode, benfCertiValidDate, benfMobile, officeTel, homeTel, benfNationatiy, taxPayerType, benfEmail, benfWorkCode, benfPostalCode, benfProvinceCode, benfCityCode, benfDistrictCode, benfAddrDetail)
                    #获取所需响应值
                    # 返回是否成功，以及错误信息
                    is_success = response['success']
                    if is_success:
                        message = ''
                    else:
                        message = response['message']
                    #断言
                    assert is_success == True


            """step9:收费方式信息录入接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_收费方式信息录入接口'  # 节点
            #dict获取数据+定义数据
            payMode = dict['policy_info']['payment_nb']['payMode']
            name = dict['policy_info']['holder_info']['name']  # 账户名(投保人姓名)
            bankCode = dict['policy_info']['payment_nb']['bankCode']  # 开户银行
            bankAccount = '64556' + str(random.randint(100000000000,999999999999)) # 银行卡号
            reserveMobile = '13556546366'  # 银行预留手机号
            #调用原子脚本
            response = API0000_nb().payModeSave(magicX, interface_ip, proposalId, payMode, name, bankCode, bankAccount, reserveMobile)
            #获取所需响应值
            # 返回是否成功，以及错误信息
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step10:险种信息录入接口"""
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_险种信息录入接口-主险1'  # 节点
            is_2risk = dict['public']['is_2risk']  #是否双主险
            sub_risk_info1 = []   #初始化-主险1附加险
            sub_risk_info2 = []   #初始化-主险2附加险

            #拆分双主险下的附加险
            if is_2risk == '1':
                for i in range(0, len(dict['policy_info']['sub_risk_info'])):
                    if dict['policy_info']['sub_risk_info'][i]['belong_mainrisk_productId'] == 'main_risk_1':
                        sub_risk_info1.append(dict['policy_info']['sub_risk_info'][i])
                    elif dict['policy_info']['sub_risk_info'][i]['belong_mainrisk_productId'] == 'main_risk_2':
                        sub_risk_info2.append(dict['policy_info']['sub_risk_info'][i])
            else:
                sub_risk_info1 = dict['policy_info']['sub_risk_info']

            ##先录入主险1
                #dict获取数据+定义数据
            productId = dict['policy_info']['main_risk_info'][0]['productId']  # 险种id
            chargeMode = dict['policy_info']['main_risk_info'][0]['chargeMode']  # 缴费频率
            coverPeriodType = dict['policy_info']['main_risk_info'][0]['coverPeriodType']  # 保障期间类型
            coveragePeriod = dict['policy_info']['main_risk_info'][0]['coveragePeriod']  # 保障期间
            chargePeriodType = dict['policy_info']['main_risk_info'][0]['chargePeriodType']  # 缴费期间类型
            chargePeriod = dict['policy_info']['main_risk_info'][0]['chargePeriod']  # 缴费期间
            sumAssured = dict['policy_info']['main_risk_info'][0]['sumAssured']  # 保额
            stdPremBf = dict['policy_info']['main_risk_info'][0]['stdPremBf']  # 保费
            units = dict['policy_info']['main_risk_info'][0]['units']  # 份数
            planFreq = dict['policy_info']['main_risk_info'][0]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
            payOption = dict['policy_info']['main_risk_info'][0]['payOption']  # 年金使用方式
            bonusYearOrAge = dict['policy_info']['main_risk_info'][0]['bonusYearOrAge']  # 领取年期或年龄
            renew = dict['policy_info']['main_risk_info'][0]['renew']  # 是否续保
            mainItemId = ''
                #调用原子脚本
            response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId, mainItemId, insuId, chargeMode, coverPeriodType, coveragePeriod, chargePeriodType, chargePeriod, units, sumAssured, stdPremBf, planFreq, payOption, bonusYearOrAge, headPartner, organId, renew)
                #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
                others = response['others']
                mainItemId1 = others['itemId']
                sys_stdPremBf = others['stdPremBf']   #保费
                sys_sumAssured = others['sumAssured']   #保额
            else:
                message = response['message']
                mainItemId1 = ''
                sys_stdPremBf = ''
                sys_sumAssured = ''
            assert is_success == True
            #存入dict
            dict['policy_info']['main_risk_info'][0]['stdPremBf'] = sys_stdPremBf  # 保费
            dict['policy_info']['main_risk_info'][0]['sumAssured'] = sys_sumAssured  # 保额
            ##录入主险1-豁免附加险
            if is_HolderExemption == '1':
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_险种信息录入接口-主险1_豁免附加险'  # 节点
                #dict获取数据+定义数据
                productId1 = sub_risk_info1[0]['productId']  # 险种id
                chargeMode1 = sub_risk_info1[0]['chargeMode']  # 缴费频率
                coverPeriodType1 = sub_risk_info1[0]['coverPeriodType']  # 保障期间类型
                coveragePeriod1 = sub_risk_info1[0]['coveragePeriod']  # 保障期间
                chargePeriodType1 = sub_risk_info1[0]['chargePeriodType']  # 缴费期间类型
                chargePeriod1 = sub_risk_info1[0]['chargePeriod']  # 缴费期间
                sumAssured1 = sub_risk_info1[0]['sumAssured']  # 保额
                stdPremBf1 = sub_risk_info1[0]['stdPremBf']  # 保费
                units1 = sub_risk_info1[0]['units']  # 份数
                planFreq1 = sub_risk_info1[0]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                payOption1 = sub_risk_info1[0]['payOption']  # 年金使用方式
                bonusYearOrAge1 = sub_risk_info1[0]['bonusYearOrAge']  # 领取年期或年龄
                renew1 = sub_risk_info1[0]['renew']  # 是否续保
                #调用原子脚本
                response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId1, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew1)   
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                    others = response['others']
                    sys_stdPremBf = others['stdPremBf']   #保费
                    sys_sumAssured = others['sumAssured']   #保额
                else:
                    message = response['message']
                    sys_stdPremBf = ''
                    sys_sumAssured = ''
                assert is_success == True
                    #存入sub_risk_info1
                sub_risk_info1[0]['stdPremBf'] = sys_stdPremBf  # 保费
                sub_risk_info1[0]['sumAssured'] = sys_sumAssured  # 保额
            ##判断存在附加险不存在豁免险时
            if is_HolderExemption == '0' and len(sub_risk_info1) > 0:
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_险种信息录入接口-非豁免的附加险'  # 节点
                for i in range(0, len(sub_risk_info1)):
                    #dict获取数据+定义数据
                    productId1 = sub_risk_info1[i]['productId']  # 险种id
                    chargeMode1 = sub_risk_info1[i]['chargeMode']  # 缴费频率
                    coverPeriodType1 = sub_risk_info1[i]['coverPeriodType']  # 保障期间类型
                    coveragePeriod1 = sub_risk_info1[i]['coveragePeriod']  # 保障期间
                    chargePeriodType1 = sub_risk_info1[i]['chargePeriodType']  # 缴费期间类型
                    chargePeriod1 = sub_risk_info1[i]['chargePeriod']  # 缴费期间
                    sumAssured1 = sub_risk_info1[i]['sumAssured']  # 保额
                    stdPremBf1 = sub_risk_info1[i]['stdPremBf']  # 保费
                    units1 = sub_risk_info1[i]['units']  # 份数
                    planFreq1 = sub_risk_info1[i]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                    payOption1 = sub_risk_info1[i]['payOption']  # 年金使用方式
                    bonusYearOrAge1 = sub_risk_info1[i]['bonusYearOrAge']  # 领取年期或年龄
                    renew1 = sub_risk_info1[i]['renew']   #是否续保
                    # 调用原子化脚本
                    response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew1) 
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        message = ''
                        others = response['others']
                        sys_stdPremBf = others['stdPremBf']   #保费
                        sys_sumAssured = others['sumAssured']   #保额
                    else:
                        message = response['message']
                        sys_stdPremBf = ''
                        sys_sumAssured = ''
                    assert is_success == True
                        #存入sub_risk_info1
                    sub_risk_info1[i]['stdPremBf'] = sys_stdPremBf  # 保费
                    sub_risk_info1[i]['sumAssured'] = sys_sumAssured  # 保额
            ##判断存在附加险也存在豁免险时
            elif is_HolderExemption == '1' and len(sub_risk_info1) > 1:
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_险种信息录入接口-非豁免的附加险'  # 节点
                for i in range(1, len(sub_risk_info1)):
                    productId1 = sub_risk_info1[i]['productId']  # 险种id
                    chargeMode1 = sub_risk_info1[i]['chargeMode']  # 缴费频率
                    coverPeriodType1 = sub_risk_info1[i]['coverPeriodType']  # 保障期间类型
                    coveragePeriod1 = dsub_risk_info1[i]['coveragePeriod']  # 保障期间
                    chargePeriodType1 = sub_risk_info1[i]['chargePeriodType']  # 缴费期间类型
                    chargePeriod1 = sub_risk_info1[i]['chargePeriod']  # 缴费期间
                    sumAssured1 = sub_risk_info1[i]['sumAssured']  # 保额
                    stdPremBf1 = sub_risk_info1[i]['stdPremBf']  # 保费
                    units1 = sub_risk_info1[i]['units']  # 份数
                    planFreq1 = sub_risk_info1[i]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                    payOption1 = sub_risk_info1[i]['payOption']  # 年金使用方式
                    bonusYearOrAge1 = sub_risk_info1[i]['bonusYearOrAge']  # 领取年期或年龄
                    renew1 = sub_risk_info1[i]['renew']  # 是否续保
                    # 调用原子化脚本
                    response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId1, mainItemId1, insuId, chargeMode1, coverPeriodType1, coveragePeriod1, chargePeriodType1, chargePeriod1, units1, sumAssured1, stdPremBf1, planFreq1, payOption1, bonusYearOrAge1, headPartner, organId, renew1) 
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        message = ''
                        others = response['others']
                        sys_stdPremBf = others['stdPremBf']   #保费
                        sys_sumAssured = others['sumAssured']   #保额
                    else:
                        message = response['message']
                        sys_stdPremBf = ''
                        sys_sumAssured = ''
                    assert is_success == True
                        #存入sub_risk_info1
                    sub_risk_info1[i]['stdPremBf'] = sys_stdPremBf  # 保费
                    sub_risk_info1[i]['sumAssured'] = sys_sumAssured  # 保额
            ##存在双主险时-录入第二个主险 
            if is_2risk == '1':
                is_success = ''
                message = ''
                unit = '节点：系统-新契约录入_险种信息录入接口-录入第二个主险'  # 节点
                productId2 = dict['policy_info']['main_risk_info'][-1]['productId']  # 险种id
                chargeMode2 = dict['policy_info']['main_risk_info'][-1]['chargeMode']  # 缴费频率
                coverPeriodType2 = dict['policy_info']['main_risk_info'][-1]['coverPeriodType']  # 保障期间类型
                coveragePeriod2 = dict['policy_info']['main_risk_info'][-1]['coveragePeriod']  # 保障期间
                chargePeriodType2 = dict['policy_info']['main_risk_info'][-1]['chargePeriodType']  # 缴费期间类型
                chargePeriod2 = dict['policy_info']['main_risk_info'][-1]['chargePeriod']  # 缴费期间
                sumAssured2 = dict['policy_info']['main_risk_info'][-1]['sumAssured']  # 保额
                stdPremBf2 = dict['policy_info']['main_risk_info'][-1]['stdPremBf']  # 保费
                units2 = dict['policy_info']['main_risk_info'][-1]['units']  # 份数
                planFreq2 = dict['policy_info']['main_risk_info'][-1]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                payOption2 = dict['policy_info']['main_risk_info'][-1]['payOption']  # 年金使用方式
                bonusYearOrAge2 = dict['policy_info']['main_risk_info'][-1]['bonusYearOrAge']  # 领取年期或年龄
                renew2 = dict['policy_info']['main_risk_info'][-1]['renew']  # 是否续保
                mainItemId2 = ''  # 主险标识
                # 调用原子化脚本      
                response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId2, mainItemId2, insuId, chargeMode2, coverPeriodType2, coveragePeriod2, chargePeriodType2, chargePeriod2, units2, sumAssured2, stdPremBf2, planFreq2, payOption2, bonusYearOrAge2, headPartner, organId, renew2) 
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                    others = response['others']
                    mainItemId2 = others['itemId']
                    sys_stdPremBf = others['stdPremBf']   #保费
                    sys_sumAssured = others['sumAssured']   #保额
                else:
                    message = response['message']
                    mainItemId2 = ''
                    sys_stdPremBf = ''
                    sys_sumAssured = ''
                assert is_success == True
                    #存入dict
                dict['policy_info']['main_risk_info'][-1]['stdPremBf'] = sys_stdPremBf  # 保费
                dict['policy_info']['main_risk_info'][-1]['sumAssured'] = sys_sumAssured  # 保额
                assert is_success == True

                #主险2附加险
                for i in range(0, len(sub_risk_info2)):
                    is_success = ''
                    message = ''
                    unit = '节点：系统-新契约录入_险种信息录入接口-双主险附加险'  # 节点
                    #dict获取数据+定义数据
                    productId3 = sub_risk_info2[i]['productId']  # 险种id
                    chargeMode3 = sub_risk_info2[i]['chargeMode']  # 缴费频率
                    coverPeriodType3 = sub_risk_info2[i]['coverPeriodType']  # 保障期间类型
                    coveragePeriod3 = sub_risk_info2[i]['coveragePeriod']  # 保障期间
                    chargePeriodType3 = sub_risk_info2[i]['chargePeriodType']  # 缴费期间类型
                    chargePeriod3 = sub_risk_info2[i]['chargePeriod']  # 缴费期间
                    sumAssured3 = sub_risk_info2[i]['sumAssured']  # 保额
                    stdPremBf3 = sub_risk_info2[i]['stdPremBf']  # 保费
                    units3 = sub_risk_info2[i]['units']  # 份数
                    planFreq3 = sub_risk_info2[i]['planFreq']  # 年金领取方式 或 年金/生存金领取方式
                    payOption3 = sub_risk_info2[i]['payOption']  # 年金使用方式
                    bonusYearOrAge3 = sub_risk_info2[i]['bonusYearOrAge']  # 领取年期或年龄
                    renew3 = sub_risk_info2[i]['renew']   #是否续保
                    # 调用原子化脚本
                    response = API0000_nb().saveProductInfo(magicX, interface_ip, proposalId, productId3, mainItemId2, insuId, chargeMode3, coverPeriodType3, coveragePeriod3, chargePeriodType3, chargePeriod3, units3, sumAssured3, stdPremBf3, planFreq3, payOption3, bonusYearOrAge3, headPartner, organId, renew3) 
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        message = ''
                        others = response['others']
                        sys_stdPremBf = others['stdPremBf']   #保费
                        sys_sumAssured = others['sumAssured']   #保额
                    else:
                        message = response['message']
                        sys_stdPremBf = ''
                        sys_sumAssured = ''
                    assert is_success == True
                        #存入sub_risk_info2
                    sub_risk_info2[i]['stdPremBf'] = sys_stdPremBf  # 保费
                    sub_risk_info2[i]['sumAssured'] = sys_sumAssured  # 保额
                assert is_success == True
            dict['policy_info']['sub_risk_info'] = sub_risk_info1 + sub_risk_info2

            """step11:新契约录入提交前规则校验接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_提交前规则校验接口'  # 节点
            #调用原子脚本
            response = API0000_nb().valiDateRule(magicX, interface_ip, proposalId)
            #获取所需响应值
            if response['success'] == True:
                is_success = True
                message = ''
            else:
                is_success = False
                message = response['message']
            #断言
            assert is_success == True  
            """step12:新契约录入提交接口"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_提交接口'  # 节点
            #调用原子脚本
            response = API0000_nb().reviewProposal(magicX, interface_ip, proposalId)
            #获取所需响应值
            if (response['success'] == True) and (response['message'] == ''):
                is_success = True
                others = ''
                message = ''
            else:
                is_success = False
                others = response['others']
                if others != None:
                    for i in range(0, len(others)):
                        if i == 0:
                            message = response['message'] + ':1.' + others[i]['issueContent']
                        else:
                            message = message + str(i+1) + '.' + others[i]['issueContent']
                else:
                    message = response['message']
            #断言
            assert is_success == True       
            
            """step13:登出系统-新契约录入用户"""
            #节点初始化
            is_success = ''
            message = ''
            unit = '节点：系统-新契约录入_登出系统'  # 节点
            #调用原子脚本
            rs = API0000_sys_otherapi().logout(magicX, interface_ip)
            #获取所需响应值
            magicX = ''
            is_success = rs[0]
            message = rs[1]
            assert is_success == True
            dict['public']['magicX'] = ''

            """存入dict"""
                #保单信息
            dict['policy_info']['proposalNo'] = proposalNo  # 投保单号
            dict['policy_info']['policyType'] = policyType  # 保单号类型
            dict['policy_info']['effectiveDate'] = effectiveDate  # 生效日期
            dict['policy_info']['organId'] = organId  # 所属机构
            dict['policy_info']['policyChannel'] = policyChannel  # 公司渠道
                #投保人信息
            dict['policy_info']['holder_info']['isMarried'] = Holder_isMarried  # 婚姻状况
            dict['policy_info']['holder_info']['height'] = Holder_height  # 身高/CM
            dict['policy_info']['holder_info']['weight'] = Holder_weight  # 体重 /KG
            dict['policy_info']['holder_info']['homeTel'] = Holder_homeTel  # 家庭电话
            dict['policy_info']['holder_info']['industry'] = Holder_industry  # 工作行业
            dict['policy_info']['holder_info']['workCompany'] = Holder_workCompany  # 工作单位/学校名称
            dict['policy_info']['holder_info']['email'] = Holder_email  # 电子邮箱
            dict['policy_info']['holder_info']['jobCode'] = Holder_jobCode  # 职业代码/名称
            dict['policy_info']['holder_info']['jobClass'] = Holder_jobClass  # 职业等级
            dict['policy_info']['holder_info']['driverLicenseType'] = Holder_driverLicenseType  # 驾照类型
            dict['policy_info']['holder_info']['nationality'] = Holder_nationality  # 国籍/地区
            dict['policy_info']['holder_info']['nationnality1'] = Holder_nationnality1  # 民族
            dict['policy_info']['holder_info']['educationId'] = Holder_educationId  # 学位
            dict['policy_info']['holder_info']['medicalInsType'] = Holder_medicalInsType  # 医保类型
            dict['policy_info']['holder_info']['incomeSource'] = Holder_incomeSource  # 收入来源
            dict['policy_info']['holder_info']['incomeSourceNote'] = Holder_incomeSourceNote  # 收入来源备注
            dict['policy_info']['holder_info']['annualIncome'] = Holder_annualIncome  # 年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['familyIncome'] = Holder_familyIncome  # 家庭年收入/万元（系统自动转换为30万元）
            dict['policy_info']['holder_info']['premBudget'] = Holder_premBudget  # 保费预算/元
            dict['policy_info']['holder_info']['residentType'] = Holder_residentType  # 居民类型
            dict['policy_info']['holder_info']['taxPayerType'] = Holder_taxPayerType  # 税收居民类型
            dict['policy_info']['holder_info']['postalCode'] = Holder_postalCode  # 邮编
            dict['policy_info']['holder_info']['provinceCode'] = Holder_provinceCode  # 省份
            dict['policy_info']['holder_info']['cityCode'] = Holder_cityCode  # 市区
            dict['policy_info']['holder_info']['districtCode'] = Holder_districtCode  # 地区
            dict['policy_info']['holder_info']['addrDetail'] = Holder_addrDetail   # 详细地址
                #被保人信息
                #①非投保人豁免，投被保人不为同一人时
            if dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '0':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                #②非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '0' and dict['policy_info']['holder_info']['sameASInsurd'] == '1':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
                #③非投保人豁免，投被保人为同一人时：
            elif dict['public']['is_HolderExemption'] == '1':
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId   #客户id
                dict['policy_info']['insured_info'][0]['insuHeight'] = insuHeight   # 身高/CM
                dict['policy_info']['insured_info'][0]['insuWeight'] = insuWeight   # 体重/KG
                dict['policy_info']['insured_info'][0]['insuMarriage'] = insuMarriage   # 婚姻状况
                dict['policy_info']['insured_info'][0]['insuWorkCode'] = insuWorkCode   # 职业代码/名称
                dict['policy_info']['insured_info'][0]['insujobClass'] = insujobClass   # 职业等级
                dict['policy_info']['insured_info'][0]['insuEmail'] = insuEmail   # 电子邮箱
                dict['policy_info']['insured_info'][0]['insuWorkType'] = insuWorkType   # 工作行业
                dict['policy_info']['insured_info'][0]['insuCompany'] = insuCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][0]['insuNationatiy'] = insuNationatiy   # 国籍/地区
                dict['policy_info']['insured_info'][0]['insunationnality1'] = insunationnality1   # 民族
                dict['policy_info']['insured_info'][0]['insumedicalInsType'] = insumedicalInsType   # 医保类型
                dict['policy_info']['insured_info'][0]['driverLicenseType'] = insudriverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][0]['homeTel'] = insuhomeTel   # 家庭电话
                dict['policy_info']['insured_info'][0]['insuincomeSource'] = insuincomeSource   # 收入来源
                dict['policy_info']['insured_info'][0]['insuincomeSourceNote'] = insuincomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][0]['annualIncome'] = insuannualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][0]['insuresidentType'] = insuresidentType   # 居民类型
                dict['policy_info']['insured_info'][0]['taxPayerType'] = insutaxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][0]['insueducationId'] = insueducationId   # 学位
                dict['policy_info']['insured_info'][0]['insuPostalCode'] = insuPostalCode   # 邮编
                dict['policy_info']['insured_info'][0]['insuProvinceCode'] = insuProvinceCode   # 省
                dict['policy_info']['insured_info'][0]['insuCityCode'] = insuCityCode   # 市
                dict['policy_info']['insured_info'][0]['insuDistrictCode'] = insuDistrictCode   # 区
                dict['policy_info']['insured_info'][0]['insuAddrDetail'] = insuAddrDetail   # 详细地址
                dict['policy_info']['insured_info'][0]['insucustomerId'] = insucustomerId1   #客户id
                dict['policy_info']['insured_info'][1]['insuHeight'] = Holder_height   # 身高/CM
                dict['policy_info']['insured_info'][1]['insuWeight'] = Holder_weight   # 体重/KG
                dict['policy_info']['insured_info'][1]['insuMarriage'] = Holder_isMarried   # 婚姻状况
                dict['policy_info']['insured_info'][1]['insuWorkCode'] = Holder_jobCode   # 职业代码/名称
                dict['policy_info']['insured_info'][1]['insujobClass'] = Holder_jobClass   # 职业等级
                dict['policy_info']['insured_info'][1]['insuEmail'] = Holder_email   # 电子邮箱
                dict['policy_info']['insured_info'][1]['insuWorkType'] = Holder_industry   # 工作行业
                dict['policy_info']['insured_info'][1]['insuCompany'] = Holder_workCompany   # 工作单位/学校名称
                dict['policy_info']['insured_info'][1]['insuNationatiy'] = Holder_nationality   # 国籍/地区
                dict['policy_info']['insured_info'][1]['insunationnality1'] = Holder_nationnality1   # 民族
                dict['policy_info']['insured_info'][1]['insumedicalInsType'] = Holder_medicalInsType   # 医保类型
                dict['policy_info']['insured_info'][1]['insudriverLicenseType'] = Holder_driverLicenseType   # 驾照类型
                dict['policy_info']['insured_info'][1]['insuhomeTel'] = Holder_homeTel   # 家庭电话
                dict['policy_info']['insured_info'][1]['insuincomeSource'] = Holder_incomeSource   # 收入来源
                dict['policy_info']['insured_info'][1]['insuincomeSourceNote'] = Holder_incomeSourceNote   # 收入来源备注
                dict['policy_info']['insured_info'][1]['annualIncome'] = Holder_annualIncome   # 年收入/万元
                dict['policy_info']['insured_info'][1]['insuresidentType'] = Holder_residentType   # 居民类型
                dict['policy_info']['insured_info'][1]['insutaxPayerType'] = Holder_taxPayerType   # 税收居民类型
                dict['policy_info']['insured_info'][1]['insueducationId'] = Holder_educationId   # 学位
                dict['policy_info']['insured_info'][1]['insuPostalCode'] = Holder_postalCode   # 邮编
                dict['policy_info']['insured_info'][1]['insuProvinceCode'] = Holder_provinceCode   # 省
                dict['policy_info']['insured_info'][1]['insuCityCode'] = Holder_cityCode   # 市
                dict['policy_info']['insured_info'][1]['insuDistrictCode'] = Holder_districtCode   # 区
                dict['policy_info']['insured_info'][1]['insuAddrDetail'] = Holder_addrDetail   # 详细地址
            dict['logInfo']['code'] = '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约录入_结束(实际耗时:%s)%s' %(API0000_diy().logger_count_time(begin_time, end_time), policyNo)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0'
            dict['logInfo']['err'] = unit + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return proposalId,dict


    """新契约复核"""
    @func_set_timeout(60)#设定函数超执行时间
    def nb_review(self,proposalId, dict):
        """
           step1:复核用户登录
           step2:新契约复核updatePropStatus接口
           step3:新契约复核提交接口
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约复核_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            magicX = ''   #初始化会话id
            """step1:登录系统-新契约复核用户"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '节点：系统-新契约复核_登录系统'  # 节点
            message = ''   #信息
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   #环境名称
            interface_ip = dict['public']['interface_ip']   #接口ip
            username = dict['public']['username_review']    #新契约复核用户
            #调用原子脚本
            rs = API0000_sys_otherapi().login(env_name,username, interface_ip)
            #获取所需响应值
            is_success = rs[0]   #是否成功标识
            message = rs[1]   #信息
            magicX = rs[2]   #会话id
            #断言
            assert is_success == True
            dict['public']['magicX'] = magicX

            """step2:登录系统-新契约复核updatePropStatus接口"""
            #节点初始化
            is_success = ''   #是否成功标识
            unit = '节点：系统-新契约复核_新契约复核updatePropStatus接口'  # 节点
            message = ''   #信息
            #调用原子脚本
            response = API0000_nb().updatePropStatus(magicX, interface_ip, proposalId)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                message = ''
            else:
                message = response['message']
            #断言
            assert is_success == True
            
            """step3:登录系统-新契约复核提交接口"""
            #节点初始化
            is_success = ''
            unit = '节点：系统-新契约复核_新契约复核提交接口'  # 节点
            message = ''
            #dict获取数据+定义数据
            personUW = 'false'   #是否强制人工核保
            #调用原子脚本
            response = API0000_nb().confirmProposal(magicX, interface_ip, proposalId, personUW)
            #获取所需响应值
            is_success = response['success']
            if is_success:
                others = response['others']   # others为1时自动核保通过；为2时人工核保
                message = ''
            else:
                others = ''
                message = response['message']
            #断言
            assert is_success == True 
            dict['logInfo']['code'] = '1' 
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约复核_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] = '0' 
            dict['logInfo']['err'] = unit + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return magicX, others, dict


    """新契约核保"""
    @func_set_timeout(30)#设定函数超执行时间
    def nb_uw(self, magicX, others, dict):
        """
           step1:判断自动核保或人工核保
           step2:新契约人工核保_核保共享池查询接口
           step3:新契约人工核保_锁定任务接口
           step4:新契约人工核保_获取保单信息接口
           step5:新契约人工核保_险种层提交核保决定接口
           step6:新契约人工核保_契约核保结论提交接口
        """
        
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约核保_开始(预计耗时:10s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            """step1:判断自动核保或人工核保"""
            if others == '1':  #新契约自核通过
                pass
            elif others == '2':   #进入新契约人工核保
                """step2:新契约人工核保_核保共享池查询接口"""
                #节点初始化
                is_success = ''
                unit = '节点：系统-新契约人工核保_核保共享池查询接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                interface_ip = dict['public']['interface_ip']   #接口ip
                policyNo = dict['policy_info']['policyNo']  # 保单号
                #调用原子脚本
                response = API0000_uw().uwPolicyQuery_nb(magicX, interface_ip, policyNo)
                #获取所需响应值
                is_success = response['isSuccess']
                if is_success:
                    resultData = response['resultData']
                    uwId = resultData[0]['uwId']
                    mainId = resultData[0]['mainId']
                    message = ''
                else:
                    uwId = ''
                    mainId = ''
                    message = ''
                    message = response['message']
                #断言
                assert is_success == True
                
                """step3:新契约人工核保_锁定任务接口"""
                #节点初始化
                is_success = ''
                unit = '节点：系统-新契约人工核保_锁定任务接口'  # 节点
                message = ''
                #调用原子脚本
                response = API0000_uw().uwLock_nb(magicX, interface_ip, mainId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                else:
                    message = response['message']
                #断言
                assert is_success == True

                """step4:新契约人工核保_获取保单信息接口"""
                #节点初始化
                is_success = ''
                unit = '节点：系统-新契约人工核保_获取保单信息接口'  # 节点
                message = ''
                #调用原子脚本
                response = API0000_uw().uw_queryPersonalPolicy(magicX, interface_ip, uwId, mainId)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    others = response['others']
                    uwProducts = others['uwProducts']   # 核保险种信息
                    message = ''
                else:
                    uwProducts = ''
                    message = response['message']
                #断言
                assert is_success == True
                
                """step5:新契约人工核保_险种层提交核保决定接口"""
                #节点初始化
                is_success = ''
                unit = '节点：系统-新契约人工核保_险种层提交核保决定接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                uwDecision = '01'  # 核保决定 01-核保通过
                decisionComment = '核保通过'  # 核保决定描述
                uwExceptTxt = ''  # 除外责任

                num = 0
                lenght = len(uwProducts)   # 获取险种个数
                for a in range(0, lenght):
                    uwItemId = uwProducts[a]['uwItemId']
                    #调用原子脚本
                    response = API0000_uw().updateUwDecision_nb(magicX, interface_ip, uwItemId, uwId, uwDecision, decisionComment, uwExceptTxt, policyNo)
                    #获取所需响应值
                    is_success = response['success']
                    if is_success:
                        num = num + 1
                if num == lenght:
                    is_success = True
                    message = ''
                else:
                    is_success = False
                    message = '险种层提交核保决定未成功'
                #断言
                assert is_success == True

                """step6:新契约人工核保_契约核保结论提交接口"""
                #节点初始化
                is_success = ''
                unit = '节点：系统-新契约人工核保_契约核保结论提交接口'  # 节点
                message = ''
                #dict获取数据+定义数据
                uwPolicyDecision = '01'  # 保单核保结论 01-标准通过
                #调用原子脚本
                response = API0000_uw().commitUwComplete(magicX, interface_ip, uwId, uwPolicyDecision)
                #获取所需响应值
                status = response['status']
                uwStatus = response['result']['uwStatus']
                if status == '0':
                    is_success = True
                    message = ''
                else:
                    is_success = False
                    message = response['message']
                """
                if status == '0':
                    if uwStatus != '04':
                        is_success = True
                        message = ''
                    else:
                        is_success = False
                        message = '核保通过。新契约进入双录，请手动进行双录处理！'
                else:
                    is_success = False
                    message = response['message']
                """
                #断言
                assert is_success == True
                dict['logInfo']['code'] == '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约核保_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict


    """新契约收费"""
    @func_set_timeout(180)#设定函数超执行时间
    def nb_payment(self, magicX, dict):
        """
           step1:柜面收付费查询接口
           step2:柜面收付费保存接口
           step3:柜面收付费审核接口
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约收费_开始(预计耗时:30s)'))
            """step1:柜面收付费查询接口"""
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            is_success = ''   #初始化是否成功标识
            unit = '节点：系统-新契约收费_柜面收付费查询接口'  # 节点
            message = ''   #初始化信息
            payment = 0   #初始化收付费金额 
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']
            interface_ip = dict['public']['interface_ip']   #接口ip
            policyNo = dict['policy_info']['policyNo']  # 保单号
            proposalNo = dict['policy_info']['proposalNo']  #投保单号

            for i in range(0, 60):  #收付费查询(每1秒查询一次，最大40次，直到查到结果便结束查询)
                #调用原子脚本
                rs = API0000_payment().FeeQuery(magicX, interface_ip, policyNo)
                is_success = rs[0]
                message = rs[1]
                resultData = rs[2]
                len_resultData = rs[3]
                if len_resultData > 0:
                    payment = float(resultData[0]['amount'])
                    break
                else:
                    payment = 0
                    is_success = False
                    message = '此保单未查询到待新契约收费记录！'
                    time.sleep(1)
                    #判断是否进入双录
                    if i in (5,10,20,30,40,50):
                        #连接数据库
                        cursor = API0000_diy().db_conf(env_name,'wift_prop')   #连接数据库
                        sql = "select proposal_status from t_proposal where PROPOSAL_NO='" + proposalNo + "'"
                        cursor.execute(sql)
                        result = cursor.fetchall()
                        proposal_status = str(result[0][0])   #新契约状态
                        if proposal_status == '46':
                            is_success = False
                            message = '经数据库查询，新契约进入双录，请手动进行双录处理！'
                            break  
            #存入dict
            dict['policy_info']['payment_nb']['payment'] = payment  # 新契约收费金额                    
            assert is_success == True
            """step2:柜面收付费保存接口"""    
            #节点初始化
            is_success = ''
            unit = '节点：系统-新契约收费_柜面收付费保存接口'  # 节点
            message = ''
            #dict获取数据+定义数据
            bankEndorse = '123123'  # 银行凭证号码
            feeConfirmDate = dict['policy_info']['applicationDate']  # 费用确认日期
            internalAccount = '08-400301040007336'  # 公司收费账号
            internalBankCode = '05'  # 收费机构
            feeChannel = 'CASH-10017-AR'  # 支付通道
            cashOrgId = '10017'  # 机构名称
            internalBankName = '中国农业银行绥化市大有支行'  # 所属银行
            #调用原子脚本
            rs = API0000_payment().CashApproveInfoSave(magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, cashOrgId, internalBankName)
            #获取所需响应值
            is_success = rs[0]
            message = rs[1]
            assert is_success == True

            """step3:柜面收付费审核接口"""    
            #节点初始化
            is_success = ''
            unit = '节点：系统-新契约收费_柜面收付费审核接口'  # 节点
            message = ''
            #调用原子脚本
            rs = API0000_payment().finishFeeFlow(magicX, interface_ip, resultData, bankEndorse, feeConfirmDate, internalAccount, internalBankCode, feeChannel, internalBankName)
            #获取所需响应值
            is_success = rs[0]
            message = rs[1]
            assert is_success == True 

            """step4:验证收费后新契约状态是否流转"""
            #节点初始化
            is_success = ''
            unit = '节点：系统-新契约收费_验证收费后节点是否流转'  # 节点
            message = ''         
            #遍历
            for i in range(0, 60):
                #连接数据库
                cursor = API0000_diy().db_conf(env_name,'wift_prop')   #连接数据库
                sql1 = "select proposal_status from t_proposal where PROPOSAL_NO='" + proposalNo + "'"
                cursor.execute(sql1)
                result1 = cursor.fetchall()
                proposal_status1 = str(result1[0][0])
                if proposal_status1 == '61':
                    is_success = False
                    message = '经数据库查询，新契约收费成功后，保单状态仍为“待收费”状态，120s节点未流转，请检查是否环境问题或者产品犹豫期是否未配置！'
                    time.sleep(2)
                    continue
                if proposal_status1 == '81':
                    is_success = True
                    message = ''
                    break
            assert is_success == True

            dict['logInfo']['code'] == '1' 
            end_time = datetime.datetime.now()   #结束时间 
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约收费_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))   
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict
            

    """新契约回执"""
    @func_set_timeout(90)#设定函数超执行时间
    def nb_receipt(self, magicX, dict):
        """
           step1:保单确认回执查询接口
           step2:保单确认回执接口
           step3:登出系统-新契约复核用户
        """
        try:
            begin_time = datetime.datetime.now()   #开始时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约回执_开始(预计耗时:30s)'))
            #初始化
            dict['logInfo']['code'] = '0'
            dict['logInfo']['msg'] = ''
            dict['logInfo']['err'] = ''
            dict['logInfo']['result'] = ''
            """step1:保单确认回执查询接口"""
            #节点初始化
            is_success = ''    #初始化是否成功标识
            unit = '节点：系统-新契约回执_保单确认回执查询接口'  # 节点
            message = ''    #初始化信息
            #dict获取数据+定义数据
            interface_ip = dict['public']['interface_ip']  # 接口ip
            policyNo = dict['policy_info']['policyNo']  # 保单号
            for i in range(0, 200):
                #调用原子脚本
                response = API0000_nb().query_policyReceiptConfirm(magicX, interface_ip, policyNo)
                #获取所需响应值
                is_success = response['isSuccess']
                if is_success:
                    message = ''
                    resultData = response['resultData']
                else:
                    message = response['message']
                    resultData = ''
                #重复发送接口请求
                if resultData != '' and resultData != None:
                    break
                else:
                    time.sleep(0.5)
            assert is_success == True 
            
            """step2:保单确认回执接口"""
            #节点初始化
            #time.sleep(2)
            is_success = ''
            unit = '节点：系统-新契约回执_保单确认回执接口'  # 节点
            message = ''
            #dict获取数据+定义数据
            proposalNo = dict['policy_info']['proposalNo']  # 投保单号
            callBackDate = dict['policy_info']['effectiveDate']  # 回执日期
            callBackWay = '01'  # 回执方式
            appointVisitStartTime = ''  # 预约回访起始日期
            appointVisitEndTime = ''  # 预约回访终止日期
            for j in range(0, 3):
                #调用原子脚本
                response = API0000_nb().policyReceiptConfirm(magicX, interface_ip, policyNo, proposalNo, callBackDate, callBackWay, appointVisitStartTime, appointVisitEndTime)
                #获取所需响应值
                is_success = response['success']
                if is_success:
                    message = ''
                    break
                else:
                    message = response['message']
                    if message == '' or message == None:
                        message = '保单确认回执接口未回执成功，建议通过数据库sql插入保单回执日期！'
                    time.sleep(1)
            #断言
            assert is_success == True    


            """step2:登出系统-新契约复核用户"""
            #节点初始化
            is_success = ''
            unit = '节点：系统-新契约复核_登出系统'  # 节点
            message = ''
            #调用原子脚本
            rs = API0000_sys_otherapi().logout(magicX, interface_ip)
            #获取所需响应值
            magicX = ''
            is_success = rs[0]
            message = rs[1]
            assert is_success == True
            dict['public']['magicX'] = ''
            
            #存入dict:
            dict['policy_info']['policy_status'] = '01'   #保单状态。01有效；02失效
            dict['policy_info']['callBackDate'] = callBackDate   #回执日期
            dict['logInfo']['code'] = '1'
            end_time = datetime.datetime.now()   #结束时间
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*新契约回执_结束(实际耗时:%s)' % API0000_diy().logger_count_time(begin_time, end_time)))
        except Exception, e:
            is_success = False
            dict['logInfo']['code'] == '0'
            dict['logInfo']['err'] = unit + str(e)
            #如果程序报错，且magicX不为空（已经登录），对用户进行解锁
            if magicX != '':
                username = dict['public']['username_review']    #新契约复核用户
                API0000_sys_otherapi().deblocking(magicX, interface_ip, username)
            #记录日志
            logging.warning(API0000_diy().text_conversion('异常位置-' + unit + ':\n')  + traceback.format_exc())
        finally:
            # 加入阻断及错误信息。存入dict['result']中
            dict = API0000_diy().result(dict, is_success, message, unit)
        return dict
            


    """定义轨迹信息+检查点信息"""
    @func_set_timeout(30)#设定函数超执行时间
    def save_check_info(self, dict):
        """
           step1:定义轨迹信息
           step2:定义检查点信息
        """
        try:
            #初始化
            unit = '定义轨迹信息+检查点信息'    #节点标识
            #获取dict数据
            """step1:定义检查点信息"""
            #dict获取数据+定义数据
            env_name = dict['public']['env_name']   # 环境名称
            #dict获取数据+定义数据
            policyNo = dict['policy_info']['policyNo']  # 保单号
            applicationDate = dict['policy_info']['applicationDate']
            #获取产品列表
            product_list = dict['policy_info']['main_risk_info'] + dict['policy_info']['sub_risk_info']   #保单产品列表
            product = []
            if dict['logInfo']['code'] == '1':
                #产品信息product_info追加到dict['check']['calc_nb']['product']中
                for i in range (0,len(product_list)):
                    productId = str(product_list[i]['productId'])   #产品id
                    productCode = API0000_diy().query_productCode(env_name, productId)   #获取产品code
                    premium = float(product_list[i]['stdPremBf'])   #获取产品保费
                    amount = float(product_list[i]['sumAssured'])   #获取产品保额
                    #产品列表中每个产品生成一个产品信息product_info
                    product_info = {'productCode':productCode,'premium':premium,'amount':amount}
                    #存入dict['check']['calc_nb']中
                    product.append(product_info)
                dict['check']['calc_nb']['product'] = product
                #新契约收费金额存入dict['check']['calc_nb']中
                dict['check']['calc_nb']['payment'] = float(dict['policy_info']['payment_nb']['payment'])   #新契约收费金额
                dict['check']['calc_nb']['apply_date'] = applicationDate   #投保日期
                dict['check']['calc_nb']['msg'] = ''   #投保日期
                """step2:定义轨迹信息"""
                #定义轨迹信息
                track_info = {'trackType':'nb', 'trackTime':applicationDate, 'trackData':{'product':product, 'payment':float(dict['policy_info']['payment_nb']['payment'])}, 'msg': ''}
                dict['track_info'].append(track_info)   #存入轨迹信息
            else:
                message = dict['logInfo']['msg']
                err = dict['logInfo']['err']
                if dict['logInfo']['msg'] == '' or dict['logInfo']['msg'] == None:
                    message = str(dict['logInfo']['err'])
                dict['check']['calc_nb']['product'] = ''
                #新契约收费金额存入dict['check']['calc_nb']中
                dict['check']['calc_nb']['payment'] = dict['policy_info']['payment_nb']['payment']   #新契约收费金额
                dict['check']['calc_nb']['apply_date'] = applicationDate   #投保日期
                dict['check']['calc_nb']['msg'] = message 
                #定义轨迹信息
                track_info = {'trackType':'nb', 'trackTime':applicationDate, 'trackData':{'product':'', 'payment':dict['policy_info']['payment_nb']['payment']}, 'msg': message}
                dict['track_info'].append(track_info)   #存入轨迹信息
        except Exception, e:
            #记录日志
            logging.warning(API0000_diy().text_conversion('&nbsp&nbsp&nbsp*定义轨迹信息+检查点信息部分脚本报错:\n')  + traceback.format_exc())
        finally:
            pass 
        return dict



















