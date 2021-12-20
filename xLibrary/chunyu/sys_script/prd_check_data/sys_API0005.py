# coding=utf-8

import os
import sys
import json
import pymysql
import datetime
import requests
import traceback
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.writer.excel import ExcelWriter
from win32com.client import Dispatch



class API004:

    def __init__(self):
        self.file_path1 = os.path.realpath(__file__)  #运行的python文件的绝对路径
        self.file_path2 = os.path.basename(__file__)  #运行的python文件的相对路径
        #获取运行的python文件的上一级绝对路径
        self.path = self.file_path1[0:len(self.file_path1)-len(self.file_path2)]
        self.dir = '/data/xWork/xResult/xunjian/'
       
    def public(self,execNo):
        #数据初始化
        dict = {'logInfo': {'err':'','msg':'','result':''}}
        dict['execNo'] = execNo,
        dict['begin_time'] = datetime.datetime.now()
        dict['finish_date'] = '2020-09-10 12:00:00'  #巡测起期
        dict['UPDATE_TIME'] = '2020-09-10 13:30:00'  #巡测止期
       
        #begin_time = datetime.datetime(2020, 9, 10, 19, 0, 1)#2020-08-09 12:01:01
        #hour = begin_time.strftime("%H")
        #if hour in ('12','13','14','15','16','17'):
            #dict['finish_date'] = (begin_time+datetime.timedelta(days=-1)).strftime("%Y-%m-%d ") + '18:00:00'  #巡测起期
            #dict['UPDATE_TIME'] = begin_time.strftime('%Y-%m-%d ') + '12:00:00'  #巡测止期
        #elif hour in ('18','19','20','21','22','23'):
            #dict['finish_date'] = begin_time.strftime('%Y-%m-%d ') + '12:00:00'  #巡测起期
            #dict['UPDATE_TIME'] = begin_time.strftime('%Y-%m-%d ') + '18:00:00'  #巡测止期
        #else :
            #dict['finish_date'] = (begin_time+datetime.timedelta(days=-1)).strftime("%Y-%m-%d ") + '12:00:00'  #巡测起期
            #dict['UPDATE_TIME'] = (begin_time+datetime.timedelta(days=-1)).strftime("%Y-%m-%d ") + '18:00:00'  #巡测止期  
        print dict['begin_time'] 
        print dict['finish_date']
        print dict['UPDATE_TIME']
  
        #支持数据巡测的全部险种（目前需要实时更新）
        dict['risk_list'] = ['1027','1212','1214','3205','3218','3236','3237','3238','3239','3240','3247','3254','3256','3257','3259','3260','3264','3265','3266','3267','3268','3271','3273','3274','8209','8220','8221','8222','8223','8224','8228','8231','8233','1212CA''3239B','3254CO','3257CA','3265CB','3267CO']
        #支持数据巡测的万能险种（目前需要实时更新）
        dict['wn_list'] = ['3247','3256','3257','3264','3265','3267','8233','3257CA','3265CB','3267CO']
        return dict

    #发送post请求
    def send_post(self, url, headers, data):
        # 定义响应response
        response = {
            "success": False,
            "message": "参数不完整"
        }
        print url
        print data
        if (not url) or (not data):
            response["message"] = "参数不完整"
        else:
            response = requests.post(url, data=data, headers=headers)
            print response.status_code

            if response.status_code == 200:
                response = response.json()
            else:
                response["message"] = "response error,status_code=" + str(response.status_code)
        return response
                

    # 获取巡检数据
    def query1(self,execNo,dict):
        #获取dict值
        finish_date = dict['finish_date']
        UPDATE_TIME = dict['UPDATE_TIME']
        #定义列表
        result = []
        # 定义url
        url = "https://aitask.9an-data.com/data/v1"
        # 定义headers
        headers = {"Content-Type": "application/json"}
        # 组装请求参数
        data = {
            "icode": "DS4data",
            "sourceId": "2020-09-08@15:52:10.518@eZnrvXpijY6z5whZxyptTr",
            "scriptParams": {
                "finish_date": finish_date,
                "UPDATE_TIME": UPDATE_TIME
            }
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = self.send_post(url, headers, data)
        a = '保单号'.decode("utf-8")
        if len(response['data']) > 2:
            print len(response['data']['cleanerResult'])
            #对查询结果长度遍历
            for i in range(0,len(response['data']['cleanerResult'])):
                result.append([])    #每个查询结果，追加一个空数组
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种ID'.decode("utf-8")])    #追加-险种ID
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种代码'.decode("utf-8")])    #追加-险种代码
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种名称'.decode("utf-8")])    #追加-险种名称
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保单号'.decode("utf-8")])    #追加-保单号 
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种保费'.decode("utf-8")])    #追加-险种保费
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种保额'.decode("utf-8")])    #追加-险种保额
                result[i].append(datetime.datetime.strptime(response['data']['cleanerResult'][i]['dataRaw']['被保人生日'.decode("utf-8")], '%Y-%m-%d'))    #追加-被保人生日
                result[i].append(datetime.datetime.strptime(response['data']['cleanerResult'][i]['dataRaw']['保单生效日'.decode("utf-8")], '%Y-%m-%d %H:%M:%S'))    #追加-保单生效日
                result[i].append(datetime.datetime.strptime(response['data']['cleanerResult'][i]['dataRaw']['退保申请日'.decode("utf-8")], '%Y-%m-%d %H:%M:%S'))    #追加-退保申请日
                result[i].append(datetime.datetime.strptime(response['data']['cleanerResult'][i]['dataRaw']['下期交费日'.decode("utf-8")], '%Y-%m-%d'))    #追加-下期交费日
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['贷款'.decode("utf-8")])    #追加-贷款
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['还款'.decode("utf-8")])    #追加-还款
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['被保人性别'.decode("utf-8")])    #追加-被保人性别
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['领取方式'.decode("utf-8")])    #追加-领取方式
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['领取年龄'.decode("utf-8")])    #追加-领取年龄
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['交费年期'.decode("utf-8")])    #追加-交费年期
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保障期间'.decode("utf-8")])    #追加-保障期间
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F72扣贷款后退保金'.decode("utf-8")])    #追加-F72扣贷款后退保金
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F70现价'.decode("utf-8")])    #追加-F70现价
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F118贷款费用'.decode("utf-8")])    #追加-F118贷款费用
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F20万能退保金'.decode("utf-8")])    #追加-F20万能退保金
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F10万能手续费'.decode("utf-8")])    #追加-F10万能手续费
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F672未到期coi'.decode("utf-8")])    #追加-F672未到期coi
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F941分红费用'.decode("utf-8")])    #追加-F941分红费用
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F892分红扣回费用'.decode("utf-8")])    #追加-F892分红扣回费用
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['F75退续期保费'.decode("utf-8")])    #追加-F75退续期保费
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['险种汇总退保金'.decode("utf-8")])    #追加-险种汇总退保金
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保单推bcp汇总费用'.decode("utf-8")])    #追加-保单推bcp汇总费用
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保单投保日'.decode("utf-8")])    #追加-保单投保日
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['满期日'.decode("utf-8")])    #追加-满期日
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['客户重要信息变更'.decode("utf-8")])    #追加-客户重要信息变更
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['减额缴清'.decode("utf-8")])    #追加-减额缴清
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['增加保额'.decode("utf-8")])    #追加-增加保额
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['减少保额'.decode("utf-8")])    #追加-减少保额
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保全回退'.decode("utf-8")])    #追加-保全回退          
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['保全申请号'.decode("utf-8")])    #追加-保全申请号
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['退保完成日'.decode("utf-8")])    #追加-退保完成日
                result[i].append(response['data']['cleanerResult'][i]['dataRaw']['申请来源'.decode("utf-8")])    #追加-申请来源
        #存入dict
        dict['result'] = result           
        return dict

    #查询险种列表
    def query2(self,execNo,dict):
        #获取dict值
        finish_date = dict['finish_date']
        UPDATE_TIME = dict['UPDATE_TIME']
        #定义列表
        result_query_diffrisk = []
        # 定义url
        url = "https://aitask.9an-data.com/data/v1"
        # 定义headers
        headers = {"Content-Type": "application/json"}
        # 组装请求参数
        data = {
            "icode": "DS4data",
            "sourceId": "2020-09-08@15:31:41.450@eaHQpudmUz1vjUvWV1WTr2",
            "scriptParams": {
                "finish_date": finish_date,
                "UPDATE_TIME": UPDATE_TIME
            }
        }
        data = json.dumps(data)
        # 调用发送post请求函数
        response = self.send_post(url, headers, data)
        if len(response['data']) > 2:
            for i in range(0,len(response['data']['cleanerResult'])):
                result_query_diffrisk.append([])
                result_query_diffrisk[i].append(response['data']['cleanerResult'][i]['dataRaw']['product_code'])

        #存入dict
        dict['result_query_diffrisk'] = result_query_diffrisk           
        return dict


    #整理数据（获取数据库查询的数据后，根据险种进行分类，每一个险种的数据生成一个数组，得到多个数组后再统一放到大数组中）
    def manage_data(self,execNo,dict):
        #获取dict值
        result = dict['result']
        result_query_diffrisk = dict['result_query_diffrisk']
        #第一步：定义一个一维空数组list。
        list = []
        #第二步：distinct出数据组不同险种代码的数量。向list一维空数组中追加相同数量的空数组，组成一个空的二维数组。设计出大框
        for ab in range(0,len(result_query_diffrisk)):
            list.append([])
        #第三步：2层for循环，把相同险种的数据一一追加到二维数组中的第二层空数组中。完成组装list
        for i in range(0,len(result_query_diffrisk)):
            for j in range(0,len(result)):
                if result_query_diffrisk[i][0] == result[j][1]:
                    list[i].append(result[j])
        #存入dict
        dict['list'] = list  
        return dict

    
    #手工excel保存的操作,保存后方可获取excel中公式计算的值。
    def auto_save(self,filename):
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(filename)
        xlBook.Save()
        xlBook.Close()


    #非万能险数据-读取excel单元格数据（处理逻辑：同险种统一录入excel中，统一读取值）
    def fwn_read_excel(self,filename,len1,fwn_readdata_list):
        # 第一步：打开工作簿
        wb = load_workbook(filename,data_only=True)
        # 第二步：选取sheet页
        sh = wb['a']
        # 第三步：读取数据存入数组中
        for aab in range(2,len1+2):
            #每次在fwn_readdata_list最后位置新建一个空数组
            fwn_readdata_list.append([])
            #excel中取值
            ce_1 = sh.cell(row = aab,column = 3)  #险种代码
            ce_2 = sh.cell(row = aab,column = 4)  #险种名称
            ce_3 = sh.cell(row = aab,column = 5)  #保单号
            ce_4 = sh.cell(row = aab,column = 19)  #现价差额
            ce_5 = sh.cell(row = aab,column = 23)  #分红费用差额
            ce_9 = sh.cell(row = aab,column = 20)  #贷款费用差额  
            #获取数组fwn_readdata_list长度
            last = len(fwn_readdata_list)
            #在fwn_readdata_list最后位置新建的空数组中追加（险种代码、险种名称、保单号、现价差额、分红费用差额、贷款费用差额 ）
            fwn_readdata_list[last-1].append(ce_1.value)
            fwn_readdata_list[last-1].append(ce_2.value)
            fwn_readdata_list[last-1].append(ce_3.value)
            fwn_readdata_list[last-1].append(ce_4.value)
            fwn_readdata_list[last-1].append(ce_5.value)
            fwn_readdata_list[last-1].append(ce_9.value)
        wb.close()


    #万能险数据-读取excel单元格数据（处理逻辑：同险种，每条数据录入excel中，一个一个读取值）
    def wn_read_excel(self,filename,uu1,len1,wn_readdata_list):
        if len1/10>uu1:
            for i in range(0,10):
                #每次在wn_readdata_list最后位置新建一个空数组
                wn_readdata_list.append([])
                # 第一步：打开工作簿
                wb = load_workbook(filename,data_only=True)
                # 第二步：选取sheet页
                sh = wb['a']
                # 第三步：读取数据
                ce_1 = sh.cell(row = 3 + i,column = 3)  #险种代码
                ce_2 = sh.cell(row = 3 + i,column = 4)  #险种名称
                ce_3 = sh.cell(row = 3 + i,column = 5)  #保单号
                ce_6 = sh.cell(row = 16,column = 4 + i*7)  #万能退保金差额
                ce_7 = sh.cell(row = 17,column = 4 + i*7)  #万能手续费差额 
                ce_8 = sh.cell(row = 18,column = 4 + i*7)  #未到期coi差额
                ce_9 = sh.cell(row = 19,column = 4 + i*7)  #贷款费用差额 
                #获取数组wn_readdata_list长度        
                last = len(wn_readdata_list)  
                #在wn_readdata_list最后位置的空数组中追加（险种代码、险种名称、保单号、万能退保金差额、万能手续费差额、未到期coi差额、贷款费用差额 ）        
                wn_readdata_list[last-1].append(ce_1.value)  
                wn_readdata_list[last-1].append(ce_2.value)
                wn_readdata_list[last-1].append(ce_3.value) 
                wn_readdata_list[last-1].append(ce_6.value) 
                wn_readdata_list[last-1].append(ce_7.value) 
                wn_readdata_list[last-1].append(ce_8.value)
                wn_readdata_list[last-1].append(ce_9.value) 
                wb.close()

        if len1/10==uu1:
            for i in range(0,len1%10):
                #每次在wn_readdata_list最后位置新建一个空数组
                wn_readdata_list.append([])
                # 第一步：打开工作簿
                wb = load_workbook(filename,data_only=True)
                # 第二步：选取sheet页
                sh = wb['a']
                # 第三步：读取数据
                ce_1 = sh.cell(row = 3 + i,column = 3)  #险种代码
                ce_2 = sh.cell(row = 3 + i,column = 4)  #险种名称
                ce_3 = sh.cell(row = 3 + i,column = 5)  #保单号
                ce_6 = sh.cell(row = 16,column = 4 + i*7)  #万能退保金差额
                ce_7 = sh.cell(row = 17,column = 4 + i*7)  #万能手续费差额 
                ce_8 = sh.cell(row = 18,column = 4 + i*7)  #未到期coi差额
                ce_9 = sh.cell(row = 19,column = 4 + i*7)  #贷款费用差额 
                #获取数组wn_readdata_list长度        
                last = len(wn_readdata_list)  
                #在wn_readdata_list最后位置的空数组中追加（险种代码、险种名称、保单号、万能退保金差额、万能手续费差额、未到期coi差额、贷款费用差额 ）        
                wn_readdata_list[last-1].append(ce_1.value)  
                wn_readdata_list[last-1].append(ce_2.value)
                wn_readdata_list[last-1].append(ce_3.value) 
                wn_readdata_list[last-1].append(ce_6.value) 
                wn_readdata_list[last-1].append(ce_7.value) 
                wn_readdata_list[last-1].append(ce_8.value)
                wn_readdata_list[last-1].append(ce_9.value) 
                wb.close()
    

    #定义写入excel中的-非万能险数据，并写入excel中
    def fwndata_write_excel(self,sh,list1):      
        len_result = len(list1)
        
        for i in range(0,len_result):  
            #第一步：设计数据格式
                #险种ID            
            if list1[i][0] != None:
                resulti0 = str(list1[i][0])
            else:
                resulti0 = None
                #险种代码。（判断4位或者大于4位，4位时转换为整型。大于4位时不变）
            if len(list1[i][1]) == 4 and list1[i][1] != None:
                resulti1 = int(list1[i][1])
            elif len(list1[i][1]) > 4 and list1[i][1] != None:
                resulti1 = list1[i][1]
            else:
                resulti1 = None
                #险种名称    
            if list1[i][2] != None:
                resulti2 = list1[i][2]
            else:
                resulti2 = None
                #保单号    
            if list1[i][3] != None:
                resulti3 = str(list1[i][3])
            else:
                resulti3 = None
                #险种保费    
            if list1[i][4] != None:
                resulti4 = float(list1[i][4])
            else:
                resulti4 = None
                #险种保额    
            if list1[i][5] != None:
                resulti5 = float(list1[i][5])
            else:
                resulti5 = None
                #贷款数据。目前贷款计算不需考虑
            if list1[i][10] != None:
                loan = list1[i][10].encode('utf-8')
                loan_money = str(loan)[33:]
                loan_year = int(str(loan)[7:11])
                loan_month = int(str(loan)[12:14])
                loan_date1 = int(str(loan)[15:17])
                loan_date = datetime.date(loan_year, loan_month, loan_date1)
                loan_money = loan_money
            elif list1[i][10] == None:
                loan_money = None
                loan_date = None  
                #被保人性别       
            if list1[i][12] != None:
                resulti12 = int(list1[i][12])
            else:
                resulti12 = None
                #领取年龄        
            if list1[i][14] != None:
                resulti14 = int(list1[i][14])
            else:
                resulti14 = None         
                #交费年期        
            if list1[i][15] != None:
                resulti15 = int(list1[i][15])                
            else:
                resulti15 = None    
                #保障期间    
            if list1[i][16] != None:
                resulti16 = int(list1[i][16])                              
            else:
                resulti16 = None         
                #F72扣贷款后退保金   
            if list1[i][17] != None:
                resulti17 = float(list1[i][17])                
            else:
                resulti17 = None   
                #F70现价    
            if list1[i][18] != None:
                resulti18 = float(list1[i][18]) 
            else:
                resulti18 = None      
                #F118贷款费用    
            if list1[i][19] != None:
                resulti19 = float(list1[i][19]) 
            else:
                resulti19 = None    
                #F20万能退保金    
            if list1[i][20] != None:
                resulti20 = float(list1[i][20])     
            else:
                resulti20 = None      
                #F10万能手续费    
            if list1[i][21] != None:
                resulti21 = float(list1[i][21])       
            else:
                resulti21 = None   
                #F672未到期coi    
            if list1[i][22] != None:
                resulti22 = float(list1[i][22]) 
            else:
                resulti22 = None    
                #F941分红费用    
            if list1[i][23] != None:
                resulti23 = float(list1[i][23])  
            else:
                resulti23 = None  
                #F892分红扣回费用    
            if list1[i][24] != None:
                resulti24 = float(list1[i][24])    
            else:
                resulti24 = None     
                #F75退续期保费   
            if list1[i][25] != None:
                resulti25 = float(list1[i][25]) 
            else:
                resulti25 = None      
                #险种汇总退保金    
            if list1[i][26] != None:
                resulti26 = float(list1[i][26])                
            else:
                resulti26 = None  
            #第二步：向excel中写入数据   
            sh['B' + str(i+2)] = resulti0 #险种ID
            sh['C' + str(i+2)] = resulti1 #险种代码        
            sh['D' + str(i+2)] = resulti2 #险种名称
            sh['E' + str(i+2)] = resulti3 #保单号
            sh['X' + str(i+2)] = resulti4 #险种保费
            sh['Y' + str(i+2)] = resulti5 #险种保额
            sh['Z' + str(i+2)] =  list1[i][6] #被保人生日
            sh['AA' + str(i+2)] = list1[i][7] #保单生效日
            sh['AB' + str(i+2)] = list1[i][8] #退保申请日
            sh['AC' + str(i+2)] = list1[i][9] #下期交费日
            sh['AD' + str(i+2)] = loan_money #贷款金额
            sh['AE' + str(i+2)] = loan_date #贷款申请日
            sh['AF' + str(i+2)] = resulti12 #被保人性别
            sh['AG' + str(i+2)] = list1[i][13] #领取方式
            sh['AH' + str(i+2)] = resulti14 #领取年龄
            sh['AI' + str(i+2)] = resulti15 #交费年期
            sh['AJ' + str(i+2)] = resulti16 #保障期间
            sh['AK' + str(i+2)] = resulti17 #F72扣贷款后退保金·
            sh['AL' + str(i+2)] = resulti18 #F70现价
            sh['AM' + str(i+2)] = resulti19 #F118贷款费用
            sh['AN' + str(i+2)] = resulti20 #F20万能退保金
            sh['AO' + str(i+2)] = resulti21 #F10万能手续费
            sh['AP' + str(i+2)] = resulti22 #F672未到期coi
            sh['AQ' + str(i+2)] = resulti23 #F941分红费用
            sh['AR' + str(i+2)] = resulti24 #F892分红扣回费用
            sh['AS' + str(i+2)] = resulti25 #F75退续期保费
            sh['AT' + str(i+2)] = resulti26 #险种汇总退保金



    #定义写入excel中的-万能险数据，并写入excel中
    def wndata_write_excel(self,sh,list1,uu1,len1):      
        if len1/10 > uu1:
            for i in range(0,10):
            #第一步：设计数据格式
                #险种ID            
                if list1[i+uu1*10][0] != None:
                    resulti0 = str(list1[i+uu1*10][0])
                else:
                    resulti0 = None
                #险种代码。（判断4位或者大于4位，4位时转换为整型。大于4位时不变）
                if len(list1[i+uu1*10][1]) == 4 and list1[i+uu1*10][1] != None:
                    resulti1 = int(list1[i+uu1*10][1])
                elif len(list1[i+uu1*10][1]) > 4 and list1[i+uu1*10][1] != None:
                    resulti1 = list1[i+uu1*10][1]
                else:
                    resulti1 = None
                #险种名称    
                if list1[i+uu1*10][2] != None:
                    resulti2 = list1[i+uu1*10][2]
                else:
                    resulti2 = None
                #保单号    
                if list1[i+uu1*10][3] != None:
                    resulti3 = str(list1[i+uu1*10][3])
                else:
                    resulti3 = None
                #险种保费    
                if list1[i+uu1*10][4] != None:
                    resulti4 = float(list1[i+uu1*10][4])
                else:
                    resulti4 = None
                #险种保额    
                if list1[i+uu1*10][5] != None:
                    resulti5 = float(list1[i+uu1*10][5])
                else:
                    resulti5 = None
                #贷款数据。目前贷款计算不需考虑
                if list1[i+uu1*10][10] != None:
                    loan = list1[i+uu1*10][10].encode('utf-8')
                    loan_money = str(loan)[33:]
                    loan_year = int(str(loan)[7:11])
                    loan_month = int(str(loan)[12:14])
                    loan_date1 = int(str(loan)[15:17])
                    loan_date = datetime.date(loan_year, loan_month, loan_date1)
                    loan_money = loan_money
                elif list1[i+uu1*10][10] == None:
                    loan_money = None
                    loan_date = None  
                #被保人性别       
                if list1[i+uu1*10][12] != None:
                    resulti12 = int(list1[i+uu1*10][12])
                else:
                    resulti12 = None
                #领取年龄        
                if list1[i+uu1*10][14] != None:
                    resulti14 = int(list1[i+uu1*10][14])
                else:
                    resulti14 = None         
                #交费年期        
                if list1[i+uu1*10][15] != None:
                    resulti15 = int(list1[i+uu1*10][15])                
                else:
                    resulti15 = None    
                    #保障期间    
                if list1[i+uu1*10][16] != None:
                    resulti16 = int(list1[i+uu1*10][16])                              
                else:
                    resulti16 = None         
                #F72扣贷款后退保金   
                if list1[i+uu1*10][17] != None:
                    resulti17 = float(list1[i+uu1*10][17])                
                else:
                    resulti17 = None   
                #F70现价    
                if list1[i+uu1*10][18] != None:
                    resulti18 = float(list1[i+uu1*10][18]) 
                else:
                    resulti18 = None      
                #F118贷款费用    
                if list1[i+uu1*10][19] != None:
                    resulti19 = float(list1[i+uu1*10][19]) 
                else:
                    resulti19 = None    
                #F20万能退保金    
                if list1[i+uu1*10][20] != None:
                    resulti20 = float(list1[i+uu1*10][20])     
                else:
                    resulti20 = None      
                #F10万能手续费    
                if list1[i+uu1*10][21] != None:
                    resulti21 = float(list1[i+uu1*10][21])       
                else:
                    resulti21 = None   
                #F672未到期coi    
                if list1[i+uu1*10][22] != None:
                    resulti22 = float(list1[i+uu1*10][22]) 
                else:
                    resulti22 = None    
                #F941分红费用    
                if list1[i+uu1*10][23] != None:
                    resulti23 = float(list1[i+uu1*10][23])  
                else:
                    resulti23 = None  
                #F892分红扣回费用    
                if list1[i+uu1*10][24] != None:
                    resulti24 = float(list1[i+uu1*10][24])    
                else:
                    resulti24 = None     
                #F75退续期保费   
                if list1[i+uu1*10][25] != None:
                    resulti25 = float(list1[i+uu1*10][25]) 
                else:
                    resulti25 = None      
                #险种汇总退保金    
                if list1[i+uu1*10][26] != None:
                    resulti26 = float(list1[i+uu1*10][26])                
                else:
                    resulti26 = None  
            #第二步：向excel中写入数据   
                sh['B' + str(i+3)] = resulti0 #险种ID
                sh['C' + str(i+3)] = resulti1 #险种代码        
                sh['D' + str(i+2)] = resulti2 #险种名称
                sh['E' + str(i+3)] = resulti3 #保单号
                sh['F' + str(i+3)] = resulti4 #险种保费
                sh['G' + str(i+3)] = resulti5 #险种保额
                sh['H' + str(i+3)] =  list1[i+uu1*10][6] #被保人生日
                sh['I' + str(i+3)] = list1[i+uu1*10][7] #保单生效日
                sh['J' + str(i+3)] = list1[i+uu1*10][8] #退保申请日
                sh['K' + str(i+3)] = list1[i+uu1*10][9] #下期交费日
                sh['L' + str(i+3)] = loan_money #贷款金额
                sh['M' + str(i+3)] = loan_date #贷款申请日
                sh['N' + str(i+3)] = resulti12 #被保人性别
                sh['O' + str(i+3)] = list1[i+uu1*10][13] #领取方式
                sh['P' + str(i+3)] = resulti14 #领取年龄
                sh['Q' + str(i+3)] = resulti15 #交费年期
                sh['R' + str(i+3)] = resulti16 #保障期间
                sh['S' + str(i+3)] = resulti17 #F72扣贷款后退保金·
                sh['T' + str(i+3)] = resulti18 #F70现价
                sh['U' + str(i+3)] = resulti19 #F118贷款费用
                sh['V' + str(i+3)] = resulti20 #F20万能退保金
                sh['W' + str(i+3)] = resulti21 #F10万能手续费
                sh['X' + str(i+3)] = resulti22 #F672未到期coi
                sh['Y' + str(i+3)] = resulti23 #F941分红费用
                sh['Z' + str(i+3)] = resulti24 #F892分红扣回费用
                sh['AA' + str(i+3)] = resulti25 #F75退续期保费
                sh['AB' + str(i+3)] = resulti26 #险种汇总退保金

        if len1/10 == uu1:
            for i in range(0,len1%10):
            #第一步：设计数据格式
                #险种ID            
                if list1[i+uu1*10][0] != None:
                    resulti0 = str(list1[i+uu1*10][0])
                else:
                    resulti0 = None
                #险种代码。（判断4位或者大于4位，4位时转换为整型。大于4位时不变）
                if len(list1[i+uu1*10][1]) == 4 and list1[i+uu1*10][1] != None:
                    resulti1 = int(list1[i+uu1*10][1])
                elif len(list1[i+uu1*10][1]) > 4 and list1[i+uu1*10][1] != None:
                    resulti1 = list1[i+uu1*10][1]
                else:
                    resulti1 = None
                #险种名称    
                if list1[i+uu1*10][2] != None:
                    resulti2 = list1[i+uu1*10][2]
                else:
                    resulti2 = None
                #保单号    
                if list1[i+uu1*10][3] != None:
                    resulti3 = str(list1[i+uu1*10][3])
                else:
                    resulti3 = None
                #险种保费    
                if list1[i+uu1*10][4] != None:
                    resulti4 = float(list1[i+uu1*10][4])
                else:
                    resulti4 = None
                #险种保额    
                if list1[i+uu1*10][5] != None:
                    resulti5 = float(list1[i+uu1*10][5])
                else:
                    resulti5 = None
                #贷款数据。目前贷款计算不需考虑
                if list1[i+uu1*10][10] != None:
                    loan = list1[i+uu1*10][10].encode('utf-8')
                    loan_money = str(loan)[33:]
                    loan_year = int(str(loan)[7:11])
                    loan_month = int(str(loan)[12:14])
                    loan_date1 = int(str(loan)[15:17])
                    loan_date = datetime.date(loan_year, loan_month, loan_date1)
                    loan_money = float(loan_money)
                elif list1[i+uu1*10][10] == None:
                    loan_money = None
                    loan_date = None  
                #被保人性别       
                if list1[i+uu1*10][12] != None:
                    resulti12 = int(list1[i+uu1*10][12])
                else:
                    resulti12 = None
                #领取年龄        
                if list1[i+uu1*10][14] != None:
                    resulti14 = int(list1[i+uu1*10][14])
                else:
                    resulti14 = None         
                #交费年期        
                if list1[i+uu1*10][15] != None:
                    resulti15 = int(list1[i+uu1*10][15])                
                else:
                    resulti15 = None    
                    #保障期间    
                if list1[i+uu1*10][16] != None:
                    resulti16 = int(list1[i+uu1*10][16])                              
                else:
                    resulti16 = None         
                #F72扣贷款后退保金   
                if list1[i+uu1*10][17] != None:
                    resulti17 = float(list1[i+uu1*10][17])                
                else:
                    resulti17 = None   
                #F70现价    
                if list1[i+uu1*10][18] != None:
                    resulti18 = float(list1[i+uu1*10][18]) 
                else:
                    resulti18 = None      
                #F118贷款费用    
                if list1[i+uu1*10][19] != None:
                    resulti19 = float(list1[i+uu1*10][19]) 
                else:
                    resulti19 = None    
                #F20万能退保金    
                if list1[i+uu1*10][20] != None:
                    resulti20 = float(list1[i+uu1*10][20])     
                else:
                    resulti20 = None      
                #F10万能手续费    
                if list1[i+uu1*10][21] != None:
                    resulti21 = float(list1[i+uu1*10][21])       
                else:
                    resulti21 = None   
                #F672未到期coi    
                if list1[i+uu1*10][22] != None:
                    resulti22 = float(list1[i+uu1*10][22]) 
                else:
                    resulti22 = None    
                #F941分红费用    
                if list1[i+uu1*10][23] != None:
                    resulti23 = float(list1[i+uu1*10][23])  
                else:
                    resulti23 = None  
                #F892分红扣回费用    
                if list1[i+uu1*10][24] != None:
                    resulti24 = float(list1[i+uu1*10][24])    
                else:
                    resulti24 = None     
                #F75退续期保费   
                if list1[i+uu1*10][25] != None:
                    resulti25 = float(list1[i+uu1*10][25]) 
                else:
                    resulti25 = None      
                #险种汇总退保金    
                if list1[i+uu1*10][26] != None:
                    resulti26 = float(list1[i+uu1*10][26])                
                else:
                    resulti26 = None  
            #第二步：向excel中写入数据   
                sh['B' + str(i+3)] = resulti0 #险种ID
                sh['C' + str(i+3)] = resulti1 #险种代码        
                sh['D' + str(i+2)] = resulti2 #险种名称
                sh['E' + str(i+3)] = resulti3 #保单号
                sh['F' + str(i+3)] = resulti4 #险种保费
                sh['G' + str(i+3)] = resulti5 #险种保额
                sh['H' + str(i+3)] =  list1[i+uu1*10][6] #被保人生日
                sh['I' + str(i+3)] = list1[i+uu1*10][7] #保单生效日
                sh['J' + str(i+3)] = list1[i+uu1*10][8] #退保申请日
                sh['K' + str(i+3)] = list1[i+uu1*10][9] #下期交费日
                sh['L' + str(i+3)] = loan_money #贷款金额
                sh['M' + str(i+3)] = loan_date #贷款申请日
                sh['N' + str(i+3)] = resulti12 #被保人性别
                sh['O' + str(i+3)] = list1[i+uu1*10][13] #领取方式
                sh['P' + str(i+3)] = resulti14 #领取年龄
                sh['Q' + str(i+3)] = resulti15 #交费年期
                sh['R' + str(i+3)] = resulti16 #保障期间
                sh['S' + str(i+3)] = resulti17 #F72扣贷款后退保金·
                sh['T' + str(i+3)] = resulti18 #F70现价
                sh['U' + str(i+3)] = resulti19 #F118贷款费用
                sh['V' + str(i+3)] = resulti20 #F20万能退保金
                sh['W' + str(i+3)] = resulti21 #F10万能手续费
                sh['X' + str(i+3)] = resulti22 #F672未到期coi
                sh['Y' + str(i+3)] = resulti23 #F941分红费用
                sh['Z' + str(i+3)] = resulti24 #F892分红扣回费用
                sh['AA' + str(i+3)] = resulti25 #F75退续期保费
                sh['AB' + str(i+3)] = resulti26 #险种汇总退保金
        
           
    #对每个险种的数据写入excel中，并获取excel结果
    def diffrisk_write_excel(self,execNo,dict):
        #定义非万能、万能数据的数组，存储从excel中读取的结果
        dict['fwn_readdata_list'] = []
        dict['wn_readdata_list'] = []
        #获取dict值
        list = dict['list']
        wn_list = dict['wn_list']
        risk_list = dict['risk_list']
        fwn_readdata_list = dict['fwn_readdata_list']
        wn_readdata_list = dict['wn_readdata_list']     
        #
        ablerisk_lisk = []
        for i in range(0,len(risk_list)):
            for n in range(0,len(list)):
                if risk_list[i] == list[n][0][1]:
                    ablerisk_lisk.append(n)
                    print str(list[n][0][1]) + '-dir:' + str(n)
        #
        for k in ablerisk_lisk:
            #第一步：根据每个险种，定义组装出对应的excel文件名
            filename = self.path + '\\excel\\' + list[k][0][1] + '.xlsx'
            print filename
            len1 = len(list[k])
            print str(list[k][0][1]) + '-len:' + str(len1)
            #第二步1：判断是否非万能险，设计出写入excel及获取公式结果的逻辑（同险种数据，统一录入excel中，统一读取结果）
            if list[k][0][1] not in wn_list:
                wb = load_workbook(filename)
                sh = wb['a']
                API004().fwndata_write_excel(sh,list[k])
                wb.save(filename) 
                wb.close()
                API004().auto_save(filename)
                API004().fwn_read_excel(filename,len1,fwn_readdata_list)

            #第二步2：判断是否万能险，设计出写入excel及获取公式结果的逻辑（同险种数据，每条数据一一录入excel中，一一读取结果）
            elif list[k][0][1] in wn_list:
                #一次最多可录入10个，循环次数=len1/10 + 1(每一次再具体计算，例如22条数据，遍历3次，第一次遍历出10条，第二次遍历出10条，第三次遍历出2条)
                for uu1 in range(0,len1/10 + 1):
                    wb = load_workbook(filename)
                    sh = wb['a']                
                    API004().wndata_write_excel(sh,list[k],uu1,len1)
                    wb.save(filename) 
                    wb.close()
                    API004().auto_save(filename)
                    API004().wn_read_excel(filename,uu1,len1,wn_readdata_list) 
        #存入dict
        dict['fwn_readdata_list'] = fwn_readdata_list
        dict['wn_readdata_list'] = wn_readdata_list
        return dict


    #每次巡检新建目录
    def mkdir(self,execNo,dict):
        dict['new_dir'] = self.dir + execNo + '/'
        os.mkdir(dict['new_dir'])
        return dict


    #生成report报告文件
    def report_excel(self,execNo,dict):
        #获取dict值
        begin_time = dict['begin_time']
        finish_date = dict['finish_date']
        UPDATE_TIME = dict['UPDATE_TIME']
        fwn_readdata_list = dict['fwn_readdata_list']
        wn_readdata_list = dict['wn_readdata_list'] 
        new_dir = dict['new_dir']        
        #定义excel名
        hour = begin_time.strftime("%H")
        if hour in ('12','13','14','15','16','17'):
            report_name = new_dir + begin_time.strftime('%Y-%m-%d ') + '12am.xlsx'  
        elif hour in ('18','19','20','21','22','23'):
            report_name = new_dir + begin_time.strftime('%Y-%m-%d ') + '18pm.xlsx' 
        else :
            report_name = new_dir + (begin_time+datetime.timedelta(days=-1)).strftime("%Y-%m-%d ") + '18pm.xlsx'   
        
        #新建excel
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'a'
            
        #设定列宽
        ws1.column_dimensions['A'].width = 10
        ws1.column_dimensions['B'].width = 36
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 16
        ws1.column_dimensions['E'].width = 14
        ws1.column_dimensions['F'].width = 16
        ws1.column_dimensions['G'].width = 16
        ws1.column_dimensions['H'].width = 14
        ws1.column_dimensions['I'].width = 14
        #设计首行字段颜色
        yellow_fill = PatternFill(fill_type='solid', fgColor="FFFF00")      
        green_fill = PatternFill(fill_type='solid', fgColor="63C251")
        ws1.cell(row=1, column=1).fill = yellow_fill
        ws1.cell(row=1, column=3).fill = yellow_fill
        for aac in range(1,10):
            ws1.cell(row=2, column=aac).fill = green_fill
        
        #设计首行固定信息
        ws1.cell(row=1,column=1).value='巡测起期'
        ws1.cell(row=1,column=2).value=dict['finish_date']
        ws1.cell(row=1,column=3).value='巡测止期'
        ws1.cell(row=1,column=4).value=dict['UPDATE_TIME']
        #设计第二行固定表头
        ws1.cell(row=2,column=1).value='险种代码'
        ws1.cell(row=2,column=2).value='险种名称'
        ws1.cell(row=2,column=3).value='保单号'
        ws1.cell(row=2,column=4).value='现价差额'
        ws1.cell(row=2,column=5).value='分红费用差额'
        ws1.cell(row=2,column=6).value='万能退保金差额'
        ws1.cell(row=2,column=7).value='万能手续费差额'
        ws1.cell(row=2,column=8).value='未到期coi差额'
        ws1.cell(row=2,column=9).value='贷款费用差额'

        #动态数据写入excel
            #非万能数据
        for kk in range(3,len(fwn_readdata_list)+3):
                #险种代码
            ws1.cell(row=kk,column=1).value=fwn_readdata_list[kk-3][0]
                #险种名称            
            ws1.cell(row=kk,column=2).value=fwn_readdata_list[kk-3][1] 
                #保单号            
            ws1.cell(row=kk,column=3).value=fwn_readdata_list[kk-3][2]
                #现价差额            
            if type(fwn_readdata_list[kk-3][3]) == int:
                ws1.cell(row=kk,column=4).value=fwn_readdata_list[kk-3][3]  
            elif  type(fwn_readdata_list[kk-3][3]) == float:
                ws1.cell(row=kk,column=4).value=round(fwn_readdata_list[kk-3][3],2)
            else:
                ws1.cell(row=kk,column=4).value=fwn_readdata_list[kk-3][3]         
                #分红费用差额                
            if type(fwn_readdata_list[kk-3][4]) == int:
                ws1.cell(row=kk,column=5).value=fwn_readdata_list[kk-3][4]  
            elif  type(fwn_readdata_list[kk-3][4]) == float:
                ws1.cell(row=kk,column=5).value=round(fwn_readdata_list[kk-3][4],2)  
            else:
                ws1.cell(row=kk,column=5).value=fwn_readdata_list[kk-3][4]
                #贷款费用差额    
            if type(fwn_readdata_list[kk-3][5]) == int:
                ws1.cell(row=kk,column=9).value=fwn_readdata_list[kk-3][5]  
            elif  type(fwn_readdata_list[kk-3][5]) == float:
                ws1.cell(row=kk,column=9).value=round(fwn_readdata_list[kk-3][5],2)  
            else:
                ws1.cell(row=kk,column=9).value=fwn_readdata_list[kk-3][5]             
            #万能数据
        for kk1 in range(len(fwn_readdata_list)+3,len(fwn_readdata_list)+len(wn_readdata_list)+3):
                #险种代码
            ws1.cell(row=kk1,column=1).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][0]
                #险种名称 
            ws1.cell(row=kk1,column=2).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][1]         
                #保单号
            ws1.cell(row=kk1,column=3).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][2]
                #万能退保金差额
            if type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][3]) == int: 
                ws1.cell(row=kk1,column=6).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][3]
            elif type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][3]) == float:
                ws1.cell(row=kk1,column=6).value=round(wn_readdata_list[kk1-len(fwn_readdata_list)-3][3],2)
            else:
                ws1.cell(row=kk1,column=6).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][3]
                #万能手续费差额
            if type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][4]) == int: 
                ws1.cell(row=kk1,column=7).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][4]
            elif type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][4]) == float:
                ws1.cell(row=kk1,column=7).value=round(wn_readdata_list[kk1-len(fwn_readdata_list)-3][4],2)
            else:
                ws1.cell(row=kk1,column=7).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][4]            
                #未到期coi差额
            if type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][5]) == int: 
                ws1.cell(row=kk1,column=8).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][5]
            elif type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][5]) == float:
                ws1.cell(row=kk1,column=8).value=round(wn_readdata_list[kk1-len(fwn_readdata_list)-3][5],2)
            else:
                ws1.cell(row=kk1,column=8).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][5]            
                #贷款费用差额
            if type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][6]) == int: 
                ws1.cell(row=kk1,column=9).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][6]
            elif type(wn_readdata_list[kk1-len(fwn_readdata_list)-3][6]) == float:
                ws1.cell(row=kk1,column=9).value=round(wn_readdata_list[kk1-len(fwn_readdata_list)-3][6],2)
            else:
                ws1.cell(row=kk1,column=9).value=wn_readdata_list[kk1-len(fwn_readdata_list)-3][6]           
        #保存excel
        wb.save(report_name)       
        #存入dict
        dict['report_name'] = report_name
        return dict
        

    #生成存储巡检数据文件
    def save_testdata(self,execNo,dict):
        #获取dict值
        list = dict['list'] 
        begin_time = dict['begin_time']
        risk_list = dict['risk_list']
        new_dir = dict['new_dir']   
        ablerisk_lisk = []
        #获取此次数据中支持巡检的险种，在全部支持巡测险种列表中的位置，存入ablerisk_lisk列表中
        for i in range(0,len(risk_list)):
            for n in range(0,len(list)):
                if risk_list[i] == list[n][0][1]:
                    ablerisk_lisk.append(n)
        #遍历此次数据中支持巡检的险种中第2个到最后一个险种，list列表中，把第二个险种到最后一个险种均追加到第一个险种所在的子数组中。
        for i in range(1,len(ablerisk_lisk)):
            for i1 in range(0,len(list[ablerisk_lisk[i]])):
                list[ablerisk_lisk[0]].append(list[ablerisk_lisk[i]][i1])

        #定义testdata文件 
        testdata_name = new_dir + 'testdata.xlsx'
        print testdata_name

        #定义sheet名
        hour = begin_time.strftime("%H")
        if hour in ('12','13','14','15','16','17'):
            sheetname = begin_time.strftime('%Y-%m-%d ') + '12 am'  
        elif hour in ('18','19','20','21','22','23'):
            sheetname = begin_time.strftime('%Y-%m-%d ') + '18 pm'  
        else :
            sheetname = (begin_time+datetime.timedelta(days=-1)).strftime("%Y-%m-%d ") + '18 pm'  

        #新建excel
        wb = Workbook()
        ws1 = wb.active
        ws1.title = sheetname

        #设计首行固定表头
        ws1.cell(row=1,column=2).value='险种ID'        
        ws1.cell(row=1,column=3).value='险种代码'
        ws1.cell(row=1,column=4).value='险种名称'
        ws1.cell(row=1,column=5).value='保单号'
        ws1.cell(row=1,column=24).value='险种保费'        
        ws1.cell(row=1,column=25).value='险种保额'        
        ws1.cell(row=1,column=26).value='被保人生日'        
        ws1.cell(row=1,column=27).value='保单生效日'      
        ws1.cell(row=1,column=28).value='退保申请日'        
        ws1.cell(row=1,column=29).value='下期交费日'        
        ws1.cell(row=1,column=30).value='贷款金额'        
        ws1.cell(row=1,column=31).value='贷款申请日' 
        ws1.cell(row=1,column=32).value='被保人性别'        
        ws1.cell(row=1,column=33).value='领取方式'        
        ws1.cell(row=1,column=34).value='领取年龄'        
        ws1.cell(row=1,column=35).value='ppp'      
        ws1.cell(row=1,column=36).value='bp'     
        ws1.cell(row=1,column=37).value='F118贷款费用'         
        ws1.cell(row=1,column=38).value='F70现价'        
        ws1.cell(row=1,column=39).value='F118贷款费用'        
        ws1.cell(row=1,column=40).value='F20万能退保金' 
        ws1.cell(row=1,column=41).value='F10万能手续费'
        ws1.cell(row=1,column=42).value='F672未到期coi'
        ws1.cell(row=1,column=43).value='F941分红费用'
        ws1.cell(row=1,column=44).value='F892分红扣回费用'
        ws1.cell(row=1,column=45).value='F75退续期保费'
        ws1.cell(row=1,column=46).value='险种汇总退保金'
        #设定列宽
        ws1.column_dimensions['D'].width = 36
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['Z'].width = 14  
        ws1.column_dimensions['AA'].width = 17    
        ws1.column_dimensions['AB'].width = 17   
        ws1.column_dimensions['AC'].width = 14 
        ws1.column_dimensions['AE'].width = 14    
        listchar = ['F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W']    
        for i in listchar:
            ws1.column_dimensions[i].width = 0.1
        #写入excel
        if len(list)>0:
            API004().fwndata_write_excel(ws1,list[ablerisk_lisk[0]])
        wb.save(testdata_name)     
        #存入dict
        dict['testdata_name'] = testdata_name
        return dict
        



#dict = API004().public(execNo)
#dict = API004().query1(execNo,dict)
#dict = API004().query2(execNo,dict)
#dict = API004().manage_data(execNo,dict)
#dict = API004().diffrisk_write_excel(execNo,dict)
#dict = API004().mkdir(execNo,dict)
#dict = API004().report_excel(execNo,dict)  
#dict = API004().save_testdata(execNo,dict)




















